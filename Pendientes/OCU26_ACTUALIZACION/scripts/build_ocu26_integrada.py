"""Construye la base candidata OCU26 integrando NUEVA sobre ACTUAL.

Aplica ÚNICAMENTE los "cambios seguros" (regla 2 de FASE 4: actual vacío +
nuevo con dato válido -> completar) calculados por merge_common, que es el
mismo motor que usa audit_ocu26_actualizacion.py (garantiza que la
auditoría y la candidata sean consistentes entre sí).

Explícitamente NO hace (por diseño, ver docs/CM1.md y el pedido de
integración):
  - No sobrescribe ningún valor no vacío (conflictos quedan registrados,
    nunca aplicados).
  - No elimina ninguna fila de ACTUAL, exista o no en NUEVA.
  - No agrega filas nuevas de CAMPANAS que llegan sin CargaID: el proyecto
    no define una función de generación de CargaID (solo un patrón
    histórico documentado 'HIST-00000001'), así que generarlo aquí sería
    inventar lógica de negocio sin evidencia. Esos registros quedan
    documentados en la auditoría (REGISTROS_NUEVOS / A_VALIDAR) para
    decisión humana.
  - No agrega elementos nuevos a MAESTRO_ELEMENTOS (en esta corrida no hay
    ninguno que califique; si en el futuro aparecieran, deben incorporarse
    en una iteración explícita, no de forma silenciosa).
  - No modifica ni input/OCU26_BASE_DATOS.xlsx ni el archivo NUEVA: ambos
    se abren en modo lectura y se verifica su SHA-256 antes/después.

Construcción: parte de una copia en memoria del workbook ACTUAL (abierto
con openpyxl), preservando 100% de su estructura (hojas, tablas, rangos,
encabezados, orden, estilos) y solo escribe valores en las celdas
puntuales identificadas como COMPLETAR_VACIO_ACTUAL. La cantidad de filas
de cada hoja en la candidata es siempre igual a la de ACTUAL.

Uso:
    python build_ocu26_integrada.py
    python build_ocu26_integrada.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

import merge_common as mc


class BuildError(Exception):
    """Error bloqueante: alguna invariante de seguridad se violó durante el build."""


def _header_row_maps(ws, table_name: str, key_column: str) -> tuple[dict[str, int], dict[Any, int], int, int]:
    """Devuelve (col_index_por_header, row_index_por_clave, min_row, max_row)."""
    t = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    col_index = {h: min_col + i for i, h in enumerate(headers)}
    key_col = col_index[key_column]
    row_index: dict[Any, int] = {}
    for r in range(min_row + 1, max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if not mc.is_blank(v):
            row_index[v] = r
    return col_index, row_index, min_row, max_row


def _apply_completions(
    target_ws,
    source_ws,
    table_name: str,
    key_column: str,
    completions: list[mc.FieldCompletion],
    label: str,
) -> list[dict[str, Any]]:
    applied: list[dict[str, Any]] = []

    t_cols, t_rows, _, _ = _header_row_maps(target_ws, table_name, key_column)
    s_cols, s_rows, _, _ = _header_row_maps(source_ws, table_name, key_column)

    for c in completions:
        if c.key not in t_rows:
            raise BuildError(f"{label}: clave {c.key!r} no encontrada en la hoja destino (ACTUAL)")
        if c.key not in s_rows:
            raise BuildError(f"{label}: clave {c.key!r} no encontrada en la hoja fuente (NUEVA)")
        if c.column not in t_cols or c.column not in s_cols:
            raise BuildError(f"{label}: columna {c.column!r} no encontrada en ambas hojas")

        target_cell = target_ws.cell(row=t_rows[c.key], column=t_cols[c.column])
        source_cell = source_ws.cell(row=s_rows[c.key], column=s_cols[c.column])

        if not mc.is_blank(target_cell.value):
            raise BuildError(
                f"{label}: intento de sobrescribir celda no vacía {c.key!r}/{c.column!r} "
                f"(valor actual={target_cell.value!r}) -> esto violaría la regla de no-sobrescritura"
            )

        target_cell.value = source_cell.value
        if source_cell.number_format and source_cell.number_format != "General":
            target_cell.number_format = source_cell.number_format

        applied.append({"key": c.key, "column": c.column, "value": source_cell.value})

    return applied


def _assert_row_count_unchanged(ws, table_name: str, expected_rows: int, label: str) -> None:
    t = ws.tables[table_name]
    _, min_row, _, max_row = range_boundaries(t.ref)
    actual_rows = max_row - min_row
    if actual_rows != expected_rows:
        raise BuildError(f"{label}: la cantidad de filas cambió inesperadamente ({actual_rows} != {expected_rows})")


def build_candidate() -> dict[str, Any]:
    sha_actual_before = mc.calculate_sha256(mc.ACTUAL_PATH)
    sha_nueva_before = mc.calculate_sha256(mc.NUEVA_PATH)

    # --- Clasificación (misma lógica que la auditoría) ---
    actual_maestro_df = mc.read_table_df(mc.ACTUAL_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    nueva_maestro_df = mc.read_table_df(mc.NUEVA_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    actual_campanas_df = mc.read_table_df(mc.ACTUAL_PATH, "CAMPANAS", "tblCampanas")
    nueva_campanas_df = mc.read_table_df(mc.NUEVA_PATH, "CAMPANAS", "tblCampanas")
    actual_parametros_df = mc.read_table_df(mc.ACTUAL_PATH, "PARAMETROS", "tblParametros")

    maestro_cls = mc.classify_maestro(actual_maestro_df, nueva_maestro_df)
    campanas_cls = mc.classify_campanas(actual_campanas_df, nueva_campanas_df)

    expected_rows = {
        "MAESTRO_ELEMENTOS": len(actual_maestro_df),
        "CAMPANAS": len(actual_campanas_df),
        "PARAMETROS": len(actual_parametros_df),
    }

    # --- Workbook destino: copia en memoria de ACTUAL (preserva TODO) ---
    target_wb = load_workbook(mc.ACTUAL_PATH, data_only=False, read_only=False, keep_vba=False)
    # --- Workbook fuente de valores: NUEVA, solo lectura ---
    source_wb = load_workbook(mc.NUEVA_PATH, data_only=True, read_only=False, keep_vba=False)

    applied_maestro = _apply_completions(
        target_wb["MAESTRO_ELEMENTOS"],
        source_wb["MAESTRO_ELEMENTOS"],
        "tblElementos",
        mc.MAESTRO_KEY,
        maestro_cls.completions,
        "MAESTRO_ELEMENTOS",
    )
    applied_campanas = _apply_completions(
        target_wb["CAMPANAS"],
        source_wb["CAMPANAS"],
        "tblCampanas",
        mc.CAMPANAS_KEY,
        campanas_cls.completions,
        "CAMPANAS",
    )

    # PARAMETROS: sin cambios (idéntico); no se toca.

    # --- Invariantes estructurales ---
    for sheet_name, table_name in mc.EXPECTED_TABLES.items():
        _assert_row_count_unchanged(target_wb[sheet_name], table_name, expected_rows[sheet_name], sheet_name)

    if target_wb.sheetnames != mc.EXPECTED_SHEETS:
        raise BuildError(f"Las hojas de la candidata no coinciden con lo esperado: {target_wb.sheetnames}")

    # --- Ningún conflicto fue aplicado: verificar que esas celdas siguen == ACTUAL original ---
    conflict_check_wb = load_workbook(mc.ACTUAL_PATH, data_only=False, read_only=False, keep_vba=False)
    for cls, sheet_name, table_name, key_col in [
        (maestro_cls, "MAESTRO_ELEMENTOS", "tblElementos", mc.MAESTRO_KEY),
        (campanas_cls, "CAMPANAS", "tblCampanas", mc.CAMPANAS_KEY),
    ]:
        if not cls.conflicts:
            continue
        cols_map, rows_map, _, _ = _header_row_maps(target_wb[sheet_name], table_name, key_col)
        orig_cols_map, orig_rows_map, _, _ = _header_row_maps(conflict_check_wb[sheet_name], table_name, key_col)
        for conflict in cls.conflicts:
            r = rows_map[conflict.key]
            oc_r = orig_rows_map[conflict.key]
            col = cols_map[conflict.column]
            oc_col = orig_cols_map[conflict.column]
            candidate_value = target_wb[sheet_name].cell(row=r, column=col).value
            original_value = conflict_check_wb[sheet_name].cell(row=oc_r, column=oc_col).value
            if not (candidate_value == original_value or (mc.is_blank(candidate_value) and mc.is_blank(original_value))):
                raise BuildError(
                    f"{sheet_name}: el conflicto en {conflict.key!r}/{conflict.column!r} fue alterado "
                    f"indebidamente (candidata={candidate_value!r}, original={original_value!r})"
                )
    conflict_check_wb.close()

    # --- Guardar candidata (nunca sobre las rutas fuente) ---
    mc.CANDIDATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(mc.CANDIDATE_PATH)
    target_wb.close()
    source_wb.close()

    # --- Confirmar que ninguna fuente fue modificada ---
    sha_actual_after = mc.calculate_sha256(mc.ACTUAL_PATH)
    sha_nueva_after = mc.calculate_sha256(mc.NUEVA_PATH)
    if sha_actual_after != sha_actual_before:
        raise BuildError("ERROR CRÍTICO: input/OCU26_BASE_DATOS.xlsx fue modificado durante el build")
    if sha_nueva_after != sha_nueva_before:
        raise BuildError("ERROR CRÍTICO: OCU26_BASE_NUEVA_RECIBIDA.xlsx fue modificado durante el build")

    excluded_new_campanas = len(campanas_cls.only_nueva_unkeyed_rows) + len(campanas_cls.only_nueva_keyed_new)
    excluded_new_maestro = len(maestro_cls.only_nueva_unkeyed_rows) + len(maestro_cls.only_nueva_keyed_new)

    return {
        "result": "BUILD_OK",
        "candidate_path": str(mc.CANDIDATE_PATH),
        "sources_intact": True,
        "actual_sha256": sha_actual_after,
        "nueva_sha256": sha_nueva_after,
        "rows_preserved": expected_rows,
        "completions_applied": {
            "MAESTRO_ELEMENTOS": len(applied_maestro),
            "CAMPANAS": len(applied_campanas),
        },
        "conflicts_not_applied": {
            "MAESTRO_ELEMENTOS": len(maestro_cls.conflicts),
            "CAMPANAS": len(campanas_cls.conflicts),
        },
        "rows_only_actual_preserved": {
            "MAESTRO_ELEMENTOS": len(maestro_cls.only_actual_keys),
            "CAMPANAS": len(campanas_cls.only_actual_keys),
        },
        "rows_only_nueva_excluded": {
            "MAESTRO_ELEMENTOS": excluded_new_maestro,
            "CAMPANAS": excluded_new_campanas,
        },
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Construye la base candidata OCU26 integrada.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = build_candidate()
    except BuildError as exc:
        if args.json:
            print(json.dumps({"result": "BUILD_ERROR", "error": str(exc)}, ensure_ascii=False, indent=2))
        else:
            print("=" * 60)
            print("OCU26 BUILD INTEGRADA - ERROR")
            print("=" * 60)
            print(str(exc))
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 BUILD INTEGRADA")
        print("=" * 60)
        print(f"Candidata: {result['candidate_path']}")
        print(f"Fuentes intactas: {result['sources_intact']}")
        print("Filas preservadas:", result["rows_preserved"])
        print("Completions aplicadas:", result["completions_applied"])
        print("Conflictos NO aplicados:", result["conflicts_not_applied"])
        print("Filas solo-ACTUAL conservadas:", result["rows_only_actual_preserved"])
        print("Filas solo-NUEVA excluidas (sin identificador válido):", result["rows_only_nueva_excluded"])

    return 0


if __name__ == "__main__":
    sys.exit(main())
