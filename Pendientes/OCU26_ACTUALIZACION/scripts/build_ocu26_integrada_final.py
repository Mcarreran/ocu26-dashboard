"""Construye la base FINAL OCU26 a partir de la candidata V2 (no de ACTUAL).

Aplica EXCLUSIVAMENTE las ediciones autorizadas en merge_final_common.
AUTHORIZED_EDITS (Decisiones 1 y 2, confirmadas por el usuario). Las
Decisiones 3 y 4 se verifican pero NO se aplican: ambas fallan sus propias
condiciones de seguridad (ver merge_final_common.verify_decision_3/4) y
quedan documentadas como BLOQUEADAS.

No modifica ni sobrescribe:
  - input/OCU26_BASE_DATOS.xlsx
  - Pendientes/OCU26_ACTUALIZACION/input/OCU26_BASE_NUEVA_RECIBIDA.xlsx
  - Pendientes/OCU26_ACTUALIZACION/output/OCU26_BASE_DATOS_INTEGRADA_CANDIDATA_V2_2026-08-18.xlsx

Antes de escribir cada campo, verifica que el valor actual coincida
exactamente con el "antes" esperado; si no coincide, aborta el build
completo sin escribir nada (no hay escritura parcial).

Uso:
    python build_ocu26_integrada_final.py
    python build_ocu26_integrada_final.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

import merge_common as mc
import merge_final_common as fc


class BuildError(Exception):
    pass


def _header_row_maps(ws, table_name: str, key_column: str) -> tuple[dict[str, int], dict[Any, int], int, int]:
    t = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    col_index = {h: min_col + i for i, h in enumerate(headers)}
    key_col = col_index[key_column]
    row_index: dict[Any, int] = {}
    for r in range(min_row + 1, max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if not mc.is_blank(v):
            row_index[v] = r
    return col_index, row_index, min_row, max_row


def build_final() -> dict[str, Any]:
    sha_actual_before = mc.calculate_sha256(fc.mc.ACTUAL_PATH)
    sha_nueva_before = mc.calculate_sha256(fc.mc.NUEVA_PATH)
    sha_v2_before = mc.calculate_sha256(fc.CANDIDATE_V2_PATH)

    campanas_df = mc.read_table_df(fc.CANDIDATE_V2_PATH, "CAMPANAS", "tblCampanas")
    maestro_df = mc.read_table_df(fc.CANDIDATE_V2_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    parametros_df = mc.read_table_df(fc.CANDIDATE_V2_PATH, "PARAMETROS", "tblParametros")

    expected_maestro_rows = len(maestro_df)
    expected_campanas_rows = len(campanas_df)
    expected_parametros_rows = len(parametros_df)

    # --- Decisiones 3 y 4: verificar (no aplicar) ---
    decision_3 = fc.verify_decision_3(campanas_df)
    decision_4 = fc.verify_decision_4(campanas_df)

    # --- Verificar "antes" esperado de cada edición autorizada ---
    campanas_by_id = campanas_df.set_index("CargaID", drop=False)
    for edit in fc.AUTHORIZED_EDITS:
        carga_id = edit["carga_id"]
        if carga_id not in campanas_by_id.index:
            raise BuildError(f"{edit['decision']}: CargaID {carga_id!r} no existe en la candidata V2")
        row = campanas_by_id.loc[carga_id]
        for col, expected_val in edit["expected_before"].items():
            actual_val = row[col]
            same = actual_val == expected_val or (mc.is_blank(actual_val) and mc.is_blank(expected_val))
            if not same:
                raise BuildError(
                    f"{edit['decision']}: {carga_id}/{col} no coincide con el 'antes' esperado "
                    f"(esperado={expected_val!r}, encontrado={actual_val!r}). Build abortado, nada se escribió."
                )

    # --- Aplicar ediciones autorizadas sobre una copia de la candidata V2 ---
    target_wb = load_workbook(fc.CANDIDATE_V2_PATH, data_only=False, read_only=False, keep_vba=False)
    ws = target_wb["CAMPANAS"]
    cols, rows, _, _ = _header_row_maps(ws, "tblCampanas", fc.CAMPANAS_KEY)

    applied_log: list[dict[str, Any]] = []
    for edit in fc.AUTHORIZED_EDITS:
        carga_id = edit["carga_id"]
        r = rows[carga_id]
        before_row = campanas_by_id.loc[carga_id]
        for col, new_val in edit["changes"].items():
            cell = ws.cell(row=r, column=cols[col])
            before_val = before_row[col]
            cell.value = new_val
            applied_log.append(
                {
                    "decision": edit["decision"],
                    "CargaID": carga_id,
                    "ElementoID": before_row["ElementoID"],
                    "campo": col,
                    "valor_anterior": None if mc.is_blank(before_val) else before_val,
                    "valor_nuevo": new_val,
                }
            )

    # --- Invariantes: cantidad de filas sin cambios en las tres hojas ---
    for sheet_name, table_name in mc.EXPECTED_TABLES.items():
        t = target_wb[sheet_name].tables[table_name]
        _, mr, _, xr = range_boundaries(t.ref)
        expected = {
            "MAESTRO_ELEMENTOS": expected_maestro_rows,
            "CAMPANAS": expected_campanas_rows,
            "PARAMETROS": expected_parametros_rows,
        }[sheet_name]
        if xr - mr != expected:
            raise BuildError(f"{sheet_name}: la cantidad de filas cambió inesperadamente ({xr - mr} != {expected})")

    if target_wb.sheetnames != mc.EXPECTED_SHEETS:
        raise BuildError(f"Las hojas de la base FINAL no coinciden con lo esperado: {target_wb.sheetnames}")

    fc.FINAL_PATH.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(fc.FINAL_PATH)
    target_wb.close()

    sha_actual_after = mc.calculate_sha256(fc.mc.ACTUAL_PATH)
    sha_nueva_after = mc.calculate_sha256(fc.mc.NUEVA_PATH)
    sha_v2_after = mc.calculate_sha256(fc.CANDIDATE_V2_PATH)
    if sha_actual_after != sha_actual_before:
        raise BuildError("ERROR CRÍTICO: input/OCU26_BASE_DATOS.xlsx fue modificado durante el build FINAL")
    if sha_nueva_after != sha_nueva_before:
        raise BuildError("ERROR CRÍTICO: OCU26_BASE_NUEVA_RECIBIDA.xlsx fue modificado durante el build FINAL")
    if sha_v2_after != sha_v2_before:
        raise BuildError("ERROR CRÍTICO: la candidata V2 fue modificada durante el build FINAL")

    return {
        "result": "BUILD_FINAL_OK",
        "final_path": str(fc.FINAL_PATH),
        "sources_intact": True,
        "actual_sha256": sha_actual_after,
        "nueva_sha256": sha_nueva_after,
        "candidate_v2_sha256": sha_v2_after,
        "rows": {
            "MAESTRO_ELEMENTOS": expected_maestro_rows,
            "CAMPANAS": expected_campanas_rows,
            "PARAMETROS": expected_parametros_rows,
        },
        "ediciones_aplicadas": applied_log,
        "decision_3": decision_3,
        "decision_4": decision_4,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Construye la base FINAL OCU26 desde la candidata V2.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = build_final()
    except BuildError as exc:
        if args.json:
            print(json.dumps({"result": "BUILD_FINAL_ERROR", "error": str(exc)}, ensure_ascii=False, indent=2))
        else:
            print("=" * 60)
            print("OCU26 BUILD FINAL - ERROR")
            print("=" * 60)
            print(str(exc))
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 BUILD FINAL")
        print("=" * 60)
        print(f"Ruta: {result['final_path']}")
        print(f"Fuentes intactas: {result['sources_intact']}")
        print("Filas:", result["rows"])
        print()
        print("Ediciones aplicadas:")
        for e in result["ediciones_aplicadas"]:
            print(f"  [{e['decision']}] {e['CargaID']} / {e['campo']}: {e['valor_anterior']!r} -> {e['valor_nuevo']!r}")
        print()
        print("Decisión 3 (YPF Pier 3):", result["decision_3"]["status"], "-", result["decision_3"].get("motivo", ""))
        print("Decisión 4 (Beyond Cuarteto):", result["decision_4"]["status"], "-", result["decision_4"].get("motivo", ""))

    return 0


if __name__ == "__main__":
    sys.exit(main())
