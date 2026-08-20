"""Construye la base FINAL_V2 OCU26 a partir de la base FINAL anterior.

Aplica EXCLUSIVAMENTE las 3 correcciones de IDCampaña confirmadas
comercialmente (ver merge_final_v2_common.CORRECTIONS), cada una tras
pasar sus verificaciones (CargaID único, Campaña/ElementoID/IDCampaña
anterior coinciden, ID nuevo compatible, ClaveNegocio sin colisión). Si
alguna corrección no pasa sus controles, esa corrección puntual queda
bloqueada y NO se escribe (no hay escritura parcial de una corrección a
medias); las demás correcciones válidas proceden igual.

No modifica ni sobrescribe:
  - input/OCU26_BASE_DATOS.xlsx
  - OCU26_BASE_NUEVA_RECIBIDA.xlsx
  - OCU26_BASE_DATOS_INTEGRADA_CANDIDATA_V2_2026-08-18.xlsx
  - OCU26_BASE_DATOS_INTEGRADA_FINAL_2026-08-18.xlsx (fuente de este pase)

Uso:
    python build_ocu26_integrada_final_v2.py
    python build_ocu26_integrada_final_v2.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

import merge_common as mc
import merge_final_v2_common as fv2


class BuildError(Exception):
    pass


def _header_row_maps(ws, table_name: str, key_column: str) -> tuple[dict[str, int], dict[Any, int], int, int]:
    t = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    col_index = {h: min_col + i for i, h in enumerate(headers)}
    key_col = col_index[key_column]
    row_index: dict[Any, int] = {}
    for r in range(min_row + 1, max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if not mc.is_blank(v):
            row_index[v] = r
    return col_index, row_index, min_row, max_row


def build_final_v2() -> dict[str, Any]:
    sha_actual_before = mc.calculate_sha256(mc.ACTUAL_PATH)
    sha_nueva_before = mc.calculate_sha256(mc.NUEVA_PATH)
    sha_final_before = mc.calculate_sha256(fv2.FINAL_PATH)

    campanas_df = mc.read_table_df(fv2.FINAL_PATH, "CAMPANAS", "tblCampanas")
    maestro_df = mc.read_table_df(fv2.FINAL_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    parametros_df = mc.read_table_df(fv2.FINAL_PATH, "PARAMETROS", "tblParametros")

    expected_rows = {
        "MAESTRO_ELEMENTOS": len(maestro_df),
        "CAMPANAS": len(campanas_df),
        "PARAMETROS": len(parametros_df),
    }

    results = fv2.verify_all_corrections(campanas_df)
    applicable = [r for r in results if r["status"] == "APLICABLE"]
    blocked = [r for r in results if r["status"] != "APLICABLE"]

    target_wb = load_workbook(fv2.FINAL_PATH, data_only=False, read_only=False, keep_vba=False)
    ws = target_wb["CAMPANAS"]
    cols, rows, _, _ = _header_row_maps(ws, "tblCampanas", fv2.CAMPANAS_KEY)

    applied_log: list[dict[str, Any]] = []
    for result in applicable:
        carga_id = result["carga_id"]
        r = rows[carga_id]

        # Re-verificar el "antes" contra el workbook que se está por escribir
        # (no solo contra el DataFrame leído antes) antes de tocar nada.
        for col, expected_val in result["expected_before"].items():
            current = ws.cell(row=r, column=cols[col]).value
            same = current == expected_val or (mc.is_blank(current) and mc.is_blank(expected_val))
            if not same:
                raise BuildError(
                    f"{result['nombre']}: {carga_id}/{col} no coincide con el 'antes' esperado justo antes de "
                    f"escribir (esperado={expected_val!r}, encontrado={current!r}). Build abortado."
                )

        for col, new_val in result["changes"].items():
            cell = ws.cell(row=r, column=cols[col])
            before_val = cell.value
            cell.value = new_val
            applied_log.append(
                {
                    "correccion": result["nombre"],
                    "CargaID": carga_id,
                    "ElementoID": result["elemento_id"],
                    "campo": col,
                    "valor_anterior": None if mc.is_blank(before_val) else before_val,
                    "valor_nuevo": new_val,
                }
            )

    for sheet_name, table_name in mc.EXPECTED_TABLES.items():
        t = target_wb[sheet_name].tables[table_name]
        _, mr, _, xr = range_boundaries(t.ref)
        if xr - mr != expected_rows[sheet_name]:
            raise BuildError(f"{sheet_name}: la cantidad de filas cambió inesperadamente ({xr - mr} != {expected_rows[sheet_name]})")

    if target_wb.sheetnames != mc.EXPECTED_SHEETS:
        raise BuildError(f"Las hojas de FINAL_V2 no coinciden con lo esperado: {target_wb.sheetnames}")

    fv2.FINAL_V2_PATH.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(fv2.FINAL_V2_PATH)
    target_wb.close()

    sha_actual_after = mc.calculate_sha256(mc.ACTUAL_PATH)
    sha_nueva_after = mc.calculate_sha256(mc.NUEVA_PATH)
    sha_final_after = mc.calculate_sha256(fv2.FINAL_PATH)
    if sha_actual_after != sha_actual_before:
        raise BuildError("ERROR CRÍTICO: input/OCU26_BASE_DATOS.xlsx fue modificado durante el build FINAL_V2")
    if sha_nueva_after != sha_nueva_before:
        raise BuildError("ERROR CRÍTICO: OCU26_BASE_NUEVA_RECIBIDA.xlsx fue modificado durante el build FINAL_V2")
    if sha_final_after != sha_final_before:
        raise BuildError("ERROR CRÍTICO: la base FINAL anterior fue modificada durante el build FINAL_V2")

    return {
        "result": "BUILD_FINAL_V2_OK",
        "final_v2_path": str(fv2.FINAL_V2_PATH),
        "sources_intact": True,
        "actual_sha256": sha_actual_after,
        "nueva_sha256": sha_nueva_after,
        "final_v1_sha256": sha_final_after,
        "rows": expected_rows,
        "corrections_applied": len(applicable),
        "corrections_blocked": len(blocked),
        "applied_log": applied_log,
        "verification_results": results,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Construye la base FINAL_V2 OCU26 desde la base FINAL anterior.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = build_final_v2()
    except BuildError as exc:
        if args.json:
            print(json.dumps({"result": "BUILD_FINAL_V2_ERROR", "error": str(exc)}, ensure_ascii=False, indent=2))
        else:
            print("=" * 60)
            print("OCU26 BUILD FINAL_V2 - ERROR")
            print("=" * 60)
            print(str(exc))
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 BUILD FINAL_V2")
        print("=" * 60)
        print(f"Ruta: {result['final_v2_path']}")
        print(f"Fuentes intactas: {result['sources_intact']}")
        print("Filas:", result["rows"])
        print(f"Correcciones aplicadas: {result['corrections_applied']} / bloqueadas: {result['corrections_blocked']}")
        print()
        for e in result["applied_log"]:
            print(f"  [{e['correccion']}] {e['CargaID']} / {e['campo']}: {e['valor_anterior']!r} -> {e['valor_nuevo']!r}")
        print()
        for r in result["verification_results"]:
            if r["status"] != "APLICABLE":
                print(f"BLOQUEADA: {r['nombre']} - {r.get('motivo')}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
