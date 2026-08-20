"""Construye la base candidata V2 OCU26, reconstruyendo desde las dos
fuentes originales (NO parte de la candidata V1).

Aplica la precedencia V2 autorizada (ver merge_v2_common.py):
  - CAMPANAS, campos operativos, filas con CargaID en común: NUEVA gana
    siempre que tenga dato (actualiza incluso sobre un valor distinto ya
    existente); un vacío de NUEVA nunca borra un dato de ACTUAL.
  - ClaveNegocio se recalcula con la fórmula canónica (validada al 100%
    contra los datos reales) cuando cambia algún campo componente, salvo
    que la nueva clave colisione con otra fila -> en ese caso se revierte
    solo el/los campo(s) componente(s) de esa fila y queda en A_VALIDAR.
  - Los 4 registros nuevos de CAMPANAS se incorporan con CargaID/ClaveNegocio
    generados de forma determinista si pasan todos los controles; si no,
    quedan en A_VALIDAR sin incorporar.
  - MAESTRO_ELEMENTOS y PARAMETROS: sin cambio de precedencia (el pedido es
    exclusivo sobre CAMPANAS/CargaID); se verifican celda por celda al final.

Nunca escribe sobre input/OCU26_BASE_DATOS.xlsx ni sobre
OCU26_BASE_NUEVA_RECIBIDA.xlsx. No sobrescribe la candidata V1.

Uso:
    python build_ocu26_integrada_v2.py
    python build_ocu26_integrada_v2.py --json
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
from copy import copy

import merge_common as mc
import merge_v2_common as v2


class BuildError(Exception):
    pass


def _header_row_maps(ws, table_name: str, key_column: str) -> tuple[dict[str, int], dict[Any, int], int, int]:
    t = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    col_index = {h: min_col + i for i, h in enumerate(headers)}
    key_col = col_index[key_column]
    row_index: dict[Any, int] = {}
    for r in range(min_row + 1, max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if not mc.is_blank(v):
            row_index[v] = r
    return col_index, row_index, min_row, max_row


def _apply_maestro_completions(target_ws, source_ws, maestro_cls) -> int:
    if not maestro_cls.completions:
        return 0
    t_cols, t_rows, _, _ = _header_row_maps(target_ws, "tblElementos", v2.MAESTRO_KEY)
    s_cols, s_rows, _, _ = _header_row_maps(source_ws, "tblElementos", v2.MAESTRO_KEY)
    count = 0
    for c in maestro_cls.completions:
        target_cell = target_ws.cell(row=t_rows[c.key], column=t_cols[c.column])
        source_cell = source_ws.cell(row=s_rows[c.key], column=s_cols[c.column])
        if not mc.is_blank(target_cell.value):
            raise BuildError(f"MAESTRO_ELEMENTOS: intento de sobrescribir celda no vacía {c.key!r}/{c.column!r}")
        target_cell.value = source_cell.value
        if source_cell.number_format and source_cell.number_format != "General":
            target_cell.number_format = source_cell.number_format
        count += 1
    return count


def _apply_campanas_updates(target_ws, source_ws, final_applied: dict[Any, dict[str, Any]]) -> int:
    t_cols, t_rows, _, _ = _header_row_maps(target_ws, "tblCampanas", v2.CAMPANAS_KEY)
    s_cols, s_rows, _, _ = _header_row_maps(source_ws, "tblCampanas", v2.CAMPANAS_KEY)
    count = 0
    for key, applied in final_applied.items():
        if key not in t_rows:
            raise BuildError(f"CAMPANAS: CargaID {key!r} no encontrado en la hoja destino")
        for col, value in applied.items():
            target_cell = target_ws.cell(row=t_rows[key], column=t_cols[col])
            if col == "ClaveNegocio":
                target_cell.value = value  # calculado, no proviene de NUEVA
            else:
                if key not in s_rows:
                    raise BuildError(f"CAMPANAS: CargaID {key!r} no encontrado en la hoja fuente (NUEVA)")
                source_cell = source_ws.cell(row=s_rows[key], column=s_cols[col])
                target_cell.value = source_cell.value
                if source_cell.number_format and source_cell.number_format != "General":
                    target_cell.number_format = source_cell.number_format
            count += 1
    return count


def _append_new_campanas_rows(target_ws, incorporated: list[dict[str, Any]]) -> int:
    if not incorporated:
        return 0

    t = target_ws.tables["tblCampanas"]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(target_ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )

    date_columns = {"FechaInicio", "FechaFin"}
    datetime_columns = {"FechaHoraCarga"}

    style_row = max_row  # última fila de datos existente, fuente de estilo
    next_row = max_row + 1

    for record in incorporated:
        for i, header in enumerate(headers):
            col = min_col + i
            cell = target_ws.cell(row=next_row, column=col)
            style_cell = target_ws.cell(row=style_row, column=col)
            cell.font = copy(style_cell.font)
            cell.border = copy(style_cell.border)
            cell.fill = copy(style_cell.fill)
            cell.alignment = copy(style_cell.alignment)

            value = record.get(header)
            cell.value = value
            if header in date_columns and isinstance(value, (dt.date, dt.datetime)):
                cell.number_format = "yyyy-mm-dd"
            elif header in datetime_columns and isinstance(value, (dt.date, dt.datetime)):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            else:
                cell.number_format = style_cell.number_format
        next_row += 1

    new_max_row = max_row + len(incorporated)
    new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
    t.ref = new_ref

    return len(incorporated)


def build_candidate_v2(run_timestamp: dt.datetime | None = None) -> dict[str, Any]:
    sha_actual_before = mc.calculate_sha256(v2.ACTUAL_PATH)
    sha_nueva_before = mc.calculate_sha256(v2.NUEVA_PATH)

    plan = v2.compute_v2_plan(run_timestamp)

    expected_maestro_rows = len(plan.actual_maestro_df)
    expected_campanas_rows_before = len(plan.actual_campanas_df)
    expected_parametros_rows = len(plan.actual_parametros_df)

    target_wb = load_workbook(v2.ACTUAL_PATH, data_only=False, read_only=False, keep_vba=False)
    source_wb = load_workbook(v2.NUEVA_PATH, data_only=True, read_only=False, keep_vba=False)

    maestro_completions_applied = _apply_maestro_completions(
        target_wb["MAESTRO_ELEMENTOS"], source_wb["MAESTRO_ELEMENTOS"], plan.maestro_cls
    )
    campanas_fields_applied = _apply_campanas_updates(
        target_wb["CAMPANAS"], source_wb["CAMPANAS"], plan.campanas_final_applied
    )
    new_rows_appended = _append_new_campanas_rows(target_wb["CAMPANAS"], plan.new_records_incorporated)

    # --- Invariantes estructurales ---
    for sheet_name in ("MAESTRO_ELEMENTOS", "PARAMETROS"):
        table_name = mc.EXPECTED_TABLES[sheet_name]
        t = target_wb[sheet_name].tables[table_name]
        _, mr, _, xr = range_boundaries(t.ref)
        expected = expected_maestro_rows if sheet_name == "MAESTRO_ELEMENTOS" else expected_parametros_rows
        if xr - mr != expected:
            raise BuildError(f"{sheet_name}: la cantidad de filas cambió inesperadamente")

    t_campanas = target_wb["CAMPANAS"].tables["tblCampanas"]
    _, mr, _, xr = range_boundaries(t_campanas.ref)
    expected_campanas_rows_after = expected_campanas_rows_before + len(plan.new_records_incorporated)
    if xr - mr != expected_campanas_rows_after:
        raise BuildError(
            f"CAMPANAS: filas finales ({xr - mr}) no coinciden con lo esperado ({expected_campanas_rows_after})"
        )

    if target_wb.sheetnames != mc.EXPECTED_SHEETS:
        raise BuildError(f"Las hojas de la candidata no coinciden con lo esperado: {target_wb.sheetnames}")

    # --- Ningún campo rechazado fue aplicado: verificar contra ACTUAL original ---
    conflict_check_wb = load_workbook(v2.ACTUAL_PATH, data_only=False, read_only=False, keep_vba=False)
    t_cols, t_rows, _, _ = _header_row_maps(target_wb["CAMPANAS"], "tblCampanas", v2.CAMPANAS_KEY)
    oc_cols, oc_rows, _, _ = _header_row_maps(conflict_check_wb["CAMPANAS"], "tblCampanas", v2.CAMPANAS_KEY)
    for record in plan.campanas_change_records:
        if record.resultado != "RECHAZADO_A_VALIDAR":
            continue
        if record.carga_id not in t_rows:
            continue
        r = t_rows[record.carga_id]
        oc_r = oc_rows[record.carga_id]
        col = t_cols[record.columna]
        oc_col = oc_cols[record.columna]
        candidate_value = target_wb["CAMPANAS"].cell(row=r, column=col).value
        original_value = conflict_check_wb["CAMPANAS"].cell(row=oc_r, column=oc_col).value
        same = candidate_value == original_value or (mc.is_blank(candidate_value) and mc.is_blank(original_value))
        if not same:
            raise BuildError(
                f"CAMPANAS: un cambio marcado RECHAZADO_A_VALIDAR fue aplicado igualmente "
                f"({record.carga_id!r}/{record.columna!r}: candidata={candidate_value!r} original={original_value!r})"
            )
    conflict_check_wb.close()

    v2.CANDIDATE_V2_PATH.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(v2.CANDIDATE_V2_PATH)
    target_wb.close()
    source_wb.close()

    sha_actual_after = mc.calculate_sha256(v2.ACTUAL_PATH)
    sha_nueva_after = mc.calculate_sha256(v2.NUEVA_PATH)
    if sha_actual_after != sha_actual_before:
        raise BuildError("ERROR CRÍTICO: input/OCU26_BASE_DATOS.xlsx fue modificado durante el build V2")
    if sha_nueva_after != sha_nueva_before:
        raise BuildError("ERROR CRÍTICO: OCU26_BASE_NUEVA_RECIBIDA.xlsx fue modificado durante el build V2")

    aplicados = [c for c in plan.campanas_change_records if c.resultado == "APLICADO"]
    rechazados = [c for c in plan.campanas_change_records if c.resultado != "APLICADO"]
    completados = [c for c in aplicados if "Completado" in c.motivo]
    actualizados = [c for c in aplicados if "Actualizado" in c.motivo or c.fuente == "CALCULADO"]

    return {
        "result": "BUILD_V2_OK",
        "candidate_path": str(v2.CANDIDATE_V2_PATH),
        "sources_intact": True,
        "actual_sha256": sha_actual_after,
        "nueva_sha256": sha_nueva_after,
        "run_timestamp": plan.run_timestamp.isoformat(),
        "rows": {
            "MAESTRO_ELEMENTOS": expected_maestro_rows,
            "CAMPANAS_antes": expected_campanas_rows_before,
            "CAMPANAS_despues": expected_campanas_rows_after,
            "PARAMETROS": expected_parametros_rows,
        },
        "maestro_completions_applied": maestro_completions_applied,
        "campanas_field_changes_applied": campanas_fields_applied,
        "campanas_change_records_total": len(plan.campanas_change_records),
        "campanas_change_records_aplicados": len(aplicados),
        "campanas_change_records_rechazados": len(rechazados),
        "campanas_completados": len(completados),
        "campanas_actualizados": len(actualizados),
        "nuevos_registros_incorporados": len(plan.new_records_incorporated),
        "nuevos_registros_incorporados_ids": [r["CargaID"] for r in plan.new_records_incorporated],
        "nuevos_registros_rechazados": len(plan.new_records_rejected),
        "nuevos_registros_rechazados_detalle": [
            {"ElementoID": r.get("ElementoID"), "motivo": r.get("motivo_rechazo")} for r in plan.new_records_rejected
        ],
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Construye la base candidata V2 OCU26.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = build_candidate_v2()
    except BuildError as exc:
        if args.json:
            print(json.dumps({"result": "BUILD_V2_ERROR", "error": str(exc)}, ensure_ascii=False, indent=2))
        else:
            print("=" * 60)
            print("OCU26 BUILD INTEGRADA V2 - ERROR")
            print("=" * 60)
            print(str(exc))
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 BUILD INTEGRADA V2")
        print("=" * 60)
        for k, val in result.items():
            print(f"{k}: {val}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
