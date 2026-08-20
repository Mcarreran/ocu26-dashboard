"""Motor compartido de cruce OCU26_BASE_DATOS.xlsx vs OCU26_BASE_NUEVA_RECIBIDA.xlsx.

Usado por audit_ocu26_actualizacion.py y build_ocu26_integrada.py para
garantizar que la auditoría y la construcción de la base candidata usen
exactamente la misma clasificación (determinismo).

Estrictamente READ-ONLY sobre ambos archivos de entrada: solo lee, nunca
escribe ni recalcula nada en input/OCU26_BASE_DATOS.xlsx ni en
Pendientes/OCU26_ACTUALIZACION/input/OCU26_BASE_NUEVA_RECIBIDA.xlsx.

Reglas de fusión (FASE 4 del pedido de integración):
1. Actual con dato + nuevo vacío -> conservar actual.
2. Actual vacío + nuevo con dato válido -> completar en la candidata.
3. Ambos iguales -> sin cambios.
4. Ambos con valores diferentes -> NUNCA sobrescribir; registrar conflicto.
5. Registro solo en la base nueva -> agregar solo si tiene identificadores
   válidos y referencias completas (para CAMPANAS esto excluye filas sin
   CargaID, porque no existe en el proyecto una función de generación de
   CargaID: solo un patrón histórico documentado 'HIST-00000001'. Inventar
   esa lógica sin evidencia de código está fuera de alcance -> A_VALIDAR).
6. Registro solo en la base actual -> conservar, nunca eliminar.
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

REPO_ROOT = Path(__file__).resolve().parents[3]
UPDATE_ROOT = Path(__file__).resolve().parents[1]

sys.path.insert(0, str(REPO_ROOT / "scripts"))
import validate_input as vi  # noqa: E402

ACTUAL_PATH = REPO_ROOT / "input" / "OCU26_BASE_DATOS.xlsx"
NUEVA_PATH = UPDATE_ROOT / "input" / "OCU26_BASE_NUEVA_RECIBIDA.xlsx"

CANDIDATE_PATH = UPDATE_ROOT / "output" / "OCU26_BASE_DATOS_INTEGRADA_CANDIDATA_2026-08-18.xlsx"
AUDIT_PATH = UPDATE_ROOT / "output" / "OCU26_AUDITORIA_CRUCE_2026-08-18.xlsx"

EXPECTED_SHEETS = vi.EXPECTED_SHEETS
EXPECTED_TABLES = vi.EXPECTED_TABLES
MAESTRO_HEADERS = vi.MAESTRO_HEADERS
CAMPANAS_HEADERS = vi.CAMPANAS_HEADERS
PARAMETROS_HEADERS = vi.PARAMETROS_HEADERS

MAESTRO_KEY = "ElementoID"
CAMPANAS_KEY = "CargaID"

# Snapshot histórico documentado en docs/CM1.md (sección 9), usado solo como
# referencia informativa en el preflight; nunca fuerza resultados.
HISTORICAL_SNAPSHOT = {
    "maestro_rows": 4338,
    "campanas_rows": 9503,
    "parametros_rows": 23,
}

calculate_sha256 = vi.calculate_sha256


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except (TypeError, ValueError):
        pass
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def read_table_df(path: Path, sheet: str, table: str) -> pd.DataFrame:
    """Lee una tabla estructurada por su rango real (ws.tables). READ-ONLY."""
    wb = load_workbook(path, data_only=True, read_only=False, keep_vba=False)
    try:
        ws = wb[sheet]
        t = ws.tables[table]
        min_col, min_row, max_col, max_row = range_boundaries(t.ref)
        headers = list(
            next(
                ws.iter_rows(
                    min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True
                )
            )
        )
        data_rows = list(
            ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True)
        )
    finally:
        wb.close()
    return pd.DataFrame(data_rows, columns=headers)


def inspect_structure(path: Path) -> dict[str, Any]:
    """Inspección estructural completa (FASE 1) de un archivo. READ-ONLY."""
    info: dict[str, Any] = {
        "path": str(path),
        "exists": path.exists(),
        "sha256": None,
        "size_bytes": None,
        "zip_ok": None,
        "vba": [],
        "external_links": [],
        "macro_enabled": None,
        "sheets": [],
        "defined_names": [],
        "formula_cells": 0,
        "error_cells": 0,
        "sheet_detail": {},
    }
    if not path.exists():
        return info

    info["size_bytes"] = path.stat().st_size
    info["sha256"] = calculate_sha256(path)

    import zipfile

    with zipfile.ZipFile(path) as zf:
        bad = zf.testzip()
        info["zip_ok"] = bad is None
        names = zf.namelist()
        info["vba"] = [n for n in names if "vbaProject" in n]
        info["external_links"] = [n for n in names if n.startswith("xl/externalLinks/")]
        ct = zf.read("[Content_Types].xml").decode("utf-8", errors="replace") if "[Content_Types].xml" in names else ""
        info["macro_enabled"] = "macroEnabled" in ct

    wb = load_workbook(path, data_only=False, read_only=False, keep_vba=False)
    info["sheets"] = list(wb.sheetnames)
    info["defined_names"] = list(wb.defined_names.keys())

    for sn in wb.sheetnames:
        ws = wb[sn]
        tnames = list(ws.tables.keys())
        table_detail = {}
        for tname in tnames:
            t = ws.tables[tname]
            min_col, min_row, max_col, max_row = range_boundaries(t.ref)
            table_detail[tname] = {
                "ref": t.ref,
                "data_rows": max_row - min_row,
                "columns": max_col - min_col + 1,
            }
        info["sheet_detail"][sn] = {
            "state": ws.sheet_state,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "tables": table_detail,
        }
        if ws.max_row and ws.max_column:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    if cell.data_type == "f":
                        info["formula_cells"] += 1
                    elif cell.data_type == "e":
                        info["error_cells"] += 1
                    elif isinstance(cell.value, str) and cell.value.strip() in vi.EXCEL_ERROR_STRINGS:
                        info["error_cells"] += 1
    wb.close()
    return info


def column_mapping(headers_a: list[str], headers_b: list[str]) -> list[dict[str, str]]:
    """Mapeo de columnas entre ACTUAL y NUEVA para una hoja (FASE 2).

    No aplica fuzzy matching. Solo compara nombre exacto de encabezado y
    posición.
    """
    rows: list[dict[str, str]] = []
    set_a, set_b = set(headers_a), set(headers_b)
    for i, h in enumerate(headers_a):
        if h in set_b:
            j = headers_b.index(h)
            clasif = "COINCIDENCIA_EXACTA" if i == j else "EQUIVALENCIA_CONFIRMADA"
            rows.append(
                {
                    "columna": h,
                    "posicion_actual": i + 1,
                    "posicion_nueva": j + 1,
                    "clasificacion": clasif,
                }
            )
        else:
            rows.append({"columna": h, "posicion_actual": i + 1, "posicion_nueva": None, "clasificacion": "SOLO_BASE_ACTUAL"})
    for j, h in enumerate(headers_b):
        if h not in set_a:
            rows.append({"columna": h, "posicion_actual": None, "posicion_nueva": j + 1, "clasificacion": "SOLO_BASE_NUEVA"})
    return rows


class FieldCompletion:
    __slots__ = ("key", "column", "actual_value", "nueva_value")

    def __init__(self, key: Any, column: str, actual_value: Any, nueva_value: Any) -> None:
        self.key = key
        self.column = column
        self.actual_value = actual_value
        self.nueva_value = nueva_value


class FieldConflict:
    __slots__ = ("key", "column", "actual_value", "nueva_value")

    def __init__(self, key: Any, column: str, actual_value: Any, nueva_value: Any) -> None:
        self.key = key
        self.column = column
        self.actual_value = actual_value
        self.nueva_value = nueva_value


class SheetClassification:
    def __init__(self, sheet_name: str, key_column: str) -> None:
        self.sheet_name = sheet_name
        self.key_column = key_column
        self.completions: list[FieldCompletion] = []
        self.conflicts: list[FieldConflict] = []
        self.only_actual_keys: list[Any] = []
        self.only_nueva_unkeyed_rows: list[dict[str, Any]] = []  # filas nuevas sin clave estable
        self.only_nueva_keyed_new: list[Any] = []  # claves nuevas válidas presentes solo en NUEVA (no aplica hoy)
        self.rows_no_change: int = 0
        self.rows_common: int = 0
        self.duplicates_actual: dict[Any, int] = {}
        self.duplicates_nueva: dict[Any, int] = {}


def _classify_keyed_sheet(
    sheet_name: str,
    key_column: str,
    actual_df: pd.DataFrame,
    nueva_df: pd.DataFrame,
    excluded_from_field_diff: Iterable[str] = (),
) -> SheetClassification:
    """Clasificación genérica por clave estable para MAESTRO_ELEMENTOS y CAMPANAS."""
    result = SheetClassification(sheet_name, key_column)

    actual_keyed = actual_df[~actual_df[key_column].apply(is_blank)].copy()
    nueva_keyed = nueva_df[~nueva_df[key_column].apply(is_blank)].copy()

    dup_a = actual_keyed[key_column][actual_keyed[key_column].duplicated(keep=False)].value_counts()
    dup_n = nueva_keyed[key_column][nueva_keyed[key_column].duplicated(keep=False)].value_counts()
    result.duplicates_actual = {k: int(v) for k, v in dup_a.items()}
    result.duplicates_nueva = {k: int(v) for k, v in dup_n.items()}

    # Para filas con clave duplicada en cualquiera de las dos bases, no se
    # puede alinear 1:1 con seguridad -> se excluyen del cruce por campo y
    # se documentan como A_VALIDAR vía duplicates_actual/duplicates_nueva.
    dup_keys = set(dup_a.index) | set(dup_n.index)

    actual_indexed = actual_keyed[~actual_keyed[key_column].isin(dup_keys)].set_index(key_column)
    nueva_indexed = nueva_keyed[~nueva_keyed[key_column].isin(dup_keys)].set_index(key_column)

    common_keys = actual_indexed.index.intersection(nueva_indexed.index)
    only_actual = actual_indexed.index.difference(nueva_indexed.index)
    only_nueva = nueva_indexed.index.difference(actual_indexed.index)

    result.rows_common = len(common_keys)
    result.only_actual_keys = list(only_actual)
    result.only_nueva_keyed_new = list(only_nueva)

    # Filas de NUEVA sin clave (no se pueden alinear de forma estable).
    unkeyed_new = nueva_df[nueva_df[key_column].apply(is_blank)]
    result.only_nueva_unkeyed_rows = unkeyed_new.to_dict(orient="records")

    a_common = actual_indexed.loc[common_keys]
    n_common = nueva_indexed.loc[common_keys]

    columns = [c for c in actual_df.columns if c != key_column and c not in excluded_from_field_diff]
    no_change_mask = pd.Series(True, index=common_keys)

    for col in columns:
        s1 = a_common[col]
        s2 = n_common[col]
        for key in common_keys:
            v1 = s1.loc[key]
            v2 = s2.loc[key]
            blank1, blank2 = is_blank(v1), is_blank(v2)
            if blank1 and blank2:
                continue
            if not blank1 and not blank2:
                if v1 == v2:
                    continue
                result.conflicts.append(FieldConflict(key, col, v1, v2))
                no_change_mask.loc[key] = False
                continue
            if blank1 and not blank2:
                result.completions.append(FieldCompletion(key, col, v1, v2))
                no_change_mask.loc[key] = False
                continue
            # not blank1 and blank2 -> conservar actual, sin acción (regla 1)

    result.rows_no_change = int(no_change_mask.sum())
    return result


def classify_maestro(actual_df: pd.DataFrame, nueva_df: pd.DataFrame) -> SheetClassification:
    return _classify_keyed_sheet("MAESTRO_ELEMENTOS", MAESTRO_KEY, actual_df, nueva_df)


def classify_campanas(actual_df: pd.DataFrame, nueva_df: pd.DataFrame) -> SheetClassification:
    return _classify_keyed_sheet("CAMPANAS", CAMPANAS_KEY, actual_df, nueva_df)


def classify_parametros(actual_df: pd.DataFrame, nueva_df: pd.DataFrame) -> dict[str, Any]:
    """PARAMETROS no tiene clave única declarada; se compara por tupla
    (Categoria, Valor) completa, que es el grano real de la hoja."""
    a_tuples = list(actual_df.itertuples(index=False, name=None))
    n_tuples = list(nueva_df.itertuples(index=False, name=None))
    set_a, set_n = set(a_tuples), set(n_tuples)
    return {
        "identical": a_tuples == n_tuples,
        "rows_actual": len(a_tuples),
        "rows_nueva": len(n_tuples),
        "only_actual": sorted(set_a - set_n),
        "only_nueva": sorted(set_n - set_a),
        "common": len(set_a & set_n),
    }


def maestro_referential_orphans(campanas_df: pd.DataFrame, maestro_ids: set[Any]) -> list[tuple[Any, Any]]:
    orphans = []
    for _, row in campanas_df.iterrows():
        eid = row["ElementoID"]
        if is_blank(eid):
            continue
        if eid not in maestro_ids:
            orphans.append((row.get(CAMPANAS_KEY), eid))
    return orphans
