"""Validación obligatoria de la base candidata OCU26 integrada.

Corre el Gate 1 canónico (scripts/validate_input.py) sobre la candidata y,
además, invariantes propias de la integración:

  - Estructura canónica preservada (hojas/tablas/encabezados/orden).
  - Cero ElementoID vacíos o duplicados (vía validate_input).
  - Cero ElementoID huérfanos en CAMPANAS (vía validate_input).
  - Cero pérdida silenciosa de filas: todo CargaID/ElementoID de ACTUAL
    sigue presente en la candidata.
  - CargaID estable: el conjunto de CargaID de la candidata es idéntico al
    de ACTUAL (no se generaron IDs nuevos, no se perdió ninguno).
  - Cero sobreescritura de valores conflictivos: cada conflicto detectado
    en el cruce sigue teniendo en la candidata el valor original de ACTUAL.
  - El conjunto de celdas que cambiaron entre ACTUAL y la candidata es
    EXACTAMENTE el conjunto de "cambios seguros" calculado por
    merge_common (ni una celda de más, ni una de menos).
  - Ausencia de fórmulas externas, errores de Excel, macros, vínculos
    externos; integridad ZIP/XLSX.
  - Apertura con data_only=False y data_only=True.
  - Hashes de ambas fuentes intactos tras la corrida.

Uso:
    python validate_ocu26_integrada.py
    python validate_ocu26_integrada.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

import merge_common as mc

vi = mc.vi

EXPECTED_ACTUAL_SHA256 = "2f165e12ad90c2f05963367a8e1717dcd1006e9ce669f976afa6e57470aca2cd"


class ValidationFailure(Exception):
    pass


def _key_value_map(ws, table_name: str, key_column: str, value_columns: list[str]) -> dict[Any, dict[str, Any]]:
    t = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    col_index = {h: i for i, h in enumerate(headers)}
    key_i = col_index[key_column]
    result: dict[Any, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True):
        key = row[key_i]
        if mc.is_blank(key):
            continue
        result[key] = {h: row[col_index[h]] for h in value_columns}
    return result


def _full_row_map(ws, table_name: str) -> dict[int, dict[str, Any]]:
    """Todas las celdas por (fila, columna) -> valor, para diff exhaustivo."""
    t = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    grid: dict[int, dict[str, Any]] = {}
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True),
        start=min_row + 1,
    ):
        grid[r_idx] = dict(zip(headers, row))
    return grid


def run_validation() -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    checks: dict[str, Any] = {}

    if not mc.CANDIDATE_PATH.exists():
        raise ValidationFailure(f"No existe la candidata: {mc.CANDIDATE_PATH}")

    sha_actual_before = mc.calculate_sha256(mc.ACTUAL_PATH)
    sha_nueva_before = mc.calculate_sha256(mc.NUEVA_PATH)

    checks["actual_sha256"] = sha_actual_before
    checks["actual_sha256_matches_expected_historico"] = sha_actual_before == EXPECTED_ACTUAL_SHA256
    if not checks["actual_sha256_matches_expected_historico"]:
        errors.append("El SHA-256 de input/OCU26_BASE_DATOS.xlsx no coincide con el esperado histórico")

    # --- Gate 1 canónico sobre la candidata ---
    candidate_validation = vi.validate_input(mc.CANDIDATE_PATH)
    checks["candidate_validate_input_result"] = candidate_validation["result"]
    checks["candidate_validate_input_errors"] = candidate_validation["errors"]
    checks["candidate_validate_input_warnings_count"] = len(candidate_validation["warnings"])
    if candidate_validation["result"] == "INVALID":
        errors.append(f"La candidata NO pasa validate_input.py (Gate 1): {candidate_validation['errors']}")

    # --- Apertura data_only=False y data_only=True ---
    try:
        wb_f = load_workbook(mc.CANDIDATE_PATH, data_only=False, read_only=False, keep_vba=False)
        wb_t = load_workbook(mc.CANDIDATE_PATH, data_only=True, read_only=False, keep_vba=False)
        checks["opens_data_only_false"] = True
        checks["opens_data_only_true"] = True
    except Exception as exc:  # noqa: BLE001
        errors.append(f"La candidata no pudo abrirse con openpyxl: {exc}")
        return _finalize(errors, warnings, checks)

    # --- Estructura canónica ---
    if wb_f.sheetnames != mc.EXPECTED_SHEETS:
        errors.append(f"Hojas de la candidata no coinciden: {wb_f.sheetnames} != {mc.EXPECTED_SHEETS}")
    for sheet_name, table_name in mc.EXPECTED_TABLES.items():
        tnames = list(wb_f[sheet_name].tables.keys())
        if tnames != [table_name]:
            errors.append(f"Hoja '{sheet_name}': tabla esperada '{table_name}', encontrada {tnames}")
    checks["structure_preserved"] = not errors

    # --- Integridad estructural (zip/vba/external links/macro/formulas/errors) ---
    cand_info = mc.inspect_structure(mc.CANDIDATE_PATH)
    checks["zip_ok"] = cand_info["zip_ok"]
    checks["vba"] = cand_info["vba"]
    checks["external_links"] = cand_info["external_links"]
    checks["macro_enabled"] = cand_info["macro_enabled"]
    checks["formula_cells"] = cand_info["formula_cells"]
    checks["error_cells"] = cand_info["error_cells"]
    if not cand_info["zip_ok"]:
        errors.append("El contenedor ZIP de la candidata está corrupto")
    if cand_info["vba"]:
        errors.append("La candidata contiene macros (vbaProject.bin)")
    if cand_info["external_links"]:
        errors.append("La candidata contiene vínculos externos")
    if cand_info["macro_enabled"]:
        errors.append("La candidata está marcada como macro-habilitada")
    if cand_info["formula_cells"]:
        errors.append(f"La candidata contiene {cand_info['formula_cells']} celda(s) con fórmulas")
    if cand_info["error_cells"]:
        errors.append(f"La candidata contiene {cand_info['error_cells']} celda(s) con errores de Excel")

    # --- Recalcular clasificación de referencia (misma lógica que build) ---
    actual_maestro_df = mc.read_table_df(mc.ACTUAL_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    nueva_maestro_df = mc.read_table_df(mc.NUEVA_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    actual_campanas_df = mc.read_table_df(mc.ACTUAL_PATH, "CAMPANAS", "tblCampanas")
    nueva_campanas_df = mc.read_table_df(mc.NUEVA_PATH, "CAMPANAS", "tblCampanas")
    actual_parametros_df = mc.read_table_df(mc.ACTUAL_PATH, "PARAMETROS", "tblParametros")
    candidate_parametros_df = mc.read_table_df(mc.CANDIDATE_PATH, "PARAMETROS", "tblParametros")

    maestro_cls = mc.classify_maestro(actual_maestro_df, nueva_maestro_df)
    campanas_cls = mc.classify_campanas(actual_campanas_df, nueva_campanas_df)

    expected_changed_cells: set[tuple[str, Any, str]] = set()
    for c in maestro_cls.completions:
        expected_changed_cells.add(("MAESTRO_ELEMENTOS", c.key, c.column))
    for c in campanas_cls.completions:
        expected_changed_cells.add(("CAMPANAS", c.key, c.column))

    # --- Diff exhaustivo candidata vs ACTUAL, celda por celda ---
    wb_actual = load_workbook(mc.ACTUAL_PATH, data_only=False, read_only=False, keep_vba=False)

    unexpected_changes: list[tuple[str, Any, str, Any, Any]] = []
    missing_completions: list[tuple[str, Any, str]] = []
    key_columns = {"MAESTRO_ELEMENTOS": mc.MAESTRO_KEY, "CAMPANAS": mc.CAMPANAS_KEY}

    for sheet_name, table_name in mc.EXPECTED_TABLES.items():
        if sheet_name == "PARAMETROS":
            continue
        key_col = key_columns[sheet_name]
        actual_grid = _full_row_map(wb_actual[sheet_name], table_name)
        cand_grid = _full_row_map(wb_f[sheet_name], table_name)

        actual_by_key = {row[key_col]: row for row in actual_grid.values() if not mc.is_blank(row[key_col])}
        cand_by_key = {row[key_col]: row for row in cand_grid.values() if not mc.is_blank(row[key_col])}

        missing_keys = set(actual_by_key) - set(cand_by_key)
        if missing_keys:
            errors.append(f"{sheet_name}: {len(missing_keys)} clave(s) de ACTUAL ausentes en la candidata: {sorted(missing_keys)[:10]}")

        for key, actual_row in actual_by_key.items():
            cand_row = cand_by_key.get(key)
            if cand_row is None:
                continue
            for col, actual_val in actual_row.items():
                if col == key_col:
                    continue
                cand_val = cand_row[col]
                same = (actual_val == cand_val) or (mc.is_blank(actual_val) and mc.is_blank(cand_val))
                if same:
                    continue
                tag = (sheet_name, key, col)
                if tag not in expected_changed_cells:
                    unexpected_changes.append((sheet_name, key, col, actual_val, cand_val))

        for tag in expected_changed_cells:
            if tag[0] != sheet_name:
                continue
            _, key, col = tag
            cand_row = cand_by_key.get(key)
            if cand_row is None:
                continue
            if mc.is_blank(cand_row[col]):
                missing_completions.append(tag)

    checks["unexpected_changed_cells"] = len(unexpected_changes)
    checks["missing_completions"] = len(missing_completions)
    if unexpected_changes:
        errors.append(
            f"{len(unexpected_changes)} celda(s) cambiaron sin estar en la lista de cambios seguros: "
            f"{unexpected_changes[:10]}"
        )
    if missing_completions:
        errors.append(f"{len(missing_completions)} completado(s) esperado(s) no se aplicaron en la candidata: {missing_completions[:10]}")

    # --- PARAMETROS: debe seguir idéntico a ACTUAL ---
    if not actual_parametros_df.equals(candidate_parametros_df):
        errors.append("PARAMETROS cambió entre ACTUAL y la candidata (debía permanecer idéntico)")
    checks["parametros_unchanged"] = actual_parametros_df.equals(candidate_parametros_df)

    # --- CargaID / ElementoID estables (mismo conjunto, sin generación de IDs nuevos) ---
    cand_campanas_df = mc.read_table_df(mc.CANDIDATE_PATH, "CAMPANAS", "tblCampanas")
    cand_maestro_df = mc.read_table_df(mc.CANDIDATE_PATH, "MAESTRO_ELEMENTOS", "tblElementos")

    actual_carga_ids = set(actual_campanas_df["CargaID"].dropna())
    cand_carga_ids = set(cand_campanas_df["CargaID"].dropna())
    checks["carga_id_set_stable"] = actual_carga_ids == cand_carga_ids
    if actual_carga_ids != cand_carga_ids:
        errors.append(
            f"El conjunto de CargaID cambió: {len(actual_carga_ids - cand_carga_ids)} perdido(s), "
            f"{len(cand_carga_ids - actual_carga_ids)} nuevo(s) no autorizado(s)"
        )

    actual_elemento_ids = set(actual_maestro_df["ElementoID"].dropna())
    cand_elemento_ids = set(cand_maestro_df["ElementoID"].dropna())
    checks["elemento_id_set_stable"] = actual_elemento_ids == cand_elemento_ids
    if actual_elemento_ids != cand_elemento_ids:
        errors.append(
            f"El conjunto de ElementoID cambió: {len(actual_elemento_ids - cand_elemento_ids)} perdido(s), "
            f"{len(cand_elemento_ids - actual_elemento_ids)} nuevo(s) no autorizado(s)"
        )

    wb_f.close()
    wb_t.close()
    wb_actual.close()

    # --- Fuentes intactas tras toda la corrida ---
    sha_actual_after = mc.calculate_sha256(mc.ACTUAL_PATH)
    sha_nueva_after = mc.calculate_sha256(mc.NUEVA_PATH)
    checks["actual_sha256_unchanged"] = sha_actual_after == sha_actual_before
    checks["nueva_sha256_unchanged"] = sha_nueva_after == sha_nueva_before
    if not checks["actual_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: input/OCU26_BASE_DATOS.xlsx cambió durante la validación")
    if not checks["nueva_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: OCU26_BASE_NUEVA_RECIBIDA.xlsx cambió durante la validación")

    return _finalize(errors, warnings, checks)


def _finalize(errors: list[str], warnings: list[str], checks: dict[str, Any]) -> dict[str, Any]:
    result = "INVALID" if errors else ("VALID_WITH_WARNINGS" if warnings else "VALID")
    return {"result": result, "errors": errors, "warnings": warnings, "checks": checks}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Valida la base candidata OCU26 integrada.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = run_validation()
    except ValidationFailure as exc:
        if args.json:
            print(json.dumps({"result": "INVALID", "errors": [str(exc)], "warnings": [], "checks": {}}, ensure_ascii=False, indent=2))
        else:
            print("RESULT: INVALID")
            print(str(exc))
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 VALIDACION DE CANDIDATA")
        print("=" * 60)
        for k, v in result["checks"].items():
            print(f"  {k}: {v}")
        print()
        print("ERRORS:")
        for e in result["errors"]:
            print(f"  - {e}")
        if not result["errors"]:
            print("  none")
        print()
        print("RESULT:", result["result"])

    return 0 if result["result"] != "INVALID" else 1


if __name__ == "__main__":
    sys.exit(main())
