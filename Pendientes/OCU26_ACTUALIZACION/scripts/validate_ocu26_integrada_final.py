"""Validación obligatoria de la base FINAL OCU26.

Controles:
  - Gate 1 canónico (scripts/validate_input.py) sobre la base FINAL.
  - Estructura canónica preservada.
  - Integridad ZIP/XLSX, ausencia de macros/vínculos externos/fórmulas/errores.
  - Apertura con data_only=False y data_only=True.
  - MAESTRO_ELEMENTOS y PARAMETROS: idénticos celda por celda a la candidata V2.
  - CAMPANAS: el conjunto de celdas que cambiaron respecto de la candidata V2
    es EXACTAMENTE el de las 8 ediciones autorizadas (4 filas x 2 campos),
    ni una celda de más ni de menos. Ningún CargaID perdido. Ningún
    ElementoID huérfano. Cero colisiones nuevas de ClaveNegocio (nadie tocó
    ClaveNegocio en este pase). Los 11 grupos de duplicados exactos y los 3
    grupos con diferencia comercial real siguen presentes sin alteración de
    identidad (mismos CargaID, mismas filas).
  - Hashes de ACTUAL, NUEVA y candidata V2 intactos antes/después.

Uso:
    python validate_ocu26_integrada_final.py
    python validate_ocu26_integrada_final.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

import merge_common as mc
import merge_final_common as fc

vi = mc.vi

EXPECTED_ACTUAL_SHA256 = "2f165e12ad90c2f05963367a8e1717dcd1006e9ce669f976afa6e57470aca2cd"


class ValidationFailure(Exception):
    pass


def _full_row_map(ws, table_name: str) -> dict[int, dict[str, Any]]:
    t = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    grid: dict[int, dict[str, Any]] = {}
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True),
        start=min_row + 1,
    ):
        grid[r_idx] = dict(zip(headers, row))
    return grid


def run_validation() -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    checks: dict[str, Any] = {}

    if not fc.FINAL_PATH.exists():
        raise ValidationFailure(f"No existe la base FINAL: {fc.FINAL_PATH}")

    sha_actual_before = mc.calculate_sha256(mc.ACTUAL_PATH)
    sha_nueva_before = mc.calculate_sha256(mc.NUEVA_PATH)
    sha_v2_before = mc.calculate_sha256(fc.CANDIDATE_V2_PATH)
    checks["actual_sha256"] = sha_actual_before
    checks["actual_sha256_matches_expected_historico"] = sha_actual_before == EXPECTED_ACTUAL_SHA256
    if not checks["actual_sha256_matches_expected_historico"]:
        errors.append("El SHA-256 de input/OCU26_BASE_DATOS.xlsx no coincide con el esperado histórico")

    candidate_validation = vi.validate_input(fc.FINAL_PATH)
    checks["final_validate_input_result"] = candidate_validation["result"]
    checks["final_validate_input_errors"] = candidate_validation["errors"]
    checks["final_validate_input_warnings_count"] = len(candidate_validation["warnings"])
    if candidate_validation["result"] == "INVALID":
        errors.append(f"La base FINAL NO pasa validate_input.py (Gate 1): {candidate_validation['errors']}")

    try:
        wb_f = load_workbook(fc.FINAL_PATH, data_only=False, read_only=False, keep_vba=False)
        wb_t = load_workbook(fc.FINAL_PATH, data_only=True, read_only=False, keep_vba=False)
        checks["opens_data_only_false"] = True
        checks["opens_data_only_true"] = True
    except Exception as exc:  # noqa: BLE001
        errors.append(f"La base FINAL no pudo abrirse con openpyxl: {exc}")
        return _finalize(errors, warnings, checks)

    if wb_f.sheetnames != mc.EXPECTED_SHEETS:
        errors.append(f"Hojas no coinciden: {wb_f.sheetnames} != {mc.EXPECTED_SHEETS}")
    for sheet_name, table_name in mc.EXPECTED_TABLES.items():
        tnames = list(wb_f[sheet_name].tables.keys())
        if tnames != [table_name]:
            errors.append(f"Hoja '{sheet_name}': tabla esperada '{table_name}', encontrada {tnames}")
    checks["structure_preserved"] = not errors

    final_info = mc.inspect_structure(fc.FINAL_PATH)
    checks["zip_ok"] = final_info["zip_ok"]
    checks["vba"] = final_info["vba"]
    checks["external_links"] = final_info["external_links"]
    checks["macro_enabled"] = final_info["macro_enabled"]
    checks["formula_cells"] = final_info["formula_cells"]
    checks["error_cells"] = final_info["error_cells"]
    if not final_info["zip_ok"]:
        errors.append("El contenedor ZIP de la base FINAL está corrupto")
    if final_info["vba"]:
        errors.append("La base FINAL contiene macros")
    if final_info["external_links"]:
        errors.append("La base FINAL contiene vínculos externos")
    if final_info["macro_enabled"]:
        errors.append("La base FINAL está marcada como macro-habilitada")
    if final_info["formula_cells"]:
        errors.append(f"La base FINAL contiene {final_info['formula_cells']} celda(s) con fórmulas")
    if final_info["error_cells"]:
        errors.append(f"La base FINAL contiene {final_info['error_cells']} celda(s) con errores de Excel")

    wb_v2 = load_workbook(fc.CANDIDATE_V2_PATH, data_only=False, read_only=False, keep_vba=False)

    # --- MAESTRO_ELEMENTOS y PARAMETROS: idénticos celda por celda a V2 ---
    for sheet_name, table_name in (("MAESTRO_ELEMENTOS", "tblElementos"), ("PARAMETROS", "tblParametros")):
        v2_grid = _full_row_map(wb_v2[sheet_name], table_name)
        final_grid = _full_row_map(wb_f[sheet_name], table_name)
        diffs = 0
        if len(v2_grid) != len(final_grid):
            errors.append(f"{sheet_name}: cantidad de filas distinta (V2={len(v2_grid)}, FINAL={len(final_grid)})")
        else:
            for r in v2_grid:
                for col, v2_val in v2_grid[r].items():
                    f_val = final_grid[r][col]
                    same = v2_val == f_val or (mc.is_blank(v2_val) and mc.is_blank(f_val))
                    if not same:
                        diffs += 1
        checks[f"{sheet_name.lower()}_cell_diffs_vs_v2"] = diffs
        if diffs:
            errors.append(f"{sheet_name}: {diffs} celda(s) distinta(s) de la candidata V2 (debía ser 0)")

    # --- CAMPANAS: diff exhaustivo vs V2, comparado contra el whitelist de ediciones autorizadas ---
    v2_campanas_grid = _full_row_map(wb_v2["CAMPANAS"], "tblCampanas")
    final_campanas_grid = _full_row_map(wb_f["CAMPANAS"], "tblCampanas")

    v2_by_carga = {row[fc.CAMPANAS_KEY]: row for row in v2_campanas_grid.values() if not mc.is_blank(row[fc.CAMPANAS_KEY])}
    final_by_carga = {row[fc.CAMPANAS_KEY]: row for row in final_campanas_grid.values() if not mc.is_blank(row[fc.CAMPANAS_KEY])}

    missing = set(v2_by_carga) - set(final_by_carga)
    extra = set(final_by_carga) - set(v2_by_carga)
    checks["campanas_missing_carga_ids"] = len(missing)
    checks["campanas_extra_carga_ids"] = len(extra)
    if missing:
        errors.append(f"CAMPANAS: {len(missing)} CargaID de la candidata V2 ausentes en FINAL: {sorted(missing)[:10]}")
    if extra:
        errors.append(f"CAMPANAS: {len(extra)} CargaID nuevos no autorizados en FINAL: {sorted(extra)[:10]}")

    expected_changes: dict[tuple[Any, str], Any] = {}
    for edit in fc.AUTHORIZED_EDITS:
        for col, new_val in edit["changes"].items():
            expected_changes[(edit["carga_id"], col)] = new_val

    unexpected_changes = []
    missing_expected = []
    for carga_id, v2_row in v2_by_carga.items():
        final_row = final_by_carga.get(carga_id)
        if final_row is None:
            continue
        for col, v2_val in v2_row.items():
            if col == fc.CAMPANAS_KEY:
                continue
            f_val = final_row[col]
            same = v2_val == f_val or (mc.is_blank(v2_val) and mc.is_blank(f_val))
            if same:
                continue
            key = (carga_id, col)
            if key in expected_changes:
                expected_val = expected_changes[key]
                if not (f_val == expected_val or (mc.is_blank(f_val) and mc.is_blank(expected_val))):
                    unexpected_changes.append((carga_id, col, "valor distinto al autorizado", f_val, expected_val))
            else:
                unexpected_changes.append((carga_id, col, "cambio no autorizado", v2_val, f_val))

    for (carga_id, col), expected_val in expected_changes.items():
        final_row = final_by_carga.get(carga_id)
        if final_row is None:
            continue
        f_val = final_row[col]
        if not (f_val == expected_val or (mc.is_blank(f_val) and mc.is_blank(expected_val))):
            missing_expected.append((carga_id, col, expected_val, f_val))

    checks["campanas_unexpected_changes"] = len(unexpected_changes)
    checks["campanas_missing_expected_changes"] = len(missing_expected)
    checks["campanas_authorized_changes_count"] = len(expected_changes)
    if unexpected_changes:
        errors.append(f"CAMPANAS: {len(unexpected_changes)} celda(s) cambiaron sin autorización: {unexpected_changes[:10]}")
    if missing_expected:
        errors.append(f"CAMPANAS: {len(missing_expected)} edición(es) autorizada(s) no se aplicaron: {missing_expected[:10]}")

    # --- Referencial e integridad ---
    maestro_ids = set(v for v in [row["ElementoID"] for row in _full_row_map(wb_f["MAESTRO_ELEMENTOS"], "tblElementos").values()] if not mc.is_blank(v))
    orphans = [cid for cid, row in final_by_carga.items() if not mc.is_blank(row["ElementoID"]) and row["ElementoID"] not in maestro_ids]
    checks["elemento_id_orphans"] = len(orphans)
    if orphans:
        errors.append(f"CAMPANAS: {len(orphans)} ElementoID huérfano(s) en FINAL: {orphans[:10]}")

    import pandas as pd

    v2_cn = pd.Series([row["ClaveNegocio"] for row in v2_by_carga.values()]).dropna()
    final_cn = pd.Series([row["ClaveNegocio"] for row in final_by_carga.values()]).dropna()
    v2_dup_groups = set(v2_cn.value_counts()[v2_cn.value_counts() > 1].index)
    final_dup_groups = set(final_cn.value_counts()[final_cn.value_counts() > 1].index)
    new_dup_groups = final_dup_groups - v2_dup_groups
    lost_dup_groups = v2_dup_groups - final_dup_groups
    checks["clave_negocio_duplicate_groups_v2"] = len(v2_dup_groups)
    checks["clave_negocio_duplicate_groups_final"] = len(final_dup_groups)
    checks["clave_negocio_new_duplicate_groups"] = len(new_dup_groups)
    if new_dup_groups:
        errors.append(f"CAMPANAS: {len(new_dup_groups)} grupo(s) NUEVO(s) de ClaveNegocio duplicada: {sorted(new_dup_groups)[:10]}")
    if lost_dup_groups:
        errors.append(f"CAMPANAS: {len(lost_dup_groups)} grupo(s) duplicado(s) preexistente(s) desaparecieron: {sorted(lost_dup_groups)[:10]}")

    rows_final = {
        "MAESTRO_ELEMENTOS": len(_full_row_map(wb_f["MAESTRO_ELEMENTOS"], "tblElementos")),
        "CAMPANAS": len(final_by_carga),
        "PARAMETROS": len(_full_row_map(wb_f["PARAMETROS"], "tblParametros")),
    }

    wb_f.close()
    wb_t.close()
    wb_v2.close()

    sha_actual_after = mc.calculate_sha256(mc.ACTUAL_PATH)
    sha_nueva_after = mc.calculate_sha256(mc.NUEVA_PATH)
    sha_v2_after = mc.calculate_sha256(fc.CANDIDATE_V2_PATH)
    checks["actual_sha256_unchanged"] = sha_actual_after == sha_actual_before
    checks["nueva_sha256_unchanged"] = sha_nueva_after == sha_nueva_before
    checks["candidate_v2_sha256_unchanged"] = sha_v2_after == sha_v2_before
    if not checks["actual_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: input/OCU26_BASE_DATOS.xlsx cambió durante la validación FINAL")
    if not checks["nueva_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: OCU26_BASE_NUEVA_RECIBIDA.xlsx cambió durante la validación FINAL")
    if not checks["candidate_v2_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: la candidata V2 cambió durante la validación FINAL")

    checks["rows_final"] = rows_final

    return _finalize(errors, warnings, checks)


def _finalize(errors: list[str], warnings: list[str], checks: dict[str, Any]) -> dict[str, Any]:
    result = "INVALID" if errors else ("VALID_WITH_WARNINGS" if warnings else "VALID")
    return {"result": result, "errors": errors, "warnings": warnings, "checks": checks}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Valida la base FINAL OCU26 integrada.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = run_validation()
    except ValidationFailure as exc:
        if args.json:
            print(json.dumps({"result": "INVALID", "errors": [str(exc)], "warnings": [], "checks": {}}, ensure_ascii=False, indent=2))
        else:
            print("RESULT: INVALID")
            print(str(exc))
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 VALIDACION DE BASE FINAL")
        print("=" * 60)
        for k, v in result["checks"].items():
            print(f"  {k}: {v}")
        print()
        print("ERRORS:")
        for e in result["errors"]:
            print(f"  - {e}")
        if not result["errors"]:
            print("  none")
        print()
        print("RESULT:", result["result"])

    return 0 if result["result"] != "INVALID" else 1


if __name__ == "__main__":
    sys.exit(main())
