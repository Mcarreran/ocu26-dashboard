"""Validación obligatoria de la base FINAL_V2 OCU26.

Controles:
  - Gate 1 canónico (scripts/validate_input.py) sobre FINAL_V2.
  - Estructura canónica preservada.
  - Integridad ZIP/XLSX, ausencia de macros/vínculos externos/fórmulas/errores.
  - Apertura con data_only=False y data_only=True.
  - MAESTRO_ELEMENTOS y PARAMETROS: idénticos celda por celda a la base
    FINAL anterior.
  - CAMPANAS: el conjunto de celdas que cambiaron respecto de FINAL es
    EXACTAMENTE el de las correcciones aplicadas (recalculado de forma
    independiente vía merge_final_v2_common.verify_all_corrections), ni una
    celda de más ni de menos. Ningún CargaID perdido. Ningún ElementoID
    huérfano. Cero colisiones nuevas de ClaveNegocio. IDCampaña 4311/4336/
    4396/4470/4480/4771 quedan asociados exactamente a quien corresponde.
  - Hashes de ACTUAL, NUEVA, candidata V2 y base FINAL intactos.

Uso:
    python validate_ocu26_integrada_final_v2.py
    python validate_ocu26_integrada_final_v2.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

import merge_common as mc
import merge_final_common as fc
import merge_final_v2_common as fv2

vi = mc.vi

EXPECTED_ACTUAL_SHA256 = "2f165e12ad90c2f05963367a8e1717dcd1006e9ce669f976afa6e57470aca2cd"


class ValidationFailure(Exception):
    pass


def _full_row_map(ws, table_name: str) -> dict[int, dict[str, Any]]:
    t = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    grid: dict[int, dict[str, Any]] = {}
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True),
        start=min_row + 1,
    ):
        grid[r_idx] = dict(zip(headers, row))
    return grid


def run_validation() -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    checks: dict[str, Any] = {}

    if not fv2.FINAL_V2_PATH.exists():
        raise ValidationFailure(f"No existe la base FINAL_V2: {fv2.FINAL_V2_PATH}")

    sha_actual_before = mc.calculate_sha256(mc.ACTUAL_PATH)
    sha_nueva_before = mc.calculate_sha256(mc.NUEVA_PATH)
    sha_v2_before = mc.calculate_sha256(fc.CANDIDATE_V2_PATH)
    sha_final_before = mc.calculate_sha256(fv2.FINAL_PATH)
    checks["actual_sha256"] = sha_actual_before
    checks["actual_sha256_matches_expected_historico"] = sha_actual_before == EXPECTED_ACTUAL_SHA256
    if not checks["actual_sha256_matches_expected_historico"]:
        errors.append("El SHA-256 de input/OCU26_BASE_DATOS.xlsx no coincide con el esperado histórico")

    final_v2_validation = vi.validate_input(fv2.FINAL_V2_PATH)
    checks["final_v2_validate_input_result"] = final_v2_validation["result"]
    checks["final_v2_validate_input_errors"] = final_v2_validation["errors"]
    checks["final_v2_validate_input_warnings_count"] = len(final_v2_validation["warnings"])
    if final_v2_validation["result"] == "INVALID":
        errors.append(f"FINAL_V2 NO pasa validate_input.py (Gate 1): {final_v2_validation['errors']}")

    try:
        wb_f = load_workbook(fv2.FINAL_V2_PATH, data_only=False, read_only=False, keep_vba=False)
        wb_t = load_workbook(fv2.FINAL_V2_PATH, data_only=True, read_only=False, keep_vba=False)
        checks["opens_data_only_false"] = True
        checks["opens_data_only_true"] = True
    except Exception as exc:  # noqa: BLE001
        errors.append(f"FINAL_V2 no pudo abrirse con openpyxl: {exc}")
        return _finalize(errors, warnings, checks)

    if wb_f.sheetnames != mc.EXPECTED_SHEETS:
        errors.append(f"Hojas no coinciden: {wb_f.sheetnames} != {mc.EXPECTED_SHEETS}")
    for sheet_name, table_name in mc.EXPECTED_TABLES.items():
        tnames = list(wb_f[sheet_name].tables.keys())
        if tnames != [table_name]:
            errors.append(f"Hoja '{sheet_name}': tabla esperada '{table_name}', encontrada {tnames}")
    checks["structure_preserved"] = not errors

    final_v2_info = mc.inspect_structure(fv2.FINAL_V2_PATH)
    checks["zip_ok"] = final_v2_info["zip_ok"]
    checks["vba"] = final_v2_info["vba"]
    checks["external_links"] = final_v2_info["external_links"]
    checks["macro_enabled"] = final_v2_info["macro_enabled"]
    checks["formula_cells"] = final_v2_info["formula_cells"]
    checks["error_cells"] = final_v2_info["error_cells"]
    if not final_v2_info["zip_ok"]:
        errors.append("El contenedor ZIP de FINAL_V2 está corrupto")
    if final_v2_info["vba"]:
        errors.append("FINAL_V2 contiene macros")
    if final_v2_info["external_links"]:
        errors.append("FINAL_V2 contiene vínculos externos")
    if final_v2_info["macro_enabled"]:
        errors.append("FINAL_V2 está marcada como macro-habilitada")
    if final_v2_info["formula_cells"]:
        errors.append(f"FINAL_V2 contiene {final_v2_info['formula_cells']} celda(s) con fórmulas")
    if final_v2_info["error_cells"]:
        errors.append(f"FINAL_V2 contiene {final_v2_info['error_cells']} celda(s) con errores de Excel")

    wb_final = load_workbook(fv2.FINAL_PATH, data_only=False, read_only=False, keep_vba=False)

    for sheet_name, table_name in (("MAESTRO_ELEMENTOS", "tblElementos"), ("PARAMETROS", "tblParametros")):
        final_grid = _full_row_map(wb_final[sheet_name], table_name)
        v2_grid = _full_row_map(wb_f[sheet_name], table_name)
        diffs = 0
        if len(final_grid) != len(v2_grid):
            errors.append(f"{sheet_name}: cantidad de filas distinta (FINAL={len(final_grid)}, FINAL_V2={len(v2_grid)})")
        else:
            for r in final_grid:
                for col, f_val in final_grid[r].items():
                    n_val = v2_grid[r][col]
                    same = f_val == n_val or (mc.is_blank(f_val) and mc.is_blank(n_val))
                    if not same:
                        diffs += 1
        checks[f"{sheet_name.lower()}_cell_diffs_vs_final"] = diffs
        if diffs:
            errors.append(f"{sheet_name}: {diffs} celda(s) distinta(s) de la base FINAL anterior (debía ser 0)")

    final_campanas_grid = _full_row_map(wb_final["CAMPANAS"], "tblCampanas")
    v2_campanas_grid = _full_row_map(wb_f["CAMPANAS"], "tblCampanas")

    final_by_carga = {row[fv2.CAMPANAS_KEY]: row for row in final_campanas_grid.values() if not mc.is_blank(row[fv2.CAMPANAS_KEY])}
    v2_by_carga = {row[fv2.CAMPANAS_KEY]: row for row in v2_campanas_grid.values() if not mc.is_blank(row[fv2.CAMPANAS_KEY])}

    missing = set(final_by_carga) - set(v2_by_carga)
    extra = set(v2_by_carga) - set(final_by_carga)
    checks["campanas_missing_carga_ids"] = len(missing)
    checks["campanas_extra_carga_ids"] = len(extra)
    if missing:
        errors.append(f"CAMPANAS: {len(missing)} CargaID de FINAL ausentes en FINAL_V2: {sorted(missing)[:10]}")
    if extra:
        errors.append(f"CAMPANAS: {len(extra)} CargaID nuevos no autorizados en FINAL_V2: {sorted(extra)[:10]}")

    # --- Whitelist recalculado de forma independiente (no se confía en applied_log del build) ---
    campanas_df_final = mc.read_table_df(fv2.FINAL_PATH, "CAMPANAS", "tblCampanas")
    verification = fv2.verify_all_corrections(campanas_df_final)
    checks["corrections_status"] = {r["nombre"]: r["status"] for r in verification}

    expected_changes: dict[tuple[Any, str], Any] = {}
    for r in verification:
        if r["status"] != "APLICABLE":
            continue
        for col, new_val in r["changes"].items():
            expected_changes[(r["carga_id"], col)] = new_val

    unexpected_changes = []
    missing_expected = []
    for carga_id, final_row in final_by_carga.items():
        v2_row = v2_by_carga.get(carga_id)
        if v2_row is None:
            continue
        for col, f_val in final_row.items():
            if col == fv2.CAMPANAS_KEY:
                continue
            n_val = v2_row[col]
            same = f_val == n_val or (mc.is_blank(f_val) and mc.is_blank(n_val))
            if same:
                continue
            key = (carga_id, col)
            if key in expected_changes:
                expected_val = expected_changes[key]
                if not (n_val == expected_val or (mc.is_blank(n_val) and mc.is_blank(expected_val))):
                    unexpected_changes.append((carga_id, col, "valor distinto al autorizado", n_val, expected_val))
            else:
                unexpected_changes.append((carga_id, col, "cambio no autorizado", f_val, n_val))

    for (carga_id, col), expected_val in expected_changes.items():
        v2_row = v2_by_carga.get(carga_id)
        if v2_row is None:
            continue
        n_val = v2_row[col]
        if not (n_val == expected_val or (mc.is_blank(n_val) and mc.is_blank(expected_val))):
            missing_expected.append((carga_id, col, expected_val, n_val))

    checks["campanas_unexpected_changes"] = len(unexpected_changes)
    checks["campanas_missing_expected_changes"] = len(missing_expected)
    checks["campanas_authorized_changes_count"] = len(expected_changes)
    if unexpected_changes:
        errors.append(f"CAMPANAS: {len(unexpected_changes)} celda(s) cambiaron sin autorización: {unexpected_changes[:10]}")
    if missing_expected:
        errors.append(f"CAMPANAS: {len(missing_expected)} corrección(es) autorizada(s) no se aplicaron: {missing_expected[:10]}")

    # --- IDCampaña quedan asociados a quien corresponde ---
    import pandas as pd

    v2_campanas_df = mc.read_table_df(fv2.FINAL_V2_PATH, "CAMPANAS", "tblCampanas")
    id_ownership = {}
    for target_id, expected_brand in [(4311, "ypf"), (4336, "samsung"), (4396, "audi"), (4470, "beeyond"), (4480, "taggify"), (4771, "disney")]:
        sub = v2_campanas_df[v2_campanas_df["IDCampaña"] == target_id]
        cliente_vals = {str(v).strip().casefold() for v in sub["Cliente"] if not mc.is_blank(v)}
        marca_vals = {str(v).strip().casefold() for v in sub["Marca"] if not mc.is_blank(v)}
        all_vals = cliente_vals | marca_vals
        matches = any(expected_brand in v for v in all_vals)
        id_ownership[target_id] = {"filas": len(sub), "cliente_presentes": sorted(cliente_vals), "marca_presentes": sorted(marca_vals), "contiene_marca_esperada": matches}
        if not matches:
            errors.append(f"IDCampaña {target_id}: no se encontró la marca esperada ({expected_brand}) ni en Cliente {sorted(cliente_vals)} ni en Marca {sorted(marca_vals)}")
    checks["id_ownership"] = id_ownership

    # --- Referencial / duplicados ---
    maestro_ids = set(v for v in [row["ElementoID"] for row in _full_row_map(wb_f["MAESTRO_ELEMENTOS"], "tblElementos").values()] if not mc.is_blank(v))
    orphans = [cid for cid, row in v2_by_carga.items() if not mc.is_blank(row["ElementoID"]) and row["ElementoID"] not in maestro_ids]
    checks["elemento_id_orphans"] = len(orphans)
    if orphans:
        errors.append(f"CAMPANAS: {len(orphans)} ElementoID huérfano(s) en FINAL_V2: {orphans[:10]}")

    final_cn = pd.Series([row["ClaveNegocio"] for row in final_by_carga.values()]).dropna()
    v2_cn = pd.Series([row["ClaveNegocio"] for row in v2_by_carga.values()]).dropna()
    final_dup_groups = set(final_cn.value_counts()[final_cn.value_counts() > 1].index)
    v2_dup_groups = set(v2_cn.value_counts()[v2_cn.value_counts() > 1].index)
    new_dup_groups = v2_dup_groups - final_dup_groups

    # Los grupos de colisión que las correcciones autorizadas resuelven a
    # propósito (una fila se muda a otro IDCampaña, dejando de compartir
    # ClaveNegocio con la otra) deben desaparecer: es el resultado esperado,
    # no una pérdida indebida.
    expected_resolved_groups = {
        r["clave_negocio_anterior"] for r in verification if r["status"] == "APLICABLE"
    } & final_dup_groups
    lost_dup_groups = final_dup_groups - v2_dup_groups
    unexpected_lost_dup_groups = lost_dup_groups - expected_resolved_groups

    checks["clave_negocio_duplicate_groups_final"] = len(final_dup_groups)
    checks["clave_negocio_duplicate_groups_final_v2"] = len(v2_dup_groups)
    checks["clave_negocio_new_duplicate_groups"] = len(new_dup_groups)
    checks["clave_negocio_groups_resolved_by_corrections"] = sorted(expected_resolved_groups)
    checks["clave_negocio_unexpected_lost_groups"] = len(unexpected_lost_dup_groups)
    if new_dup_groups:
        errors.append(f"CAMPANAS: {len(new_dup_groups)} grupo(s) NUEVO(s) de ClaveNegocio duplicada: {sorted(new_dup_groups)[:10]}")
    if unexpected_lost_dup_groups:
        errors.append(f"CAMPANAS: {len(unexpected_lost_dup_groups)} grupo(s) duplicado(s) preexistente(s) desaparecieron sin autorización: {sorted(unexpected_lost_dup_groups)[:10]}")

    rows_final = {
        "MAESTRO_ELEMENTOS": len(_full_row_map(wb_f["MAESTRO_ELEMENTOS"], "tblElementos")),
        "CAMPANAS": len(v2_by_carga),
        "PARAMETROS": len(_full_row_map(wb_f["PARAMETROS"], "tblParametros")),
    }

    wb_f.close()
    wb_t.close()
    wb_final.close()

    sha_actual_after = mc.calculate_sha256(mc.ACTUAL_PATH)
    sha_nueva_after = mc.calculate_sha256(mc.NUEVA_PATH)
    sha_v2_after = mc.calculate_sha256(fc.CANDIDATE_V2_PATH)
    sha_final_after = mc.calculate_sha256(fv2.FINAL_PATH)
    checks["actual_sha256_unchanged"] = sha_actual_after == sha_actual_before
    checks["nueva_sha256_unchanged"] = sha_nueva_after == sha_nueva_before
    checks["candidate_v2_sha256_unchanged"] = sha_v2_after == sha_v2_before
    checks["final_v1_sha256_unchanged"] = sha_final_after == sha_final_before
    if not checks["actual_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: input/OCU26_BASE_DATOS.xlsx cambió durante la validación FINAL_V2")
    if not checks["nueva_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: OCU26_BASE_NUEVA_RECIBIDA.xlsx cambió durante la validación FINAL_V2")
    if not checks["candidate_v2_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: la candidata V2 cambió durante la validación FINAL_V2")
    if not checks["final_v1_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: la base FINAL anterior cambió durante la validación FINAL_V2")

    checks["rows_final"] = rows_final

    return _finalize(errors, warnings, checks)


def _finalize(errors: list[str], warnings: list[str], checks: dict[str, Any]) -> dict[str, Any]:
    result = "INVALID" if errors else ("VALID_WITH_WARNINGS" if warnings else "VALID")
    return {"result": result, "errors": errors, "warnings": warnings, "checks": checks}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Valida la base FINAL_V2 OCU26.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = run_validation()
    except ValidationFailure as exc:
        if args.json:
            print(json.dumps({"result": "INVALID", "errors": [str(exc)], "warnings": [], "checks": {}}, ensure_ascii=False, indent=2))
        else:
            print("RESULT: INVALID")
            print(str(exc))
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 VALIDACION DE BASE FINAL_V2")
        print("=" * 60)
        for k, v in result["checks"].items():
            print(f"  {k}: {v}")
        print()
        print("ERRORS:")
        for e in result["errors"]:
            print(f"  - {e}")
        if not result["errors"]:
            print("  none")
        print()
        print("RESULT:", result["result"])

    return 0 if result["result"] != "INVALID" else 1


if __name__ == "__main__":
    sys.exit(main())
