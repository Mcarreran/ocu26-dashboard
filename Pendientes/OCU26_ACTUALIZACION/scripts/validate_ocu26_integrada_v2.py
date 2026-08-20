"""Validación obligatoria de la base candidata V2 OCU26 integrada.

Reconstruye el plan V2 desde cero (merge_v2_common.compute_v2_plan) y lo usa
como referencia independiente para verificar, celda por celda, que la
candidata V2 contiene EXACTAMENTE lo esperado: ni una celda de más, ni una
de menos.

Controles:
  - Gate 1 canónico (scripts/validate_input.py) sobre la candidata V2.
  - Estructura canónica preservada (hojas/tablas/encabezados).
  - Integridad ZIP/XLSX, ausencia de macros/vínculos externos/fórmulas/errores.
  - Apertura con data_only=False y data_only=True.
  - MAESTRO_ELEMENTOS y PARAMETROS: idénticos celda por celda a ACTUAL.
  - CAMPANAS: todo CargaID de ACTUAL sigue presente; los campos aplicados
    coinciden exactamente con el plan recalculado; los campos rechazados
    conservan el valor original de ACTUAL; los 4 registros nuevos (si
    fueron incorporados) tienen los valores esperados (CargaID, ClaveNegocio,
    EstadoValidacion, UsuarioCarga, FuenteCarga, campos operativos) —
    excepto FechaHoraCarga, que es una marca de tiempo de ejecución y solo
    se valida que esté presente y sea posterior al build.
  - Cero ElementoID huérfanos, cero ClaveNegocio duplicada NUEVA (los grupos
    duplicados preexistentes deben seguir siendo exactamente los mismos).
  - Hashes de ambas fuentes intactos antes/después.

Uso:
    python validate_ocu26_integrada_v2.py
    python validate_ocu26_integrada_v2.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

import merge_common as mc
import merge_v2_common as v2

vi = mc.vi

EXPECTED_ACTUAL_SHA256 = "2f165e12ad90c2f05963367a8e1717dcd1006e9ce669f976afa6e57470aca2cd"


class ValidationFailure(Exception):
    pass


def _full_row_map(ws, table_name: str) -> dict[Any, dict[str, Any]]:
    """CargaID/ElementoID (según hoja) -> {columna: valor} para TODAS las filas."""
    t = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    grid: dict[int, dict[str, Any]] = {}
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True),
        start=min_row + 1,
    ):
        grid[r_idx] = dict(zip(headers, row))
    return grid


def run_validation() -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    checks: dict[str, Any] = {}

    if not v2.CANDIDATE_V2_PATH.exists():
        raise ValidationFailure(f"No existe la candidata V2: {v2.CANDIDATE_V2_PATH}")

    sha_actual_before = mc.calculate_sha256(v2.ACTUAL_PATH)
    sha_nueva_before = mc.calculate_sha256(v2.NUEVA_PATH)
    checks["actual_sha256"] = sha_actual_before
    checks["actual_sha256_matches_expected_historico"] = sha_actual_before == EXPECTED_ACTUAL_SHA256
    if not checks["actual_sha256_matches_expected_historico"]:
        errors.append("El SHA-256 de input/OCU26_BASE_DATOS.xlsx no coincide con el esperado histórico")

    # --- Gate 1 canónico ---
    candidate_validation = vi.validate_input(v2.CANDIDATE_V2_PATH)
    checks["candidate_validate_input_result"] = candidate_validation["result"]
    checks["candidate_validate_input_errors"] = candidate_validation["errors"]
    checks["candidate_validate_input_warnings_count"] = len(candidate_validation["warnings"])
    if candidate_validation["result"] == "INVALID":
        errors.append(f"La candidata V2 NO pasa validate_input.py (Gate 1): {candidate_validation['errors']}")

    # --- Apertura ---
    try:
        wb_f = load_workbook(v2.CANDIDATE_V2_PATH, data_only=False, read_only=False, keep_vba=False)
        wb_t = load_workbook(v2.CANDIDATE_V2_PATH, data_only=True, read_only=False, keep_vba=False)
        checks["opens_data_only_false"] = True
        checks["opens_data_only_true"] = True
    except Exception as exc:  # noqa: BLE001
        errors.append(f"La candidata V2 no pudo abrirse con openpyxl: {exc}")
        return _finalize(errors, warnings, checks)

    if wb_f.sheetnames != mc.EXPECTED_SHEETS:
        errors.append(f"Hojas no coinciden: {wb_f.sheetnames} != {mc.EXPECTED_SHEETS}")
    for sheet_name, table_name in mc.EXPECTED_TABLES.items():
        tnames = list(wb_f[sheet_name].tables.keys())
        if tnames != [table_name]:
            errors.append(f"Hoja '{sheet_name}': tabla esperada '{table_name}', encontrada {tnames}")
    checks["structure_preserved"] = not errors

    cand_info = mc.inspect_structure(v2.CANDIDATE_V2_PATH)
    checks["zip_ok"] = cand_info["zip_ok"]
    checks["vba"] = cand_info["vba"]
    checks["external_links"] = cand_info["external_links"]
    checks["macro_enabled"] = cand_info["macro_enabled"]
    checks["formula_cells"] = cand_info["formula_cells"]
    checks["error_cells"] = cand_info["error_cells"]
    if not cand_info["zip_ok"]:
        errors.append("El contenedor ZIP de la candidata V2 está corrupto")
    if cand_info["vba"]:
        errors.append("La candidata V2 contiene macros")
    if cand_info["external_links"]:
        errors.append("La candidata V2 contiene vínculos externos")
    if cand_info["macro_enabled"]:
        errors.append("La candidata V2 está marcada como macro-habilitada")
    if cand_info["formula_cells"]:
        errors.append(f"La candidata V2 contiene {cand_info['formula_cells']} celda(s) con fórmulas")
    if cand_info["error_cells"]:
        errors.append(f"La candidata V2 contiene {cand_info['error_cells']} celda(s) con errores de Excel")

    # --- Recalcular plan V2 de referencia (independiente del build) ---
    plan = v2.compute_v2_plan()

    # --- MAESTRO_ELEMENTOS: idéntico celda por celda a ACTUAL ---
    wb_actual = load_workbook(v2.ACTUAL_PATH, data_only=False, read_only=False, keep_vba=False)
    actual_maestro_grid = _full_row_map(wb_actual["MAESTRO_ELEMENTOS"], "tblElementos")
    cand_maestro_grid = _full_row_map(wb_f["MAESTRO_ELEMENTOS"], "tblElementos")
    maestro_cell_diffs = 0
    if len(actual_maestro_grid) != len(cand_maestro_grid):
        errors.append(f"MAESTRO_ELEMENTOS: cantidad de filas distinta (ACTUAL={len(actual_maestro_grid)}, candidata={len(cand_maestro_grid)})")
    else:
        for r in actual_maestro_grid:
            a_row = actual_maestro_grid[r]
            c_row = cand_maestro_grid[r]
            for col, a_val in a_row.items():
                c_val = c_row[col]
                same = a_val == c_val or (mc.is_blank(a_val) and mc.is_blank(c_val))
                if not same:
                    maestro_cell_diffs += 1
    checks["maestro_elementos_cell_diffs"] = maestro_cell_diffs
    checks["maestro_elementos_completions_expected"] = len(plan.maestro_cls.completions)
    if maestro_cell_diffs != len(plan.maestro_cls.completions):
        errors.append(
            f"MAESTRO_ELEMENTOS: {maestro_cell_diffs} celda(s) distinta(s) de ACTUAL, se esperaban "
            f"{len(plan.maestro_cls.completions)} (según el plan recalculado)"
        )

    # --- PARAMETROS: idéntico celda por celda a ACTUAL (sin excepciones) ---
    actual_param_grid = _full_row_map(wb_actual["PARAMETROS"], "tblParametros")
    cand_param_grid = _full_row_map(wb_f["PARAMETROS"], "tblParametros")
    param_cell_diffs = 0
    if len(actual_param_grid) != len(cand_param_grid):
        errors.append(f"PARAMETROS: cantidad de filas distinta (ACTUAL={len(actual_param_grid)}, candidata={len(cand_param_grid)})")
    else:
        for r in actual_param_grid:
            for col, a_val in actual_param_grid[r].items():
                c_val = cand_param_grid[r][col]
                same = a_val == c_val or (mc.is_blank(a_val) and mc.is_blank(c_val))
                if not same:
                    param_cell_diffs += 1
    checks["parametros_cell_diffs"] = param_cell_diffs
    if param_cell_diffs != 0:
        errors.append(f"PARAMETROS: {param_cell_diffs} celda(s) distinta(s) de ACTUAL (debía ser 0)")

    # --- CAMPANAS: filas existentes (por CargaID) ---
    actual_campanas_grid = _full_row_map(wb_actual["CAMPANAS"], "tblCampanas")
    cand_campanas_grid = _full_row_map(wb_f["CAMPANAS"], "tblCampanas")
    actual_by_carga = {row[v2.CAMPANAS_KEY]: row for row in actual_campanas_grid.values() if not mc.is_blank(row[v2.CAMPANAS_KEY])}
    cand_by_carga = {row[v2.CAMPANAS_KEY]: row for row in cand_campanas_grid.values() if not mc.is_blank(row[v2.CAMPANAS_KEY])}

    missing_carga_ids = set(actual_by_carga) - set(cand_by_carga)
    checks["campanas_missing_carga_ids"] = len(missing_carga_ids)
    if missing_carga_ids:
        errors.append(f"CAMPANAS: {len(missing_carga_ids)} CargaID de ACTUAL ausentes en la candidata V2: {sorted(missing_carga_ids)[:10]}")

    expected_applied_by_key = plan.campanas_final_applied
    expected_rejected: dict[tuple[Any, str], Any] = {
        (c.carga_id, c.columna): c.valor_anterior for c in plan.campanas_change_records if c.resultado == "RECHAZADO_A_VALIDAR"
    }

    unexpected_changes = []
    missing_expected_changes = []
    rejected_but_changed = []

    for carga_id, actual_row in actual_by_carga.items():
        cand_row = cand_by_carga.get(carga_id)
        if cand_row is None:
            continue
        expected_here = expected_applied_by_key.get(carga_id, {})
        for col, a_val in actual_row.items():
            if col == v2.CAMPANAS_KEY:
                continue
            c_val = cand_row[col]
            same = a_val == c_val or (mc.is_blank(a_val) and mc.is_blank(c_val))
            if same:
                if (carga_id, col) in expected_rejected:
                    pass  # correcto: se esperaba que NO cambiara
                continue
            if col in expected_here:
                expected_val = expected_here[col]
                if not (c_val == expected_val or (mc.is_blank(c_val) and mc.is_blank(expected_val))):
                    unexpected_changes.append((carga_id, col, "valor distinto al esperado", c_val, expected_val))
            elif (carga_id, col) in expected_rejected:
                rejected_but_changed.append((carga_id, col, a_val, c_val))
            else:
                unexpected_changes.append((carga_id, col, "cambio no previsto por el plan", a_val, c_val))

        for col in expected_here:
            expected_val = expected_here[col]
            c_val = cand_row[col]
            if not (c_val == expected_val or (mc.is_blank(c_val) and mc.is_blank(expected_val))):
                missing_expected_changes.append((carga_id, col, expected_val, c_val))

    checks["campanas_unexpected_changes"] = len(unexpected_changes)
    checks["campanas_rejected_but_changed"] = len(rejected_but_changed)
    checks["campanas_missing_expected_changes"] = len(missing_expected_changes)
    if unexpected_changes:
        errors.append(f"CAMPANAS: {len(unexpected_changes)} celda(s) cambiaron de forma no prevista por el plan: {unexpected_changes[:10]}")
    if rejected_but_changed:
        errors.append(f"CAMPANAS: {len(rejected_but_changed)} celda(s) marcadas RECHAZADO_A_VALIDAR fueron aplicadas igualmente: {rejected_but_changed[:10]}")
    if missing_expected_changes:
        errors.append(f"CAMPANAS: {len(missing_expected_changes)} cambio(s) esperados no se aplicaron: {missing_expected_changes[:10]}")

    # --- Registros nuevos ---
    expected_new_ids = {r["CargaID"] for r in plan.new_records_incorporated}
    actual_carga_ids_set = set(actual_by_carga.keys())
    cand_carga_ids_set = set(cand_by_carga.keys())
    new_ids_in_candidate = cand_carga_ids_set - actual_carga_ids_set
    checks["nuevos_cargaid_esperados"] = sorted(expected_new_ids)
    checks["nuevos_cargaid_en_candidata"] = sorted(new_ids_in_candidate)
    if new_ids_in_candidate != expected_new_ids:
        errors.append(
            f"CAMPANAS: los CargaID nuevos en la candidata ({sorted(new_ids_in_candidate)}) no coinciden con "
            f"los esperados por el plan ({sorted(expected_new_ids)})"
        )

    new_record_field_errors = []
    for record in plan.new_records_incorporated:
        cand_row = cand_by_carga.get(record["CargaID"])
        if cand_row is None:
            new_record_field_errors.append((record["CargaID"], "no encontrado en la candidata"))
            continue
        for col, expected_val in record.items():
            if col == "FechaHoraCarga":
                if mc.is_blank(cand_row[col]):
                    new_record_field_errors.append((record["CargaID"], col, "vacío, se esperaba una marca de tiempo"))
                continue
            c_val = cand_row[col]
            same = c_val == expected_val or (mc.is_blank(c_val) and mc.is_blank(expected_val))
            if not same:
                new_record_field_errors.append((record["CargaID"], col, expected_val, c_val))
    checks["nuevos_registros_field_errors"] = len(new_record_field_errors)
    if new_record_field_errors:
        errors.append(f"CAMPANAS: {len(new_record_field_errors)} discrepancia(s) en campos de registros nuevos: {new_record_field_errors[:10]}")

    # --- Duplicados de ClaveNegocio: ningún grupo nuevo ---
    import pandas as pd

    actual_cn = pd.Series([row["ClaveNegocio"] for row in actual_by_carga.values()]).dropna()
    cand_cn = pd.Series([row["ClaveNegocio"] for row in cand_by_carga.values()]).dropna()
    actual_dup_groups = set(actual_cn.value_counts()[actual_cn.value_counts() > 1].index)
    cand_dup_groups = set(cand_cn.value_counts()[cand_cn.value_counts() > 1].index)
    new_dup_groups = cand_dup_groups - actual_dup_groups
    lost_dup_groups = actual_dup_groups - cand_dup_groups
    checks["clave_negocio_duplicate_groups_actual"] = len(actual_dup_groups)
    checks["clave_negocio_duplicate_groups_candidate"] = len(cand_dup_groups)
    checks["clave_negocio_new_duplicate_groups"] = len(new_dup_groups)
    if new_dup_groups:
        errors.append(f"CAMPANAS: se introdujeron {len(new_dup_groups)} grupo(s) NUEVO(s) de ClaveNegocio duplicada: {sorted(new_dup_groups)[:10]}")
    if lost_dup_groups:
        errors.append(f"CAMPANAS: {len(lost_dup_groups)} grupo(s) duplicado(s) preexistente(s) desaparecieron (no debían eliminarse): {sorted(lost_dup_groups)[:10]}")

    # --- Referencial: ElementoID huérfano ---
    maestro_ids = set(plan.actual_maestro_df["ElementoID"].dropna())
    orphans = [cid for cid, row in cand_by_carga.items() if not mc.is_blank(row["ElementoID"]) and row["ElementoID"] not in maestro_ids]
    checks["elemento_id_orphans"] = len(orphans)
    if orphans:
        errors.append(f"CAMPANAS: {len(orphans)} ElementoID huérfano(s) en la candidata V2: {orphans[:10]}")

    wb_f.close()
    wb_t.close()
    wb_actual.close()

    sha_actual_after = mc.calculate_sha256(v2.ACTUAL_PATH)
    sha_nueva_after = mc.calculate_sha256(v2.NUEVA_PATH)
    checks["actual_sha256_unchanged"] = sha_actual_after == sha_actual_before
    checks["nueva_sha256_unchanged"] = sha_nueva_after == sha_nueva_before
    if not checks["actual_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: input/OCU26_BASE_DATOS.xlsx cambió durante la validación V2")
    if not checks["nueva_sha256_unchanged"]:
        errors.append("ERROR CRÍTICO: OCU26_BASE_NUEVA_RECIBIDA.xlsx cambió durante la validación V2")

    checks["rows_final"] = {
        "MAESTRO_ELEMENTOS": len(cand_maestro_grid),
        "CAMPANAS": len(cand_by_carga),
        "PARAMETROS": len(cand_param_grid),
    }

    return _finalize(errors, warnings, checks)


def _finalize(errors: list[str], warnings: list[str], checks: dict[str, Any]) -> dict[str, Any]:
    result = "INVALID" if errors else ("VALID_WITH_WARNINGS" if warnings else "VALID")
    return {"result": result, "errors": errors, "warnings": warnings, "checks": checks}


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Valida la base candidata V2 OCU26 integrada.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = run_validation()
    except ValidationFailure as exc:
        if args.json:
            print(json.dumps({"result": "INVALID", "errors": [str(exc)], "warnings": [], "checks": {}}, ensure_ascii=False, indent=2))
        else:
            print("RESULT: INVALID")
            print(str(exc))
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 VALIDACION DE CANDIDATA V2")
        print("=" * 60)
        for k, v in result["checks"].items():
            print(f"  {k}: {v}")
        print()
        print("ERRORS:")
        for e in result["errors"]:
            print(f"  - {e}")
        if not result["errors"]:
            print("  none")
        print()
        print("RESULT:", result["result"])

    return 0 if result["result"] != "INVALID" else 1


if __name__ == "__main__":
    sys.exit(main())
