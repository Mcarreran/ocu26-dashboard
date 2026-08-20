# -*- coding: utf-8 -*-
"""Genera la auditoría de integración YPF -> OCU26 (READ-ONLY sobre las 3
fuentes; no construye ni modifica la candidata).

Hojas generadas en OCU26_AUDITORIA_INTEGRACION_YPF_2026-08-18.xlsx:
  RESUMEN, MAPEO_COLUMNAS, ELEMENTOS_YA_EXISTENTES, ELEMENTOS_NUEVOS,
  CONFLICTOS_ELEMENTOS, CAMPAÑAS_YA_EXISTENTES, CAMPAÑAS_NUEVAS,
  CONFLICTOS_CAMPAÑAS, DUPLICADOS_EVITADOS, APIE_30943,
  INTEGRIDAD_REFERENCIAL, A_VALIDAR

Uso:
    python audit_ocu26_ypf.py
    python audit_ocu26_ypf.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from typing import Any

import pandas as pd
from openpyxl import Workbook

import merge_ypf_common as m


def _autosize(ws, df: pd.DataFrame) -> None:
    for i, col in enumerate(df.columns, start=1):
        try:
            max_len = max([len(str(col))] + [len(str(v)) for v in df[col].head(500)])
        except Exception:
            max_len = 12
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max(max_len + 2, 10), 60)


def _write_df(wb: Workbook, sheet_name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([str(v) if isinstance(v, (dict, list)) else v for v in row])
    _autosize(ws, df)


def build_audit(plan: m.YpfPlan) -> dict[str, pd.DataFrame]:
    sheets: dict[str, pd.DataFrame] = {}

    # --- RESUMEN ---
    ec = plan.expected_counts
    resumen_rows = [
        ("Fecha de generación", plan.run_timestamp.isoformat()),
        ("--- FUENTES ---", ""),
        ("OCU26 FINAL_V2 (ruta)", str(m.FINAL_V2_PATH)),
        ("OCU26 FINAL_V2 (SHA-256)", m.calculate_sha256(m.FINAL_V2_PATH)),
        ("YPF Etapa 1 (ruta)", str(m.YPF_ETAPA1_PATH)),
        ("YPF Etapa 1 (SHA-256)", m.calculate_sha256(m.YPF_ETAPA1_PATH)),
        ("YPF Etapa 2 (ruta)", str(m.YPF_ETAPA2_PATH)),
        ("YPF Etapa 2 (SHA-256)", m.calculate_sha256(m.YPF_ETAPA2_PATH)),
        ("--- MAESTRO_ELEMENTOS ---", ""),
        ("MAESTRO_ELEMENTOS antes", ec["MAESTRO_ELEMENTOS_antes"]),
        ("YPF BASE ELEMENTOS (catálogo autorizado)", len(plan.elementos.ypf_autorizados)),
        ("Elementos YPF ya existentes (idénticos)", len(plan.elementos.identicos)),
        ("Elementos YPF ya existentes (con datos completados)", len(plan.elementos.con_completions)),
        ("Elementos YPF ya existentes (con conflicto documentado)", len(plan.elementos.con_conflictos)),
        ("Elementos YPF nuevos agregados", ec["MAESTRO_ELEMENTOS_nuevos"]),
        ("Elementos solo-OCU26 (no tocados)", len(plan.elementos.solo_ocu26)),
        ("ElementoID YPF inválidos (no matchean patrón)", len(plan.elementos.ypf_elementoid_invalido)),
        ("ElementoID YPF duplicados dentro de BASE ELEMENTOS", len(plan.elementos.ypf_duplicados)),
        ("Campos completados (total celdas)", len(plan.elementos.completions)),
        ("Conflictos documentados (total celdas)", len(plan.elementos.conflicts)),
        ("MAESTRO_ELEMENTOS después", ec["MAESTRO_ELEMENTOS_despues"]),
        ("--- CAMPANAS (reemplazo de bloque autorizado) ---", ""),
        ("CAMPANAS antes (FINAL_V2)", ec["CAMPANAS_antes"]),
        ("Bloque legacy YPF retirado (IDCampaña 10000-10009)", ec["CAMPANAS_legacy_retiradas"]),
        ("  de las cuales REEMPLAZADA_POR_YPF_ETAPA2_EXACTA", sum(1 for r in plan.legacy_rows if r.categoria == "REEMPLAZADA_POR_YPF_ETAPA2_EXACTA")),
        ("  de las cuales LEGACY_EXCLUIDA_POR_REGLA", sum(1 for r in plan.legacy_rows if r.categoria == "LEGACY_EXCLUIDA_POR_REGLA")),
        ("  de las cuales LEGACY_FUERA_CATALOGO", sum(1 for r in plan.legacy_rows if r.categoria == "LEGACY_FUERA_CATALOGO")),
        ("  de las cuales LEGACY_APIE_30943_DIGITAL_INVALIDA", sum(1 for r in plan.legacy_rows if r.categoria == "LEGACY_APIE_30943_DIGITAL_INVALIDA")),
        ("  de las cuales LEGACY_YPF_NO_RECONCILIADA", sum(1 for r in plan.legacy_rows if r.categoria == "LEGACY_YPF_NO_RECONCILIADA")),
        ("  de las cuales LEGACY_ELEMENTO_INVALIDO", sum(1 for r in plan.legacy_rows if r.categoria == "LEGACY_ELEMENTO_INVALIDO")),
        ("Campañas no-YPF protegidas (sin cambios)", ec["CAMPANAS_protegidas_no_ypf"]),
        ("Campañas YPF Etapa 2 recibidas (BASE CAMPAÑAS)", len(plan.ypf_camp_df)),
        ("Campañas YPF Etapa 2 insertadas", ec["CAMPANAS_ypf_insertadas"]),
        ("Campañas YPF Etapa 2 rechazadas (A_VALIDAR)", ec["CAMPANAS_ypf_rechazadas"]),
        ("CAMPANAS después (resultado de la unión, no de la suma)", ec["CAMPANAS_despues"]),
        ("Fórmula de verificación", f"{ec['CAMPANAS_antes']} - {ec['CAMPANAS_legacy_retiradas']} + {ec['CAMPANAS_ypf_insertadas']} = {ec['CAMPANAS_despues']}"),
        ("--- PARAMETROS ---", ""),
        ("PARAMETROS (sin cambios)", ec["PARAMETROS"]),
    ]
    sheets["RESUMEN"] = pd.DataFrame(resumen_rows, columns=["Campo", "Valor"])

    # --- MAPEO_COLUMNAS ---
    mapeo_rows = []
    for h in m.MAESTRO_HEADERS:
        en_ypf = h in plan.ypf_elem_df.columns and h not in m.MAESTRO_COLUMNS_NOT_IN_YPF
        mapeo_rows.append({
            "Hoja": "MAESTRO_ELEMENTOS", "ColumnaOCU26": h,
            "ColumnaFuenteYPF": h if en_ypf else None,
            "Clasificacion": "MAPEO_DIRECTO" if en_ypf else "SIN_DATO_YPF (queda en blanco para elementos nuevos)",
        })
    campanas_map = {
        "TipoCargaDeclarado": "Medio", "Observaciones": "ObservacionesComercial",
        "FechaIndefinida": None, "ModalidadPauta": None, "PROGRAMATICA": None,
        "TipoExclusividad": None, "HoraInicio": None, "HoraFin": None, "CANJE": None, "FilaOrigen": None,
        "CargaID": None, "ClaveNegocio": None, "FechaHoraCarga": None, "UsuarioCarga": None,
        "FuenteCarga": None, "EstadoValidacion": None, "ObservacionValidacion": None,
    }
    for h in m.CAMPANAS_HEADERS:
        if h in campanas_map:
            src = campanas_map[h]
            clasif = "MAPEO_RENOMBRADO" if src else "GENERADO_TECNICO" if h in {"CargaID","ClaveNegocio","FechaHoraCarga","UsuarioCarga","FuenteCarga","EstadoValidacion"} else "SIN_DATO_YPF (queda en blanco)"
        elif h in plan.ypf_camp_df.columns:
            src, clasif = h, "MAPEO_DIRECTO"
        else:
            src, clasif = None, "SIN_DATO_YPF (queda en blanco)"
        mapeo_rows.append({"Hoja": "CAMPANAS", "ColumnaOCU26": h, "ColumnaFuenteYPF": src, "Clasificacion": clasif})
    sheets["MAPEO_COLUMNAS"] = pd.DataFrame(mapeo_rows)

    # --- ELEMENTOS_YA_EXISTENTES ---
    rows = []
    for eid in sorted(plan.elementos.comunes, key=str):
        if eid in plan.elementos.identicos:
            tag = "YA_EXISTE_IDENTICO"
        elif eid in plan.elementos.con_completions:
            tag = "YA_EXISTE_CON_DATOS_FALTANTES"
        else:
            tag = "CONFLICTO"
        rows.append({"ElementoID": eid, "Clasificacion": tag})
    sheets["ELEMENTOS_YA_EXISTENTES"] = pd.DataFrame(rows)

    # --- ELEMENTOS_NUEVOS ---
    sheets["ELEMENTOS_NUEVOS"] = pd.DataFrame(plan.elementos.nuevos_registros)

    # --- CONFLICTOS_ELEMENTOS ---
    sheets["CONFLICTOS_ELEMENTOS"] = pd.DataFrame(
        [{"ElementoID": c.elemento_id, "Columna": c.columna, "ValorOCU26": c.valor_ocu26, "ValorYPF": c.valor_ypf}
         for c in plan.elementos.conflicts]
    )

    # --- CAMPAÑAS_YA_EXISTENTES (bloque legacy reemplazado, con motivo) ---
    sheets["CAMPAÑAS_YA_EXISTENTES"] = pd.DataFrame(
        [{
            "CargaID_retirado": r.carga_id, "IDCampaña_retirado": r.id_campana, "Campaña": r.campana,
            "ElementoID": r.elemento_id, "FechaInicio": r.fecha_inicio, "FechaFin": r.fecha_fin,
            "Categoria": r.categoria, "Motivo": r.motivo,
        } for r in plan.legacy_rows]
    )

    # --- CAMPAÑAS_NUEVAS ---
    sheets["CAMPAÑAS_NUEVAS"] = pd.DataFrame(plan.new_campanas_incorporated)

    # --- CONFLICTOS_CAMPAÑAS (rechazadas / A_VALIDAR) ---
    sheets["CONFLICTOS_CAMPAÑAS"] = pd.DataFrame(plan.new_campanas_rejected)

    # --- DUPLICADOS_EVITADOS ---
    dup_rows = []
    for eid in sorted(plan.elementos.identicos, key=str):
        dup_rows.append({"Tipo": "ELEMENTO_IDENTICO_NO_DUPLICADO", "Clave": eid, "Detalle": "ElementoID ya existía idéntico en OCU26; no se duplicó."})
    for k, v in plan.elementos.ypf_duplicados.items():
        dup_rows.append({"Tipo": "ELEMENTOID_DUPLICADO_EN_FUENTE_YPF", "Clave": k, "Detalle": f"Aparece {v} veces en BASE ELEMENTOS; excluido del cruce por ambigüedad."})
    sheets["DUPLICADOS_EVITADOS"] = pd.DataFrame(dup_rows)

    # --- APIE_30943 ---
    apie_maestro = plan.ocu_maestro_df[plan.ocu_maestro_df[m.MAESTRO_KEY].astype(str).str.startswith("30943 - ")]
    apie_rows = []
    for _, r in apie_maestro.iterrows():
        eid = r[m.MAESTRO_KEY]
        tipo = m.elemento_id_tipo(eid)
        origen = "YPF_AUTORIZADO" if eid in plan.elementos.ypf_autorizados else "PREEXISTENTE_OCU26_NO_YPF"
        apie_rows.append({"ElementoID": eid, "Tipo": tipo, "Medio": r["Medio"], "Origen": origen})
    for r in plan.legacy_rows:
        if r.categoria == "LEGACY_APIE_30943_DIGITAL_INVALIDA":
            apie_rows.append({"ElementoID": r.elemento_id, "Tipo": m.elemento_id_tipo(r.elemento_id), "Medio": "Digital (campaña legacy retirada)", "Origen": f"CAMPANA_LEGACY_RETIRADA CargaID={r.carga_id}"})
    sheets["APIE_30943"] = pd.DataFrame(apie_rows)

    # --- INTEGRIDAD_REFERENCIAL ---
    maestro_ids_final = set(plan.ocu_maestro_df[m.MAESTRO_KEY].dropna()) | plan.elementos.solo_ypf
    new_orphans = [r for r in plan.new_campanas_incorporated if r["ElementoID"] not in maestro_ids_final]
    remaining_orphans = plan.remaining_campanas_df[~plan.remaining_campanas_df["ElementoID"].isin(maestro_ids_final) & plan.remaining_campanas_df["ElementoID"].notna()]
    integridad_rows = [
        {"Control": "CAMPANAS nuevas con ElementoID inexistente en MAESTRO_ELEMENTOS final", "Cantidad": len(new_orphans)},
        {"Control": "CAMPANAS protegidas (no-YPF) con ElementoID inexistente en MAESTRO_ELEMENTOS final", "Cantidad": len(remaining_orphans)},
        {"Control": "CAMPANAS nuevas digitales sobre ElementoID FB", "Cantidad": sum(1 for r in plan.new_campanas_incorporated if m.elemento_id_tipo(r["ElementoID"]) == "FB")},
        {"Control": "CAMPANAS nuevas digitales sobre APIE 30943", "Cantidad": sum(1 for r in plan.new_campanas_incorporated if m.es_elemento_digital_30943(r["ElementoID"]))},
        {"Control": "ElementoID vacío o duplicado en MAESTRO_ELEMENTOS final", "Cantidad": 0 if not plan.elementos.ypf_duplicados else len(plan.elementos.ypf_duplicados)},
    ]
    sheets["INTEGRIDAD_REFERENCIAL"] = pd.DataFrame(integridad_rows)

    # --- A_VALIDAR ---
    a_validar_rows = list(plan.elementos.ypf_elementoid_invalido)
    for k, v in plan.elementos.ypf_duplicados.items():
        a_validar_rows.append({"tipo": "ELEMENTOID_DUPLICADO", "ElementoID": k, "ocurrencias": v})
    for r in plan.new_campanas_rejected:
        a_validar_rows.append({"tipo": "CAMPANA_RECHAZADA", **r})
    for r in plan.legacy_rows:
        if r.categoria in ("LEGACY_YPF_NO_RECONCILIADA", "LEGACY_ELEMENTO_INVALIDO"):
            a_validar_rows.append({
                "tipo": r.categoria, "CargaID": r.carga_id, "IDCampaña": r.id_campana,
                "ElementoID": r.elemento_id, "Motivo": r.motivo,
            })
    sheets["A_VALIDAR"] = pd.DataFrame(a_validar_rows)

    return sheets


def write_audit_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    order = [
        "RESUMEN", "MAPEO_COLUMNAS", "ELEMENTOS_YA_EXISTENTES", "ELEMENTOS_NUEVOS",
        "CONFLICTOS_ELEMENTOS", "CAMPAÑAS_YA_EXISTENTES", "CAMPAÑAS_NUEVAS",
        "CONFLICTOS_CAMPAÑAS", "DUPLICADOS_EVITADOS", "APIE_30943",
        "INTEGRIDAD_REFERENCIAL", "A_VALIDAR",
    ]
    for name in order:
        df = sheets.get(name, pd.DataFrame())
        if df.empty and list(df.columns) == []:
            df = pd.DataFrame([{"info": "sin filas"}])
        _write_df(wb, name, df)
    m.AUDIT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(m.AUDIT_PATH)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Genera la auditoría de integración YPF -> OCU26.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    plan = m.compute_plan()
    sheets = build_audit(plan)
    write_audit_workbook(sheets)

    result = {
        "result": "AUDIT_YPF_OK",
        "audit_path": str(m.AUDIT_PATH),
        "expected_counts": plan.expected_counts,
        "sheets": {k: len(v) for k, v in sheets.items()},
    }

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 + YPF AUDITORÍA")
        print("=" * 60)
        print(f"Ruta: {result['audit_path']}")
        for k, v in result["expected_counts"].items():
            print(f"{k}: {v}")
        print()
        for k, v in result["sheets"].items():
            print(f"  hoja {k}: {v} filas")

    return 0


if __name__ == "__main__":
    sys.exit(main())
