# -*- coding: utf-8 -*-
"""Construye la base candidata OCU26 + YPF (no reemplaza input/OCU26_BASE_DATOS.xlsx).

Fuentes (READ-ONLY, nunca se escriben):
  - OCU26 FINAL_V2 (Pendientes/OCU26_ACTUALIZACION/output/...)
  - YPF Etapa 1 (catálogo de elementos)
  - YPF Etapa 2 (campañas validadas)

Operación (ver merge_ypf_common.py para el detalle de cada regla):
  1. MAESTRO_ELEMENTOS: completa campos vacíos de los 2.900 elementos YPF ya
     existentes (nunca sobrescribe no-vacíos), agrega los 983 elementos
     solo-YPF adaptados al esquema OCU26. Nunca borra nada.
  2. CAMPANAS: retira las 7.790 filas del bloque histórico YPF
     [IDCampaña 10000-10009] (decisión de negocio autorizada), conserva
     intactas las 1.717 filas no-YPF restantes, inserta las 13.616 filas de
     BASE CAMPAÑAS (YPF Etapa 2) con CargaID/ClaveNegocio generados.
  3. PARAMETROS: sin cambios.

Uso:
    python build_ocu26_ypf.py
    python build_ocu26_ypf.py --json
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import sys
from copy import copy
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

import merge_ypf_common as m


class BuildError(Exception):
    pass


def _header_row_maps(ws, table_name: str, key_column: str) -> tuple[dict[str, int], dict[Any, int], int, int]:
    t = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    col_index = {h: min_col + i for i, h in enumerate(headers)}
    key_col = col_index[key_column]
    row_index: dict[Any, int] = {}
    for r in range(min_row + 1, max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if not m.is_blank(v):
            row_index[v] = r
    return col_index, row_index, min_row, max_row


def _apply_maestro_completions(ws, completions: list) -> int:
    if not completions:
        return 0
    cols, rows, _, _ = _header_row_maps(ws, "tblElementos", m.MAESTRO_KEY)
    count = 0
    for c in completions:
        cell = ws.cell(row=rows[c.elemento_id], column=cols[c.columna])
        if not m.is_blank(cell.value):
            raise BuildError(
                f"MAESTRO_ELEMENTOS: intento de completar una celda no vacía "
                f"({c.elemento_id!r}/{c.columna!r} ya tiene {cell.value!r})"
            )
        cell.value = c.valor_nuevo
        count += 1
    return count


def _append_new_maestro_rows(ws, nuevos_registros: list[dict[str, Any]]) -> int:
    if not nuevos_registros:
        return 0
    t = ws.tables["tblElementos"]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    style_row = max_row
    next_row = max_row + 1
    for record in nuevos_registros:
        for i, header in enumerate(headers):
            col = min_col + i
            cell = ws.cell(row=next_row, column=col)
            style_cell = ws.cell(row=style_row, column=col)
            cell.font = copy(style_cell.font)
            cell.border = copy(style_cell.border)
            cell.fill = copy(style_cell.fill)
            cell.alignment = copy(style_cell.alignment)
            cell.number_format = style_cell.number_format
            cell.value = record.get(header)
        next_row += 1
    new_max_row = max_row + len(nuevos_registros)
    t.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"
    return len(nuevos_registros)


def _capture_campanas_rows(ws) -> tuple[list[list[dict[str, Any]]], list[str], int, int, int, int]:
    t = ws.tables["tblCampanas"]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    captured: list[list[dict[str, Any]]] = []
    for r in range(min_row + 1, max_row + 1):
        row_cells = []
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row_cells.append(
                {
                    "value": cell.value,
                    "font": copy(cell.font),
                    "border": copy(cell.border),
                    "fill": copy(cell.fill),
                    "alignment": copy(cell.alignment),
                    "number_format": cell.number_format,
                }
            )
        captured.append(row_cells)
    return captured, headers, min_col, min_row, max_col, max_row


def _write_campanas_row(ws, row_idx: int, min_col: int, row_cells: list[dict[str, Any]]) -> None:
    for i, cd in enumerate(row_cells):
        cell = ws.cell(row=row_idx, column=min_col + i)
        cell.value = cd["value"]
        cell.font = cd["font"]
        cell.border = cd["border"]
        cell.fill = cd["fill"]
        cell.alignment = cd["alignment"]
        cell.number_format = cd["number_format"]


DATE_COLUMNS = {"FechaInicio", "FechaFin"}
DATETIME_COLUMNS = {"FechaHoraCarga"}


def _rebuild_campanas_sheet(ws, plan: m.YpfPlan) -> tuple[int, int, int]:
    captured, headers, min_col, min_row, max_col, max_row = _capture_campanas_rows(ws)
    n_original = max_row - min_row
    if len(captured) != n_original or n_original != len(plan.ocu_campanas_df):
        raise BuildError(
            f"CAMPANAS: la cantidad de filas capturadas ({len(captured)}) no coincide con lo esperado "
            f"({n_original} / plan={len(plan.ocu_campanas_df)})"
        )

    id_campana_series = plan.ocu_campanas_df["IDCampaña"]
    keep_flags = ~((id_campana_series >= m.LEGACY_ID_MIN) & (id_campana_series <= m.LEGACY_ID_MAX))
    keep_flags = keep_flags.tolist()

    kept_rows = [captured[i] for i in range(n_original) if keep_flags[i]]
    if len(kept_rows) != len(plan.remaining_campanas_df):
        raise BuildError(
            f"CAMPANAS: filas conservadas ({len(kept_rows)}) no coincide con lo esperado "
            f"({len(plan.remaining_campanas_df)})"
        )

    template_row = captured[-1]

    cursor = min_row + 1
    for row_cells in kept_rows:
        _write_campanas_row(ws, cursor, min_col, row_cells)
        cursor += 1

    for record in plan.new_campanas_incorporated:
        new_cells = []
        for i, header in enumerate(headers):
            tmpl = template_row[i]
            value = record.get(header)
            nf = tmpl["number_format"]
            if header in DATE_COLUMNS and isinstance(value, (dt.date, dt.datetime)):
                nf = "yyyy-mm-dd"
            elif header in DATETIME_COLUMNS and isinstance(value, (dt.date, dt.datetime)):
                nf = "yyyy-mm-dd hh:mm:ss"
            new_cells.append(
                {
                    "value": value,
                    "font": copy(tmpl["font"]),
                    "border": copy(tmpl["border"]),
                    "fill": copy(tmpl["fill"]),
                    "alignment": copy(tmpl["alignment"]),
                    "number_format": nf,
                }
            )
        _write_campanas_row(ws, cursor, min_col, new_cells)
        cursor += 1

    final_max_row = cursor - 1
    t = ws.tables["tblCampanas"]
    t.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{final_max_row}"

    return len(kept_rows), len(plan.new_campanas_incorporated), final_max_row - min_row


def build_candidate(run_timestamp: dt.datetime | None = None) -> dict[str, Any]:
    sha_final_v2_before = m.calculate_sha256(m.FINAL_V2_PATH)
    sha_ypf1_before = m.calculate_sha256(m.YPF_ETAPA1_PATH)
    sha_ypf2_before = m.calculate_sha256(m.YPF_ETAPA2_PATH)

    plan = m.compute_plan(run_timestamp)

    if plan.new_campanas_rejected:
        # No se detiene el build por esto (las filas rechazadas simplemente no
        # se insertan y quedan documentadas), pero se deja constancia clara.
        pass

    target_wb = load_workbook(m.FINAL_V2_PATH, data_only=False, read_only=False, keep_vba=False)

    if target_wb.sheetnames != m.mc.EXPECTED_SHEETS:
        raise BuildError(f"Las hojas de FINAL_V2 no coinciden con lo esperado: {target_wb.sheetnames}")

    expected_parametros_rows = len(plan.ocu_parametros_df)

    maestro_completions_applied = _apply_maestro_completions(
        target_wb["MAESTRO_ELEMENTOS"], plan.elementos.completions
    )
    maestro_new_applied = _append_new_maestro_rows(
        target_wb["MAESTRO_ELEMENTOS"], plan.elementos.nuevos_registros
    )

    kept_count, new_count, campanas_final_rows = _rebuild_campanas_sheet(target_wb["CAMPANAS"], plan)

    # --- Invariantes estructurales ---
    t_maestro = target_wb["MAESTRO_ELEMENTOS"].tables["tblElementos"]
    _, mr, _, xr = range_boundaries(t_maestro.ref)
    expected_maestro_rows = plan.expected_counts["MAESTRO_ELEMENTOS_despues"]
    if xr - mr != expected_maestro_rows:
        raise BuildError(f"MAESTRO_ELEMENTOS: filas finales ({xr - mr}) != esperado ({expected_maestro_rows})")

    t_param = target_wb["PARAMETROS"].tables["tblParametros"]
    _, mr, _, xr = range_boundaries(t_param.ref)
    if xr - mr != expected_parametros_rows:
        raise BuildError(f"PARAMETROS: la cantidad de filas cambió inesperadamente ({xr - mr} != {expected_parametros_rows})")

    expected_campanas_rows = plan.expected_counts["CAMPANAS_despues"]
    if campanas_final_rows != expected_campanas_rows:
        raise BuildError(f"CAMPANAS: filas finales ({campanas_final_rows}) != esperado ({expected_campanas_rows})")

    # --- Controles de integridad sobre el resultado final ---
    # ElementoID: sin vacíos ni duplicados en MAESTRO_ELEMENTOS final
    cols_m, _, _, _ = _header_row_maps(target_wb["MAESTRO_ELEMENTOS"], "tblElementos", m.MAESTRO_KEY)
    eid_col = cols_m[m.MAESTRO_KEY]
    t_maestro2 = target_wb["MAESTRO_ELEMENTOS"].tables["tblElementos"]
    _, mr2, _, xr2 = range_boundaries(t_maestro2.ref)
    all_eids = [target_wb["MAESTRO_ELEMENTOS"].cell(row=r, column=eid_col).value for r in range(mr2 + 1, xr2 + 1)]
    if any(m.is_blank(e) for e in all_eids):
        raise BuildError("MAESTRO_ELEMENTOS: hay ElementoID vacío en la candidata final")
    if len(set(all_eids)) != len(all_eids):
        raise BuildError("MAESTRO_ELEMENTOS: hay ElementoID duplicado en la candidata final")
    final_maestro_ids = set(all_eids)

    # CAMPANAS: referencial + FB digital + APIE 30943 digital + ClaveNegocio única
    cols_c, _, minr_c, _ = _header_row_maps(target_wb["CAMPANAS"], "tblCampanas", m.CAMPANAS_KEY)
    t_camp = target_wb["CAMPANAS"].tables["tblCampanas"]
    _, mrc, _, xrc = range_boundaries(t_camp.ref)
    eid_col_c = cols_c["ElementoID"]
    clave_col_c = cols_c["ClaveNegocio"]
    cargaid_col_c = cols_c["CargaID"]

    new_region_start = minr_c + kept_count + 1  # primera fila perteneciente a las 13.616 nuevas de YPF

    orphan_count = 0
    fb_digital_count = 0
    apie_30943_digital_count = 0
    clave_seen: dict[Any, int] = {}
    clave_seen_new: dict[Any, int] = {}
    carga_seen: dict[Any, int] = {}
    for r in range(mrc + 1, xrc + 1):
        eid = target_wb["CAMPANAS"].cell(row=r, column=eid_col_c).value
        if m.is_blank(eid):
            orphan_count += 1
        elif eid not in final_maestro_ids:
            orphan_count += 1
        if not m.is_blank(eid) and m.elemento_id_tipo(eid) == "FB":
            fb_digital_count += 1
        if not m.is_blank(eid) and m.es_elemento_digital_30943(eid):
            apie_30943_digital_count += 1
        clave = target_wb["CAMPANAS"].cell(row=r, column=clave_col_c).value
        if not m.is_blank(clave):
            clave_seen[clave] = clave_seen.get(clave, 0) + 1
            if r >= new_region_start:
                clave_seen_new[clave] = clave_seen_new.get(clave, 0) + 1
        carga = target_wb["CAMPANAS"].cell(row=r, column=cargaid_col_c).value
        if not m.is_blank(carga):
            carga_seen[carga] = carga_seen.get(carga, 0) + 1

    if orphan_count:
        raise BuildError(f"CAMPANAS: {orphan_count} fila(s) con ElementoID vacío o inexistente en MAESTRO_ELEMENTOS final")
    if fb_digital_count:
        raise BuildError(f"CAMPANAS: {fb_digital_count} fila(s) de campaña digital asignada a un ElementoID FB (prohibido)")
    if apie_30943_digital_count:
        raise BuildError(f"CAMPANAS: {apie_30943_digital_count} fila(s) de campaña digital sobre APIE 30943 (prohibido)")
    # Solo se bloquea por ClaveNegocio duplicada si la duplicación involucra al menos una fila
    # NUEVA (YPF); los grupos duplicados exactos ya existentes en FINAL_V2 (ej. Adidas 4322,
    # documentados como REPETICION_VALIDA) son datos protegidos ajenos a YPF y no se tocan.
    clave_dups = {k: v for k, v in clave_seen.items() if v > 1 and clave_seen_new.get(k, 0) > 0}
    preexisting_clave_dups = {k: v for k, v in clave_seen.items() if v > 1 and clave_seen_new.get(k, 0) == 0}
    if clave_dups:
        raise BuildError(f"CAMPANAS: {len(clave_dups)} ClaveNegocio duplicada(s) en la candidata final: {list(clave_dups.items())[:5]}")
    carga_dups = {k: v for k, v in carga_seen.items() if v > 1}
    if carga_dups:
        raise BuildError(f"CAMPANAS: {len(carga_dups)} CargaID duplicado(s) en la candidata final: {list(carga_dups.items())[:5]}")

    # APIE 30943: exactamente 2 elementos FB, cero digitales, en el MAESTRO_ELEMENTOS final
    apie_30943_elems = sorted(e for e in final_maestro_ids if str(e).startswith("30943 - "))
    apie_30943_fb = [e for e in apie_30943_elems if m.elemento_id_tipo(e) == "FB"]
    apie_30943_digital = [e for e in apie_30943_elems if m.elemento_id_tipo(e) in m.YPF_DIGITAL_TYPES]
    # Nota: los elementos digitales preexistentes de OCU26 bajo 30943 (no autorizados por YPF)
    # se conservan sin cambios por protección FASE 6; se reportan pero no bloquean el build.

    if target_wb.sheetnames != m.mc.EXPECTED_SHEETS:
        raise BuildError(f"Las hojas de la candidata no coinciden con lo esperado: {target_wb.sheetnames}")

    m.CANDIDATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(m.CANDIDATE_PATH)
    target_wb.close()

    sha_final_v2_after = m.calculate_sha256(m.FINAL_V2_PATH)
    sha_ypf1_after = m.calculate_sha256(m.YPF_ETAPA1_PATH)
    sha_ypf2_after = m.calculate_sha256(m.YPF_ETAPA2_PATH)
    if sha_final_v2_after != sha_final_v2_before:
        raise BuildError("ERROR CRÍTICO: OCU26 FINAL_V2 fue modificado durante el build")
    if sha_ypf1_after != sha_ypf1_before:
        raise BuildError("ERROR CRÍTICO: YPF Etapa 1 fue modificado durante el build")
    if sha_ypf2_after != sha_ypf2_before:
        raise BuildError("ERROR CRÍTICO: YPF Etapa 2 fue modificado durante el build")

    return {
        "result": "BUILD_YPF_OK",
        "candidate_path": str(m.CANDIDATE_PATH),
        "sources_intact": True,
        "final_v2_sha256": sha_final_v2_after,
        "ypf_etapa1_sha256": sha_ypf1_after,
        "ypf_etapa2_sha256": sha_ypf2_after,
        "run_timestamp": plan.run_timestamp.isoformat(),
        "expected_counts": plan.expected_counts,
        "maestro_completions_applied": maestro_completions_applied,
        "maestro_new_rows_applied": maestro_new_applied,
        "campanas_kept": kept_count,
        "campanas_new": new_count,
        "campanas_rejected": len(plan.new_campanas_rejected),
        "apie_30943_fb_final": apie_30943_fb,
        "apie_30943_digital_preexistente_no_ypf": apie_30943_digital,
        "legacy_rows_retiradas": len(plan.legacy_rows),
        "clavenegocio_duplicados_preexistentes_no_ypf": preexisting_clave_dups,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Construye la base candidata OCU26 + YPF.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = build_candidate()
    except BuildError as exc:
        if args.json:
            print(json.dumps({"result": "BUILD_YPF_ERROR", "error": str(exc)}, ensure_ascii=False, indent=2))
        else:
            print("=" * 60)
            print("OCU26 + YPF BUILD - ERROR")
            print("=" * 60)
            print(str(exc))
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 + YPF BUILD")
        print("=" * 60)
        for k, v in result.items():
            print(f"{k}: {v}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
