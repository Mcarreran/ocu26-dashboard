# -*- coding: utf-8 -*-
"""Construye la versión FINAL de la base OCU26 + YPF a partir de la
CANDIDATA ya generada (no re-deriva de las 3 fuentes originales; mismo
patrón que Pendientes/OCU26_ACTUALIZACION/scripts/build_ocu26_integrada_final.py).

Corrección puntual autorizada: retira de MAESTRO_ELEMENTOS los elementos
digitales legacy de APIE 30943 que no pertenecen al catálogo YPF Etapa 1
validado, una vez verificado que ninguna fila de CAMPANAS los referencia.
No modifica CAMPANAS ni PARAMETROS. No modifica la CANDIDATA ni ninguna de
las 3 fuentes originales.

Uso:
    python build_ocu26_ypf_final.py
    python build_ocu26_ypf_final.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from copy import copy
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter

import merge_ypf_common as m


class BuildError(Exception):
    pass


def _remove_maestro_rows(ws, eids_a_retirar: set[Any]) -> int:
    t = ws.tables["tblElementos"]
    min_col, min_row, max_col, max_row = range_boundaries(t.ref)
    headers = list(
        next(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True))
    )
    eid_idx = headers.index(m.MAESTRO_KEY)

    captured = []
    for r in range(min_row + 1, max_row + 1):
        row_cells = []
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row_cells.append(
                {
                    "value": cell.value,
                    "font": copy(cell.font),
                    "border": copy(cell.border),
                    "fill": copy(cell.fill),
                    "alignment": copy(cell.alignment),
                    "number_format": cell.number_format,
                }
            )
        captured.append(row_cells)

    kept = [row for row in captured if row[eid_idx]["value"] not in eids_a_retirar]
    removed_count = len(captured) - len(kept)

    cursor = min_row + 1
    for row_cells in kept:
        for i, cd in enumerate(row_cells):
            cell = ws.cell(row=cursor, column=min_col + i)
            cell.value = cd["value"]
            cell.font = cd["font"]
            cell.border = cd["border"]
            cell.fill = cd["fill"]
            cell.alignment = cd["alignment"]
            cell.number_format = cd["number_format"]
        cursor += 1

    # Limpiar las filas sobrantes al final del rango original (la tabla se
    # achica: hay menos filas que antes).
    for r in range(cursor, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None

    new_max_row = cursor - 1
    t.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{new_max_row}"

    return removed_count


def build_final(run_timestamp=None) -> dict[str, Any]:
    sha_final_v2_before = m.calculate_sha256(m.FINAL_V2_PATH)
    sha_ypf1_before = m.calculate_sha256(m.YPF_ETAPA1_PATH)
    sha_ypf2_before = m.calculate_sha256(m.YPF_ETAPA2_PATH)
    sha_candidate_before = m.calculate_sha256(m.CANDIDATE_PATH)

    plan = m.compute_final_removal_plan(run_timestamp)
    eids_a_retirar = {r[m.MAESTRO_KEY] for r in plan.elementos_a_retirar}

    if sorted(eids_a_retirar) != [
        "30943 - MB - 1", "30943 - MB - 2",
        "30943 - PPUNTER - 1", "30943 - PPUNTER - 2", "30943 - PPUNTER - 3", "30943 - PPUNTER - 4", "30943 - PPUNTER - 5",
        "30943 - TT - 1", "30943 - TT - 10", "30943 - TT - 2", "30943 - TT - 3", "30943 - TT - 4",
        "30943 - TT - 5", "30943 - TT - 6", "30943 - TT - 7", "30943 - TT - 8", "30943 - TT - 9",
    ]:
        raise BuildError(f"El conjunto de elementos a retirar no es el esperado (17 de APIE 30943): {sorted(eids_a_retirar)}")

    target_wb = load_workbook(m.CANDIDATE_PATH, data_only=False, read_only=False, keep_vba=False)
    if target_wb.sheetnames != m.mc.EXPECTED_SHEETS:
        raise BuildError(f"Las hojas de la candidata no coinciden con lo esperado: {target_wb.sheetnames}")

    removed_count = _remove_maestro_rows(target_wb["MAESTRO_ELEMENTOS"], eids_a_retirar)
    if removed_count != 17:
        raise BuildError(f"Se retiraron {removed_count} filas, se esperaban exactamente 17")

    # --- Invariantes: CAMPANAS y PARAMETROS deben permanecer con la misma cantidad de filas ---
    t_maestro = target_wb["MAESTRO_ELEMENTOS"].tables["tblElementos"]
    _, mr, _, xr = range_boundaries(t_maestro.ref)
    expected_maestro_rows = plan.expected_counts["MAESTRO_ELEMENTOS_despues"]
    if xr - mr != expected_maestro_rows:
        raise BuildError(f"MAESTRO_ELEMENTOS: filas finales ({xr - mr}) != esperado ({expected_maestro_rows})")

    t_camp = target_wb["CAMPANAS"].tables["tblCampanas"]
    _, mrc, _, xrc = range_boundaries(t_camp.ref)
    if xrc - mrc != plan.expected_counts["CAMPANAS"]:
        raise BuildError(f"CAMPANAS: la cantidad de filas cambió inesperadamente ({xrc - mrc} != {plan.expected_counts['CAMPANAS']})")

    t_param = target_wb["PARAMETROS"].tables["tblParametros"]
    _, mrp, _, xrp = range_boundaries(t_param.ref)
    if xrp - mrp != plan.expected_counts["PARAMETROS"]:
        raise BuildError(f"PARAMETROS: la cantidad de filas cambió inesperadamente ({xrp - mrp} != {plan.expected_counts['PARAMETROS']})")

    # --- Controles sobre el resultado final ---
    t2 = target_wb["MAESTRO_ELEMENTOS"].tables["tblElementos"]
    minc2, minr2, maxc2, maxr2 = range_boundaries(t2.ref)
    headers2 = list(
        next(target_wb["MAESTRO_ELEMENTOS"].iter_rows(min_row=minr2, max_row=minr2, min_col=minc2, max_col=maxc2, values_only=True))
    )
    eid_col2 = minc2 + headers2.index(m.MAESTRO_KEY)
    all_eids = [target_wb["MAESTRO_ELEMENTOS"].cell(row=r, column=eid_col2).value for r in range(minr2 + 1, maxr2 + 1)]
    if any(m.is_blank(e) for e in all_eids):
        raise BuildError("MAESTRO_ELEMENTOS: hay ElementoID vacío en la FINAL")
    if len(set(all_eids)) != len(all_eids):
        raise BuildError("MAESTRO_ELEMENTOS: hay ElementoID duplicado en la FINAL")
    if eids_a_retirar & set(all_eids):
        raise BuildError("MAESTRO_ELEMENTOS: alguno de los elementos a retirar sigue presente en la FINAL")
    final_maestro_ids = set(all_eids)

    apie_30943_final = sorted(e for e in final_maestro_ids if str(e).startswith("30943 - "))
    if apie_30943_final != ["30943 - FB - 1", "30943 - FB - 2"]:
        raise BuildError(f"APIE 30943 en la FINAL no quedó como exactamente 2 FB: {apie_30943_final}")

    t3 = target_wb["CAMPANAS"].tables["tblCampanas"]
    minc3, minr3, maxc3, maxr3 = range_boundaries(t3.ref)
    headers3 = list(
        next(target_wb["CAMPANAS"].iter_rows(min_row=minr3, max_row=minr3, min_col=minc3, max_col=maxc3, values_only=True))
    )
    eid_col3 = minc3 + headers3.index("ElementoID")
    orphan_count = 0
    apie_30943_camp_count = 0
    for r in range(minr3 + 1, maxr3 + 1):
        eid = target_wb["CAMPANAS"].cell(row=r, column=eid_col3).value
        if m.is_blank(eid) or eid not in final_maestro_ids:
            orphan_count += 1
        if not m.is_blank(eid) and str(eid).startswith("30943 - "):
            apie_30943_camp_count += 1
    if orphan_count:
        raise BuildError(f"CAMPANAS: {orphan_count} fila(s) con ElementoID huérfano tras el retiro")
    if apie_30943_camp_count:
        raise BuildError(f"CAMPANAS: {apie_30943_camp_count} fila(s) referencian APIE 30943 (se esperaba 0)")

    m.FINAL_PATH.parent.mkdir(parents=True, exist_ok=True)
    target_wb.save(m.FINAL_PATH)
    target_wb.close()

    sha_final_v2_after = m.calculate_sha256(m.FINAL_V2_PATH)
    sha_ypf1_after = m.calculate_sha256(m.YPF_ETAPA1_PATH)
    sha_ypf2_after = m.calculate_sha256(m.YPF_ETAPA2_PATH)
    sha_candidate_after = m.calculate_sha256(m.CANDIDATE_PATH)
    if sha_final_v2_after != sha_final_v2_before:
        raise BuildError("ERROR CRÍTICO: OCU26 FINAL_V2 fue modificado durante el build FINAL")
    if sha_ypf1_after != sha_ypf1_before:
        raise BuildError("ERROR CRÍTICO: YPF Etapa 1 fue modificado durante el build FINAL")
    if sha_ypf2_after != sha_ypf2_before:
        raise BuildError("ERROR CRÍTICO: YPF Etapa 2 fue modificado durante el build FINAL")
    if sha_candidate_after != sha_candidate_before:
        raise BuildError("ERROR CRÍTICO: la CANDIDATA fue modificada durante el build FINAL")

    return {
        "result": "BUILD_YPF_FINAL_OK",
        "final_path": str(m.FINAL_PATH),
        "candidate_path": str(m.CANDIDATE_PATH),
        "sources_intact": True,
        "final_v2_sha256": sha_final_v2_after,
        "ypf_etapa1_sha256": sha_ypf1_after,
        "ypf_etapa2_sha256": sha_ypf2_after,
        "candidate_sha256": sha_candidate_after,
        "run_timestamp": plan.run_timestamp.isoformat(),
        "expected_counts": plan.expected_counts,
        "elementos_retirados": sorted(eids_a_retirar),
        "apie_30943_final": apie_30943_final,
        "campanas_apie_30943": apie_30943_camp_count,
        "campanas_huerfanas": orphan_count,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Construye la versión FINAL de OCU26 + YPF desde la candidata.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = build_final()
    except (BuildError, ValueError, FileNotFoundError) as exc:
        if args.json:
            print(json.dumps({"result": "BUILD_YPF_FINAL_ERROR", "error": str(exc)}, ensure_ascii=False, indent=2))
        else:
            print("=" * 60)
            print("OCU26 + YPF FINAL BUILD - ERROR")
            print("=" * 60)
            print(str(exc))
        return 1

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 + YPF FINAL BUILD")
        print("=" * 60)
        for k, v in result.items():
            print(f"{k}: {v}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
