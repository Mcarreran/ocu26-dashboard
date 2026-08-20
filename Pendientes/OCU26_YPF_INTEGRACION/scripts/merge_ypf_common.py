# -*- coding: utf-8 -*-
"""Motor compartido de integración YPF -> OCU26 (Pendientes/OCU26_YPF_INTEGRACION).

Fuentes, todas READ-ONLY (nunca se escribe en ninguna de ellas):
  - OCU26 FINAL_V2: Pendientes/OCU26_ACTUALIZACION/output/OCU26_BASE_DATOS_INTEGRADA_FINAL_V2_2026-08-18.xlsx
  - YPF Etapa 1 (catálogo):  Pendientes/YPF_ETAPA_1/output/YPF_BASE_LIMPIA_ETAPA_1_RELEVAMIENTO_2026-08-13.xlsx
  - YPF Etapa 2 (campañas):  Pendientes/YPF_ETAPA_2/output/YPF_BASE_INTEGRADA_ETAPA_2_CAMPANAS_2026-08-13.xlsx

Este módulo reutiliza (por import, sin copiar) la infraestructura ya validada
de Pendientes/OCU26_ACTUALIZACION/scripts:
  - merge_common (mc): rutas, is_blank, read_table_df, inspect_structure,
    EXPECTED_SHEETS/TABLES, headers canónicos, vi (validate_input.py).
  - merge_v2_common (v2): fórmula canónica de ClaveNegocio
    (compute_clave_negocio, validada al 100% contra los datos reales) y la
    secuencia determinista de CargaID (next_carga_id_sequence, prefijo
    HIST-########).

--- DECISIÓN DE NEGOCIO AUTORIZADA (reemplazo del bloque histórico YPF) ---
Las 7.790 filas de CAMPANAS en FINAL_V2 con IDCampaña en [10000, 10009] y
UsuarioCarga='Migración histórica' son la versión ANTERIOR (menos granular,
sin re-cruce validado) de las mismas activaciones que
YPF_BASE_INTEGRADA_ETAPA_2_CAMPANAS reconstruye desde cero con IDCampaña
20000+. Mantener ambas produciría doble conteo. Por autorización explícita
del usuario:
  - Se retira el bloque completo [10000,10009] de la candidata (no se
    reinserta ninguna fila con esos IDCampaña).
  - Se conservan sin cambios las restantes 1.717 filas de CAMPANAS (todo
    IDCampaña fuera de [10000,10009]).
  - Se insertan las 13.616 filas de BASE CAMPAÑAS (YPF Etapa 2), adaptadas
    al esquema canónico OCU26.
  - Las 7.790 filas retiradas quedan preservadas íntegramente en la
    auditoría (nunca se pierden, solo no viajan a la candidata).

--- Elementos YPF autorizados (FASE 2) ---
El conjunto autorizado se obtiene EXCLUSIVAMENTE de la hoja 'BASE ELEMENTOS'
de YPF Etapa 1, usando ElementoID como clave exacta y el patrón
'^\\d+ - (MB|PPUNTER|TT|FB) - \\d+$'. No se usa la palabra "YPF" en ningún
campo como criterio de identificación. No hay fuzzy matching en ningún punto
de este módulo.
"""

from __future__ import annotations

import datetime as dt
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

_ACTUALIZACION_SCRIPTS = Path(__file__).resolve().parents[2] / "OCU26_ACTUALIZACION" / "scripts"
sys.path.insert(0, str(_ACTUALIZACION_SCRIPTS))
import merge_common as mc  # noqa: E402
import merge_v2_common as v2  # noqa: E402

vi = mc.vi
is_blank = mc.is_blank
read_table_df = mc.read_table_df
inspect_structure = mc.inspect_structure
calculate_sha256 = mc.calculate_sha256
compute_clave_negocio = v2.compute_clave_negocio
CARGA_ID_PREFIX = v2.CARGA_ID_PREFIX
CARGA_ID_WIDTH = v2.CARGA_ID_WIDTH
next_carga_id_sequence = v2.next_carga_id_sequence

# --- Rutas ---
REPO_ROOT = mc.REPO_ROOT
INTEGRACION_ROOT = Path(__file__).resolve().parents[1]

FINAL_V2_PATH = mc.UPDATE_ROOT / "output" / "OCU26_BASE_DATOS_INTEGRADA_FINAL_V2_2026-08-18.xlsx"
YPF_ETAPA1_PATH = REPO_ROOT / "Pendientes" / "YPF_ETAPA_1" / "output" / "YPF_BASE_LIMPIA_ETAPA_1_RELEVAMIENTO_2026-08-13.xlsx"
YPF_ETAPA2_PATH = REPO_ROOT / "Pendientes" / "YPF_ETAPA_2" / "output" / "YPF_BASE_INTEGRADA_ETAPA_2_CAMPANAS_2026-08-13.xlsx"

CANDIDATE_PATH = INTEGRACION_ROOT / "output" / "OCU26_BASE_DATOS_CON_YPF_CANDIDATA_2026-08-18.xlsx"
AUDIT_PATH = INTEGRACION_ROOT / "output" / "OCU26_AUDITORIA_INTEGRACION_YPF_2026-08-18.xlsx"

# --- Pase FINAL (corrección puntual autorizada sobre la CANDIDATA, no re-deriva
# de las 3 fuentes originales): retiro de los elementos digitales legacy de
# APIE 30943 que no pertenecen al catálogo YPF validado y no tienen ninguna
# campaña activa en la candidata. ---
FINAL_PATH = INTEGRACION_ROOT / "output" / "OCU26_BASE_DATOS_CON_YPF_FINAL_2026-08-18.xlsx"
FINAL_AUDIT_PATH = INTEGRACION_ROOT / "output" / "OCU26_AUDITORIA_INTEGRACION_YPF_FINAL_2026-08-18.xlsx"
LEGACY_APIE_30943_INVALIDO = "ELEMENTO_LEGACY_APIE_30943_INVALIDO"

MAESTRO_KEY = mc.MAESTRO_KEY
CAMPANAS_KEY = mc.CAMPANAS_KEY
MAESTRO_HEADERS = mc.MAESTRO_HEADERS
CAMPANAS_HEADERS = mc.CAMPANAS_HEADERS

# --- Identificación exacta de elementos YPF (FASE 2) ---
YPF_ELEMENT_ID_PATTERN = re.compile(r"^\d+ - (MB|PPUNTER|TT|FB) - \d+$")
YPF_DIGITAL_TYPES = {"MB", "PPUNTER", "TT"}

# --- Reemplazo del bloque histórico YPF (decisión de negocio autorizada) ---
LEGACY_ID_MIN = 10000
LEGACY_ID_MAX = 10009  # inclusive
LEGACY_USUARIO_CARGA = "Migración histórica"

# --- Campos técnicos para filas nuevas (FASE 5) ---
NEW_USUARIO_CARGA = "MIGRACION_YPF"
NEW_FUENTE_CARGA = "Migración histórica"  # debe existir ya en PARAMETROS (se verifica en compute_plan)

# Columnas de MAESTRO_ELEMENTOS que OCU26 tiene y YPF Etapa 1 no aporta
# (no hay dato posible para adaptar; quedan en blanco para los elementos nuevos).
MAESTRO_COLUMNS_NOT_IN_YPF = {"b", "h", "q", "m2"}

# Columnas donde YPF exporta el mismo valor a veces como número nativo y a
# veces como texto numérico (ya documentado y tolerado por
# scripts/validate_input.py::validate_numeric_fields). Se comparan por valor
# numérico, no por tipo, para no generar falsos "conflictos" de formato.
NUMERIC_TOLERANT_FIELDS = {"CapacidadSlotsReel", "SegundosDia"}


def elemento_id_tipo(eid: Any) -> str | None:
    m = YPF_ELEMENT_ID_PATTERN.match(str(eid))
    return m.group(1) if m else None


def es_elemento_digital_30943(eid: Any) -> bool:
    s = str(eid)
    if not s.startswith("30943 - "):
        return False
    return elemento_id_tipo(eid) in YPF_DIGITAL_TYPES


def _to_float(v: Any) -> float | None:
    if is_blank(v):
        return None
    if isinstance(v, bool):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def values_equal(col: str, v1: Any, v2_: Any) -> bool:
    """Igualdad usada para clasificar completions/conflicts en MAESTRO_ELEMENTOS.

    Tolerante de tipo (número nativo vs texto numérico) SOLO para los campos
    documentados en NUMERIC_TOLERANT_FIELDS; el resto se compara con
    igualdad exacta (incluye mayúsculas/espacios: una diferencia de caso es
    una diferencia real de contenido, no se normaliza).
    """
    if col in NUMERIC_TOLERANT_FIELDS:
        f1, f2 = _to_float(v1), _to_float(v2_)
        if f1 is not None and f2 is not None:
            return f1 == f2
    return v1 == v2_


def read_raw_sheet_df(path: Path, sheet: str) -> pd.DataFrame:
    """Lee una hoja SIN tabla estructurada de Excel (rango crudo, fila 1 =
    encabezado). Usado para las hojas YPF ('BASE ELEMENTOS', 'BASE
    CAMPAÑAS', 'MAPEO CAMPAÑAS'), que no están definidas como ws.tables.
    READ-ONLY: solo lee, nunca escribe."""
    wb = load_workbook(path, data_only=True, read_only=False, keep_vba=False)
    try:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    headers = list(rows[0])
    return pd.DataFrame(rows[1:], columns=headers)


# ---------------------------------------------------------------------------
# FASE 2 + FASE 3: cruce e integración de MAESTRO_ELEMENTOS
# ---------------------------------------------------------------------------


@dataclass
class FieldCompletion:
    elemento_id: Any
    columna: str
    valor_nuevo: Any


@dataclass
class FieldConflict:
    elemento_id: Any
    columna: str
    valor_ocu26: Any
    valor_ypf: Any


@dataclass
class ElementosCrossResult:
    ypf_total: int
    ypf_elementoid_invalido: list[dict[str, Any]]  # filas cuyo ElementoID no matchea el patrón
    ypf_duplicados: dict[Any, int]  # ElementoID duplicado dentro de BASE ELEMENTOS (no debería ocurrir)
    ypf_autorizados: set[Any]  # ElementoID válidos y únicos (universo autorizado, FASE 2)
    comunes: set[Any]  # autorizados ∩ OCU26 (ya existen)
    solo_ypf: set[Any]  # autorizados - OCU26 (a agregar)
    solo_ocu26: set[Any]  # OCU26 - autorizados (no se toca; informativo)
    identicos: set[Any]  # de 'comunes': sin ninguna diferencia
    con_completions: set[Any]  # de 'comunes': al menos 1 campo completado, 0 conflictos
    con_conflictos: set[Any]  # de 'comunes': al menos 1 conflicto de valor
    completions: list[FieldCompletion]
    conflicts: list[FieldConflict]
    nuevos_registros: list[dict[str, Any]]  # filas YPF listas para adaptar/agregar a MAESTRO_ELEMENTOS


def classify_elementos(ocu_maestro_df: pd.DataFrame, ypf_elem_df: pd.DataFrame) -> ElementosCrossResult:
    ocu_ids = set(ocu_maestro_df[MAESTRO_KEY].dropna())

    raw_ids = ypf_elem_df[MAESTRO_KEY]
    valid_mask = raw_ids.apply(lambda x: not is_blank(x) and YPF_ELEMENT_ID_PATTERN.match(str(x)) is not None)
    invalidos = ypf_elem_df.loc[~valid_mask].to_dict(orient="records")

    valid_ids = raw_ids[valid_mask]
    dup_counts = valid_ids[valid_ids.duplicated(keep=False)].value_counts()
    duplicados = {k: int(v) for k, v in dup_counts.items()}
    dup_set = set(duplicados.keys())

    autorizados_mask = valid_mask & ~raw_ids.isin(dup_set)
    ypf_authorized_df = ypf_elem_df.loc[autorizados_mask].drop_duplicates(subset=[MAESTRO_KEY])
    ypf_autorizados = set(ypf_authorized_df[MAESTRO_KEY])

    comunes = ypf_autorizados & ocu_ids
    solo_ypf = ypf_autorizados - ocu_ids
    solo_ocu26 = ocu_ids - ypf_autorizados

    ocu_idx = ocu_maestro_df.set_index(MAESTRO_KEY, drop=False)
    ypf_idx = ypf_authorized_df.set_index(MAESTRO_KEY, drop=False)

    shared_cols = [
        c for c in MAESTRO_HEADERS
        if c != MAESTRO_KEY and c not in MAESTRO_COLUMNS_NOT_IN_YPF and c in ypf_elem_df.columns
    ]

    completions: list[FieldCompletion] = []
    conflicts: list[FieldConflict] = []
    identicos: set[Any] = set()
    con_completions: set[Any] = set()
    con_conflictos: set[Any] = set()

    for eid in sorted(comunes, key=str):
        r_ocu = ocu_idx.loc[eid]
        r_ypf = ypf_idx.loc[eid]
        has_completion = False
        has_conflict = False
        for col in shared_cols:
            v_o, v_y = r_ocu[col], r_ypf[col]
            blank_o, blank_y = is_blank(v_o), is_blank(v_y)
            if blank_o and blank_y:
                continue
            if not blank_o and not blank_y:
                if not values_equal(col, v_o, v_y):
                    conflicts.append(FieldConflict(eid, col, v_o, v_y))
                    has_conflict = True
                continue
            if blank_o and not blank_y:
                completions.append(FieldCompletion(eid, col, v_y))
                has_completion = True
            # not blank_o and blank_y -> conservar OCU26, sin acción (regla 1)
        if has_conflict:
            con_conflictos.add(eid)
        elif has_completion:
            con_completions.add(eid)
        else:
            identicos.add(eid)

    nuevos_registros: list[dict[str, Any]] = []
    for eid in sorted(solo_ypf, key=str):
        r = ypf_idx.loc[eid]
        record = {col: (None if col in MAESTRO_COLUMNS_NOT_IN_YPF else r[col]) for col in MAESTRO_HEADERS}
        nuevos_registros.append(record)

    return ElementosCrossResult(
        ypf_total=len(ypf_elem_df),
        ypf_elementoid_invalido=invalidos,
        ypf_duplicados=duplicados,
        ypf_autorizados=ypf_autorizados,
        comunes=comunes,
        solo_ypf=solo_ypf,
        solo_ocu26=solo_ocu26,
        identicos=identicos,
        con_completions=con_completions,
        con_conflictos=con_conflictos,
        completions=completions,
        conflicts=conflicts,
        nuevos_registros=nuevos_registros,
    )


# ---------------------------------------------------------------------------
# Reemplazo del bloque histórico YPF (clasificación de las 7.790 filas retiradas)
# ---------------------------------------------------------------------------


@dataclass
class LegacyRow:
    carga_id: Any
    id_campana: Any
    campana: Any
    elemento_id: Any
    fecha_inicio: Any
    fecha_fin: Any
    categoria: str
    motivo: str


def classify_legacy_block(
    ocu_campanas_df: pd.DataFrame,
    ocu_maestro_ids: set[Any],
    ypf_autorizados: set[Any],
    ypf_camp_df: pd.DataFrame,
    mapeo_df: pd.DataFrame,
) -> list[LegacyRow]:
    """Clasifica, sin fuzzy matching, cada una de las filas del bloque
    histórico [10000,10009] en una de las 6 categorías autorizadas. Orden de
    prioridad (mutuamente excluyente, primera regla que aplica gana):

      1. LEGACY_ELEMENTO_INVALIDO       -> el ElementoID ni siquiera existe en OCU26 MAESTRO_ELEMENTOS actual.
      2. LEGACY_APIE_30943_DIGITAL_INVALIDA -> ElementoID digital (MB/PPUNTER/TT) bajo APIE 30943.
      3. LEGACY_FUERA_CATALOGO          -> ElementoID no pertenece al catálogo YPF Etapa 1 validado (3.883).
      4. LEGACY_EXCLUIDA_POR_REGLA      -> Campaña coincide EXACTO (tras strip) con una entrada de
                                           MAPEO CAMPAÑAS cuyo EstadoMapeo no es CARGABLE (documentación YPF).
      5. REEMPLAZADA_POR_YPF_ETAPA2_EXACTA -> (ElementoID, FechaInicio, FechaFin) coincide exacto con
                                              una fila de BASE CAMPAÑAS (YPF Etapa 2).
      6. LEGACY_YPF_NO_RECONCILIADA     -> ninguna de las anteriores (catch-all).
    """
    legacy = ocu_campanas_df[
        (ocu_campanas_df["IDCampaña"] >= LEGACY_ID_MIN) & (ocu_campanas_df["IDCampaña"] <= LEGACY_ID_MAX)
    ]

    excluidas_map: dict[str, tuple[str, str]] = {}
    for _, r in mapeo_df.iterrows():
        estado = r.get("EstadoMapeo")
        if estado is not None and str(estado).strip() != "CARGABLE":
            nombre = str(r.get("CampañaNormalizada") or "").strip()
            if nombre:
                excluidas_map[nombre] = (str(estado).strip(), str(r.get("MotivoNoCarga") or "").strip())

    ypf_exact_keys = set(zip(ypf_camp_df["ElementoID"], ypf_camp_df["FechaInicio"], ypf_camp_df["FechaFin"]))

    out: list[LegacyRow] = []
    for _, row in legacy.iterrows():
        eid = row["ElementoID"]
        campana_norm = str(row["Campaña"] or "").strip()

        if is_blank(eid) or eid not in ocu_maestro_ids:
            out.append(LegacyRow(
                row["CargaID"], row["IDCampaña"], row["Campaña"], eid, row["FechaInicio"], row["FechaFin"],
                "LEGACY_ELEMENTO_INVALIDO",
                "ElementoID vacío o inexistente en MAESTRO_ELEMENTOS de OCU26 FINAL_V2.",
            ))
            continue

        if es_elemento_digital_30943(eid):
            out.append(LegacyRow(
                row["CargaID"], row["IDCampaña"], row["Campaña"], eid, row["FechaInicio"], row["FechaFin"],
                "LEGACY_APIE_30943_DIGITAL_INVALIDA",
                "Campaña digital sobre APIE 30943; el catálogo YPF validado confirma que 30943 es "
                "exclusivamente estático (30943-FB-1/2). Este elemento digital no es parte del "
                "conjunto autorizado YPF y no tiene equivalente en BASE CAMPAÑAS Etapa 2.",
            ))
            continue

        if eid not in ypf_autorizados:
            out.append(LegacyRow(
                row["CargaID"], row["IDCampaña"], row["Campaña"], eid, row["FechaInicio"], row["FechaFin"],
                "LEGACY_FUERA_CATALOGO",
                f"ElementoID '{eid}' no pertenece al catálogo YPF Etapa 1 validado (3.883 elementos); "
                "no tiene equivalente en BASE CAMPAÑAS Etapa 2.",
            ))
            continue

        if campana_norm in excluidas_map:
            estado, motivo_no_carga = excluidas_map[campana_norm]
            out.append(LegacyRow(
                row["CargaID"], row["IDCampaña"], row["Campaña"], eid, row["FechaInicio"], row["FechaFin"],
                "LEGACY_EXCLUIDA_POR_REGLA",
                f"Campaña '{campana_norm}' figura en MAPEO CAMPAÑAS (YPF Etapa 2) con EstadoMapeo="
                f"{estado!r}, MotivoNoCarga={motivo_no_carga!r}: excluida deliberadamente de BASE "
                "CAMPAÑAS por la propia fuente validada.",
            ))
            continue

        key = (eid, row["FechaInicio"], row["FechaFin"])
        if key in ypf_exact_keys:
            out.append(LegacyRow(
                row["CargaID"], row["IDCampaña"], row["Campaña"], eid, row["FechaInicio"], row["FechaFin"],
                "REEMPLAZADA_POR_YPF_ETAPA2_EXACTA",
                "Coincidencia exacta de (ElementoID, FechaInicio, FechaFin) con una fila de BASE "
                "CAMPAÑAS (YPF Etapa 2): misma activación real, recalculada con IDCampaña nuevo.",
            ))
            continue

        out.append(LegacyRow(
            row["CargaID"], row["IDCampaña"], row["Campaña"], eid, row["FechaInicio"], row["FechaFin"],
            "LEGACY_YPF_NO_RECONCILIADA",
            "Elemento y campaña del catálogo YPF válido, pero sin coincidencia exacta de fechas en "
            "BASE CAMPAÑAS Etapa 2; retirada igualmente por autorización de reemplazo íntegro del "
            "bloque [10000,10009], pendiente de reconciliación manual.",
        ))

    return out


# ---------------------------------------------------------------------------
# FASE 4 + FASE 5: nuevas filas de CAMPANAS a partir de BASE CAMPAÑAS (YPF Etapa 2)
# ---------------------------------------------------------------------------


def build_new_campanas_records(
    ypf_camp_df: pd.DataFrame,
    maestro_ids_final: set[Any],
    remaining_clave_negocio: set[Any],
    next_num_start: int,
    run_timestamp: dt.datetime,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Adapta cada fila de BASE CAMPAÑAS (YPF Etapa 2) al esquema canónico
    CAMPANAS y genera sus campos técnicos. Orden determinístico:
    (IDCampaña, ElementoID, FechaInicio, FechaFin) ascendente, antes de
    asignar CargaID secuenciales."""
    sorted_df = ypf_camp_df.sort_values(
        by=["IDCampaña", "ElementoID", "FechaInicio", "FechaFin"], kind="mergesort"
    ).reset_index(drop=True)

    incorporated: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    seen_pairs: set[tuple[Any, Any]] = set()
    seen_clave: set[str] = set()
    next_num = next_num_start

    for _, row in sorted_df.iterrows():
        eid = row["ElementoID"]
        reasons: list[str] = []

        tipo = elemento_id_tipo(eid)
        if is_blank(eid):
            reasons.append("ElementoID vacío")
        elif eid not in maestro_ids_final:
            reasons.append(f"ElementoID '{eid}' no existe en el MAESTRO_ELEMENTOS final (integridad referencial)")
        elif tipo == "FB":
            reasons.append("Prohibido: campaña digital asignada a un elemento FB (estático)")
        elif tipo not in YPF_DIGITAL_TYPES:
            reasons.append(f"ElementoID '{eid}' no es MB/PPUNTER/TT (único inventario digital YPF autorizado)")

        if not is_blank(eid) and es_elemento_digital_30943(eid):
            reasons.append("Prohibido: campaña digital sobre APIE 30943 (estación exclusivamente estática)")

        if is_blank(row.get("IDCampaña")):
            reasons.append("IDCampaña vacío")
        if is_blank(row.get("FechaInicio")):
            reasons.append("FechaInicio vacía")
        if is_blank(row.get("FechaFin")):
            reasons.append("FechaFin vacía")
        estado = row.get("Estado")
        if is_blank(estado):
            reasons.append("Estado vacío")
        elif estado not in vi.ESTADO_VALIDACION_VALUES and estado not in {"Reservada", "Activa", "Finalizada", "Cancelado"}:
            # Estado de CAMPANAS (comercial) valida contra PARAMETROS, no contra ESTADO_VALIDACION_VALUES;
            # se deja explícito para no depender de un dominio ausente en este contexto.
            pass

        pair = (row.get("IDCampaña"), eid)
        if pair in seen_pairs:
            reasons.append(f"Par (IDCampaña, ElementoID)={pair} duplicado dentro del propio lote YPF")

        clave = None
        if not reasons:
            fecha_indefinida = "No"
            clave = compute_clave_negocio(
                row["IDCampaña"], eid, row["FechaInicio"], row["FechaFin"], fecha_indefinida,
                None, None,
            )
            if clave in remaining_clave_negocio:
                reasons.append(f"ClaveNegocio ya existe en las campañas OCU26 conservadas: {clave!r}")
            elif clave in seen_clave:
                reasons.append(f"ClaveNegocio duplicada entre los propios registros nuevos YPF: {clave!r}")

        if reasons:
            rejected.append({**row.to_dict(), "motivo_rechazo": "; ".join(reasons)})
            continue

        seen_pairs.add(pair)
        seen_clave.add(clave)
        carga_id = f"{CARGA_ID_PREFIX}{next_num:0{CARGA_ID_WIDTH}d}"
        next_num += 1

        record = {h: None for h in CAMPANAS_HEADERS}
        record.update({
            "CargaID": carga_id,
            "ClaveNegocio": clave,
            "FechaHoraCarga": run_timestamp,
            "UsuarioCarga": NEW_USUARIO_CARGA,
            "FuenteCarga": NEW_FUENTE_CARGA,
            "EstadoValidacion": "OK",
            "ObservacionValidacion": None,
            "IDCampaña": row["IDCampaña"],
            "Campaña": row["Campaña"],
            "Cliente": row["Cliente"],
            "Marca": row["Marca"],
            "Agencia": row["Agencia"],
            "Proveedor": row["Proveedor"],
            "ElementoID": eid,
            "TipoCargaDeclarado": row["Medio"],
            "FechaInicio": row["FechaInicio"],
            "FechaFin": row["FechaFin"],
            "FechaIndefinida": "No",
            "Estado": row["Estado"],
            "CantidadUnidades": row["CantidadUnidades"],
            "DuracionSpotSeg": row["DuracionSpotSeg"],
            "SalidasVendidas": row["SalidasVendidas"],
            "ModalidadPauta": None,
            "PROGRAMATICA": None,
            "TipoExclusividad": None,
            "HoraInicio": None,
            "HoraFin": None,
            "CANJE": None,
            "Observaciones": row.get("ObservacionesComercial"),
            "FilaOrigen": None,
        })
        incorporated.append(record)

    return incorporated, rejected


# ---------------------------------------------------------------------------
# Plan completo (usado tanto por audit_ocu26_ypf.py como por build_ocu26_ypf.py)
# ---------------------------------------------------------------------------


@dataclass
class YpfPlan:
    run_timestamp: dt.datetime
    ocu_maestro_df: pd.DataFrame
    ocu_campanas_df: pd.DataFrame
    ocu_parametros_df: pd.DataFrame
    ypf_elem_df: pd.DataFrame
    ypf_camp_df: pd.DataFrame
    mapeo_df: pd.DataFrame
    elementos: ElementosCrossResult
    legacy_rows: list[LegacyRow]
    remaining_campanas_df: pd.DataFrame
    new_campanas_incorporated: list[dict[str, Any]]
    new_campanas_rejected: list[dict[str, Any]]
    expected_counts: dict[str, int]


def compute_plan(run_timestamp: dt.datetime | None = None) -> YpfPlan:
    if run_timestamp is None:
        run_timestamp = dt.datetime.now()

    ocu_maestro_df = read_table_df(FINAL_V2_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    ocu_campanas_df = read_table_df(FINAL_V2_PATH, "CAMPANAS", "tblCampanas")
    ocu_parametros_df = read_table_df(FINAL_V2_PATH, "PARAMETROS", "tblParametros")

    ypf_elem_df = read_raw_sheet_df(YPF_ETAPA1_PATH, "BASE ELEMENTOS")
    ypf_camp_df = read_raw_sheet_df(YPF_ETAPA2_PATH, "BASE CAMPAÑAS")
    mapeo_df = read_raw_sheet_df(YPF_ETAPA2_PATH, "MAPEO CAMPAÑAS")

    fuente_domain = set(ocu_parametros_df.loc[ocu_parametros_df["Categoria"] == "FuenteCarga", "Valor"])
    if NEW_FUENTE_CARGA not in fuente_domain:
        raise ValueError(f"FuenteCarga {NEW_FUENTE_CARGA!r} no está en el dominio permitido de PARAMETROS: {fuente_domain}")

    elementos = classify_elementos(ocu_maestro_df, ypf_elem_df)

    ocu_maestro_ids = set(ocu_maestro_df[MAESTRO_KEY].dropna())
    maestro_ids_final = ocu_maestro_ids | elementos.solo_ypf

    legacy_rows = classify_legacy_block(
        ocu_campanas_df, ocu_maestro_ids, elementos.ypf_autorizados, ypf_camp_df, mapeo_df
    )

    remaining_campanas_df = ocu_campanas_df[
        ~((ocu_campanas_df["IDCampaña"] >= LEGACY_ID_MIN) & (ocu_campanas_df["IDCampaña"] <= LEGACY_ID_MAX))
    ].copy()

    next_num = next_carga_id_sequence(ocu_campanas_df)
    remaining_clave_negocio = set(remaining_campanas_df["ClaveNegocio"].dropna())

    incorporated, rejected = build_new_campanas_records(
        ypf_camp_df, maestro_ids_final, remaining_clave_negocio, next_num, run_timestamp
    )

    expected_counts = {
        "MAESTRO_ELEMENTOS_antes": len(ocu_maestro_df),
        "MAESTRO_ELEMENTOS_nuevos": len(elementos.solo_ypf),
        "MAESTRO_ELEMENTOS_despues": len(ocu_maestro_df) + len(elementos.solo_ypf),
        "CAMPANAS_antes": len(ocu_campanas_df),
        "CAMPANAS_legacy_retiradas": len(legacy_rows),
        "CAMPANAS_protegidas_no_ypf": len(remaining_campanas_df),
        "CAMPANAS_ypf_insertadas": len(incorporated),
        "CAMPANAS_ypf_rechazadas": len(rejected),
        "CAMPANAS_despues": len(remaining_campanas_df) + len(incorporated),
        "PARAMETROS": len(ocu_parametros_df),
    }

    return YpfPlan(
        run_timestamp=run_timestamp,
        ocu_maestro_df=ocu_maestro_df,
        ocu_campanas_df=ocu_campanas_df,
        ocu_parametros_df=ocu_parametros_df,
        ypf_elem_df=ypf_elem_df,
        ypf_camp_df=ypf_camp_df,
        mapeo_df=mapeo_df,
        elementos=elementos,
        legacy_rows=legacy_rows,
        remaining_campanas_df=remaining_campanas_df,
        new_campanas_incorporated=incorporated,
        new_campanas_rejected=rejected,
        expected_counts=expected_counts,
    )


# ---------------------------------------------------------------------------
# Pase FINAL: retiro autorizado de los elementos digitales legacy de APIE
# 30943 (no pertenecen al catálogo YPF validado, cero campañas activas).
# Opera EXCLUSIVAMENTE sobre CANDIDATE_PATH (no re-deriva de las 3 fuentes
# originales), igual que el patrón ya establecido en
# Pendientes/OCU26_ACTUALIZACION/scripts/merge_final_common.py.
# ---------------------------------------------------------------------------


@dataclass
class FinalRemovalPlan:
    run_timestamp: dt.datetime
    candidate_maestro_df: pd.DataFrame
    candidate_campanas_df: pd.DataFrame
    candidate_parametros_df: pd.DataFrame
    ypf_autorizados: set[Any]
    elementos_a_retirar: list[dict[str, Any]]  # filas completas de MAESTRO_ELEMENTOS a retirar
    expected_counts: dict[str, int]


def compute_final_removal_plan(run_timestamp: dt.datetime | None = None) -> FinalRemovalPlan:
    """Identifica, sobre la CANDIDATA ya construida, los elementos digitales
    de APIE 30943 que no pertenecen al catálogo YPF Etapa 1 validado. Verifica
    (no asume) que ninguna fila de CAMPANAS de la candidata los referencia
    antes de proponer su retiro."""
    if run_timestamp is None:
        run_timestamp = dt.datetime.now()

    if not CANDIDATE_PATH.exists():
        raise FileNotFoundError(f"No existe la candidata a corregir: {CANDIDATE_PATH}")

    candidate_maestro_df = read_table_df(CANDIDATE_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    candidate_campanas_df = read_table_df(CANDIDATE_PATH, "CAMPANAS", "tblCampanas")
    candidate_parametros_df = read_table_df(CANDIDATE_PATH, "PARAMETROS", "tblParametros")

    ypf_elem_df = read_raw_sheet_df(YPF_ETAPA1_PATH, "BASE ELEMENTOS")
    ocu_maestro_df_original = read_table_df(FINAL_V2_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    elementos = classify_elementos(ocu_maestro_df_original, ypf_elem_df)
    ypf_autorizados = elementos.ypf_autorizados

    eids_a_retirar = sorted(
        e for e in candidate_maestro_df[MAESTRO_KEY]
        if str(e).startswith("30943 - ") and elemento_id_tipo(e) in YPF_DIGITAL_TYPES and e not in ypf_autorizados
    )

    referencing = candidate_campanas_df[candidate_campanas_df["ElementoID"].isin(eids_a_retirar)]
    if not referencing.empty:
        raise ValueError(
            f"No se puede retirar: {len(referencing)} fila(s) de CAMPANAS en la candidata referencian "
            f"alguno de estos ElementoID: {sorted(referencing['ElementoID'].unique())}. "
            "Retiro abortado (no se toca nada)."
        )

    idx = candidate_maestro_df.set_index(MAESTRO_KEY, drop=False)
    elementos_a_retirar = [idx.loc[e].to_dict() for e in eids_a_retirar]

    expected_counts = {
        "MAESTRO_ELEMENTOS_antes": len(candidate_maestro_df),
        "MAESTRO_ELEMENTOS_retirados": len(elementos_a_retirar),
        "MAESTRO_ELEMENTOS_despues": len(candidate_maestro_df) - len(elementos_a_retirar),
        "CAMPANAS": len(candidate_campanas_df),
        "PARAMETROS": len(candidate_parametros_df),
    }

    return FinalRemovalPlan(
        run_timestamp=run_timestamp,
        candidate_maestro_df=candidate_maestro_df,
        candidate_campanas_df=candidate_campanas_df,
        candidate_parametros_df=candidate_parametros_df,
        ypf_autorizados=ypf_autorizados,
        elementos_a_retirar=elementos_a_retirar,
        expected_counts=expected_counts,
    )
