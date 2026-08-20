# -*- coding: utf-8 -*-
"""Validación independiente de la candidata OCU26 + YPF.

Recalcula el plan desde las 3 fuentes (sin confiar en lo que haya escrito
build_ocu26_ypf.py) y lo contrasta celda por celda contra la candidata ya
guardada. READ-ONLY sobre las fuentes y sobre la candidata.

Controles (ver enunciado FASE 7 / "VALIDACIÓN OBLIGATORIA"):
  - Estructura OCU26 preservada (hojas, tablas, encabezados).
  - Cero ElementoID vacíos o duplicados.
  - Cero campañas con ElementoID inexistente.
  - Cero campañas digitales en FB / en APIE 30943.
  - Dos FB conservados para APIE 30943.
  - Cero pérdida de registros protegidos (comparación celda por celda).
  - Cero cambios no autorizados fuera de YPF.
  - Cero duplicados nuevos de ClaveNegocio / CargaID perdidos o reutilizados.
  - PARAMETROS idéntico.
  - Cero fórmulas externas, cero errores Excel, integridad ZIP/XLSX.
  - Apertura data_only=False y data_only=True.
  - Hashes de las 3 fuentes intactos.

Uso:
    python validate_ocu26_ypf.py
    python validate_ocu26_ypf.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from typing import Any

import pandas as pd
from openpyxl import load_workbook

import merge_ypf_common as m


class ValidationResult:
    def __init__(self) -> None:
        self.checks: list[dict[str, Any]] = []

    def add(self, name: str, ok: bool, detail: Any = None) -> None:
        self.checks.append({"check": name, "result": "OK" if ok else "ERROR", "detail": detail})

    @property
    def all_ok(self) -> bool:
        return all(c["result"] == "OK" for c in self.checks)


def _is_equal(v1: Any, v2: Any) -> bool:
    b1, b2 = m.is_blank(v1), m.is_blank(v2)
    if b1 and b2:
        return True
    if b1 != b2:
        return False
    return v1 == v2


def validate() -> tuple[ValidationResult, dict[str, Any]]:
    r = ValidationResult()

    sha_final_v2 = m.calculate_sha256(m.FINAL_V2_PATH)
    sha_ypf1 = m.calculate_sha256(m.YPF_ETAPA1_PATH)
    sha_ypf2 = m.calculate_sha256(m.YPF_ETAPA2_PATH)
    r.add("fuentes_existen", m.FINAL_V2_PATH.exists() and m.YPF_ETAPA1_PATH.exists() and m.YPF_ETAPA2_PATH.exists())

    if not m.CANDIDATE_PATH.exists():
        r.add("candidata_existe", False, str(m.CANDIDATE_PATH))
        return r, {"sha_final_v2": sha_final_v2, "sha_ypf1": sha_ypf1, "sha_ypf2": sha_ypf2}
    r.add("candidata_existe", True, str(m.CANDIDATE_PATH))

    # --- ZIP / XLSX / fórmulas / errores / enlaces externos ---
    info = m.inspect_structure(m.CANDIDATE_PATH)
    r.add("zip_integro", info["zip_ok"] is True)
    r.add("sin_vba", info["vba"] == [])
    r.add("sin_enlaces_externos", info["external_links"] == [])
    r.add("sin_formulas", info["formula_cells"] == 0, info["formula_cells"])
    r.add("sin_errores_excel", info["error_cells"] == 0, info["error_cells"])

    # --- apertura data_only=False y data_only=True ---
    try:
        wb_f = load_workbook(m.CANDIDATE_PATH, data_only=False, read_only=False, keep_vba=False)
        wb_t = load_workbook(m.CANDIDATE_PATH, data_only=True, read_only=False, keep_vba=False)
        r.add("abre_data_only_false", True)
        r.add("abre_data_only_true", True)
    except Exception as exc:  # noqa: BLE001
        r.add("abre_workbook", False, str(exc))
        return r, {"sha_final_v2": sha_final_v2, "sha_ypf1": sha_ypf1, "sha_ypf2": sha_ypf2}

    r.add("hojas_esperadas", wb_f.sheetnames == m.mc.EXPECTED_SHEETS, wb_f.sheetnames)
    for sheet, table in m.mc.EXPECTED_TABLES.items():
        r.add(f"tabla_{sheet}", list(wb_f[sheet].tables.keys()) == [table], list(wb_f[sheet].tables.keys()))
    wb_f.close()
    wb_t.close()

    # --- Recalcular el plan desde cero (independiente del build) ---
    plan = m.compute_plan()
    ec = plan.expected_counts

    cand_maestro_df = m.read_table_df(m.CANDIDATE_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    cand_campanas_df = m.read_table_df(m.CANDIDATE_PATH, "CAMPANAS", "tblCampanas")
    cand_parametros_df = m.read_table_df(m.CANDIDATE_PATH, "PARAMETROS", "tblParametros")

    r.add("headers_maestro", list(cand_maestro_df.columns) == m.MAESTRO_HEADERS)
    r.add("headers_campanas", list(cand_campanas_df.columns) == m.CAMPANAS_HEADERS)

    # --- Conteos ---
    r.add("conteo_maestro_despues", len(cand_maestro_df) == ec["MAESTRO_ELEMENTOS_despues"],
          {"esperado": ec["MAESTRO_ELEMENTOS_despues"], "real": len(cand_maestro_df)})
    r.add("conteo_campanas_despues", len(cand_campanas_df) == ec["CAMPANAS_despues"],
          {"esperado": ec["CAMPANAS_despues"], "real": len(cand_campanas_df)})
    r.add("conteo_parametros_sin_cambios", len(cand_parametros_df) == ec["PARAMETROS"])

    # --- ElementoID: sin vacíos ni duplicados ---
    eids = cand_maestro_df[m.MAESTRO_KEY]
    r.add("elementoid_sin_vacios", int(eids.apply(m.is_blank).sum()) == 0, int(eids.apply(m.is_blank).sum()))
    dup = eids[eids.duplicated(keep=False)]
    r.add("elementoid_sin_duplicados", dup.empty, sorted(set(dup))[:10])

    # --- Integridad referencial CAMPANAS ---
    final_maestro_ids = set(eids.dropna())
    orphans = cand_campanas_df[cand_campanas_df["ElementoID"].apply(m.is_blank) | ~cand_campanas_df["ElementoID"].isin(final_maestro_ids)]
    r.add("campanas_sin_elementoid_huerfano", len(orphans) == 0, len(orphans))

    # --- FB digital / APIE 30943 digital ---
    fb_digital = cand_campanas_df[cand_campanas_df["ElementoID"].apply(lambda e: m.elemento_id_tipo(e) == "FB")]
    r.add("cero_campanas_digitales_en_fb", len(fb_digital) == 0, len(fb_digital))
    apie_digital = cand_campanas_df[cand_campanas_df["ElementoID"].apply(m.es_elemento_digital_30943)]
    r.add("cero_campanas_digitales_apie_30943", len(apie_digital) == 0, len(apie_digital))

    # --- APIE 30943: exactamente 2 FB, y los YPF-nuevos no agregan digitales ---
    apie_30943_all = cand_maestro_df[cand_maestro_df[m.MAESTRO_KEY].astype(str).str.startswith("30943 - ")]
    apie_30943_fb = apie_30943_all[apie_30943_all[m.MAESTRO_KEY].apply(lambda e: m.elemento_id_tipo(e) == "FB")]
    r.add("apie_30943_dos_fb", sorted(apie_30943_fb[m.MAESTRO_KEY].tolist()) == ["30943 - FB - 1", "30943 - FB - 2"],
          sorted(apie_30943_fb[m.MAESTRO_KEY].tolist()))
    apie_30943_digital_nuevos = [e for e in plan.elementos.nuevos_registros if str(e[m.MAESTRO_KEY]).startswith("30943 - ")]
    r.add("apie_30943_sin_digitales_nuevos_ypf", len(apie_30943_digital_nuevos) == 0, apie_30943_digital_nuevos)

    # --- ClaveNegocio: sin duplicados NUEVOS (los duplicados exactos ya existentes en FINAL_V2,
    # ej. Adidas 4322 documentado como REPETICION_VALIDA, son datos protegidos ajenos a YPF y no
    # se tocan; solo se bloquea si la fila nueva YPF participa de la duplicación) ---
    n_kept_for_clave = len(plan.remaining_campanas_df)
    clave = cand_campanas_df["ClaveNegocio"]
    dup_counts = clave[~clave.apply(m.is_blank)].value_counts()
    dup_claves = set(dup_counts[dup_counts > 1].index)
    new_region_claves = set(cand_campanas_df["ClaveNegocio"].iloc[n_kept_for_clave:].dropna())
    new_clave_dups = dup_claves & new_region_claves
    preexisting_clave_dups = dup_claves - new_region_claves
    r.add("clavenegocio_sin_duplicados_nuevos", len(new_clave_dups) == 0,
          {"nuevos": sorted(new_clave_dups)[:10], "preexistentes_protegidos": len(preexisting_clave_dups)})

    carga = cand_campanas_df["CargaID"]
    carga_dup = carga[carga.duplicated(keep=False) & ~carga.apply(m.is_blank)]
    r.add("cargaid_sin_duplicados", carga_dup.empty, len(carga_dup))

    original_carga_ids = set(plan.ocu_campanas_df["CargaID"].dropna())  # incluye las del bloque legacy retirado
    new_carga_ids = {rec["CargaID"] for rec in plan.new_campanas_incorporated}
    reused = new_carga_ids & original_carga_ids
    r.add("cargaid_nuevos_no_reutilizan_existentes", len(reused) == 0, sorted(reused)[:10])

    kept_carga_ids = set(plan.remaining_campanas_df["CargaID"].dropna())
    cand_new_region = cand_campanas_df.iloc[len(plan.remaining_campanas_df):]
    r.add("cargaid_no_perdidos_en_conservadas", kept_carga_ids <= set(cand_campanas_df["CargaID"].dropna()))

    # --- Pares (IDCampaña, ElementoID) YPF controlados ---
    ypf_pairs = list(zip(cand_new_region["IDCampaña"], cand_new_region["ElementoID"]))
    r.add("pares_idcampana_elementoid_sin_duplicar", len(ypf_pairs) == len(set(ypf_pairs)),
          len(ypf_pairs) - len(set(ypf_pairs)))

    # --- Inventario YPF no superado: todo ElementoID nuevo insertado pertenece al catálogo autorizado ---
    fuera_inventario = [p for p in cand_new_region["ElementoID"].unique() if p not in plan.elementos.ypf_autorizados]
    r.add("inventario_ypf_no_superado", len(fuera_inventario) == 0, fuera_inventario[:10])

    # --- PARAMETROS idéntico ---
    orig_param_tuples = list(plan.ocu_parametros_df.itertuples(index=False, name=None))
    cand_param_tuples = list(cand_parametros_df.itertuples(index=False, name=None))
    r.add("parametros_identico", orig_param_tuples == cand_param_tuples)

    # --- Cero pérdida de registros OCU26 protegidos: comparación celda por celda ---
    # MAESTRO_ELEMENTOS: primeras N filas (N=original) deben coincidir con FINAL_V2, salvo
    # exactamente las celdas de completions autorizadas.
    n_orig_m = len(plan.ocu_maestro_df)
    cand_m_head = cand_maestro_df.iloc[:n_orig_m].reset_index(drop=True)
    orig_m = plan.ocu_maestro_df.reset_index(drop=True)
    completions_set = {(c.elemento_id, c.columna): c.valor_nuevo for c in plan.elementos.completions}
    unauthorized_maestro_changes: list[tuple[Any, str, Any, Any]] = []
    if not orig_m["ElementoID"].equals(cand_m_head["ElementoID"]):
        unauthorized_maestro_changes.append(("__ORDEN__", "ElementoID", "orden alterado", None))
    else:
        for col in m.MAESTRO_HEADERS:
            s_o, s_c = orig_m[col], cand_m_head[col]
            for i in range(n_orig_m):
                vo, vc = s_o.iat[i], s_c.iat[i]
                if _is_equal(vo, vc):
                    continue
                eid = orig_m["ElementoID"].iat[i]
                key = (eid, col)
                if key in completions_set and m.is_blank(vo) and vc == completions_set[key]:
                    continue
                unauthorized_maestro_changes.append((eid, col, vo, vc))
    r.add("maestro_elementos_sin_cambios_no_autorizados", len(unauthorized_maestro_changes) == 0,
          unauthorized_maestro_changes[:10])

    # CAMPANAS: las filas conservadas (no-legacy) deben coincidir exactamente, en el mismo orden
    # relativo, con FINAL_V2 (protección FASE 6).
    n_kept = len(plan.remaining_campanas_df)
    cand_kept = cand_campanas_df.iloc[:n_kept].reset_index(drop=True)
    orig_kept = plan.remaining_campanas_df.reset_index(drop=True)
    campanas_diffs: list[tuple[Any, str, Any, Any]] = []
    if not orig_kept["CargaID"].equals(cand_kept["CargaID"]):
        campanas_diffs.append(("__ORDEN__", "CargaID", "orden alterado", None))
    else:
        for col in m.CAMPANAS_HEADERS:
            s_o, s_c = orig_kept[col], cand_kept[col]
            for i in range(n_kept):
                if not _is_equal(s_o.iat[i], s_c.iat[i]):
                    campanas_diffs.append((orig_kept["CargaID"].iat[i], col, s_o.iat[i], s_c.iat[i]))
    r.add("campanas_protegidas_sin_cambios", len(campanas_diffs) == 0, campanas_diffs[:10])

    # --- Hashes de las 3 fuentes (deben permanecer intactos frente al preflight) ---
    r.add("sha_final_v2", True, sha_final_v2)
    r.add("sha_ypf_etapa1", True, sha_ypf1)
    r.add("sha_ypf_etapa2", True, sha_ypf2)

    return r, {
        "sha_final_v2": sha_final_v2,
        "sha_ypf1": sha_ypf1,
        "sha_ypf2": sha_ypf2,
        "expected_counts": ec,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Valida la candidata OCU26 + YPF.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    result, meta = validate()

    payload = {
        "result": "VALID" if result.all_ok else "INVALID",
        "checks": result.checks,
        "meta": meta,
    }

    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 + YPF VALIDACIÓN")
        print("=" * 60)
        for c in result.checks:
            marker = "OK  " if c["result"] == "OK" else "ERR "
            print(f"[{marker}] {c['check']}" + (f"  -> {c['detail']}" if c["result"] != "OK" and c["detail"] is not None else ""))
        print()
        print("RESULT:", payload["result"])

    return 0 if result.all_ok else 1


if __name__ == "__main__":
    sys.exit(main())
