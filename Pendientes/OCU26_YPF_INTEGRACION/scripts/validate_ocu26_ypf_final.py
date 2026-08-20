# -*- coding: utf-8 -*-
"""Validación independiente de la versión FINAL de OCU26 + YPF.

Recalcula el retiro desde la CANDIDATA (sin confiar en build_ocu26_ypf_final.py)
y lo contrasta celda por celda contra la FINAL ya guardada. READ-ONLY sobre
la candidata, la FINAL y las 3 fuentes originales.

Uso:
    python validate_ocu26_ypf_final.py
    python validate_ocu26_ypf_final.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from typing import Any

from openpyxl import load_workbook

import merge_ypf_common as m


class ValidationResult:
    def __init__(self) -> None:
        self.checks: list[dict[str, Any]] = []

    def add(self, name: str, ok: bool, detail: Any = None) -> None:
        self.checks.append({"check": name, "result": "OK" if ok else "ERROR", "detail": detail})

    @property
    def all_ok(self) -> bool:
        return all(c["result"] == "OK" for c in self.checks)


def _is_equal(v1: Any, v2: Any) -> bool:
    b1, b2 = m.is_blank(v1), m.is_blank(v2)
    if b1 and b2:
        return True
    if b1 != b2:
        return False
    return v1 == v2


def validate() -> tuple[ValidationResult, dict[str, Any]]:
    r = ValidationResult()

    sha_final_v2 = m.calculate_sha256(m.FINAL_V2_PATH)
    sha_ypf1 = m.calculate_sha256(m.YPF_ETAPA1_PATH)
    sha_ypf2 = m.calculate_sha256(m.YPF_ETAPA2_PATH)
    sha_candidate = m.calculate_sha256(m.CANDIDATE_PATH)
    r.add("fuentes_y_candidata_existen", m.FINAL_V2_PATH.exists() and m.YPF_ETAPA1_PATH.exists()
          and m.YPF_ETAPA2_PATH.exists() and m.CANDIDATE_PATH.exists())

    if not m.FINAL_PATH.exists():
        r.add("final_existe", False, str(m.FINAL_PATH))
        return r, {"sha_final_v2": sha_final_v2, "sha_ypf1": sha_ypf1, "sha_ypf2": sha_ypf2, "sha_candidate": sha_candidate}
    r.add("final_existe", True, str(m.FINAL_PATH))

    info = m.inspect_structure(m.FINAL_PATH)
    r.add("zip_integro", info["zip_ok"] is True)
    r.add("sin_vba", info["vba"] == [])
    r.add("sin_enlaces_externos", info["external_links"] == [])
    r.add("sin_formulas", info["formula_cells"] == 0, info["formula_cells"])
    r.add("sin_errores_excel", info["error_cells"] == 0, info["error_cells"])

    try:
        wb_f = load_workbook(m.FINAL_PATH, data_only=False, read_only=False, keep_vba=False)
        wb_t = load_workbook(m.FINAL_PATH, data_only=True, read_only=False, keep_vba=False)
        r.add("abre_data_only_false", True)
        r.add("abre_data_only_true", True)
    except Exception as exc:  # noqa: BLE001
        r.add("abre_workbook", False, str(exc))
        return r, {"sha_final_v2": sha_final_v2, "sha_ypf1": sha_ypf1, "sha_ypf2": sha_ypf2, "sha_candidate": sha_candidate}

    r.add("hojas_esperadas", wb_f.sheetnames == m.mc.EXPECTED_SHEETS, wb_f.sheetnames)
    for sheet, table in m.mc.EXPECTED_TABLES.items():
        r.add(f"tabla_{sheet}", list(wb_f[sheet].tables.keys()) == [table], list(wb_f[sheet].tables.keys()))
    wb_f.close()
    wb_t.close()

    plan = m.compute_final_removal_plan()
    ec = plan.expected_counts

    final_maestro_df = m.read_table_df(m.FINAL_PATH, "MAESTRO_ELEMENTOS", "tblElementos")
    final_campanas_df = m.read_table_df(m.FINAL_PATH, "CAMPANAS", "tblCampanas")
    final_parametros_df = m.read_table_df(m.FINAL_PATH, "PARAMETROS", "tblParametros")

    r.add("headers_maestro", list(final_maestro_df.columns) == m.MAESTRO_HEADERS)
    r.add("headers_campanas", list(final_campanas_df.columns) == m.CAMPANAS_HEADERS)

    r.add("conteo_maestro_5304", len(final_maestro_df) == 5304, len(final_maestro_df))
    r.add("conteo_campanas_15333", len(final_campanas_df) == 15333, len(final_campanas_df))
    r.add("conteo_parametros_23", len(final_parametros_df) == 23, len(final_parametros_df))
    r.add("conteo_maestro_coincide_plan", len(final_maestro_df) == ec["MAESTRO_ELEMENTOS_despues"])

    # --- ElementoID: sin vacíos ni duplicados ---
    eids = final_maestro_df[m.MAESTRO_KEY]
    r.add("elementoid_sin_vacios", int(eids.apply(m.is_blank).sum()) == 0)
    dup = eids[eids.duplicated(keep=False)]
    r.add("elementoid_sin_duplicados", dup.empty, sorted(set(dup))[:10])

    # --- Los 17 retirados ya no están; nada más cambió de tamaño ---
    eids_retirados = {row[m.MAESTRO_KEY] for row in plan.elementos_a_retirar}
    r.add("elementos_retirados_ausentes", eids_retirados.isdisjoint(set(eids)), sorted(eids_retirados & set(eids)))

    # --- APIE 30943: exactamente 2, ambos FB ---
    apie_30943 = sorted(e for e in eids if str(e).startswith("30943 - "))
    r.add("apie_30943_dos_fb", apie_30943 == ["30943 - FB - 1", "30943 - FB - 2"], apie_30943)

    # --- Integridad referencial: cero huérfanos, cero campañas en APIE 30943 ---
    final_ids_set = set(eids.dropna())
    orphans = final_campanas_df[final_campanas_df["ElementoID"].apply(m.is_blank) | ~final_campanas_df["ElementoID"].isin(final_ids_set)]
    r.add("campanas_sin_elementoid_huerfano", len(orphans) == 0, len(orphans))
    apie_camp = final_campanas_df[final_campanas_df["ElementoID"].astype(str).str.startswith("30943 - ")]
    r.add("cero_campanas_apie_30943", len(apie_camp) == 0, len(apie_camp))

    # --- PARAMETROS idéntico a la candidata ---
    cand_param_tuples = list(plan.candidate_parametros_df.itertuples(index=False, name=None))
    final_param_tuples = list(final_parametros_df.itertuples(index=False, name=None))
    r.add("parametros_identico_a_candidata", cand_param_tuples == final_param_tuples)

    # --- CAMPANAS: 100% idéntico a la candidata (no se tocó) ---
    cand_campanas_tuples = list(plan.candidate_campanas_df.itertuples(index=False, name=None))
    final_campanas_tuples = list(final_campanas_df.itertuples(index=False, name=None))
    campanas_diffs = 0
    if len(cand_campanas_tuples) == len(final_campanas_tuples):
        for a, b in zip(cand_campanas_tuples, final_campanas_tuples):
            if not all(_is_equal(x, y) for x, y in zip(a, b)):
                campanas_diffs += 1
    else:
        campanas_diffs = -1
    r.add("campanas_identica_a_candidata", campanas_diffs == 0, campanas_diffs)

    # --- MAESTRO_ELEMENTOS: todo lo que no fue retirado debe coincidir exactamente
    # con la candidata, en el mismo orden relativo (cero cambios no autorizados) ---
    cand_kept = plan.candidate_maestro_df[~plan.candidate_maestro_df[m.MAESTRO_KEY].isin(eids_retirados)].reset_index(drop=True)
    final_reset = final_maestro_df.reset_index(drop=True)
    unauthorized = []
    if not cand_kept[m.MAESTRO_KEY].equals(final_reset[m.MAESTRO_KEY]):
        unauthorized.append(("__ORDEN_O_CONTENIDO__", None, None, None))
    else:
        for col in m.MAESTRO_HEADERS:
            s_c, s_f = cand_kept[col], final_reset[col]
            for i in range(len(cand_kept)):
                if not _is_equal(s_c.iat[i], s_f.iat[i]):
                    unauthorized.append((cand_kept[m.MAESTRO_KEY].iat[i], col, s_c.iat[i], s_f.iat[i]))
    r.add("maestro_sin_cambios_no_autorizados", len(unauthorized) == 0, unauthorized[:10])

    # --- Hashes ---
    sha_final_v2_after = m.calculate_sha256(m.FINAL_V2_PATH)
    sha_ypf1_after = m.calculate_sha256(m.YPF_ETAPA1_PATH)
    sha_ypf2_after = m.calculate_sha256(m.YPF_ETAPA2_PATH)
    sha_candidate_after = m.calculate_sha256(m.CANDIDATE_PATH)
    r.add("sha_final_v2_intacto", sha_final_v2_after == sha_final_v2, sha_final_v2_after)
    r.add("sha_ypf_etapa1_intacto", sha_ypf1_after == sha_ypf1, sha_ypf1_after)
    r.add("sha_ypf_etapa2_intacto", sha_ypf2_after == sha_ypf2, sha_ypf2_after)
    r.add("sha_candidata_intacta", sha_candidate_after == sha_candidate, sha_candidate_after)

    return r, {
        "sha_final_v2": sha_final_v2_after,
        "sha_ypf1": sha_ypf1_after,
        "sha_ypf2": sha_ypf2_after,
        "sha_candidate": sha_candidate_after,
        "sha_final": m.calculate_sha256(m.FINAL_PATH),
        "expected_counts": ec,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Valida la versión FINAL de OCU26 + YPF.")
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    result, meta = validate()

    payload = {
        "result": "VALID" if result.all_ok else "INVALID",
        "checks": result.checks,
        "meta": meta,
    }

    if args.json:
        print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))
    else:
        print("=" * 60)
        print("OCU26 + YPF FINAL VALIDACIÓN")
        print("=" * 60)
        for c in result.checks:
            marker = "OK  " if c["result"] == "OK" else "ERR "
            print(f"[{marker}] {c['check']}" + (f"  -> {c['detail']}" if c["result"] != "OK" and c["detail"] is not None else ""))
        print()
        print("RESULT:", payload["result"])

    return 0 if result.all_ok else 1


if __name__ == "__main__":
    sys.exit(main())
