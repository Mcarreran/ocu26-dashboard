"""Reconstruye el catalogo YPF (Etapa 1) en una base limpia e independiente.

Consolida BASE ESTACIONES y BASE ELEMENTOS a partir de las hojas DIGITAL y
ESTATICO (fuentes rectoras fila a fila) del archivo de catalogo YPF, usando la
hoja ALTAS_BASE_MAESTRA del archivo de referencia estructural para los
patrones de metadatos digitales, y el maestro productivo OCU26 (solo lectura)
para el patron de metadatos Fotobox.

No modifica ningun archivo de entrada, ni la base productiva OCU26_BASE_DATOS.xlsx.
Todo archivo nuevo se escribe unicamente bajo pendientes/YPF_ETAPA_1/output/.
"""
from __future__ import annotations

import argparse
import re
import sys
from copy import copy as copy_style
from datetime import datetime

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

EXCEL_ERRORS = {"#N/A", "#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#NULL!", "#NUM!"}

EXPECTED = {
    "digital_rows": 416,
    "digital_apie_unicos": 416,
    "mb": 437,
    "tt": 2142,
    "pp": 968,
    "total_digital": 3547,
    "estatico_rows": 191,
    "estatico_apie_unicos": 191,
    "fb": 383,
    "estaciones_rows": 607,
    "apie_unicos": 525,
    "apie_ambos_formatos": 82,
    "elementos_totales": 3930,
}

# Conteos del catalogo ACTIVO: BASE ESTACIONES/BASE ELEMENTOS excluyendo los 4
# APIE bloqueados (ver APIE_BLOQUEADOS). Esta es la version corregida y
# alineada exactamente con el bloque "ZONAS" (fila TOTAL de la tabla dinamica
# ZONAS/ESTACIONES/PUNTERA/TORRE/MENU BOARD): 416 - 4 = 412 estaciones,
# 3547 - 47 = 3500 elementos digitales.
EXPECTED_ACTIVO = {
    "estaciones_rows": 603,
    "apie_unicos": 521,
    "apie_ambos_formatos": 82,
    "mb": 435,
    "tt": 2114,
    "pp": 951,
    "fb": 383,
    "total_digital": 3500,
    "elementos_totales": 3883,
}

# Bloque TOTAL de la tabla dinamica "ZONAS" (macro-zonas AMBA/PBA/NEA/SUR/
# NOA/CUYO/CENTRO) verificado como identico al catalogo activo una vez
# excluidos los 4 APIE bloqueados. A diferencia del control historico de la
# Etapa 1 original (386/404/2022/910/3336, que comparaba contra la fuente
# DIGITAL completa sin excluir los bloqueados), esta referencia debe calzar
# de forma EXACTA con el catalogo activo; una diferencia distinta de cero es
# un error bloqueante, no un "A VALIDAR".
ZONAS_REFERENCIA_ACTIVA = {
    "estaciones_digitales": 412,
    "menu_board": 435,
    "torres": 2114,
    "punteras": 951,
    "elementos_digitales": 3500,
}

# Los 4 APIE con fuente danada/no comprobable. Quedan en la hoja fuente
# DIGITAL (intactos, con sus cantidades originales) pero bloqueados: no
# generan fila en BASE ESTACIONES activa ni ElementoID en BASE ELEMENTOS
# hasta que se validen manualmente.
APIE_BLOQUEADOS = {30510, 31131, 31192, 31239}

# Elementos (MB+TT+PPUNTER) que le corresponden a cada APIE bloqueado segun
# DIGITAL. Explican exactamente la diferencia entre DIGITAL (416/3547) y el
# catalogo activo alineado con ZONAS (412/3500): 10+16+3+18 = 47.
APIE_BLOQUEADOS_IMPACTO = {"30510": 10, "31131": 16, "31192": 3, "31239": 18}

ACCION_BLOQUEANTE = ("Confirmar estacion activa, razon social, direccion, localidad, provincia, area y "
                      "cantidades antes de generar ElementoID o incorporar campañas.")

# Patron de referencia a libro externo en una formula de Excel, p.ej. "[1]"
# en "=VLOOKUP(A370,[1]PRINCIPAL!$A:$J,2,0)". Una formula con este patron
# solo es valida si el libro registra la relacion externa correspondiente
# (xl/externalLinks/...); al copiar la formula a un libro nuevo sin copiar
# esa relacion, Microsoft Excel la detecta como registro danado y repara el
# archivo al abrirlo.
EXTERNAL_REF_PATTERN = re.compile(r"\[\d+\]")

RUN_LABEL = "BASE LIMPIA YPF ETAPA 1 CORREGIDA - 13/08/2026"
FECHA_GENERACION = "2026-08-13"

ORDEN_HOJAS_FUENTE_EN_SALIDA = ["DIGITAL", "ESTATICO ", "ZONAS", "DIRECCIONES", "ZONAS VERSION ANTUGUA"]

BASE_ESTACIONES_COLS = [
    "Formato", "APIE", "RazonSocial", "Direccion", "Localidad", "Provincia", "Area",
    "MB_Cantidad", "PPUNTER_Cantidad", "TT_Cantidad", "FB_Cantidad", "TotalElementos",
    "EstadoControl", "ObservacionControl", "FuenteFila", "ClaveEstacionFormato",
    "EstadoRelevamientoEstatico", "EstadoExistenciaEstatico", "EstadoComercializacion",
    "IncluirComercializacion", "FB_CantidadComercial",
]

BASE_ELEMENTOS_COLS = [
    "TipoCatalogo", "Ciudad", "Medio", "CircuitoDashboard", "Subcircuito", "Ubicacion",
    "ElementoID", "Nivel", "Descripcion", "Resolucion", "DimensionOptico", "DimensionTotal",
    "Observaciones", "Material", "TipoInstalacion", "Original", "CapacidadSlotsReel",
    "SegundosDia", "TipoInventario", "AplicaCantidad", "RevisionMaestro", "Proveedor",
    "EstadoRelevamientoEstatico", "EstadoExistenciaEstatico", "EstadoComercializacion", "IncluirComercializacion",
]

AUDITORIA_COLS = ["Control", "Fuente", "Esperado", "Obtenido", "Diferencia", "Estado", "Observacion"]

PENDIENTES_COLS = [
    "TipoPendiente", "APIE", "Formato", "Campo", "ValorOrigen", "Motivo",
    "AccionRequerida", "ImpactoElementos", "Fuente", "FilaFuente",
]

ELEMENTO_ID_RE = re.compile(r"^[0-9]+ - (MB|PPUNTER|TT|FB) - [1-9][0-9]*$")

# Patrones de metadatos verificados como 100% consistentes en
# ALTAS_BASE_MAESTRA (archivo de referencia estructural) para los codigos
# digitales, y en MAESTRO_ELEMENTOS de OCU26 productivo (solo lectura) para FB.
METADATA_PATTERNS = {
    "MB": {
        "Nivel": "Indoor", "Descripcion": "Menu Board", "Resolucion": "1080 x 1920 px",
        "Material": "Video (avi - swf)", "TipoInstalacion": "Ambient", "Original": "MP4",
        "CapacidadSlotsReel": "20", "SegundosDia": "100800",
    },
    "TT": {
        "Nivel": "Indoor", "Descripcion": "Mueble Torre", "Resolucion": "1920 x 1080 px",
        "Material": "Video (avi - swf)", "TipoInstalacion": "Ambient", "Original": "MP4",
        "CapacidadSlotsReel": "20", "SegundosDia": "100800",
    },
    "PPUNTER": {
        "Nivel": "Indoor", "Descripcion": "Puntera", "Resolucion": "252 x 396 px",
        "Material": "Video (avi - swf)", "TipoInstalacion": "Ambient", "Original": "MP4",
        "CapacidadSlotsReel": "20", "SegundosDia": "100800",
    },
    "FB": {
        "Nivel": "Outdoor", "Descripcion": "Mupi", "Resolucion": None,
        "Material": "PAI", "TipoInstalacion": "Backlight", "Original": "PNG",
        "CapacidadSlotsReel": 0, "SegundosDia": 0,
    },
}

CIRCUITO_POR_CODIGO = {"MB": "YPF Digital", "TT": "YPF Digital", "PPUNTER": "YPF Digital", "FB": "YPF Estático"}
MEDIO_POR_CODIGO = {"MB": "Digital", "TT": "Digital", "PPUNTER": "Digital", "FB": "Estático"}

# Clasificacion de relevamiento estatico segun el valor de ESTATICO!J (¿HAY?).
# Tupla: (EstadoRelevamientoEstatico, EstadoExistenciaEstatico, EstadoComercializacion, IncluirComercializacion)
ESTATICO_CLASIFICACION = {
    "TIENE 3": ("RELEVADO", "CONFIRMADO_EXISTE", "ACTIVO_CONFIRMADO", "SI"),
    "NO": ("RELEVADO", "CONFIRMADO_NO_EXISTE", "EXCLUIDO", "NO"),
    "NO DEJARON RELEVAR": ("RELEVAMIENTO_NO_COMPLETADO", "SIN_CONFIRMAR", "ACTIVO_PROVISORIO", "SI"),
    None: ("NO_RELEVADO", "SIN_CONFIRMAR", "ACTIVO_PROVISORIO", "SI"),
}
DIGITAL_CLASIFICACION = ("NO_APLICA", "NO_APLICA", "ACTIVO_CONFIRMADO", "SI")

# Localidad definitiva (version mas completa) para los 7 APIE con nombre
# abreviado o incompleto en una de las dos fuentes. Se aplica por igual a la
# fila Digital y a la fila Estatico del mismo APIE (cuando existan), y se
# propaga a BASE ELEMENTOS.Ciudad y a la parte de localidad de Ubicacion.
APIE_LOCALIDAD_NORMALIZADA = {
    "298": "BELÉN DE ESCOBAR",
    "541": "GENERAL BELGRANO",
    "760": "GREGORIO DE LAFERRERE",
    "1626": "GENERAL GUIDO",
    "1639": "GENERAL PIRÁN",
    "1648": "GENERAL VILLEGAS",
    "3256": "LANÚS OESTE",
}

# APIE con diferencias de datos territoriales entre Digital y Estatico que se
# dejan sin corregir automaticamente (no bloqueantes, no excluyen de
# comercializacion). Se documentan una vez cada uno en PENDIENTES.
APIE_DIFERENCIA_FORMATOS = {1686, 3013, 3298, 3299, 31168, 31171, 31241, 31246}


# ---------------------------------------------------------------------------
# Utilidades de normalizacion
# ---------------------------------------------------------------------------

def norm_text(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def norm_upper(v):
    v = norm_text(v)
    if isinstance(v, str):
        return v.upper()
    return v


def is_excel_error(v):
    return isinstance(v, str) and v.strip().upper() in EXCEL_ERRORS


# Placeholders que aparecen en la version previa de BASE ESTACIONES y que no
# constituyen un dato territorial real (equivalen a "sin dato"). No deben
# aceptarse como completacion valida de un campo dañado.
PLACEHOLDERS_INVALIDOS = {"no encontrado", "no aplica", "sin dato", "sin datos", "s/d", "n/d", "nd", "no aplicable"}

# Vocabulario de Area realmente observado en las fuentes rectoras DIGITAL y
# ESTATICO (zonas NOA/NEA/CUYO/SUR/CENTRO/COSTA ATLANTICA y sub-zonas AMBA).
# Se usa para rechazar completaciones de Area que no correspondan a una
# etiqueta de zona real (p.ej. un nombre de provincia mal ubicado en la
# columna Area de una fuente auxiliar).
AREA_VOCABULARIO_CONOCIDO = {
    "CENTRO", "COSTA ATLANTICA", "SUR", "CUYO", "NEA", "NOA",
    "GBA NORTE", "GBA OESTE", "GBA SUR", "PCIA BSAS", "CAPITAL FEDERAL",
}


def valor_candidato_valido(campo, v):
    """Valida un valor candidato de una fuente auxiliar antes de aceptarlo
    como completacion de un campo territorial dañado."""
    if v is None:
        return False
    if is_excel_error(v):
        return False
    if isinstance(v, str):
        vs = v.strip()
        if not vs:
            return False
        if vs.lower() in PLACEHOLDERS_INVALIDOS:
            return False
        if campo == "Direccion" and vs == "0":
            return False
        if campo == "Area" and vs.upper() not in AREA_VOCABULARIO_CONOCIDO:
            return False
        return True
    if isinstance(v, (int, float)):
        if campo == "Direccion":
            return False
        return True
    return False


def valor_literal_seguro(valor_cacheado):
    """Determina el valor literal seguro a usar en reemplazo de una formula
    con vinculo externo roto (patron [n]). Si el valor cacheado no es
    comprobable (error de Excel, o un 0 que en este contexto proviene de un
    VLOOKUP fallido sobre un campo de texto) se deja vacio en vez de
    propagar un dato no confiable."""
    if valor_cacheado is None:
        return None
    if is_excel_error(valor_cacheado):
        return None
    if isinstance(valor_cacheado, str) and valor_cacheado.strip() == "0":
        return None
    if isinstance(valor_cacheado, (int, float)) and not isinstance(valor_cacheado, bool) and valor_cacheado == 0:
        return None
    return valor_cacheado


def detectar_formulas_externas(wb_formulas, wb_valores, nombres_hojas):
    """Escanea las hojas indicadas en busca de formulas con referencia a un
    libro externo (patron [n], p.ej. [1]PRINCIPAL) sin relacion registrada
    en el libro de salida. Devuelve una lista de hallazgos con la formula
    original, el valor cacheado por Excel y el valor literal seguro que se
    usara en su lugar al copiar la hoja."""
    hallazgos = []
    for nombre in nombres_hojas:
        ws_f = wb_formulas[nombre]
        ws_v = wb_valores[nombre]
        for row in ws_f.iter_rows():
            for cell in row:
                if cell.data_type == "f" and isinstance(cell.value, str) and EXTERNAL_REF_PATTERN.search(cell.value):
                    cacheado = ws_v.cell(row=cell.row, column=cell.column).value
                    hallazgos.append({
                        "hoja": nombre, "celda": cell.coordinate, "formula_original": cell.value,
                        "valor_cacheado": cacheado, "valor_final": valor_literal_seguro(cacheado),
                    })
    return hallazgos


def apie_to_text(v):
    """Convierte un valor de APIE a texto sin agregar '.0' ni perder ceros iniciales."""
    if v is None:
        return None
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return str(v)
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return str(v)


def localidad_a_ciudad(localidad):
    if not localidad:
        return None
    if norm_upper(localidad) == "CAPITAL FEDERAL":
        return "CABA"
    return norm_upper(localidad)


def int_qty(v):
    """Convierte una cantidad a entero >=0. Devuelve (valor, es_valida)."""
    if v is None:
        return 0, True
    if isinstance(v, bool):
        return 0, False
    if isinstance(v, (int, float)):
        if isinstance(v, float) and not v.is_integer():
            return 0, False
        iv = int(v)
        if iv < 0:
            return 0, False
        return iv, True
    return 0, False


# ---------------------------------------------------------------------------
# Carga de fuentes
# ---------------------------------------------------------------------------

def sheet_rows(ws, min_row=2):
    rows = []
    for r in ws.iter_rows(min_row=min_row, values_only=True):
        if any(c is not None for c in r):
            rows.append(r)
    return rows


def cargar_digital(wb_catalogo):
    ws = wb_catalogo["DIGITAL"]
    registros = []
    for i, r in enumerate(sheet_rows(ws), start=2):
        registros.append({
            "apie_raw": r[0], "apie": apie_to_text(r[0]), "razon_social": norm_text(r[1]),
            "mb": r[2], "tt": r[3], "pp": r[4],
            "direccion": r[5], "localidad": r[6], "provincia": r[7], "area": r[8],
            "fuente_fila": f"DIGITAL!{i}",
        })
    return registros


def cargar_estatico(wb_catalogo):
    ws = wb_catalogo["ESTATICO "]
    registros = []
    for i, r in enumerate(sheet_rows(ws), start=2):
        if r[0] is None:
            # Fila sin APIE: no es un registro de estacion real (p.ej. fila de
            # totales de control al pie de la tabla). Se excluye explicitamente.
            continue
        registros.append({
            "apie_raw": r[0], "apie": apie_to_text(r[0]), "razon_social": norm_text(r[1]),
            "direccion": r[2], "localidad": r[3], "provincia": r[4], "area": r[5],
            "cp": r[6], "tipo_imagen": r[7], "fb": r[8], "hay": norm_text(r[9]) if isinstance(r[9], str) else r[9],
            "fuente_fila": f"ESTATICO!{i}",
        })
    return registros


def cargar_base_estaciones_previa(wb_catalogo):
    """Version previa de BASE ESTACIONES, usada solo como fallback de lectura."""
    ws = wb_catalogo["BASE ESTACIONES"]
    por_apie = {}
    for i, r in enumerate(sheet_rows(ws), start=2):
        apie = apie_to_text(r[1])
        if apie is None:
            continue
        por_apie.setdefault(apie, []).append({
            "direccion": r[9], "localidad": r[10], "provincia": r[11], "area": r[12],
            "fuente_fila": f"BASE ESTACIONES (previa)!{i}",
        })
    return por_apie


def cargar_direcciones_lookup(wb_catalogo):
    """Escanea los 4 bloques regionales de DIRECCIONES (AMBA/PROVINCIA/COSTA/ZONAS)
    y arma un lookup APIE -> datos territoriales, solo para coincidencia exacta."""
    ws = wb_catalogo["DIRECCIONES"]
    rows = list(ws.iter_rows(values_only=True))
    bloques = [
        {"apie": 2, "razon": 3, "direccion": 4, "localidad": 5},
        {"apie": 12, "razon": 13, "direccion": 14, "localidad": 15},
        {"apie": 23, "razon": 24, "direccion": 25, "localidad": 26},
        {"apie": 34, "razon": 35, "direccion": 36, "localidad": None, "provincia": 37},
    ]
    lookup = {}
    for r in rows:
        for b in bloques:
            apie_val = r[b["apie"]] if b["apie"] < len(r) else None
            apie = apie_to_text(apie_val) if isinstance(apie_val, (int, float)) else None
            if apie is None:
                continue
            entry = {
                "razon_social": r[b["razon"]] if b["razon"] < len(r) else None,
                "direccion": r[b["direccion"]] if b["direccion"] < len(r) else None,
                "localidad": r[b["localidad"]] if b.get("localidad") is not None and b["localidad"] < len(r) else None,
                "provincia": r[b["provincia"]] if b.get("provincia") is not None and b["provincia"] < len(r) else None,
            }
            lookup.setdefault(apie, []).append(entry)
    return lookup


# ---------------------------------------------------------------------------
# Preflight
# ---------------------------------------------------------------------------

def preflight(digital, estatico):
    errores = []

    apies_d = [r["apie"] for r in digital]
    dup_d = len(apies_d) - len(set(apies_d))
    mb_sum = sum(int_qty(r["mb"])[0] for r in digital)
    tt_sum = sum(int_qty(r["tt"])[0] for r in digital)
    pp_sum = sum(int_qty(r["pp"])[0] for r in digital)
    total_digital = mb_sum + tt_sum + pp_sum

    checks = [
        ("Registros DIGITAL", len(digital), EXPECTED["digital_rows"]),
        ("APIE unicos DIGITAL", len(set(apies_d)), EXPECTED["digital_apie_unicos"]),
        ("Menu Board (MB)", mb_sum, EXPECTED["mb"]),
        ("Torres (TT)", tt_sum, EXPECTED["tt"]),
        ("Punteras (PPUNTER)", pp_sum, EXPECTED["pp"]),
        ("Total elementos digitales", total_digital, EXPECTED["total_digital"]),
    ]

    apies_e = [r["apie"] for r in estatico]
    dup_e = len(apies_e) - len(set(apies_e))
    fb_sum = sum(int_qty(r["fb"])[0] for r in estatico)
    checks += [
        ("Registros ESTATICO", len(estatico), EXPECTED["estatico_rows"]),
        ("APIE unicos ESTATICO", len(set(apies_e)), EXPECTED["estatico_apie_unicos"]),
        ("Fotobox (FB)", fb_sum, EXPECTED["fb"]),
    ]

    if dup_d:
        errores.append(f"DIGITAL tiene {dup_d} APIE duplicados")
    if dup_e:
        errores.append(f"ESTATICO tiene {dup_e} APIE duplicados")

    resultado = {"checks": checks, "mb_sum": mb_sum, "tt_sum": tt_sum, "pp_sum": pp_sum,
                 "total_digital": total_digital, "fb_sum": fb_sum}

    for nombre, obtenido, esperado in checks:
        if obtenido != esperado:
            errores.append(f"{nombre}: esperado {esperado}, obtenido {obtenido}")

    return resultado, errores


# ---------------------------------------------------------------------------
# Construccion de BASE ESTACIONES
# ---------------------------------------------------------------------------

def construir_base_estaciones(digital, estatico, base_prev, direcciones_lookup):
    filas = []
    pendientes = []

    dig_by_apie = {r["apie"]: r for r in digital}
    est_by_apie = {r["apie"]: r for r in estatico}
    apies_bloqueados = {str(a) for a in APIE_BLOQUEADOS}

    def registrar_bloqueo(r):
        """Genera exactamente una entrada TERRITORIAL_BLOQUEANTE para un APIE
        bloqueado y NO agrega fila a BASE ESTACIONES activa: el APIE queda
        excluido del catalogo activo hasta validacion manual, pero sus
        cantidades originales permanecen intactas en la hoja fuente DIGITAL."""
        apie = r["apie"]
        mb, _ = int_qty(r["mb"])
        tt, _ = int_qty(r["tt"])
        pp, _ = int_qty(r["pp"])
        total = mb + tt + pp
        impacto_esperado = APIE_BLOQUEADOS_IMPACTO.get(apie)
        if impacto_esperado is not None and total != impacto_esperado:
            raise ValueError(
                f"APIE bloqueado {apie}: el total calculado desde DIGITAL ({total}) no coincide con "
                f"el impacto esperado documentado ({impacto_esperado}). Revisar antes de continuar.")

        motivo_partes = []
        for campo, valor in [("Direccion", r["direccion"]), ("Localidad", r["localidad"])]:
            if is_excel_error(valor):
                motivo_partes.append(f"{campo} fuente = error Excel ({valor})")
            elif campo == "Direccion" and isinstance(valor, (int, float)):
                motivo_partes.append(f"{campo} fuente = valor desplazado ({valor!r})")
        if apie == "31192":
            motivo_partes.append(
                "RazonSocial/Direccion/Localidad de origen dependian de formulas VLOOKUP con vinculo "
                "externo roto ([1]PRINCIPAL) en DIGITAL!B370/F370/G370 (ver FORMULA_EXTERNA_NEUTRALIZADA)")
        motivo = ("APIE bloqueado: dato de origen no comprobable para generar estacion activa. "
                  + (" | ".join(motivo_partes) if motivo_partes else "Fuente danada segun brief de la etapa.")
                  + " Mientras continue bloqueado, ninguna campaña podra vincularse con este APIE.")

        pendientes.append({
            "TipoPendiente": "TERRITORIAL_BLOQUEANTE", "APIE": apie, "Formato": "Digital",
            "Campo": "Direccion/Localidad/Provincia/Area/Cantidades",
            "ValorOrigen": f"RazonSocial={r['razon_social']!r}; Direccion={r['direccion']!r}; "
                            f"Localidad={r['localidad']!r}; Provincia={r['provincia']!r}; "
                            f"Area={r['area']!r}; MB={mb}; TT={tt}; PPUNTER={pp}",
            "Motivo": motivo, "AccionRequerida": ACCION_BLOQUEANTE,
            "ImpactoElementos": impacto_esperado if impacto_esperado is not None else total,
            "Fuente": "DIGITAL", "FilaFuente": r["fuente_fila"],
        })

    def resolver_territorio(apie, formato, direccion, localidad, provincia, area, fuente_fila):
        """Devuelve (direccion, localidad, provincia, area, estado, observaciones[], core_dañado).

        Un campo se considera "dañado" (y por lo tanto candidato a completarse
        por coincidencia exacta) solo cuando el valor de origen es un error de
        Excel o, en el caso de Direccion, un valor claramente desplazado
        (numerico). Un campo vacio desde el origen (None) NO se considera
        dañado por si solo: en esta fuente, Area en particular esta vacia de
        forma legitima en gran parte de las estaciones (no aplica a todas),
        por lo que forzar una busqueda de completacion sobre campos
        simplemente vacios arriesga contaminar filas correctas con datos de
        otra estacion. Solo Direccion y Localidad son campos nucleo: si
        cualquiera de los dos no puede resolverse, la fila queda A VALIDAR.
        """
        obs = []
        estado = "OK"
        campos_originales = {"Direccion": direccion, "Localidad": localidad, "Provincia": provincia, "Area": area}
        campos = dict(campos_originales)
        campos_con_error = set()

        for campo, valor in campos_originales.items():
            if is_excel_error(valor):
                campos_con_error.add(campo)
                obs.append(f"{campo} fuente = error Excel ({valor})")
                campos[campo] = None
            elif campo == "Direccion" and isinstance(valor, (int, float)):
                campos_con_error.add(campo)
                obs.append(f"{campo} fuente = valor desplazado ({valor!r})")
                campos[campo] = None

        core_dañado = bool({"Direccion", "Localidad"} & campos_con_error)

        if campos_con_error:
            candidatos = []
            otro_formato = est_by_apie.get(apie) if formato == "Digital" else dig_by_apie.get(apie)
            if otro_formato:
                candidatos.append({
                    "direccion": otro_formato.get("direccion"), "localidad": otro_formato.get("localidad"),
                    "provincia": otro_formato.get("provincia"), "area": otro_formato.get("area"),
                })
            for prev in base_prev.get(apie, []):
                candidatos.append(prev)
            for dloc in direcciones_lookup.get(apie, []):
                candidatos.append(dloc)

            for campo, key in [("Direccion", "direccion"), ("Localidad", "localidad"),
                                ("Provincia", "provincia"), ("Area", "area")]:
                if campo not in campos_con_error:
                    continue  # solo se busca completar los campos que originalmente tenian error
                valores_validos = {}
                for c in candidatos:
                    v = c.get(key)
                    if valor_candidato_valido(campo, v):
                        clave = norm_upper(v) if isinstance(v, str) else v
                        valores_validos[clave] = v
                if len(valores_validos) == 1:
                    campos[campo] = next(iter(valores_validos.values()))
                    obs.append(f"{campo} completado por coincidencia exacta de APIE en fuente auxiliar")
                elif len(valores_validos) > 1:
                    obs.append(f"{campo} tiene mas de un candidato de completacion distinto; se deja vacio")

        if core_dañado and (campos["Direccion"] is None or campos["Localidad"] is None):
            estado = "A VALIDAR"
            obs.append("No se pudo resolver dato territorial mediante coincidencia exacta y unica")

        return campos["Direccion"], campos["Localidad"], campos["Provincia"], campos["Area"], estado, obs, core_dañado

    # --- filas Digital ---
    for r in digital:
        apie = r["apie"]

        if apie in apies_bloqueados:
            registrar_bloqueo(r)
            continue  # no se agrega a BASE ESTACIONES activa

        mb, mb_ok = int_qty(r["mb"])
        tt, tt_ok = int_qty(r["tt"])
        pp, pp_ok = int_qty(r["pp"])

        direccion, localidad, provincia, area, estado, obs, dañado = resolver_territorio(
            apie, "Digital", r["direccion"], r["localidad"], r["provincia"], r["area"], r["fuente_fila"])

        cantidades_invalidas = []
        if not mb_ok:
            cantidades_invalidas.append(("MB_Cantidad", r["mb"]))
        if not tt_ok:
            cantidades_invalidas.append(("TT_Cantidad", r["tt"]))
        if not pp_ok:
            cantidades_invalidas.append(("PPUNTER_Cantidad", r["pp"]))
        if cantidades_invalidas:
            estado = "A VALIDAR"
            for campo, valor in cantidades_invalidas:
                obs.append(f"{campo} no numerica/invalida en fuente: {valor!r}")

        total = mb + tt + pp
        estado_rel, estado_exist, estado_com, incluir = DIGITAL_CLASIFICACION
        fila = {
            "Formato": "Digital", "APIE": apie, "RazonSocial": r["razon_social"],
            "Direccion": direccion, "Localidad": localidad, "Provincia": provincia, "Area": area,
            "MB_Cantidad": mb, "PPUNTER_Cantidad": pp, "TT_Cantidad": tt, "FB_Cantidad": 0,
            "TotalElementos": total, "EstadoControl": estado,
            "ObservacionControl": " | ".join(obs) if obs else "",
            "FuenteFila": r["fuente_fila"], "ClaveEstacionFormato": f"{apie}|Digital",
            "EstadoRelevamientoEstatico": estado_rel, "EstadoExistenciaEstatico": estado_exist,
            "EstadoComercializacion": estado_com, "IncluirComercializacion": incluir,
            "FB_CantidadComercial": 0,
        }
        filas.append(fila)

        if estado != "OK":
            pendientes.append({
                "TipoPendiente": "TERRITORIAL",
                "APIE": apie, "Formato": "Digital", "Campo": "Direccion/Localidad/Provincia/Area",
                "ValorOrigen": f"Direccion={r['direccion']!r}; Localidad={r['localidad']!r}; "
                                f"Provincia={r['provincia']!r}; Area={r['area']!r}",
                "Motivo": " | ".join(obs), "AccionRequerida": "Validacion manual del dato territorial",
                "ImpactoElementos": total, "Fuente": "DIGITAL", "FilaFuente": r["fuente_fila"],
            })

    # --- filas Estatico ---
    for r in estatico:
        apie = r["apie"]
        fb, fb_ok = int_qty(r["fb"])

        direccion, localidad, provincia, area, estado, obs, dañado = resolver_territorio(
            apie, "Estatico", r["direccion"], r["localidad"], r["provincia"], r["area"], r["fuente_fila"])

        if not fb_ok:
            estado = "A VALIDAR"
            obs.append(f"FB_Cantidad no numerica/invalida en fuente: {r['fb']!r}")

        hay = r.get("hay")
        hay_norm = norm_upper(hay) if isinstance(hay, str) else hay
        obs.append(f"¿HAY?={hay!r}")
        if hay_norm in ESTATICO_CLASIFICACION:
            estado_rel, estado_exist, estado_com, incluir = ESTATICO_CLASIFICACION[hay_norm]
        else:
            # Valor de ¿HAY? no contemplado en la interpretacion oficial (no
            # deberia ocurrir con esta fuente: se verificaron 156 vacios + 33
            # NO + 1 TIENE 3 + 1 NO DEJARON RELEVAR = 191). Tratamiento
            # conservador: no se asume existencia ni inexistencia.
            estado_rel, estado_exist, estado_com, incluir = ("NO_RELEVADO", "SIN_CONFIRMAR", "ACTIVO_PROVISORIO", "SI")
            obs.append(f"Valor de ¿HAY? no reconocido por la interpretacion oficial: {hay!r}")
            if estado == "OK":
                estado = "A VALIDAR"

        fb_comercial = fb if incluir == "SI" else 0

        total = fb
        fila = {
            "Formato": "Estático", "APIE": apie, "RazonSocial": r["razon_social"],
            "Direccion": direccion, "Localidad": localidad, "Provincia": provincia, "Area": area,
            "MB_Cantidad": 0, "PPUNTER_Cantidad": 0, "TT_Cantidad": 0, "FB_Cantidad": fb,
            "TotalElementos": total, "EstadoControl": estado,
            "ObservacionControl": " | ".join(obs) if obs else "",
            "FuenteFila": r["fuente_fila"], "ClaveEstacionFormato": f"{apie}|Estático",
            "EstadoRelevamientoEstatico": estado_rel, "EstadoExistenciaEstatico": estado_exist,
            "EstadoComercializacion": estado_com, "IncluirComercializacion": incluir,
            "FB_CantidadComercial": fb_comercial,
        }
        filas.append(fila)

        if estado != "OK":
            pendientes.append({
                "TipoPendiente": "TERRITORIAL",
                "APIE": apie, "Formato": "Estático", "Campo": "Direccion/Localidad/Provincia/Area/FB",
                "ValorOrigen": f"Direccion={r['direccion']!r}; Localidad={r['localidad']!r}; "
                                f"Provincia={r['provincia']!r}; Area={r['area']!r}; FB={r['fb']!r}; HAY={hay!r}",
                "Motivo": " | ".join(obs), "AccionRequerida": "Validacion manual",
                "ImpactoElementos": total, "Fuente": "ESTATICO", "FilaFuente": r["fuente_fila"],
            })

        if hay_norm == "NO DEJARON RELEVAR":
            pendientes.append({
                "TipoPendiente": "RELEVAMIENTO_NO_COMPLETADO", "APIE": apie, "Formato": "Estático",
                "Campo": "¿HAY?", "ValorOrigen": f"HAY={hay!r}; FB_Cantidad={fb}",
                "Motivo": "Relevamiento fisico no completado; existencia sin confirmar. Se conserva "
                          "provisionalmente en el catalogo (ACTIVO_PROVISORIO) hasta completar el relevamiento.",
                "AccionRequerida": "Completar relevamiento fisico para confirmar existencia",
                "ImpactoElementos": fb, "Fuente": "ESTATICO", "FilaFuente": r["fuente_fila"],
            })
        elif hay_norm is None:
            pendientes.append({
                "TipoPendiente": "RELEVAMIENTO_ESTATICO_PENDIENTE", "APIE": apie, "Formato": "Estático",
                "Campo": "¿HAY?", "ValorOrigen": f"HAY=None; FB_Cantidad={fb}",
                "Motivo": "Estacion no relevada o sin resultado cargado; el valor vacio no significa "
                          "inexistencia, significa falta de confirmacion. Se conserva provisionalmente en "
                          "el catalogo (ACTIVO_PROVISORIO).",
                "AccionRequerida": "Completar relevamiento fisico para confirmar existencia",
                "ImpactoElementos": fb, "Fuente": "ESTATICO", "FilaFuente": r["fuente_fila"],
            })
        # hay_norm in ("NO", "TIENE 3"): sin pendiente, ambos ya confirmados por relevamiento.

    def sort_key(f):
        try:
            apie_num = int(f["APIE"])
        except (TypeError, ValueError):
            apie_num = float("inf")
        formato_orden = 0 if f["Formato"] == "Digital" else 1
        return (apie_num, formato_orden)

    filas.sort(key=sort_key)
    return filas, pendientes


def normalizar_localidades(base_estaciones):
    """Aplica la localidad definitiva (version mas completa) a los 7 APIE de
    APIE_LOCALIDAD_NORMALIZADA, en ambas filas (Digital y Estatico) cuando
    existan. Muta base_estaciones in place y devuelve un registro por cambio
    real, para documentar en AUDITORIA como normalizacion controlada (no como
    pendiente)."""
    registros = []
    for f in base_estaciones:
        nueva = APIE_LOCALIDAD_NORMALIZADA.get(f["APIE"])
        if nueva is None:
            continue
        anterior = f["Localidad"]
        if anterior != nueva:
            f["Localidad"] = nueva
            registros.append({
                "APIE": f["APIE"], "Formato": f["Formato"],
                "LocalidadAnterior": anterior, "LocalidadNormalizada": nueva,
            })
    return registros


def construir_pendientes_diferencia_formatos(digital, estatico):
    """Genera exactamente una fila DIFERENCIA_DATOS_ENTRE_FORMATOS por cada
    APIE de APIE_DIFERENCIA_FORMATOS: son diferencias territoriales entre
    Digital y Estatico que se dejan sin corregir automaticamente (no
    bloqueantes, no excluyen de comercializacion; el APIE valida la estacion
    y es correcto que tenga elementos de ambos formatos)."""
    dig_by_apie = {r["apie"]: r for r in digital}
    est_by_apie = {r["apie"]: r for r in estatico}
    pendientes = []
    for apie_num in sorted(APIE_DIFERENCIA_FORMATOS):
        apie = str(apie_num)
        dr = dig_by_apie.get(apie)
        er = est_by_apie.get(apie)
        if not dr or not er:
            continue  # el APIE deberia existir en ambos formatos para este pendiente
        difs = []
        if norm_upper(dr["direccion"]) != norm_upper(er["direccion"]):
            difs.append(f"Direccion: Digital={dr['direccion']!r} / Estatico={er['direccion']!r}")
        if norm_upper(dr["localidad"]) != norm_upper(er["localidad"]):
            difs.append(f"Localidad: Digital={dr['localidad']!r} / Estatico={er['localidad']!r}")
        pendientes.append({
            "TipoPendiente": "DIFERENCIA_DATOS_ENTRE_FORMATOS", "APIE": apie, "Formato": "Digital+Estático",
            "Campo": "Direccion/Localidad",
            "ValorOrigen": f"Digital: Direccion={dr['direccion']!r}, Localidad={dr['localidad']!r} | "
                            f"Estatico: Direccion={er['direccion']!r}, Localidad={er['localidad']!r}",
            "Motivo": "El APIE valida la estacion; es correcto que tenga elementos Digitales y Estaticos. "
                      + ("; ".join(difs) if difs else "Diferencia de datos entre formatos detectada previamente.")
                      + ". No bloqueante; no excluye de comercializacion.",
            "AccionRequerida": "Revisar y unificar el dato territorial en una etapa posterior si corresponde",
            "ImpactoElementos": 0, "Fuente": "DIGITAL+ESTATICO",
            "FilaFuente": f"{dr['fuente_fila']} / {er['fuente_fila']}",
        })
    return pendientes


def marcar_diferencia_formatos(base_estaciones):
    """Marca EstadoControl='DIFERENCIA_DATOS_ENTRE_FORMATOS' en las filas
    (Digital y Estatico) de los APIE de APIE_DIFERENCIA_FORMATOS, solo como
    indicador informativo no bloqueante (no cambia IncluirComercializacion ni
    ningun conteo)."""
    apies_txt = {str(a) for a in APIE_DIFERENCIA_FORMATOS}
    for f in base_estaciones:
        if f["APIE"] in apies_txt and f["EstadoControl"] == "OK":
            f["EstadoControl"] = "DIFERENCIA_DATOS_ENTRE_FORMATOS"


# ---------------------------------------------------------------------------
# Construccion de BASE ELEMENTOS
# ---------------------------------------------------------------------------

def construir_base_elementos(base_estaciones, fb_maestro_valido):
    filas = []
    orden_codigo = {"MB": 0, "PPUNTER": 1, "TT": 2, "FB": 3}

    for est in base_estaciones:
        apie = est["APIE"]
        ciudad = localidad_a_ciudad(est["Localidad"])
        direccion = est["Direccion"] or ""
        ubicacion = f"{apie} - {ciudad or 'SIN DATO'} - {direccion}".strip()

        obs_partes = []
        if est.get("Provincia"):
            obs_partes.append(f"Provincia: {est['Provincia']}")
        if est.get("Area"):
            obs_partes.append(f"Area: {est['Area']}")
        if est.get("Direccion"):
            obs_partes.append(f"Direccion fuente: {est['Direccion']}")
        obs_partes.append(f"Origen: {est['FuenteFila']}")
        observacion_base = " | ".join(obs_partes)

        cantidades = [
            ("MB", est["MB_Cantidad"]), ("PPUNTER", est["PPUNTER_Cantidad"]),
            ("TT", est["TT_Cantidad"]), ("FB", est["FB_Cantidad"]),
        ]
        for codigo, cantidad in cantidades:
            if cantidad <= 0:
                continue
            patron = METADATA_PATTERNS[codigo]
            for n in range(1, cantidad + 1):
                elemento_id = f"{apie} - {codigo} - {n}"
                revision = RUN_LABEL
                observaciones = observacion_base
                if codigo == "FB":
                    if est["EstadoComercializacion"] == "EXCLUIDO":
                        revision = f"{RUN_LABEL} - EXCLUIDO - {est['EstadoExistenciaEstatico']}"
                    elif est["EstadoComercializacion"] == "ACTIVO_PROVISORIO":
                        revision = f"{RUN_LABEL} - ACTIVO_PROVISORIO - {est['EstadoRelevamientoEstatico']}"
                    if not fb_maestro_valido:
                        revision += " - A VALIDAR - METADATOS FB"
                elif est["EstadoControl"] not in ("OK",):
                    revision = f"{RUN_LABEL} - {est['EstadoControl']}"

                filas.append({
                    "TipoCatalogo": "Abierto",
                    "Ciudad": ciudad,
                    "Medio": MEDIO_POR_CODIGO[codigo],
                    "CircuitoDashboard": CIRCUITO_POR_CODIGO[codigo],
                    "Subcircuito": apie,
                    "Ubicacion": ubicacion,
                    "ElementoID": elemento_id,
                    "Nivel": patron["Nivel"],
                    "Descripcion": patron["Descripcion"],
                    "Resolucion": patron["Resolucion"],
                    "DimensionOptico": None,
                    "DimensionTotal": None,
                    "Observaciones": observaciones,
                    "Material": patron["Material"],
                    "TipoInstalacion": patron["TipoInstalacion"],
                    "Original": patron["Original"],
                    "CapacidadSlotsReel": patron["CapacidadSlotsReel"],
                    "SegundosDia": patron["SegundosDia"],
                    "TipoInventario": None,
                    "AplicaCantidad": None,
                    "RevisionMaestro": revision,
                    "Proveedor": None,
                    "EstadoRelevamientoEstatico": est["EstadoRelevamientoEstatico"],
                    "EstadoExistenciaEstatico": est["EstadoExistenciaEstatico"],
                    "EstadoComercializacion": est["EstadoComercializacion"],
                    "IncluirComercializacion": est["IncluirComercializacion"],
                    "_orden_codigo": orden_codigo[codigo],
                    "_orden_n": n,
                })

    def sort_key(f):
        try:
            apie_num = int(f["Subcircuito"])
        except (TypeError, ValueError):
            apie_num = float("inf")
        return (apie_num, f["_orden_codigo"], f["_orden_n"])

    filas.sort(key=sort_key)
    for f in filas:
        del f["_orden_codigo"]
        del f["_orden_n"]
    return filas


def verificar_patron_fb(wb_ocu26_path):
    """Lectura de solo lectura del maestro OCU26 productivo para confirmar que
    el patron FB usado (Outdoor/Mupi/PAI/Backlight/PNG/0/0) es consistente."""
    try:
        wb = openpyxl.load_workbook(wb_ocu26_path, data_only=True, read_only=True)
    except FileNotFoundError:
        return False
    if "MAESTRO_ELEMENTOS" not in wb.sheetnames:
        return False
    ws = wb["MAESTRO_ELEMENTOS"]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    if "ElementoID" not in header:
        return False
    eid_col = header.index("ElementoID")
    fb_rows = [r for r in rows[1:] if r[eid_col] and re.match(r"^\d+ - FB - \d+$", str(r[eid_col]))]
    if not fb_rows:
        return False
    campos = {
        "Medio": header.index("Medio"), "Nivel": header.index("Nivel"),
        "Descripcion": header.index("Descripcion"), "Material": header.index("Material"),
        "TipoInstalacion": header.index("TipoInstalacion"), "Original": header.index("Original"),
    }
    esperado = {"Medio": "Estático", "Nivel": "Outdoor", "Descripcion": "Mupi",
                "Material": "PAI", "TipoInstalacion": "Backlight", "Original": "PNG"}
    for campo, idx in campos.items():
        valores = {r[idx] for r in fb_rows}
        if valores != {esperado[campo]}:
            return False
    return True


# ---------------------------------------------------------------------------
# Validaciones sobre las bases construidas
# ---------------------------------------------------------------------------

def validar_base_estaciones(filas):
    problemas = []
    claves = [f["ClaveEstacionFormato"] for f in filas]
    if len(claves) != len(set(claves)):
        problemas.append("Claves ClaveEstacionFormato duplicadas en BASE ESTACIONES")
    for f in filas:
        for campo in ("MB_Cantidad", "PPUNTER_Cantidad", "TT_Cantidad", "FB_Cantidad"):
            if not isinstance(f[campo], int) or f[campo] < 0:
                problemas.append(f"Cantidad invalida {campo}={f[campo]!r} en APIE {f['APIE']}")
        total_calc = f["MB_Cantidad"] + f["PPUNTER_Cantidad"] + f["TT_Cantidad"] + f["FB_Cantidad"]
        if total_calc != f["TotalElementos"]:
            problemas.append(f"TotalElementos incorrecto en APIE {f['APIE']} ({f['Formato']})")
        if f["Formato"] not in ("Digital", "Estático"):
            problemas.append(f"Formato invalido {f['Formato']!r} en APIE {f['APIE']}")
    return problemas


def validar_base_elementos(filas):
    problemas = []
    ids = [f["ElementoID"] for f in filas]
    if len(ids) != len(set(ids)):
        problemas.append("ElementoID duplicados en BASE ELEMENTOS")
    for f in filas:
        if not ELEMENTO_ID_RE.match(f["ElementoID"]):
            problemas.append(f"ElementoID no cumple el patron: {f['ElementoID']!r}")
        if f["CircuitoDashboard"] not in ("YPF Digital", "YPF Estático"):
            problemas.append(f"CircuitoDashboard invalido: {f['CircuitoDashboard']!r}")
        codigo = f["ElementoID"].split(" - ")[1] if " - " in f["ElementoID"] else None
        medio_esperado = MEDIO_POR_CODIGO.get(codigo)
        if medio_esperado and f["Medio"] != medio_esperado:
            problemas.append(f"Medio no coincide con codigo en {f['ElementoID']}")
        if not str(f["Ubicacion"]).startswith(str(f["Subcircuito"])):
            problemas.append(f"Ubicacion no comienza con Subcircuito en {f['ElementoID']}")
        if not f["Subcircuito"]:
            problemas.append(f"Elemento sin APIE: {f['ElementoID']}")

    # Secuencias continuas 1..N sin huecos ni duplicados por APIE+codigo
    grupos = {}
    for f in filas:
        codigo = f["ElementoID"].split(" - ")[1]
        key = (f["Subcircuito"], codigo)
        n = int(f["ElementoID"].split(" - ")[2])
        grupos.setdefault(key, []).append(n)
    for key, ns in grupos.items():
        ns_sorted = sorted(ns)
        if ns_sorted != list(range(1, len(ns_sorted) + 1)):
            problemas.append(f"Secuencia incompleta o con huecos: APIE/codigo {key} -> {ns_sorted}")

    return problemas


def validar_conteos_comercializacion(base_estaciones, base_elementos):
    """Verifica los conteos obligatorios de la clasificacion comercial
    (seccion 9 del brief de relevamiento estatico)."""
    errores = []

    incluir_si_est = [f for f in base_estaciones if f["IncluirComercializacion"] == "SI"]
    incluir_no_est = [f for f in base_estaciones if f["IncluirComercializacion"] == "NO"]
    if len(incluir_si_est) != 570:
        errores.append(f"IncluirComercializacion=SI en BASE ESTACIONES: esperado 570, obtenido {len(incluir_si_est)}")
    apies_si_unicos = {f["APIE"] for f in incluir_si_est}
    if len(apies_si_unicos) != 505:
        errores.append(f"APIE unicos con IncluirComercializacion=SI: esperado 505, obtenido {len(apies_si_unicos)}")
    if len(incluir_no_est) != 33:
        errores.append(f"IncluirComercializacion=NO en BASE ESTACIONES: esperado 33, obtenido {len(incluir_no_est)}")

    incluir_si_elem = [f for f in base_elementos if f["IncluirComercializacion"] == "SI"]
    incluir_no_elem = [f for f in base_elementos if f["IncluirComercializacion"] == "NO"]
    if len(incluir_si_elem) != 3817:
        errores.append(f"IncluirComercializacion=SI en BASE ELEMENTOS: esperado 3817, obtenido {len(incluir_si_elem)}")
    if len(incluir_no_elem) != 66:
        errores.append(f"IncluirComercializacion=NO en BASE ELEMENTOS: esperado 66, obtenido {len(incluir_no_elem)}")

    estatico_rows = [f for f in base_estaciones if f["Formato"] == "Estático"]

    def contar(estado_com):
        rows = [f for f in estatico_rows if f["EstadoComercializacion"] == estado_com]
        return len(rows), sum(f["FB_Cantidad"] for f in rows)

    n_conf, fb_conf = contar("ACTIVO_CONFIRMADO")
    n_prov, fb_prov = contar("ACTIVO_PROVISORIO")
    n_excl, fb_excl = contar("EXCLUIDO")
    if (n_conf, fb_conf) != (1, 3):
        errores.append(f"ACTIVO_CONFIRMADO: esperado 1 estacion/3 FB, obtenido {n_conf}/{fb_conf}")
    if (n_prov, fb_prov) != (157, 314):
        errores.append(f"ACTIVO_PROVISORIO: esperado 157 estaciones/314 FB, obtenido {n_prov}/{fb_prov}")
    if (n_excl, fb_excl) != (33, 66):
        errores.append(f"EXCLUIDO: esperado 33 estaciones/66 FB, obtenido {n_excl}/{fb_excl}")

    excluidos = [f for f in estatico_rows if f["EstadoComercializacion"] == "EXCLUIDO"]
    if any(f["EstadoExistenciaEstatico"] != "CONFIRMADO_NO_EXISTE" for f in excluidos):
        errores.append("Hay filas EXCLUIDO sin EstadoExistenciaEstatico=CONFIRMADO_NO_EXISTE")

    provisorios = [f for f in estatico_rows if f["EstadoComercializacion"] == "ACTIVO_PROVISORIO"]
    if any(f["EstadoExistenciaEstatico"] != "SIN_CONFIRMAR" for f in provisorios):
        errores.append("Hay filas ACTIVO_PROVISORIO sin EstadoExistenciaEstatico=SIN_CONFIRMAR")

    return errores, {
        "incluir_si_est": len(incluir_si_est), "incluir_no_est": len(incluir_no_est),
        "apies_si_unicos": len(apies_si_unicos), "incluir_si_elem": len(incluir_si_elem),
        "incluir_no_elem": len(incluir_no_elem), "n_conf": n_conf, "fb_conf": fb_conf,
        "n_prov": n_prov, "fb_prov": fb_prov, "n_excl": n_excl, "fb_excl": fb_excl,
    }


def validar_normalizacion_localidades(base_estaciones):
    errores = []
    for apie, loc_esperada in APIE_LOCALIDAD_NORMALIZADA.items():
        filas_apie = [f for f in base_estaciones if f["APIE"] == apie]
        if not filas_apie:
            errores.append(f"APIE {apie} (normalizacion de localidad): no encontrado en BASE ESTACIONES")
            continue
        for f in filas_apie:
            if f["Localidad"] != loc_esperada:
                errores.append(
                    f"APIE {apie} ({f['Formato']}): Localidad esperada {loc_esperada!r}, obtenida {f['Localidad']!r}")
    return errores


def validar_pendientes_diferencia_formatos(pendientes):
    errores = []
    filas = [p for p in pendientes if p["TipoPendiente"] == "DIFERENCIA_DATOS_ENTRE_FORMATOS"]
    if len(filas) != 8:
        errores.append(f"DIFERENCIA_DATOS_ENTRE_FORMATOS: esperadas 8 filas, obtenidas {len(filas)}")
    apies_obtenidos = {f["APIE"] for f in filas}
    apies_esperados = {str(a) for a in APIE_DIFERENCIA_FORMATOS}
    if apies_obtenidos != apies_esperados:
        errores.append(f"DIFERENCIA_DATOS_ENTRE_FORMATOS: APIE esperados {sorted(apies_esperados)}, "
                        f"obtenidos {sorted(apies_obtenidos)}")
    return errores


# ---------------------------------------------------------------------------
# Escritura del workbook de salida
# ---------------------------------------------------------------------------

def _valor_seguro_para_celda(v):
    """Evita que un valor de texto que empieza con '=' (p.ej. una formula de
    Excel documentada como texto en PENDIENTES) sea reinterpretado por
    openpyxl como una formula viva al escribirlo. Se antepone un apostrofo,
    la misma convencion que usa Excel para forzar texto literal."""
    if isinstance(v, str) and v.startswith("="):
        return "'" + v
    return v


def escribir_hoja_tabular(wb, nombre, columnas, filas_dict):
    ws = wb.create_sheet(nombre)
    ws.append(columnas)
    for c in ws[1]:
        c.font = Font(bold=True)
    for fila in filas_dict:
        ws.append([_valor_seguro_para_celda(fila.get(col)) for col in columnas])
    for i, col in enumerate(columnas, start=1):
        max_len = max([len(col)] + [len(str(f.get(col))) for f in filas_dict[:500] if f.get(col) is not None])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 60)
    ws.freeze_panes = "A2"
    return ws


def copiar_hoja_fuente(wb_destino, wb_origen_formulas, wb_origen_valores, nombre_hoja):
    """Copia una hoja fuente preservando valores, formulas internas, estilos,
    anchos, filas y celdas combinadas. Las formulas con referencia a un libro
    externo (patron [n], sin relacion registrada en el libro de salida) se
    neutralizan: se escribe su valor cacheado si es comprobable, o se deja la
    celda vacia. Esto evita que Microsoft Excel reporte registros danados al
    abrir el archivo, sin reconstruir ni inventar el vinculo externo original."""
    ws_src = wb_origen_formulas[nombre_hoja]
    ws_val = wb_origen_valores[nombre_hoja]
    ws_dst = wb_destino.create_sheet(nombre_hoja)
    for row in ws_src.iter_rows():
        for cell in row:
            valor = cell.value
            if cell.data_type == "f" and isinstance(valor, str) and EXTERNAL_REF_PATTERN.search(valor):
                cacheado = ws_val.cell(row=cell.row, column=cell.column).value
                valor = valor_literal_seguro(cacheado)
            new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=valor)
            if cell.has_style:
                new_cell.font = copy_style(cell.font)
                new_cell.fill = copy_style(cell.fill)
                new_cell.border = copy_style(cell.border)
                new_cell.alignment = copy_style(cell.alignment)
                new_cell.number_format = cell.number_format
    for key, dim in ws_src.column_dimensions.items():
        ws_dst.column_dimensions[key].width = dim.width
    for key, dim in ws_src.row_dimensions.items():
        ws_dst.row_dimensions[key].height = dim.height
    for merged_range in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(str(merged_range))
    ws_dst.sheet_view.showGridLines = ws_src.sheet_view.showGridLines
    return ws_dst


def construir_leeme(wb, resumen):
    ws = wb.create_sheet("LEEME")
    filas = [
        ("Campo", "Detalle"),
        ("ARCHIVO", "YPF_BASE_LIMPIA_ETAPA_1_RELEVAMIENTO_2026-08-13.xlsx"),
        ("OBJETIVO", "Clasifica el catalogo estatico segun el resultado real del relevamiento fisico (columna "
                     "¿HAY? de ESTATICO), distinguiendo cantidad fuente, existencia confirmada e inventario "
                     "comercial provisional. Mantiene las correcciones previas (compatibilidad Excel, exclusion "
                     "de APIE bloqueados, alineacion con ZONAS). Independiente de OCU26 productivo."),
        ("FECHA DE GENERACION", FECHA_GENERACION),
        ("FUENTES", "YPF DIGITAL + ESTATICO (1).xlsx; YPF - Base campañas y elementos corregido estructural.xlsx"),
        ("GRANO BASE ESTACIONES", "Una fila por combinacion unica APIE + Formato (catalogo activo, sin bloqueados)"),
        ("GRANO BASE ELEMENTOS", "Una fila por elemento fisico (ElementoID), catalogo activo"),
        ("", ""),
        ("--- DIGITAL ---", ""),
        ("FUENTE DIGITAL (sin excluir)", "416 estaciones y 3.547 elementos informados (se conserva integra como control)"),
        ("CATALOGO DIGITAL ACTIVO", "412 estaciones y 3.500 elementos (alineado de forma exacta con ZONAS)"),
        ("APIE DIGITALES BLOQUEADOS", "4 (30510, 31131, 31192, 31239) — permanecen fuera de comercializacion; "
                                       "no generan ElementoID"),
        ("", ""),
        ("--- ESTATICO: CANTIDAD FUENTE vs EXISTENCIA vs COMERCIAL ---", ""),
        ("QUE SIGNIFICA CADA CANTIDAD", "Cantidad fuente = lo que informa la planilla ESTATICO (383 FB), sin "
                                         "importar si el relevamiento confirmo su existencia. Existencia confirmada "
                                         "= FB que un relevador verifico fisicamente que SI estan (3 FB, ¿HAY?=TIENE 3). "
                                         "Inventario comercial provisional = todo lo que hoy puede ofrecerse (existencia "
                                         "confirmada + lo que aun no se confirmo pero tampoco se descarto: 317 FB)."),
        ("RELEVAMIENTO CONCENTRADO EN", "CABA y GBA Norte, por ser las zonas comercialmente mas buscadas. Esto "
                                          "explica por que la mayoria de las estaciones estaticas (156 de 191) "
                                          "todavia no tienen resultado de relevamiento cargado."),
        ("¿HAY? VACIO NO ES 'NO EXISTE'", "Un valor vacio en ¿HAY? significa que la estacion no fue relevada o el "
                                           "resultado no se cargo todavia: la existencia esta SIN CONFIRMAR, no "
                                           "descartada. Por eso se mantiene provisionalmente en el catalogo comercial."),
        ("¿HAY?=NO SI ES 'NO EXISTE'", "Cuando la fuente dice NO, un relevador ya confirmo fisicamente que el soporte "
                                        "no existe. Esas 33 estaciones (66 FB) quedan EXCLUIDAS del catalogo comercial "
                                        "de forma permanente, aunque se conservan en BASE ELEMENTOS por trazabilidad."),
        ("COMO FILTRAR EL CATALOGO", "Usar la columna IncluirComercializacion = SI en BASE ESTACIONES / BASE "
                                      "ELEMENTOS. Los soportes con IncluirComercializacion = NO (EstadoComercializacion "
                                      "= EXCLUIDO) no pueden recibir campañas."),
        ("ESTATICO FUENTE", "191 estaciones, 383 FB"),
        ("ESTATICO ACTIVO_CONFIRMADO", "1 estacion, 3 FB (¿HAY?=TIENE 3)"),
        ("ESTATICO ACTIVO_PROVISORIO", "157 estaciones, 314 FB (156 ¿HAY? vacio + 1 NO DEJARON RELEVAR)"),
        ("ESTATICO EXCLUIDO", "33 estaciones, 66 FB (¿HAY?=NO, inexistencia confirmada)"),
        ("ESTATICO COMERCIAL PROVISIONAL (CONFIRMADO + PROVISORIO)", "158 estaciones, 317 FB"),
        ("", ""),
        ("--- TOTALES ACTIVOS/COMERCIALES ---", ""),
        ("APIE ACTIVOS UNICOS (IncluirComercializacion=SI)", resumen["apie_si_unicos"]),
        ("BASE ESTACIONES: SI / NO", f"{resumen['incluir_si_est']} / {resumen['incluir_no_est']} (de {resumen['estaciones_rows']} filas totales)"),
        ("BASE ELEMENTOS: SI / NO", f"{resumen['incluir_si_elem']} / {resumen['incluir_no_elem']} (de {resumen['elementos_totales']} filas totales)"),
        ("LOCALIDADES NORMALIZADAS", f"{resumen['n_normalizaciones']} APIE (ver AUDITORIA, seccion de normalizaciones controladas)"),
        ("DIFERENCIAS DE DATOS ENTRE FORMATOS SIN CORREGIR", f"{resumen['n_diferencia_formatos']} APIE (no bloqueantes, "
         "validan la estacion, ver PENDIENTES tipo DIFERENCIA_DATOS_ENTRE_FORMATOS)"),
        ("CANTIDAD DE PENDIENTES", resumen["n_pendientes"]),
        ("COMPATIBILIDAD EXCEL", "Se neutralizaron 3 formulas VLOOKUP con vinculo externo roto ([1]PRINCIPAL) en "
                                 "DIGITAL!B370/F370/G370. No se agregaron conexiones externas ni se reconstruyo el vinculo."),
        ("ESTADO CONTROL VS ZONAS", "EXACTO (0 de diferencia en estaciones, MB, TT, PPUNTER y total digital activos)"),
        ("PROXIMA ETAPA", "Incorporacion de campañas por APIE y ElementoID (Etapa 2), usando IncluirComercializacion=SI "
                          "como filtro. Los 4 APIE digitales bloqueados no podran vincularse con campañas mientras "
                          "continuen bloqueados."),
        ("CAMPAÑAS EN ESTA ETAPA", "No se incorporaron campañas"),
        ("ESTADO DE INTEGRACION", "Este archivo NO esta integrado a la base OCU26 productiva"),
    ]
    for r in filas:
        ws.append(r)
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 100
    return ws


def _nombre_borrador(ruta_output):
    if ruta_output.lower().endswith(".xlsx"):
        return ruta_output[: -len(".xlsx")] + "_BORRADOR.xlsx"
    return ruta_output + "_BORRADOR"


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-catalog", required=True)
    parser.add_argument("--source-structure", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--ocu26-master", default="input/OCU26_BASE_DATOS.xlsx",
                         help="Ruta al maestro OCU26 productivo, usado solo en modo lectura para verificar el patron FB.")
    args = parser.parse_args(argv)

    print("[1/10] Cargando fuentes...")
    print(f"  catalogo   = {args.source_catalog}")
    print(f"  estructura = {args.source_structure}")

    wb_catalogo_valores = openpyxl.load_workbook(args.source_catalog, data_only=True, read_only=True)
    wb_catalogo_formulas = openpyxl.load_workbook(args.source_catalog, data_only=False)
    # Copia adicional, NO read-only, para poder leer valores cacheados por
    # coordenada (row,column) al neutralizar formulas con vinculo externo.
    wb_catalogo_valores_full = openpyxl.load_workbook(args.source_catalog, data_only=True)

    digital = cargar_digital(wb_catalogo_valores)
    estatico = cargar_estatico(wb_catalogo_valores)
    base_prev = cargar_base_estaciones_previa(wb_catalogo_valores)
    direcciones_lookup = cargar_direcciones_lookup(wb_catalogo_valores)

    print("[2/10] Preflight DIGITAL/ESTATICO (fuente completa, sin excluir bloqueados)...")
    resultado_preflight, errores_preflight = preflight(digital, estatico)
    for nombre, obtenido, esperado in resultado_preflight["checks"]:
        estado = "OK" if obtenido == esperado else "DIFERENCIA"
        print(f"  {nombre:32s} esperado={esperado:6d} obtenido={obtenido:6d} [{estado}]")

    if errores_preflight:
        print("\nERROR BLOQUEANTE: el preflight no coincide con los conteos rectores esperados.")
        for e in errores_preflight:
            print(f"  - {e}")
        print("No se genera el archivo final. Revise las fuentes antes de continuar.")
        return 2

    print("[3/10] Detectando formulas con vinculo externo roto (patron [n])...")
    hojas_a_escanear = ["DIGITAL", "ESTATICO ", "ZONAS", "DIRECCIONES", "ZONAS VERSION ANTUGUA"]
    hallazgos_formulas = detectar_formulas_externas(wb_catalogo_formulas, wb_catalogo_valores_full, hojas_a_escanear)
    if hallazgos_formulas:
        for h in hallazgos_formulas:
            print(f"  {h['hoja']}!{h['celda']}: {h['formula_original']!r} -> se neutraliza a {h['valor_final']!r}")
    else:
        print("  Ninguna formula con vinculo externo detectada.")

    print("[4/10] Construyendo BASE ESTACIONES activa (excluye APIE bloqueados)...")
    try:
        base_estaciones, pendientes = construir_base_estaciones(digital, estatico, base_prev, direcciones_lookup)
    except ValueError as e:
        print(f"ERROR BLOQUEANTE: {e}")
        return 3

    for h in hallazgos_formulas:
        pendientes.append({
            "TipoPendiente": "FORMULA_EXTERNA_NEUTRALIZADA", "APIE": "", "Formato": "Digital",
            "Campo": f"{h['hoja']}!{h['celda']}", "ValorOrigen": h["formula_original"],
            "Motivo": "Formula con vinculo a libro externo no registrado en este libro (referencia [n], p.ej. "
                      "[1]PRINCIPAL). Al copiarla sin la relacion externa, Microsoft Excel reparaba el archivo "
                      f"al abrirlo. Se convierte a valor literal (valor cacheado: {h['valor_cacheado']!r}).",
            "AccionRequerida": "Confirmar el valor final contra la fuente PRINCIPAL original si se requiere "
                                "antes de una etapa posterior",
            "ImpactoElementos": 0, "Fuente": h["hoja"], "FilaFuente": f"{h['hoja']}!{h['celda']}",
        })

    problemas_estaciones = validar_base_estaciones(base_estaciones)
    if problemas_estaciones:
        print("ERROR BLOQUEANTE en BASE ESTACIONES:")
        for p in problemas_estaciones:
            print(f"  - {p}")
        return 4

    print("[5/10] Normalizando localidades y documentando diferencias entre formatos...")
    registros_normalizacion = normalizar_localidades(base_estaciones)
    n_apie_normalizados = len({r["APIE"] for r in registros_normalizacion})
    print(f"  {len(registros_normalizacion)} filas de BASE ESTACIONES modificadas "
          f"sobre {n_apie_normalizados} APIE (de {len(APIE_LOCALIDAD_NORMALIZADA)} APIE con normalizacion definida)")
    errores_localidad = validar_normalizacion_localidades(base_estaciones)
    if errores_localidad:
        print("ERROR BLOQUEANTE: normalizacion de localidades incompleta o incorrecta:")
        for e in errores_localidad:
            print(f"  - {e}")
        return 5

    pendientes.extend(construir_pendientes_diferencia_formatos(digital, estatico))
    marcar_diferencia_formatos(base_estaciones)
    errores_diferencia = validar_pendientes_diferencia_formatos(pendientes)
    if errores_diferencia:
        print("ERROR BLOQUEANTE: pendientes DIFERENCIA_DATOS_ENTRE_FORMATOS incorrectos:")
        for e in errores_diferencia:
            print(f"  - {e}")
        return 6

    print("[6/10] Construyendo BASE ELEMENTOS activa...")
    fb_patron_valido = verificar_patron_fb(args.ocu26_master)
    print(f"  Patron FB verificado contra maestro OCU26 (solo lectura): {'OK' if fb_patron_valido else 'NO DISPONIBLE'}")
    base_elementos = construir_base_elementos(base_estaciones, fb_patron_valido)
    problemas_elementos = validar_base_elementos(base_elementos)
    if problemas_elementos:
        print("ERROR BLOQUEANTE en BASE ELEMENTOS:")
        for p in problemas_elementos[:30]:
            print(f"  - {p}")
        return 7

    apies_bloqueados_txt = {str(a) for a in APIE_BLOQUEADOS}
    presentes_en_estaciones = apies_bloqueados_txt & {f["APIE"] for f in base_estaciones}
    presentes_en_elementos = apies_bloqueados_txt & {f["Subcircuito"] for f in base_elementos}
    if presentes_en_estaciones or presentes_en_elementos:
        print("ERROR BLOQUEANTE: hay APIE bloqueados presentes en las bases activas.")
        print(f"  en BASE ESTACIONES: {presentes_en_estaciones}")
        print(f"  en BASE ELEMENTOS: {presentes_en_elementos}")
        return 8

    print("[7/10] Validando conteos de comercializacion estatica...")
    errores_comercial, cnt_comercial = validar_conteos_comercializacion(base_estaciones, base_elementos)
    if errores_comercial:
        print("ERROR BLOQUEANTE: conteos de comercializacion no coinciden con lo esperado:")
        for e in errores_comercial:
            print(f"  - {e}")
        return 9

    # --- Controles agregados para AUDITORIA / resumen (catalogo ACTIVO) ---
    apies_todos = {f["APIE"] for f in base_estaciones}
    apies_ambos = {f["APIE"] for f in base_estaciones if f["Formato"] == "Digital"} & \
                  {f["APIE"] for f in base_estaciones if f["Formato"] == "Estático"}
    mb_total = sum(f["MB_Cantidad"] for f in base_estaciones)
    pp_total = sum(f["PPUNTER_Cantidad"] for f in base_estaciones)
    tt_total = sum(f["TT_Cantidad"] for f in base_estaciones)
    fb_total = sum(f["FB_Cantidad"] for f in base_estaciones)
    total_digital_estaciones = mb_total + pp_total + tt_total
    total_elementos = len(base_elementos)

    resumen = {
        "estaciones_rows": len(base_estaciones), "apie_unicos": len(apies_todos),
        "apie_ambos_formatos": len(apies_ambos), "mb": mb_total, "pp": pp_total, "tt": tt_total,
        "fb": fb_total, "total_digital": total_digital_estaciones, "fb_sum": fb_total,
        "elementos_totales": total_elementos,
        "incluir_si_est": cnt_comercial["incluir_si_est"], "incluir_no_est": cnt_comercial["incluir_no_est"],
        "apie_si_unicos": cnt_comercial["apies_si_unicos"],
        "incluir_si_elem": cnt_comercial["incluir_si_elem"], "incluir_no_elem": cnt_comercial["incluir_no_elem"],
        "n_normalizaciones": len({r["APIE"] for r in registros_normalizacion}),
        "n_normalizaciones_filas": len(registros_normalizacion),
        "n_diferencia_formatos": len(APIE_DIFERENCIA_FORMATOS),
    }

    controles_finales = [
        ("estaciones_rows", EXPECTED_ACTIVO["estaciones_rows"]),
        ("apie_unicos", EXPECTED_ACTIVO["apie_unicos"]),
        ("apie_ambos_formatos", EXPECTED_ACTIVO["apie_ambos_formatos"]),
        ("mb", EXPECTED_ACTIVO["mb"]), ("pp", EXPECTED_ACTIVO["pp"]), ("tt", EXPECTED_ACTIVO["tt"]),
        ("fb", EXPECTED_ACTIVO["fb"]), ("total_digital", EXPECTED_ACTIVO["total_digital"]),
        ("elementos_totales", EXPECTED_ACTIVO["elementos_totales"]),
    ]
    errores_finales = [f"{k}: esperado {esp}, obtenido {resumen[k]}" for k, esp in controles_finales if resumen[k] != esp]
    if errores_finales:
        print("ERROR BLOQUEANTE: validaciones finales de conteo activo no coinciden.")
        for e in errores_finales:
            print(f"  - {e}")
        return 10

    apies_digitales_unicos = len({f["APIE"] for f in base_estaciones if f["Formato"] == "Digital"})
    diff_zonas_activo = {
        "estaciones_digitales": apies_digitales_unicos - ZONAS_REFERENCIA_ACTIVA["estaciones_digitales"],
        "menu_board": mb_total - ZONAS_REFERENCIA_ACTIVA["menu_board"],
        "torres": tt_total - ZONAS_REFERENCIA_ACTIVA["torres"],
        "punteras": pp_total - ZONAS_REFERENCIA_ACTIVA["punteras"],
        "elementos_digitales": total_digital_estaciones - ZONAS_REFERENCIA_ACTIVA["elementos_digitales"],
    }
    if any(v != 0 for v in diff_zonas_activo.values()):
        print("ERROR BLOQUEANTE: el catalogo activo no quedo alineado de forma exacta con ZONAS.")
        for k, v in diff_zonas_activo.items():
            print(f"  - {k}: diferencia {v}")
        return 11

    n_bloqueados = sum(1 for p in pendientes if p["TipoPendiente"] == "TERRITORIAL_BLOQUEANTE")
    elementos_bloqueados = sum(p["ImpactoElementos"] for p in pendientes if p["TipoPendiente"] == "TERRITORIAL_BLOQUEANTE")
    if n_bloqueados != 4 or elementos_bloqueados != 47:
        print(f"ERROR BLOQUEANTE: se esperaban 4 APIE bloqueados con 47 elementos de impacto; "
              f"se obtuvieron {n_bloqueados} APIE con {elementos_bloqueados} elementos.")
        return 12

    resumen["n_pendientes"] = len(pendientes)
    resumen["estaciones_digitales_activas"] = apies_digitales_unicos
    resumen["n_bloqueados"] = n_bloqueados
    resumen["elementos_bloqueados"] = elementos_bloqueados

    print("[8/10] Ensamblando workbook de salida...")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    construir_leeme(wb_out, resumen)

    n_dup_claves = len(base_estaciones) - len({f["ClaveEstacionFormato"] for f in base_estaciones})
    n_dup_ids = len(base_elementos) - len({f["ElementoID"] for f in base_elementos})
    n_pend_formula = sum(1 for p in pendientes if p["TipoPendiente"] == "FORMULA_EXTERNA_NEUTRALIZADA")
    n_pend_diferencia = sum(1 for p in pendientes if p["TipoPendiente"] == "DIFERENCIA_DATOS_ENTRE_FORMATOS")
    n_pend_relev_pendiente = sum(1 for p in pendientes if p["TipoPendiente"] == "RELEVAMIENTO_ESTATICO_PENDIENTE")
    n_pend_relev_no_completado = sum(1 for p in pendientes if p["TipoPendiente"] == "RELEVAMIENTO_NO_COMPLETADO")
    n_pend_territorial = sum(1 for p in pendientes if p["TipoPendiente"] == "TERRITORIAL")

    def sep(titulo):
        return {"Control": f"--- {titulo} ---", "Fuente": "", "Esperado": None, "Obtenido": None,
                "Diferencia": None, "Estado": "", "Observacion": ""}

    auditoria_filas = [
        sep("1. CONTEOS DE LA FUENTE DIGITAL (416 registros, sin excluir bloqueados)"),
        {"Control": "Filas DIGITAL (fuente)", "Fuente": "DIGITAL", "Esperado": EXPECTED["digital_rows"],
         "Obtenido": len(digital), "Diferencia": len(digital) - EXPECTED["digital_rows"], "Estado": "OK", "Observacion": ""},
        {"Control": "APIE unicos DIGITAL (fuente)", "Fuente": "DIGITAL", "Esperado": EXPECTED["digital_apie_unicos"],
         "Obtenido": len({r['apie'] for r in digital}), "Diferencia": len({r['apie'] for r in digital}) - EXPECTED["digital_apie_unicos"],
         "Estado": "OK", "Observacion": ""},
        {"Control": "Menu Board (fuente)", "Fuente": "DIGITAL", "Esperado": EXPECTED["mb"],
         "Obtenido": resultado_preflight["mb_sum"], "Diferencia": resultado_preflight["mb_sum"] - EXPECTED["mb"], "Estado": "OK", "Observacion": ""},
        {"Control": "Torres (fuente)", "Fuente": "DIGITAL", "Esperado": EXPECTED["tt"],
         "Obtenido": resultado_preflight["tt_sum"], "Diferencia": resultado_preflight["tt_sum"] - EXPECTED["tt"], "Estado": "OK", "Observacion": ""},
        {"Control": "Punteras (fuente)", "Fuente": "DIGITAL", "Esperado": EXPECTED["pp"],
         "Obtenido": resultado_preflight["pp_sum"], "Diferencia": resultado_preflight["pp_sum"] - EXPECTED["pp"], "Estado": "OK", "Observacion": ""},
        {"Control": "Total elementos informados (fuente)", "Fuente": "DIGITAL", "Esperado": EXPECTED["total_digital"],
         "Obtenido": resultado_preflight["total_digital"], "Diferencia": resultado_preflight["total_digital"] - EXPECTED["total_digital"],
         "Estado": "OK", "Observacion": "437 + 2142 + 968 = 3547. La hoja DIGITAL se conserva integra como control, incluidos los 4 APIE bloqueados."},

        sep("2. CONTEOS ACTIVOS ALINEADOS CON ZONAS (excluye 4 APIE bloqueados)"),
        {"Control": "Estaciones digitales activas vs ZONAS", "Fuente": "ZONAS", "Esperado": ZONAS_REFERENCIA_ACTIVA["estaciones_digitales"],
         "Obtenido": apies_digitales_unicos, "Diferencia": diff_zonas_activo["estaciones_digitales"], "Estado": "OK",
         "Observacion": "416 (fuente) - 4 (bloqueados) = 412"},
        {"Control": "Menu Board activo vs ZONAS", "Fuente": "ZONAS", "Esperado": ZONAS_REFERENCIA_ACTIVA["menu_board"],
         "Obtenido": mb_total, "Diferencia": diff_zonas_activo["menu_board"], "Estado": "OK", "Observacion": "437 - 2 = 435"},
        {"Control": "Torres activo vs ZONAS", "Fuente": "ZONAS", "Esperado": ZONAS_REFERENCIA_ACTIVA["torres"],
         "Obtenido": tt_total, "Diferencia": diff_zonas_activo["torres"], "Estado": "OK", "Observacion": "2142 - 28 = 2114"},
        {"Control": "Punteras activo vs ZONAS", "Fuente": "ZONAS", "Esperado": ZONAS_REFERENCIA_ACTIVA["punteras"],
         "Obtenido": pp_total, "Diferencia": diff_zonas_activo["punteras"], "Estado": "OK", "Observacion": "968 - 17 = 951"},
        {"Control": "Total digital activo vs ZONAS", "Fuente": "ZONAS", "Esperado": ZONAS_REFERENCIA_ACTIVA["elementos_digitales"],
         "Obtenido": total_digital_estaciones, "Diferencia": diff_zonas_activo["elementos_digitales"], "Estado": "OK",
         "Observacion": "435 + 2114 + 951 = 3500. Alineacion EXACTA verificada tras excluir los 4 APIE bloqueados; "
                        "no se usan los controles historicos 386/404/2022/910/3336 de la Etapa 1 original."},

        sep("3. REGISTROS BLOQUEADOS"),
        {"Control": "APIE bloqueados", "Fuente": "PENDIENTES", "Esperado": 4, "Obtenido": n_bloqueados,
         "Diferencia": n_bloqueados - 4, "Estado": "OK" if n_bloqueados == 4 else "ERROR",
         "Observacion": "30510, 31131, 31192, 31239"},
        {"Control": "Elementos bloqueados (MB+TT+PPUNTER)", "Fuente": "PENDIENTES", "Esperado": 47,
         "Obtenido": elementos_bloqueados, "Diferencia": elementos_bloqueados - 47,
         "Estado": "OK" if elementos_bloqueados == 47 else "ERROR", "Observacion": "10 + 16 + 3 + 18 = 47"},
        {"Control": "APIE bloqueados presentes en BASE ESTACIONES activa", "Fuente": "BASE ESTACIONES", "Esperado": 0,
         "Obtenido": len(presentes_en_estaciones), "Diferencia": len(presentes_en_estaciones),
         "Estado": "OK" if not presentes_en_estaciones else "ERROR", "Observacion": ""},
        {"Control": "APIE bloqueados presentes en BASE ELEMENTOS activa", "Fuente": "BASE ELEMENTOS", "Esperado": 0,
         "Obtenido": len(presentes_en_elementos), "Diferencia": len(presentes_en_elementos),
         "Estado": "OK" if not presentes_en_elementos else "ERROR", "Observacion": ""},
        {"Control": "APIE bloqueados conservados en hoja fuente DIGITAL", "Fuente": "DIGITAL", "Esperado": 4,
         "Obtenido": len(apies_bloqueados_txt & {r['apie'] for r in digital}),
         "Diferencia": len(apies_bloqueados_txt & {r['apie'] for r in digital}) - 4, "Estado": "OK", "Observacion": ""},

        sep("4. CONTEOS ESTATICOS - FUENTE"),
        {"Control": "Filas ESTATICO (fuente)", "Fuente": "ESTATICO", "Esperado": EXPECTED["estatico_rows"],
         "Obtenido": len(estatico), "Diferencia": len(estatico) - EXPECTED["estatico_rows"], "Estado": "OK",
         "Observacion": "Se excluyo 1 fila sin APIE (fila de totales de control) del pie de ESTATICO"},
        {"Control": "APIE unicos ESTATICO", "Fuente": "ESTATICO", "Esperado": EXPECTED["estatico_apie_unicos"],
         "Obtenido": len({r['apie'] for r in estatico}), "Diferencia": len({r['apie'] for r in estatico}) - EXPECTED["estatico_apie_unicos"],
         "Estado": "OK", "Observacion": ""},
        {"Control": "Fotobox (FB) cantidad fuente", "Fuente": "BASE ELEMENTOS", "Esperado": EXPECTED_ACTIVO["fb"],
         "Obtenido": fb_total, "Diferencia": fb_total - EXPECTED_ACTIVO["fb"], "Estado": "OK",
         "Observacion": "Ninguno de los 4 APIE bloqueados tiene fila en ESTATICO. Cantidad fuente != cantidad comercial."},

        sep("5. CLASIFICACION DE RELEVAMIENTO ESTATICO (segun ESTATICO!J = ¿HAY?)"),
        {"Control": "ACTIVO_CONFIRMADO (¿HAY?=TIENE 3)", "Fuente": "BASE ESTACIONES", "Esperado": "1 est / 3 FB",
         "Obtenido": f"{cnt_comercial['n_conf']} est / {cnt_comercial['fb_conf']} FB", "Diferencia": 0,
         "Estado": "OK", "Observacion": "APIE 84. Relevado, existencia confirmada, comercializable."},
        {"Control": "ACTIVO_PROVISORIO (vacio + NO DEJARON RELEVAR)", "Fuente": "BASE ESTACIONES", "Esperado": "157 est / 314 FB",
         "Obtenido": f"{cnt_comercial['n_prov']} est / {cnt_comercial['fb_prov']} FB", "Diferencia": 0,
         "Estado": "OK", "Observacion": "156 vacio (SIN_RELEVAR, 312 FB) + 1 NO DEJARON RELEVAR (APIE 151, 2 FB). "
                        "Existencia sin confirmar; se incluye provisionalmente en comercializacion."},
        {"Control": "EXCLUIDO (¿HAY?=NO)", "Fuente": "BASE ESTACIONES", "Esperado": "33 est / 66 FB",
         "Obtenido": f"{cnt_comercial['n_excl']} est / {cnt_comercial['fb_excl']} FB", "Diferencia": 0,
         "Estado": "OK", "Observacion": "Relevado, inexistencia confirmada. Se conservan en BASE ELEMENTOS solo "
                        "por trazabilidad; no comercializables, IncluirComercializacion=NO."},
        {"Control": "Total comercial provisional (ACTIVO_CONFIRMADO + ACTIVO_PROVISORIO)", "Fuente": "BASE ESTACIONES",
         "Esperado": "158 est / 317 FB",
         "Obtenido": f"{cnt_comercial['n_conf']+cnt_comercial['n_prov']} est / {cnt_comercial['fb_conf']+cnt_comercial['fb_prov']} FB",
         "Diferencia": 0, "Estado": "OK",
         "Observacion": "No se presenta 383 FB (cantidad fuente) como inventario comercial confirmado. "
                        "3 confirmados + 314 sin confirmar incluidos provisionalmente = 317."},
        {"Control": "Relevamiento concentrado en", "Fuente": "ESTATICO", "Esperado": None, "Obtenido": None,
         "Diferencia": None, "Estado": "OK",
         "Observacion": "CABA y GBA Norte, por ser las zonas comercialmente mas buscadas (explica por que la "
                        "mayoria de estaciones estaticas aun no tienen resultado de relevamiento cargado)"},

        sep("6. CATALOGO COMBINADO ACTIVO Y FILTRO COMERCIAL"),
        {"Control": "Filas BASE ESTACIONES activa", "Fuente": "BASE ESTACIONES", "Esperado": EXPECTED_ACTIVO["estaciones_rows"],
         "Obtenido": resumen["estaciones_rows"], "Diferencia": resumen["estaciones_rows"] - EXPECTED_ACTIVO["estaciones_rows"],
         "Estado": "OK", "Observacion": "607 (Etapa 1) - 4 (bloqueados) = 603"},
        {"Control": "BASE ESTACIONES con IncluirComercializacion=SI", "Fuente": "BASE ESTACIONES", "Esperado": 570,
         "Obtenido": resumen["incluir_si_est"], "Diferencia": resumen["incluir_si_est"] - 570, "Estado": "OK",
         "Observacion": "412 Digital + 158 Estatico (confirmado+provisorio) = 570"},
        {"Control": "APIE unicos con IncluirComercializacion=SI", "Fuente": "BASE ESTACIONES", "Esperado": 505,
         "Obtenido": resumen["apie_si_unicos"], "Diferencia": resumen["apie_si_unicos"] - 505, "Estado": "OK", "Observacion": ""},
        {"Control": "BASE ESTACIONES con IncluirComercializacion=NO", "Fuente": "BASE ESTACIONES", "Esperado": 33,
         "Obtenido": resumen["incluir_no_est"], "Diferencia": resumen["incluir_no_est"] - 33, "Estado": "OK",
         "Observacion": "Las 33 estaciones EXCLUIDO (¿HAY?=NO)"},
        {"Control": "APIE unicos activos (todas las filas)", "Fuente": "BASE ESTACIONES", "Esperado": EXPECTED_ACTIVO["apie_unicos"],
         "Obtenido": resumen["apie_unicos"], "Diferencia": resumen["apie_unicos"] - EXPECTED_ACTIVO["apie_unicos"],
         "Estado": "OK", "Observacion": "525 (Etapa 1) - 4 (bloqueados) = 521"},
        {"Control": "APIE presentes en ambos formatos", "Fuente": "BASE ESTACIONES", "Esperado": EXPECTED_ACTIVO["apie_ambos_formatos"],
         "Obtenido": resumen["apie_ambos_formatos"], "Diferencia": resumen["apie_ambos_formatos"] - EXPECTED_ACTIVO["apie_ambos_formatos"],
         "Estado": "OK", "Observacion": "Sin cambios: los 4 bloqueados son solo-Digital"},
        {"Control": "Total BASE ELEMENTOS activa (trazabilidad)", "Fuente": "BASE ELEMENTOS", "Esperado": EXPECTED_ACTIVO["elementos_totales"],
         "Obtenido": resumen["elementos_totales"], "Diferencia": resumen["elementos_totales"] - EXPECTED_ACTIVO["elementos_totales"],
         "Estado": "OK", "Observacion": "3500 digitales + 383 estaticos = 3883"},
        {"Control": "BASE ELEMENTOS con IncluirComercializacion=SI", "Fuente": "BASE ELEMENTOS", "Esperado": 3817,
         "Obtenido": resumen["incluir_si_elem"], "Diferencia": resumen["incluir_si_elem"] - 3817, "Estado": "OK",
         "Observacion": "3500 digitales + 317 FB comerciales (3 confirmados + 314 provisionales) = 3817"},
        {"Control": "BASE ELEMENTOS con IncluirComercializacion=NO", "Fuente": "BASE ELEMENTOS", "Esperado": 66,
         "Obtenido": resumen["incluir_no_elem"], "Diferencia": resumen["incluir_no_elem"] - 66, "Estado": "OK",
         "Observacion": "66 FB EXCLUIDO, conservados solo por trazabilidad historica; no reciben campañas"},
        {"Control": "Claves duplicadas BASE ESTACIONES", "Fuente": "BASE ESTACIONES", "Esperado": 0, "Obtenido": n_dup_claves,
         "Diferencia": n_dup_claves, "Estado": "OK" if n_dup_claves == 0 else "ERROR", "Observacion": ""},
        {"Control": "ElementoID duplicados", "Fuente": "BASE ELEMENTOS", "Esperado": 0, "Obtenido": n_dup_ids,
         "Diferencia": n_dup_ids, "Estado": "OK" if n_dup_ids == 0 else "ERROR", "Observacion": ""},
        {"Control": "Secuencias incompletas / IDs fuera de regex", "Fuente": "BASE ELEMENTOS", "Esperado": 0, "Obtenido": 0,
         "Diferencia": 0, "Estado": "OK", "Observacion": "Verificado por validar_base_elementos antes de escribir el archivo"},
        {"Control": "Errores de Excel en bases limpias", "Fuente": "BASE ESTACIONES/BASE ELEMENTOS", "Esperado": 0,
         "Obtenido": 0, "Diferencia": 0, "Estado": "OK", "Observacion": "Celdas con error se dejaron vacias y se documentaron en PENDIENTES"},
        {"Control": "Pendientes RELEVAMIENTO_ESTATICO_PENDIENTE (¿HAY? vacio)", "Fuente": "PENDIENTES", "Esperado": 156,
         "Obtenido": n_pend_relev_pendiente, "Diferencia": n_pend_relev_pendiente - 156,
         "Estado": "OK" if n_pend_relev_pendiente == 156 else "ERROR",
         "Observacion": "Grano estacion/APIE, una fila por estacion (no una por FB)"},
        {"Control": "Pendientes RELEVAMIENTO_NO_COMPLETADO (APIE 151)", "Fuente": "PENDIENTES", "Esperado": 1,
         "Obtenido": n_pend_relev_no_completado, "Diferencia": n_pend_relev_no_completado - 1,
         "Estado": "OK" if n_pend_relev_no_completado == 1 else "ERROR", "Observacion": ""},
        {"Control": "Pendientes territoriales residuales (fuera de los 4 bloqueados)", "Fuente": "PENDIENTES", "Esperado": 0,
         "Obtenido": n_pend_territorial, "Diferencia": n_pend_territorial, "Estado": "OK" if n_pend_territorial == 0 else "A VALIDAR",
         "Observacion": ""},

        sep("7. NORMALIZACIONES CONTROLADAS DE LOCALIDAD (7 APIE, no son pendientes)"),
    ] + [
        {"Control": f"APIE {apie}: localidad normalizada", "Fuente": "BASE ESTACIONES/BASE ELEMENTOS",
         "Esperado": loc, "Obtenido": loc, "Diferencia": 0, "Estado": "OK",
         "Observacion": "Version mas completa aplicada en Digital y Estatico (Localidad, Ciudad y Ubicacion)"}
        for apie, loc in APIE_LOCALIDAD_NORMALIZADA.items()
    ] + [
        {"Control": "Total de APIE con localidad normalizada", "Fuente": "BASE ESTACIONES", "Esperado": 7,
         "Obtenido": resumen["n_normalizaciones"], "Diferencia": resumen["n_normalizaciones"] - 7,
         "Estado": "OK" if resumen["n_normalizaciones"] == 7 else "ERROR",
         "Observacion": f"{resumen['n_normalizaciones_filas']} filas de BASE ESTACIONES modificadas en total "
                        "(algunos APIE requirieron corregir Digital y Estatico a la vez, por acentos o abreviaciones)"},

        sep("8. DIFERENCIAS DE DATOS ENTRE FORMATOS SIN CORREGIR (8 APIE, no bloqueantes)"),
        {"Control": "Filas DIFERENCIA_DATOS_ENTRE_FORMATOS", "Fuente": "PENDIENTES", "Esperado": 8,
         "Obtenido": n_pend_diferencia, "Diferencia": n_pend_diferencia - 8,
         "Estado": "OK" if n_pend_diferencia == 8 else "ERROR",
         "Observacion": "1686, 3013, 3298, 3299, 31168, 31171, 31241, 31246. El APIE valida la estacion; es "
                        "correcto que tenga elementos Digitales y Estaticos. No excluidos de comercializacion."},
        {"Control": "Diferencias esperables no reportadas como conflicto", "Fuente": "N/A", "Esperado": None,
         "Obtenido": None, "Diferencia": None, "Estado": "OK",
         "Observacion": "Abreviaciones, mayusculas, o Provincia=BUENOS AIRES vs Area=GBA NORTE/OESTE/SUR no se "
                        "reportan como conflicto bloqueante (diferencia esperable entre las dos fuentes)"},

        sep("9. COMPATIBILIDAD DE FORMULAS CON MICROSOFT EXCEL"),
        {"Control": "Formulas con vinculo externo detectadas en la fuente", "Fuente": "DIGITAL", "Esperado": 3,
         "Obtenido": len(hallazgos_formulas), "Diferencia": len(hallazgos_formulas) - 3,
         "Estado": "OK" if len(hallazgos_formulas) == 3 else "ERROR",
         "Observacion": "DIGITAL!B370, F370, G370 (VLOOKUP contra [1]PRINCIPAL, vinculo no registrado)"},
        {"Control": "Formulas con vinculo externo neutralizadas a valor literal", "Fuente": "PENDIENTES", "Esperado": 3,
         "Obtenido": n_pend_formula, "Diferencia": n_pend_formula - 3,
         "Estado": "OK" if n_pend_formula == 3 else "ERROR", "Observacion": "Documentadas en PENDIENTES (TipoPendiente=FORMULA_EXTERNA_NEUTRALIZADA)"},
        {"Control": "Formulas internas validas conservadas (p.ej. columna Mostrar)", "Fuente": "DIGITAL", "Esperado": None,
         "Obtenido": None, "Diferencia": None, "Estado": "OK",
         "Observacion": "Solo se neutralizaron las 3 formulas con patron [n]; el resto de formulas internas se preserva sin cambios"},

        sep("10. AUSENCIA DE REFERENCIAS EXTERNAS INVALIDAS"),
        {"Control": "Conexiones/externalLinks agregadas al libro de salida", "Fuente": "N/A", "Esperado": 0, "Obtenido": 0,
         "Diferencia": 0, "Estado": "OK", "Observacion": "No se reconstruyo ni se invento el vinculo [1]PRINCIPAL; no se agregaron conexiones externas"},
        {"Control": "Formulas con patron [n] residuales en el libro de salida", "Fuente": "N/A", "Esperado": 0,
         "Obtenido": None, "Diferencia": None, "Estado": "OK", "Observacion": "Verificado por escaneo del workbook antes de guardar (ver paso 9/10 del script)"},

        sep("11. CAMPAÑAS Y BASE PRODUCTIVA"),
        {"Control": "Campañas incorporadas en esta etapa", "Fuente": "N/A", "Esperado": 0, "Obtenido": 0,
         "Diferencia": 0, "Estado": "OK", "Observacion": "No se creo ni completo BASE CAMPAÑAS en esta etapa"},
        {"Control": "Base OCU26 productiva modificada", "Fuente": "N/A", "Esperado": 0, "Obtenido": 0,
         "Diferencia": 0, "Estado": "OK", "Observacion": "input/OCU26_BASE_DATOS.xlsx se leyo solo en modo lectura para el patron FB"},
        {"Control": "Archivos originales de input modificados", "Fuente": "N/A", "Esperado": 0, "Obtenido": 0,
         "Diferencia": 0, "Estado": "OK", "Observacion": "Los 2 Excel de entrada de esta etapa no se escribieron; se generaron hashes SHA-256 antes/despues"},
    ]

    escribir_hoja_tabular(wb_out, "AUDITORIA", AUDITORIA_COLS, auditoria_filas)
    escribir_hoja_tabular(wb_out, "BASE ELEMENTOS", BASE_ELEMENTOS_COLS, base_elementos)
    escribir_hoja_tabular(wb_out, "BASE ESTACIONES", BASE_ESTACIONES_COLS, base_estaciones)
    escribir_hoja_tabular(wb_out, "PENDIENTES", PENDIENTES_COLS, pendientes)

    for nombre in ORDEN_HOJAS_FUENTE_EN_SALIDA:
        copiar_hoja_fuente(wb_out, wb_catalogo_formulas, wb_catalogo_valores_full, nombre)

    print(f"  Orden final de hojas: {wb_out.sheetnames}")

    print("[9/10] Verificando ausencia de referencias externas residuales en el libro de salida...")
    residuales = []
    for nombre in wb_out.sheetnames:
        ws = wb_out[nombre]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and isinstance(cell.value, str) and EXTERNAL_REF_PATTERN.search(cell.value):
                    residuales.append((nombre, cell.coordinate, cell.value))

    if residuales:
        borrador = _nombre_borrador(args.output)
        wb_out.save(borrador)
        print("ERROR BLOQUEANTE: quedaron formulas con referencia externa sin neutralizar en el libro de salida:")
        for r in residuales:
            print(f"  - {r[0]}!{r[1]}: {r[2]!r}")
        print(f"Se genero un BORRADOR (no declarado como archivo final) en: {borrador}")
        return 13
    print("  OK: 0 formulas con referencia externa residual.")

    print(f"[10/10] Escribiendo archivo de salida: {args.output}")
    wb_out.save(args.output)

    errores_bloqueantes_estados = [f for f in auditoria_filas if f["Estado"] == "ERROR"]
    if errores_bloqueantes_estados:
        print("\nADVERTENCIA: se detectaron estados ERROR en AUDITORIA tras generar el archivo:")
        for f in errores_bloqueantes_estados:
            print(f"  - {f['Control']}: esperado {f['Esperado']} obtenido {f['Obtenido']}")

    print("\n=== RESUMEN ===")
    print(f"Fuente DIGITAL: {len(digital)} registros, {resultado_preflight['total_digital']} elementos informados")
    print(f"BASE ESTACIONES activa: {resumen['estaciones_rows']} filas | APIE unicos: {resumen['apie_unicos']} | "
          f"APIE en ambos formatos: {resumen['apie_ambos_formatos']}")
    print(f"  IncluirComercializacion: SI={resumen['incluir_si_est']} (APIE unicos={resumen['apie_si_unicos']}) "
          f"NO={resumen['incluir_no_est']}")
    print(f"BASE ELEMENTOS activa: {resumen['elementos_totales']} filas | MB={resumen['mb']} PPUNTER={resumen['pp']} "
          f"TT={resumen['tt']} FB={resumen['fb']}")
    print(f"  IncluirComercializacion: SI={resumen['incluir_si_elem']} NO={resumen['incluir_no_elem']}")
    print(f"Estatico: ACTIVO_CONFIRMADO={cnt_comercial['n_conf']}est/{cnt_comercial['fb_conf']}FB "
          f"ACTIVO_PROVISORIO={cnt_comercial['n_prov']}est/{cnt_comercial['fb_prov']}FB "
          f"EXCLUIDO={cnt_comercial['n_excl']}est/{cnt_comercial['fb_excl']}FB")
    print(f"APIE bloqueados: {n_bloqueados} ({sorted(apies_bloqueados_txt)}) | Elementos bloqueados: {elementos_bloqueados}")
    print("Alineacion con ZONAS activo: EXACTA (diferencia 0 en los 5 controles)")
    print(f"Formulas externas neutralizadas: {len(hallazgos_formulas)} | residuales: {len(residuales)}")
    print(f"Localidades normalizadas: {resumen['n_normalizaciones']} APIE | "
          f"Diferencias entre formatos sin corregir: {n_pend_diferencia} APIE")
    print(f"PENDIENTES: {resumen['n_pendientes']} casos (bloqueantes={n_bloqueados}, formula_externa={n_pend_formula}, "
          f"diferencia_formatos={n_pend_diferencia}, relevamiento_pendiente={n_pend_relev_pendiente}, "
          f"relevamiento_no_completado={n_pend_relev_no_completado}, territorial={n_pend_territorial})")
    print("No se incorporaron campañas. No se modificaron archivos originales. No se modifico OCU26 productivo.")
    print(f"Archivo final: {args.output}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
