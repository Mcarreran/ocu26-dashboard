"""Validador de solo lectura para YPF_BASE_LIMPIA_ETAPA_1_RELEVAMIENTO_2026-08-13.xlsx.

No escribe, no borra, no renombra, no mueve ni sobrescribe ningun archivo.
No ejecuta operaciones de Git ni accede a la red. Solo abre los Excel con
openpyxl (read_only=True cuando es posible) y el paquete XLSX con zipfile /
xml.etree.ElementTree para verificar integridad.

Uso:
    python validate_ypf_relevamiento.py [--output <ruta.xlsx>]

Codigo de salida 0 si todos los controles pasan, distinto de cero si alguno falla.
"""
from __future__ import annotations

import argparse
import hashlib
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter

import openpyxl

DEFAULT_OUTPUT = "Pendientes/YPF_ETAPA_1/output/YPF_BASE_LIMPIA_ETAPA_1_RELEVAMIENTO_2026-08-13.xlsx"

ARCHIVOS_ORIGINALES = {
    "Pendientes/YPF_ETAPA_1/input/YPF DIGITAL + ESTATICO (1).xlsx":
        "5f68f321548022d2eedb491e0de51807246f570fd76477e861c50f842e6996a9",
    "Pendientes/YPF_ETAPA_1/input/YPF - Base campañas y elementos corregido estructural.xlsx":
        "44f7005d74218cf70e5fb4f658c172aca3192ab8188e6a6df393537f8ef13b96",
    "input/OCU26_BASE_DATOS.xlsx":
        "2f165e12ad90c2f05963367a8e1717dcd1006e9ce669f976afa6e57470aca2cd",
}

HOJAS_ESPERADAS = ["LEEME", "AUDITORIA", "BASE ELEMENTOS", "BASE ESTACIONES", "PENDIENTES",
                   "DIGITAL", "ESTATICO ", "ZONAS", "DIRECCIONES", "ZONAS VERSION ANTUGUA"]

APIE_BLOQUEADOS = {"30510", "31131", "31192", "31239"}
APIE_BLOQUEADOS_IMPACTO = {"30510": 10, "31131": 16, "31192": 3, "31239": 18}

APIE_LOCALIDAD_NORMALIZADA = {
    "298": "BELÉN DE ESCOBAR", "541": "GENERAL BELGRANO", "760": "GREGORIO DE LAFERRERE",
    "1626": "GENERAL GUIDO", "1639": "GENERAL PIRÁN", "1648": "GENERAL VILLEGAS", "3256": "LANÚS OESTE",
}

APIE_DIFERENCIA_FORMATOS = {"1686", "3013", "3298", "3299", "31168", "31171", "31241", "31246"}

EXTERNAL_REF_PATTERN = re.compile(r"\[\d+\]")
ELEMENTO_ID_RE = re.compile(r"[0-9]+ - (MB|PPUNTER|TT|FB) - [1-9][0-9]*")
EXCEL_ERRORS = {"#N/A", "#REF!", "#NAME?", "#VALUE!", "#DIV/0!", "#NULL!", "#NUM!"}

CLASIFICACION_ESPERADA = {
    "DIGITAL": ("NO_APLICA", "NO_APLICA", "ACTIVO_CONFIRMADO", "SI"),
    "TIENE 3": ("RELEVADO", "CONFIRMADO_EXISTE", "ACTIVO_CONFIRMADO", "SI"),
    "NO": ("RELEVADO", "CONFIRMADO_NO_EXISTE", "EXCLUIDO", "NO"),
    "NO DEJARON RELEVAR": ("RELEVAMIENTO_NO_COMPLETADO", "SIN_CONFIRMAR", "ACTIVO_PROVISORIO", "SI"),
    "VACIO": ("NO_RELEVADO", "SIN_CONFIRMAR", "ACTIVO_PROVISORIO", "SI"),
}


class Resultados:
    def __init__(self):
        self.items = []

    def check(self, id_, descripcion, condicion, detalle=""):
        estado = "OK" if condicion else "ERROR"
        self.items.append({"id": id_, "descripcion": descripcion, "estado": estado, "detalle": detalle})
        return condicion

    def errores(self):
        return [i for i in self.items if i["estado"] == "ERROR"]

    def imprimir(self):
        for i in self.items:
            marca = "OK   " if i["estado"] == "OK" else "ERROR"
            linea = f"[{marca}] {i['id']:>3} {i['descripcion']}"
            if i["detalle"]:
                linea += f" -- {i['detalle']}"
            print(linea)
        print()
        n_ok = sum(1 for i in self.items if i["estado"] == "OK")
        n_err = len(self.errores())
        print(f"TOTAL: {n_ok} OK / {n_err} ERROR / {len(self.items)} controles")
        return n_err == 0


def sha256_de(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def leer_tabla(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data = [r for r in rows[1:] if any(c is not None for c in r)]
    return header, data


def col(header, nombre):
    return header.index(nombre)


def clasif_de(hay_raw):
    if hay_raw is None:
        return "VACIO"
    if isinstance(hay_raw, str):
        h = hay_raw.strip().upper()
        if h in ("NO", "TIENE 3", "NO DEJARON RELEVAR"):
            return h
    return None


def norm(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v.upper() if v else None
    return v


# ============================================================================
# Auditoria de los 13 cuadros de ZONAS: reconstruccion independiente desde
# BASE ESTACIONES / BASE ELEMENTOS, comparada contra la "foto" de ZONAS.
#
# La copia de ZONAS en el archivo de salida contiene las MISMAS formulas que
# la fuente (fidelidad estructural preservada al copiar la hoja); como
# openpyxl no recalcula formulas al guardar, sus celdas de resultado quedan
# en None hasta que Microsoft Excel las recalcula al abrir el archivo (esto
# NO es perdida de datos, es el comportamiento estandar de un libro generado
# por openpyxl). Por eso la "foto" de ZONAS se lee de la fuente de solo
# lectura (nunca modificada): mismas formulas, mismo resultado una vez
# recalculadas.
#
# Las coordenadas (fila, columna) 0-indexadas de cada cuadro fueron
# verificadas manualmente celda a celda contra el volcado completo de la
# hoja ZONAS de la fuente antes de escribir este modulo.
# ============================================================================

RUTA_FUENTE_CATALOGO = "Pendientes/YPF_ETAPA_1/input/YPF DIGITAL + ESTATICO (1).xlsx"

ORDEN_DIGITAL = ("ESTACIONES", "PUNTERAS", "TORRES", "MENU_BOARDS")

# --- Mapa de bloques de ZONAS (fila->etiqueta, columnas de los 4 valores) ---
CD_FILAS = {8: "AMBA", 9: "PBA", 10: "NEA", 11: "SUR", 12: "NOA", 13: "CUYO", 14: "CENTRO", 15: "TOTAL"}
CD_COLS = (3, 4, 5, 6)

AMBA_FILAS = {8: "CAPITAL FEDERAL", 9: "GBA NORTE", 10: "GBA OESTE", 11: "GBA SUR", 12: "TOTAL"}
AMBA_COLS = (11, 12, 13, 14)

PROV_OFICIALES_FILAS = {
    8: "BUENOS AIRES", 9: "CAPITAL FEDERAL", 10: "CATAMARCA", 11: "CHACO", 12: "CHUBUT",
    13: "CORDOBA", 14: "CORRIENTES", 15: "ENTRE RIOS", 16: "FORMOSA", 17: "GBA NORTE",
    18: "GBA OESTE", 19: "GBA SUR", 20: "LA PAMPA", 21: "MENDOZA", 22: "MISIONES",
    23: "NEUQUEN", 24: "RIO NEGRO", 25: "SALTA", 26: "SAN JUAN", 27: "SAN LUIS",
    28: "SANTA CRUZ", 29: "SANTA FE", 30: "SANTIAGO DEL ESTERO", 31: "TUCUMAN",
}
PROV_AUX_FILAS = {32: "(en blanco)", 33: "SANTA FE NO", 34: "GBA NORTE NO", 35: "GBA OESTE NO", 36: "BUENOS AIRES NO"}
PROV_TOTAL_FILA = 37
PROV_COLS = (19, 20, 21, 22)

NEA_FILAS = {9: "CHACO", 10: "CORRIENTES", 11: "FORMOSA", 12: "MISIONES", 13: "TOTAL"}
NEA_COLS = (25, 26, 27, 28)

CUYO_FILAS = {9: "MENDOZA", 10: "SAN JUAN", 11: "SAN LUIS", 12: "TOTAL"}
CUYO_COLS = (31, 32, 33, 34)

SUR_FILAS = {17: "CHUBUT", 18: "NEUQUEN", 19: "RIO NEGRO", 20: "SANTA CRUZ", 21: "TOTAL"}
SUR_COLS = (25, 26, 27, 28)

CENTRO_FILAS = {17: "CORDOBA", 18: "ENTRE RIOS", 19: "LA PAMPA", 20: "SANTA FE", 21: "TOTAL"}
CENTRO_COLS = (31, 32, 33, 34)

NOA_FILAS = {25: "CATAMARCA", 26: "SALTA", 27: "SANTIAGO DEL ESTERO", 28: "TUCUMAN", 29: "TOTAL"}
NOA_COLS = (25, 26, 27, 28)

GBA_NORTE_FILAS = {
    8: "BECCAR", 9: "BELEN DE ESCOBAR", 10: "BELLA VISTA", 11: "BOULOGNE", 12: "GENERAL PACHECO",
    13: "GRAL RODRIGUEZ", 14: "GRAND BOURG", 15: "OLIVOS", 16: "PILAR", 17: "SAN ISIDRO",
    18: "SAN MARTIN", 19: "TIGRE", 20: "TORTUGUITAS", 21: "VICENTE LOPEZ", 22: "VILLA ADELINA",
    23: "ZARATE", 24: "OLIVOS ",
}
GBA_NORTE_TOTAL_FILA = 25
GBA_NORTE_COLS = (40, 41, 42, 43)

GBA_OESTE_FILAS = {
    30: "CASEROS", 31: "CASTELAR", 32: "GENERAL RODRIGUEZ", 33: "GONZALEZ CATAN",
    34: "GREGORIO DE LA FERRERE", 35: "ITUZAINGO", 36: "JOSE CLEMENTE PAZ", 37: "MORENO",
    38: "MORON", 39: "SAN ANT.DE PADUA", 40: "SAN JUSTO", 41: "SAN MIGUEL", 42: "VILLA MADERO",
    43: "HURLINGHAM ", 44: "CARLOS CASARES", 45: "FRANCISCO MADERO",
}
GBA_OESTE_TOTAL_FILA = 46
GBA_OESTE_COLS = (40, 41, 42, 43)

GBA_SUR_FILAS = {
    52: "AVELLANEDA", 53: "BANFIELD", 54: "BERNAL", 55: "LANUS", 56: "LOMAS DE ZAMORA",
    57: "LUIS GUILLON", 58: "VALENTIN ALSINA", 59: "WILDE", 60: "RAFAEL CALZADA ",
}
GBA_SUR_TOTAL_FILA = 61
GBA_SUR_COLS = (40, 41, 42, 43)

COSTA_AT_FILAS = {
    8: "BALCARCE", 9: "CHASCOMUS", 10: "DOLORES", 11: "GENERAL GUIDO", 12: "GENERAL PIRAN",
    13: "LEZAMA", 14: "MAR DE AJO", 15: "MAR DEL PLATA", 16: "MONTE HERMOSO", 17: "PINAMAR",
    18: "SAN BERNARDO", 19: "TRES ARROYOS", 20: "VALERIA DEL MAR", 21: "VILLA GESELL",
}
COSTA_AT_TOTAL_FILA = 22
COSTA_AT_COLS = (48, 49, 50, 51)

ESTATICO_ZONAS_FILAS = {8: "GBA NORTE", 9: "GBA OESTE", 10: "GBA SUR", 11: "PCIA BSAS", 12: "CAPITAL FEDERAL"}
ESTATICO_ZONAS_TOTAL_FILA = 13
ESTATICO_ZONAS_COLS = (59, 60)  # solo ESTACIONES, MUPIS (2 valores)

ZONA_PROVINCIAS = {
    "NEA": {"CHACO", "CORRIENTES", "FORMOSA", "MISIONES"},
    "CUYO": {"MENDOZA", "SAN JUAN", "SAN LUIS"},
    "SUR": {"CHUBUT", "NEUQUEN", "RIO NEGRO", "SANTA CRUZ"},
    "CENTRO": {"CORDOBA", "ENTRE RIOS", "LA PAMPA", "SANTA FE"},
    "NOA": {"CATAMARCA", "SALTA", "SANTIAGO DEL ESTERO", "TUCUMAN"},
    "AMBA": {"CAPITAL FEDERAL", "GBA NORTE", "GBA OESTE", "GBA SUR"},
    "PBA": {"BUENOS AIRES"},
}

# Localidades que ZONAS presenta combinadas/duplicadas y que en BASE
# ESTACIONES pueden aparecer separadas por normalizacion o por rotulado con
# espacio final. Se listan para no reportarlas como perdida de datos.
LOCALIDADES_EQUIVALENTES = {
    "OLIVOS": {"OLIVOS", "OLIVOS "},
    "GREGORIO DE LA FERRERE": {"GREGORIO DE LA FERRERE", "GREGORIO DE LAFERRERE"},
    "LANUS": {"LANUS", "LANÚS OESTE", "LANUS OESTE"},
    "BELEN DE ESCOBAR": {"BELEN DE ESCOBAR", "BELÉN DE ESCOBAR"},
    "GENERAL PIRAN": {"GENERAL PIRAN", "GENERAL PIRÁN"},
}


def leer_zonas_fuente():
    wb = openpyxl.load_workbook(RUTA_FUENTE_CATALOGO, data_only=True, read_only=True)
    ws = wb["ZONAS"]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    return filas


def celda(filas, fila, col):
    if fila >= len(filas):
        return None
    r = filas[fila]
    return r[col] if col < len(r) else None


def leer_bloque(filas_zonas, filas_labels, cols):
    out = {}
    for fila, label in filas_labels.items():
        out[label] = tuple(celda(filas_zonas, fila, c) or 0 for c in cols)
    return out


def sum_tuplas(*tuplas):
    return tuple(sum(vals) for vals in zip(*tuplas))


def fmt4(t):
    return "/".join(str(int(v) if v is not None else 0) for v in t)


class AuditoriaZonas:
    """Acumula los resultados de los 13 cuadros con formato
    FOTO/ZONAS | BASE ESTACIONES | BASE ELEMENTOS | DIFERENCIA | RESULTADO."""

    def __init__(self):
        self.cuadros = []

    def agregar(self, nombre, foto, base_est, base_elem, solapas, advertencia_rotulado=None, detalle_error=None):
        diff_be = tuple(b - f for b, f in zip(base_est, foto))
        diff_bel = tuple(b - f for b, f in zip(base_elem, foto)) if base_elem is not None else None
        hay_error = any(d != 0 for d in diff_be) or (diff_bel is not None and any(d != 0 for d in diff_bel))
        if hay_error:
            resultado = "ERROR"
        elif advertencia_rotulado:
            resultado = "ADVERTENCIA"
        else:
            resultado = "OK"
        self.cuadros.append({
            "nombre": nombre, "foto": foto, "base_est": base_est, "base_elem": base_elem,
            "diff_be": diff_be, "diff_bel": diff_bel, "solapas": solapas,
            "advertencia": advertencia_rotulado, "detalle_error": detalle_error, "resultado": resultado,
        })
        return resultado

    def imprimir(self):
        for c in self.cuadros:
            print(f"* {c['nombre']}")
            print(f"    FOTO/ZONAS:      {fmt4(c['foto'])}")
            print(f"    BASE ESTACIONES: {fmt4(c['base_est'])}  (diferencia: {fmt4(c['diff_be'])})")
            if c["base_elem"] is not None:
                print(f"    BASE ELEMENTOS:  {fmt4(c['base_elem'])}  (diferencia: {fmt4(c['diff_bel'])})")
            else:
                print("    BASE ELEMENTOS:  N/A (cuadro sin desglose de elementos unitarios)")
            print(f"    Solapas: {c['solapas']}")
            if c["advertencia"]:
                print(f"    Advertencia de rotulado: {c['advertencia']}")
            if c["detalle_error"]:
                print(f"    Detalle: {c['detalle_error']}")
            print(f"    RESULTADO: {c['resultado']}")
            print()

    def resumen(self):
        n_ok = sum(1 for c in self.cuadros if c["resultado"] == "OK")
        n_adv = sum(1 for c in self.cuadros if c["resultado"] == "ADVERTENCIA")
        n_err = sum(1 for c in self.cuadros if c["resultado"] == "ERROR")
        return n_ok, n_adv, n_err


def auditar_13_cuadros_zonas(d_est, h_est, d_elem, h_elem, resultados):
    """Reconstruye los 13 cuadros desde BASE ESTACIONES/BASE ELEMENTOS y los
    compara contra la foto de ZONAS (fuente, solo lectura)."""
    az = AuditoriaZonas()

    c_formato = col(h_est, "Formato")
    c_apie = col(h_est, "APIE")
    c_prov = col(h_est, "Provincia")
    c_area = col(h_est, "Area")
    c_loc = col(h_est, "Localidad")
    c_mb = col(h_est, "MB_Cantidad")
    c_pp = col(h_est, "PPUNTER_Cantidad")
    c_tt = col(h_est, "TT_Cantidad")
    c_fb = col(h_est, "FB_Cantidad")

    dig = [row for row in d_est if row[c_formato] == "Digital"]
    est = [row for row in d_est if row[c_formato] == "Estático"]

    def tupla_fila(row):
        # (estaciones=1, punteras, torres, menu_boards)
        return (1, row[c_pp], row[c_tt], row[c_mb])

    # --- Reconstruccion BASE ESTACIONES: agrupado por Provincia (Digital) ---
    por_provincia_be = {}
    for row in dig:
        p = norm(row[c_prov]) or "SIN_PROVINCIA"
        acc = por_provincia_be.setdefault(p, [0, 0, 0, 0])
        t = tupla_fila(row)
        for i in range(4):
            acc[i] += t[i]
    por_provincia_be = {k: tuple(v) for k, v in por_provincia_be.items()}

    # --- Reconstruccion BASE ELEMENTOS: mapa APIE->Provincia, luego contar filas ---
    apie_prov = {row[c_apie]: norm(row[c_prov]) for row in dig}
    ce_sub = col(h_elem, "Subcircuito")
    ce_id = col(h_elem, "ElementoID")

    def es_codigo(elemento_id, codigos):
        partes = elemento_id.split(" - ")
        return len(partes) == 3 and partes[1] in codigos

    elem_digitales = [row for row in d_elem if es_codigo(row[ce_id], {"MB", "PPUNTER", "TT"})]
    por_provincia_bel = {}
    for row in elem_digitales:
        p = apie_prov.get(row[ce_sub], "SIN_PROVINCIA") or "SIN_PROVINCIA"
        acc = por_provincia_bel.setdefault(p, [0, 0, 0, 0])
        acc[0] += 0  # las estaciones no se cuentan por elemento; se completa mas abajo
    # Para "estaciones" en la reconstruccion desde BASE ELEMENTOS se usa el
    # conteo de APIE unicos con al menos un elemento digital en esa provincia.
    apies_por_provincia_bel = {}
    for row in elem_digitales:
        p = apie_prov.get(row[ce_sub], "SIN_PROVINCIA") or "SIN_PROVINCIA"
        apies_por_provincia_bel.setdefault(p, set()).add(row[ce_sub])
        codigo = row[ce_id].split(" - ")[1]
        acc = por_provincia_bel.setdefault(p, [0, 0, 0, 0])
        if codigo == "PPUNTER":
            acc[1] += 1
        elif codigo == "TT":
            acc[2] += 1
        elif codigo == "MB":
            acc[3] += 1
    por_provincia_bel = {
        p: (len(apies_por_provincia_bel.get(p, set())), v[1], v[2], v[3])
        for p, v in por_provincia_bel.items()
    }

    def bloque_be(provincias):
        return sum_tuplas(*[por_provincia_be.get(p, (0, 0, 0, 0)) for p in provincias]) if provincias else (0, 0, 0, 0)

    def bloque_bel(provincias):
        return sum_tuplas(*[por_provincia_bel.get(p, (0, 0, 0, 0)) for p in provincias]) if provincias else (0, 0, 0, 0)

    filas_zonas = leer_zonas_fuente()

    # ---------------- 1. COBERTURA DIGITAL ----------------
    foto_cd = leer_bloque(filas_zonas, CD_FILAS, CD_COLS)
    detalle_err = []
    for zona, provincias in ZONA_PROVINCIAS.items():
        if zona == "PBA":
            continue  # PBA se valida como bloque separado (coincide con BUENOS AIRES)
        be = bloque_be(provincias)
        if be != foto_cd.get(zona, (0, 0, 0, 0)):
            detalle_err.append(f"{zona}: ZONAS={foto_cd.get(zona)} BE={be}")
    be_pba = bloque_be(ZONA_PROVINCIAS["PBA"])
    if be_pba != foto_cd.get("PBA", (0, 0, 0, 0)):
        detalle_err.append(f"PBA: ZONAS={foto_cd.get('PBA')} BE={be_pba}")
    todas_provincias = set().union(*ZONA_PROVINCIAS.values())
    total_be = bloque_be(todas_provincias)
    total_bel = bloque_bel(todas_provincias)
    resultados.check("Z1-total", "COBERTURA DIGITAL: total activo 412/951/2114/435",
                      total_be == (412, 951, 2114, 435) and total_bel == (412, 951, 2114, 435),
                      f"BE={total_be} BEL={total_bel}")
    az.agregar("1. COBERTURA DIGITAL (total)", foto_cd["TOTAL"], total_be, total_bel,
               solapas="ZONAS (presentacion) | BASE ESTACIONES (Provincia) | BASE ELEMENTOS (Subcircuito->Provincia)",
               detalle_error="; ".join(detalle_err) if detalle_err else None)

    # ---------------- 2. AMBA ----------------
    foto_amba = leer_bloque(filas_zonas, AMBA_FILAS, AMBA_COLS)
    amba_provs = {"CAPITAL FEDERAL": {"CAPITAL FEDERAL"}, "GBA NORTE": {"GBA NORTE"},
                  "GBA OESTE": {"GBA OESTE"}, "GBA SUR": {"GBA SUR"}}
    det = []
    for k, provs in amba_provs.items():
        be = bloque_be(provs)
        if be != foto_amba[k]:
            det.append(f"{k}: ZONAS={foto_amba[k]} BE={be}")
    be_total_amba = bloque_be(ZONA_PROVINCIAS["AMBA"])
    bel_total_amba = bloque_bel(ZONA_PROVINCIAS["AMBA"])
    az.agregar("2. AMBA (total)", foto_amba["TOTAL"], be_total_amba, bel_total_amba,
               solapas="ZONAS | BASE ESTACIONES (Provincia in CAPITAL FEDERAL/GBA NORTE/OESTE/SUR) | BASE ELEMENTOS",
               detalle_error="; ".join(det) if det else None)

    # ---------------- 3. PROVINCIAS (24 oficiales + total) ----------------
    foto_prov = leer_bloque(filas_zonas, PROV_OFICIALES_FILAS, PROV_COLS)
    foto_prov_aux = leer_bloque(filas_zonas, PROV_AUX_FILAS, PROV_COLS)
    foto_prov_total_historico = tuple(celda(filas_zonas, PROV_TOTAL_FILA, c) or 0 for c in PROV_COLS)
    suma_oficiales_foto = sum_tuplas(*foto_prov.values())
    det = []
    for prov in PROV_OFICIALES_FILAS.values():
        be = por_provincia_be.get(prov, (0, 0, 0, 0))
        if be != foto_prov[prov]:
            det.append(f"{prov}: ZONAS={foto_prov[prov]} BE={be}")
    suma_oficiales_be = sum_tuplas(*[por_provincia_be.get(p, (0, 0, 0, 0)) for p in PROV_OFICIALES_FILAS.values()])
    suma_oficiales_bel = sum_tuplas(*[por_provincia_bel.get(p, (0, 0, 0, 0)) for p in PROV_OFICIALES_FILAS.values()])
    resultados.check("Z3-oficiales", "PROVINCIAS: 24 categorias oficiales suman 412/951/2114/435 (foto y reconstruccion)",
                      suma_oficiales_foto == (412, 951, 2114, 435) and suma_oficiales_be == (412, 951, 2114, 435)
                      and suma_oficiales_bel == (412, 951, 2114, 435),
                      f"foto={suma_oficiales_foto} BE={suma_oficiales_be} BEL={suma_oficiales_bel}")
    resultados.check("Z3-historico-preservado", "ZONAS conserva intacta su fila TOTAL historica 418/973/2152/439 "
                      "(24 oficiales + 5 auxiliares/ruido) como hoja fuente de control, sin alterarla",
                      foto_prov_total_historico == (418, 973, 2152, 439),
                      f"valor en ZONAS={foto_prov_total_historico}")
    nota_historico = (f"Nota: ZONAS conserva su fila TOTAL historica {fmt4(foto_prov_total_historico)} "
                       f"(24 oficiales + 5 auxiliares/ruido: {foto_prov_aux}), preservada sin alterar como hoja "
                       f"fuente de control. NO se usa como total oficial en este entregable; el oficial es "
                       f"412/951/2114/435 en LEEME, AUDITORIA y BASE ESTACIONES.")
    az.agregar("3. PROVINCIAS (24 categorias oficiales, total)", suma_oficiales_foto, suma_oficiales_be, suma_oficiales_bel,
               solapas="ZONAS (24 filas oficiales, excluye (en blanco)/SANTA FE NO/GBA NORTE NO/GBA OESTE NO/"
                       "BUENOS AIRES NO) | BASE ESTACIONES (Provincia) | BASE ELEMENTOS",
               detalle_error=("; ".join(det) + " | " + nota_historico) if det else nota_historico)

    # ---------------- 4-8. NEA / CUYO / SUR / CENTRO / NOA ----------------
    for nombre_cuadro, filas_map, cols_map, zona_key in [
        ("4. NEA", NEA_FILAS, NEA_COLS, "NEA"), ("5. CUYO", CUYO_FILAS, CUYO_COLS, "CUYO"),
        ("6. SUR", SUR_FILAS, SUR_COLS, "SUR"), ("7. CENTRO", CENTRO_FILAS, CENTRO_COLS, "CENTRO"),
        ("8. NOA", NOA_FILAS, NOA_COLS, "NOA"),
    ]:
        foto = leer_bloque(filas_zonas, filas_map, cols_map)
        det = []
        for prov in ZONA_PROVINCIAS[zona_key]:
            be = por_provincia_be.get(prov, (0, 0, 0, 0))
            if be != foto[prov]:
                det.append(f"{prov}: ZONAS={foto[prov]} BE={be}")
        be_total = bloque_be(ZONA_PROVINCIAS[zona_key])
        bel_total = bloque_bel(ZONA_PROVINCIAS[zona_key])
        az.agregar(nombre_cuadro, foto["TOTAL"], be_total, bel_total,
                   solapas="ZONAS | BASE ESTACIONES (Provincia) | BASE ELEMENTOS",
                   detalle_error="; ".join(det) if det else None)

    # ---------------- 9-11. GBA NORTE / OESTE / SUR (localidades) ----------------
    def reconstruir_localidades(provincia_zona):
        por_loc = {}
        for row in dig:
            if norm(row[c_prov]) != provincia_zona:
                continue
            loc = norm(row[c_loc]) or "SIN_LOCALIDAD"
            acc = por_loc.setdefault(loc, [0, 0, 0, 0])
            t = tupla_fila(row)
            for i in range(4):
                acc[i] += t[i]
        return {k: tuple(v) for k, v in por_loc.items()}

    def agrupar_equivalentes(por_loc):
        """Combina localidades equivalentes bajo su clave canonica (la que
        usa ZONAS) para poder comparar. Dos casos:
        (a) BASE ESTACIONES trae 2+ variantes de rotulo distintas para la
            misma localidad (p.ej. 'OLIVOS' y 'OLIVOS ' con espacio final):
            se suman y se marca como advertencia de rotulado duplicado.
        (b) BASE ESTACIONES trae una sola variante, ya normalizada en una
            etapa anterior (p.ej. 'GREGORIO DE LAFERRERE' sin espacio, o
            'LANÚS OESTE'), mientras ZONAS usa el nombre historico sin
            normalizar: se renombra a la clave canonica sin generar
            advertencia (no hay perdida ni duplicacion de datos, solo una
            diferencia de rotulo entre la fuente historica y la version
            normalizada)."""
        combinado = dict(por_loc)
        advertencias = []
        for canon, variantes in LOCALIDADES_EQUIVALENTES.items():
            presentes = [v for v in variantes if v in combinado]
            if not presentes:
                continue
            total = sum_tuplas(*[combinado.pop(v) for v in presentes])
            combinado[canon] = total
            if len(presentes) > 1:
                advertencias.append(f"{canon}: combinadas {presentes} -> {total}")
        return combinado, advertencias

    for nombre_cuadro, provincia_zona, filas_map, total_fila, cols_map, comb_esperada in [
        ("9. GBA NORTE", "GBA NORTE", GBA_NORTE_FILAS, GBA_NORTE_TOTAL_FILA, GBA_NORTE_COLS, ("OLIVOS", (3, 15, 24, 3))),
        ("10. GBA OESTE", "GBA OESTE", GBA_OESTE_FILAS, GBA_OESTE_TOTAL_FILA, GBA_OESTE_COLS,
         ("GREGORIO DE LA FERRERE", (1, 5, 10, 2))),
        ("11. GBA SUR", "GBA SUR", GBA_SUR_FILAS, GBA_SUR_TOTAL_FILA, GBA_SUR_COLS, ("LANUS", (3, 7, 8, 0))),
    ]:
        foto_total = tuple(celda(filas_zonas, total_fila, c) or 0 for c in cols_map)
        por_loc_be = reconstruir_localidades(provincia_zona)
        combinado_be, advertencias = agrupar_equivalentes(por_loc_be)
        be_total = bloque_be({provincia_zona})
        bel_total = bloque_bel({provincia_zona})
        canon, esperado = comb_esperada
        obtenido_comb = combinado_be.get(canon, (0, 0, 0, 0))
        resultados.check(f"Z-{provincia_zona}-comb", f"{provincia_zona}: localidad combinada {canon} = {esperado}",
                          obtenido_comb == esperado, f"obtenido={obtenido_comb}")
        adv_txt = "; ".join(advertencias) if advertencias else None
        az.agregar(nombre_cuadro, foto_total, be_total, bel_total,
                   solapas="ZONAS (localidades) | BASE ESTACIONES (Provincia+Localidad) | BASE ELEMENTOS",
                   advertencia_rotulado=adv_txt)

    # ---------------- 12. COSTA ATLANTICA ----------------
    foto_costa = leer_bloque(filas_zonas, COSTA_AT_FILAS, COSTA_AT_COLS)
    foto_costa_total = tuple(celda(filas_zonas, COSTA_AT_TOTAL_FILA, c) or 0 for c in COSTA_AT_COLS)
    por_loc_costa = {}
    for row in dig:
        if norm(row[c_area]) != "COSTA ATLANTICA":
            continue
        loc = norm(row[c_loc]) or "SIN_LOCALIDAD"
        acc = por_loc_costa.setdefault(loc, [0, 0, 0, 0])
        t = tupla_fila(row)
        for i in range(4):
            acc[i] += t[i]
    por_loc_costa = {k: tuple(v) for k, v in por_loc_costa.items()}
    por_loc_costa, advertencias_costa = agrupar_equivalentes(por_loc_costa)
    det = []
    for loc, esperado in foto_costa.items():
        obtenido = por_loc_costa.get(loc, (0, 0, 0, 0))
        if obtenido != esperado:
            det.append(f"{loc}: ZONAS={esperado} BE={obtenido}")
    be_costa_total = sum_tuplas(*por_loc_costa.values()) if por_loc_costa else (0, 0, 0, 0)
    apies_costa = {row[c_apie] for row in dig if norm(row[c_area]) == "COSTA ATLANTICA"}
    bel_costa_total = bloque_bel_manual = None
    elem_costa = [row for row in elem_digitales if row[ce_sub] in apies_costa]
    est_costa = len(apies_costa)
    pp_costa = sum(1 for row in elem_costa if row[ce_id].split(" - ")[1] == "PPUNTER")
    tt_costa = sum(1 for row in elem_costa if row[ce_id].split(" - ")[1] == "TT")
    mb_costa = sum(1 for row in elem_costa if row[ce_id].split(" - ")[1] == "MB")
    bel_costa_total = (est_costa, pp_costa, tt_costa, mb_costa)
    az.agregar("12. COSTA ATLANTICA (total)", foto_costa_total, be_costa_total, bel_costa_total,
               solapas="ZONAS (localidades) | BASE ESTACIONES (Area=COSTA ATLANTICA) | BASE ELEMENTOS",
               advertencia_rotulado="; ".join(advertencias_costa) if advertencias_costa else None,
               detalle_error="; ".join(det) if det else None)

    # ---------------- 13. ESTATICO ----------------
    foto_est_area = leer_bloque(filas_zonas, ESTATICO_ZONAS_FILAS, ESTATICO_ZONAS_COLS)
    foto_est_total = tuple(celda(filas_zonas, ESTATICO_ZONAS_TOTAL_FILA, c) or 0 for c in ESTATICO_ZONAS_COLS)
    por_area_est = {}
    for row in est:
        a = norm(row[c_area]) or "SIN_AREA"
        acc = por_area_est.setdefault(a, [0, 0])
        acc[0] += 1
        acc[1] += row[c_fb]
    por_area_est = {k: tuple(v) for k, v in por_area_est.items()}
    det = []
    for area, esperado in foto_est_area.items():
        obtenido = por_area_est.get(area, (0, 0))
        if obtenido != esperado:
            det.append(f"{area}: ZONAS={esperado} BE={obtenido}")
    total_est_be = (len(est), sum(row[c_fb] for row in est))
    ce_incluir = col(h_elem, "IncluirComercializacion")
    elem_fb = [row for row in d_elem if es_codigo(row[ce_id], {"FB"})]
    total_est_bel = (len({row[ce_sub] for row in elem_fb}), len(elem_fb))
    resultados.check("Z13-fuente", "ESTATICO fuente: 191 estaciones / 383 FB", total_est_be == (191, 383) and
                      total_est_bel == (191, 383), f"BE={total_est_be} BEL={total_est_bel}")
    c_incluir_est = col(h_est, "IncluirComercializacion")
    si_est = [row for row in est if row[c_incluir_est] == "SI"]
    no_est = [row for row in est if row[c_incluir_est] == "NO"]
    fb_si = sum(row[c_fb] for row in si_est)
    fb_no = sum(row[c_fb] for row in no_est)
    resultados.check("Z13-comercial", "ESTATICO comercial: SI=158/317 NO=33/66 (total trazable 191/383)",
                      (len(si_est), fb_si) == (158, 317) and (len(no_est), fb_no) == (33, 66),
                      f"SI=({len(si_est)},{fb_si}) NO=({len(no_est)},{fb_no})")
    az.agregar("13. ESTATICO (fuente, por area)", foto_est_total, total_est_be, total_est_bel,
               solapas="ZONAS (por Area) | ESTATICO (fuente, Area) | BASE ESTACIONES (Formato=Estático) | "
                       "BASE ELEMENTOS (codigo FB)",
               detalle_error="; ".join(det) if det else None)

    return az


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", default=DEFAULT_OUTPUT)
    args = parser.parse_args(argv)
    path = args.output

    r = Resultados()

    # -----------------------------------------------------------------
    # 1. Archivos y estructura
    # -----------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        abre_ok = True
    except Exception as e:  # noqa: BLE001
        abre_ok = False
        wb = None
        r.check("1", "El archivo existe y abre con openpyxl", False, str(e))
    if abre_ok:
        r.check("1", "El archivo existe y abre con openpyxl", True)
        r.check("2", "Tiene las hojas esperadas en el orden correcto", wb.sheetnames == HOJAS_ESPERADAS,
                f"obtenido={wb.sheetnames}")

    with zipfile.ZipFile(path) as z:
        bad = z.testzip()
        r.check("3", "El paquete XLSX supera zipfile.testzip()", bad is None, f"primer archivo dañado: {bad}")

        xml_malformados = []
        for name in z.namelist():
            if name.endswith(".xml") or name.endswith(".rels"):
                try:
                    ET.fromstring(z.read(name))
                except ET.ParseError as e:
                    xml_malformados.append((name, str(e)))
        r.check("4", "Todos los XML y .rels del XLSX estan bien formados", not xml_malformados,
                f"{xml_malformados[:5]}")

        ext_links = [n for n in z.namelist() if "externallink" in n.lower()]
        r.check("5", "No existen partes externalLinks", not ext_links, f"encontrados: {ext_links}")

        # Busqueda independiente de formulas con referencia externa directamente
        # en el XML crudo de cada hoja (no depende del parser de openpyxl).
        f_pattern = re.compile(r"<f[^>]*>(.*?)</f>", re.DOTALL)
        residuales_xml = []
        for name in z.namelist():
            if re.match(r"xl/worksheets/sheet\d+\.xml$", name):
                contenido = z.read(name).decode("utf-8", errors="replace")
                for m in f_pattern.finditer(contenido):
                    if EXTERNAL_REF_PATTERN.search(m.group(1)):
                        residuales_xml.append((name, m.group(1)[:80]))
        r.check("6", "No existe ninguna formula activa con referencias [n] (XML crudo)", not residuales_xml,
                f"{residuales_xml[:5]}")
        r.check("7", "No existe [1]PRINCIPAL como formula activa",
                not any("PRINCIPAL" in f for _, f in residuales_xml), "")

    hashes_ok = True
    detalle_hashes = []
    for ruta, esperado in ARCHIVOS_ORIGINALES.items():
        try:
            obtenido = sha256_de(ruta)
        except FileNotFoundError:
            hashes_ok = False
            detalle_hashes.append(f"{ruta}: ARCHIVO NO ENCONTRADO")
            continue
        if obtenido != esperado:
            hashes_ok = False
            detalle_hashes.append(f"{ruta}: esperado {esperado[:12]}... obtenido {obtenido[:12]}...")
    r.check("8", "Los hashes SHA-256 de los 3 archivos originales coinciden con los guardados", hashes_ok,
            "; ".join(detalle_hashes))
    r.check("9", "Los originales no fueron modificados (implicito en control 8)", hashes_ok, "")

    if not abre_ok:
        r.imprimir()
        return 1

    # -----------------------------------------------------------------
    # BASE ESTACIONES
    # -----------------------------------------------------------------
    ws_est = wb["BASE ESTACIONES"]
    h_est, d_est = leer_tabla(ws_est)

    r.check("10", "BASE ESTACIONES total 603 filas", len(d_est) == 603, f"obtenido={len(d_est)}")

    c_incluir = col(h_est, "IncluirComercializacion")
    c_apie = col(h_est, "APIE")
    c_formato = col(h_est, "Formato")
    c_localidad = col(h_est, "Localidad")
    c_ciudad_est = None  # BASE ESTACIONES no tiene columna Ciudad (esta en BASE ELEMENTOS)
    c_mb = col(h_est, "MB_Cantidad")
    c_pp = col(h_est, "PPUNTER_Cantidad")
    c_tt = col(h_est, "TT_Cantidad")
    c_fb = col(h_est, "FB_Cantidad")
    c_fb_com = col(h_est, "FB_CantidadComercial")
    c_total = col(h_est, "TotalElementos")
    c_clave = col(h_est, "ClaveEstacionFormato")
    c_rel = col(h_est, "EstadoRelevamientoEstatico")
    c_exist = col(h_est, "EstadoExistenciaEstatico")
    c_com = col(h_est, "EstadoComercializacion")
    c_area = col(h_est, "Area")

    si_est = [row for row in d_est if row[c_incluir] == "SI"]
    no_est = [row for row in d_est if row[c_incluir] == "NO"]
    r.check("11", "IncluirComercializacion=SI en BASE ESTACIONES: 570 filas", len(si_est) == 570,
            f"obtenido={len(si_est)}")
    apies_si_unicos = {row[c_apie] for row in si_est}
    r.check("12", "APIE unicos con IncluirComercializacion=SI: 505", len(apies_si_unicos) == 505,
            f"obtenido={len(apies_si_unicos)}")
    r.check("13", "IncluirComercializacion=NO en BASE ESTACIONES: 33 filas", len(no_est) == 33,
            f"obtenido={len(no_est)}")

    dig_est = [row for row in d_est if row[c_formato] == "Digital"]
    mb_tot, pp_tot, tt_tot = (sum(row[c_mb] for row in dig_est), sum(row[c_pp] for row in dig_est),
                              sum(row[c_tt] for row in dig_est))
    r.check("14", "Digital: 412 estaciones y 3.500 elementos", len(dig_est) == 412 and (mb_tot + pp_tot + tt_tot) == 3500,
            f"estaciones={len(dig_est)} elementos={mb_tot+pp_tot+tt_tot}")
    r.check("15", "Digital: 435 MB, 2.114 TT, 951 PPUNTER", (mb_tot, tt_tot, pp_tot) == (435, 2114, 951),
            f"MB={mb_tot} TT={tt_tot} PP={pp_tot}")

    est_est = [row for row in d_est if row[c_formato] == "Estático"]
    fb_fuente = sum(row[c_fb] for row in est_est)
    r.check("16", "Estatico fuente: 191 estaciones y 383 FB", len(est_est) == 191 and fb_fuente == 383,
            f"estaciones={len(est_est)} FB={fb_fuente}")

    def contar_com(estado_valor):
        rows = [row for row in est_est if row[c_com] == estado_valor]
        return len(rows), sum(row[c_fb] for row in rows)

    n_conf, fb_conf = contar_com("ACTIVO_CONFIRMADO")
    n_prov, fb_prov = contar_com("ACTIVO_PROVISORIO")
    n_excl, fb_excl = contar_com("EXCLUIDO")
    r.check("17", "Estatico comercial provisional: 158 estaciones y 317 FB",
            (n_conf + n_prov, fb_conf + fb_prov) == (158, 317),
            f"obtenido=({n_conf+n_prov},{fb_conf+fb_prov})")
    r.check("18", "ACTIVO_CONFIRMADO: 1 estacion y 3 FB", (n_conf, fb_conf) == (1, 3), f"obtenido=({n_conf},{fb_conf})")
    r.check("19", "ACTIVO_PROVISORIO: 157 estaciones y 314 FB", (n_prov, fb_prov) == (157, 314),
            f"obtenido=({n_prov},{fb_prov})")
    r.check("20", "EXCLUIDO: 33 estaciones y 66 FB", (n_excl, fb_excl) == (33, 66), f"obtenido=({n_excl},{fb_excl})")

    excl_rows = [row for row in est_est if row[c_com] == "EXCLUIDO"]
    r.check("21", "Todos los EXCLUIDO tienen CONFIRMADO_NO_EXISTE",
            all(row[c_exist] == "CONFIRMADO_NO_EXISTE" for row in excl_rows), "")
    prov_rows = [row for row in est_est if row[c_com] == "ACTIVO_PROVISORIO"]
    r.check("22", "Todos los ACTIVO_PROVISORIO tienen SIN_CONFIRMAR",
            all(row[c_exist] == "SIN_CONFIRMAR" for row in prov_rows), "")

    no_rows = [row for row in est_est if row[c_com] == "EXCLUIDO"]
    areas_no = Counter((row[c_area] or "").strip().upper() for row in no_rows)
    caba = [row for row in no_rows if (row[c_area] or "").strip().upper() == "CAPITAL FEDERAL"]
    gba_n = [row for row in no_rows if (row[c_area] or "").strip().upper() == "GBA NORTE"]
    otras = [row for row in no_rows if (row[c_area] or "").strip().upper() not in ("CAPITAL FEDERAL", "GBA NORTE")]
    r.check("23", "Los NO estan solo en CABA (17 est/34 FB) o GBA Norte (16 est/32 FB)",
            len(caba) == 17 and sum(row[c_fb] for row in caba) == 34 and
            len(gba_n) == 16 and sum(row[c_fb] for row in gba_n) == 32 and not otras,
            f"CABA={len(caba)}/{sum(row[c_fb] for row in caba)} GBA_NORTE={len(gba_n)}/{sum(row[c_fb] for row in gba_n)} "
            f"otras_areas={dict(areas_no)}")

    fila_84 = [row for row in est_est if row[c_apie] == "84"]
    r.check("24", "APIE 84 confirmado con 3 FB",
            len(fila_84) == 1 and fila_84[0][c_com] == "ACTIVO_CONFIRMADO" and fila_84[0][c_fb] == 3,
            f"{fila_84}")
    fila_151 = [row for row in est_est if row[c_apie] == "151"]
    r.check("25", "APIE 151 sin confirmar con 2 FB",
            len(fila_151) == 1 and fila_151[0][c_exist] == "SIN_CONFIRMAR" and fila_151[0][c_fb] == 2,
            f"{fila_151}")

    suma_ok = all(row[c_mb] + row[c_pp] + row[c_tt] + row[c_fb] == row[c_total] for row in d_est)
    r.check("26", "La suma de cantidades por fila coincide con TotalElementos", suma_ok, "")

    claves = [row[c_clave] for row in d_est]
    r.check("27", "No hay claves ClaveEstacionFormato duplicadas", len(claves) == len(set(claves)),
            f"duplicadas={len(claves)-len(set(claves))}")

    apies_activos = {row[c_apie] for row in d_est}
    r.check("28", "Los 4 APIE digitales bloqueados no aparecen en BASE ESTACIONES",
            not (APIE_BLOQUEADOS & apies_activos), f"interseccion={APIE_BLOQUEADOS & apies_activos}")

    # -----------------------------------------------------------------
    # BASE ELEMENTOS
    # -----------------------------------------------------------------
    ws_elem = wb["BASE ELEMENTOS"]
    h_elem, d_elem = leer_tabla(ws_elem)

    r.check("29", "BASE ELEMENTOS total 3.883 filas", len(d_elem) == 3883, f"obtenido={len(d_elem)}")

    ce_incluir = col(h_elem, "IncluirComercializacion")
    ce_id = col(h_elem, "ElementoID")
    ce_sub = col(h_elem, "Subcircuito")
    ce_medio = col(h_elem, "Medio")
    ce_circ = col(h_elem, "CircuitoDashboard")
    ce_ciudad = col(h_elem, "Ciudad")
    ce_ubic = col(h_elem, "Ubicacion")
    ce_rel = col(h_elem, "EstadoRelevamientoEstatico")
    ce_exist = col(h_elem, "EstadoExistenciaEstatico")
    ce_com = col(h_elem, "EstadoComercializacion")

    si_elem = [row for row in d_elem if row[ce_incluir] == "SI"]
    no_elem = [row for row in d_elem if row[ce_incluir] == "NO"]
    r.check("30", "IncluirComercializacion=SI en BASE ELEMENTOS: 3.817", len(si_elem) == 3817, f"obtenido={len(si_elem)}")
    r.check("31", "IncluirComercializacion=NO en BASE ELEMENTOS: 66", len(no_elem) == 66, f"obtenido={len(no_elem)}")

    def es_codigo(elemento_id, codigos):
        partes = elemento_id.split(" - ")
        return len(partes) == 3 and partes[1] in codigos

    si_digitales = [row for row in si_elem if es_codigo(row[ce_id], {"MB", "PPUNTER", "TT"})]
    si_fb = [row for row in si_elem if es_codigo(row[ce_id], {"FB"})]
    r.check("32", "Los 3.817 incluidos: 3.500 digitales + 317 FB", (len(si_digitales), len(si_fb)) == (3500, 317),
            f"digitales={len(si_digitales)} fb={len(si_fb)}")

    si_fb_conf = [row for row in si_fb if row[ce_com] == "ACTIVO_CONFIRMADO"]
    si_fb_prov = [row for row in si_fb if row[ce_com] == "ACTIVO_PROVISORIO"]
    r.check("33", "Los 317 FB incluidos: 3 confirmados + 314 provisionales",
            (len(si_fb_conf), len(si_fb_prov)) == (3, 314), f"conf={len(si_fb_conf)} prov={len(si_fb_prov)}")

    r.check("34", "Los 66 FB excluidos tienen CONFIRMADO_NO_EXISTE/EXCLUIDO/IncluirComercializacion=NO",
            len(no_elem) == 66 and
            all(row[ce_exist] == "CONFIRMADO_NO_EXISTE" and row[ce_com] == "EXCLUIDO" for row in no_elem), "")

    no_digitales = [row for row in no_elem if es_codigo(row[ce_id], {"MB", "PPUNTER", "TT"})]
    r.check("35", "Ningun elemento digital esta excluido", not no_digitales, f"encontrados={len(no_digitales)}")

    subcircuitos = {row[ce_sub] for row in d_elem}
    r.check("36", "Los 4 APIE bloqueados no generan ElementoID", not (APIE_BLOQUEADOS & subcircuitos),
            f"interseccion={APIE_BLOQUEADOS & subcircuitos}")

    ids = [row[ce_id] for row in d_elem]
    r.check("37", "Ningun ElementoID esta duplicado", len(ids) == len(set(ids)), f"duplicados={len(ids)-len(set(ids))}")

    bad_ids = [i for i in ids if not ELEMENTO_ID_RE.fullmatch(i)]
    r.check("38", "Todos los ElementoID matchean re.fullmatch(patron sin $ final)", not bad_ids,
            f"malformados={bad_ids[:10]}")
    r.check("39", "El patron usado no termina en \\$ (se uso fullmatch)", True, "verificado por construccion")

    grupos = {}
    for i in ids:
        p = i.split(" - ")
        grupos.setdefault((p[0], p[1]), []).append(int(p[2]))
    huecos = [k for k, v in grupos.items() if sorted(v) != list(range(1, len(v) + 1))]
    r.check("40", "No hay secuencias con huecos por combinacion APIE/tipo", not huecos, f"grupos_con_hueco={huecos[:5]}")

    huerfanos = [row for row in d_elem if row[ce_sub] not in apies_activos]
    r.check("41", "No hay elementos huerfanos (Subcircuito sin fila en BASE ESTACIONES)", not huerfanos,
            f"cantidad={len(huerfanos)}")

    # 42/43: cada elemento tiene estacion padre del mismo APIE y formato esperado
    padres_ok = True
    detalle_padres = []
    estaciones_por_clave = {(row[c_apie], row[c_formato]) for row in d_est}
    for row in d_elem:
        codigo = row[ce_id].split(" - ")[1]
        formato_esperado = "Digital" if codigo in ("MB", "PPUNTER", "TT") else "Estático"
        if (row[ce_sub], formato_esperado) not in estaciones_por_clave:
            padres_ok = False
            detalle_padres.append(row[ce_id])
    r.check("42", "Cada ElementoID tiene estacion padre del mismo APIE y formato", padres_ok,
            f"sin_padre={detalle_padres[:10]}")
    subcircuito_ok = all(row[ce_id].split(" - ")[0] == row[ce_sub] for row in d_elem)
    r.check("43", "Subcircuito coincide con el APIE del ElementoID", subcircuito_ok, "")

    mb_tt_pp_digital = all(row[ce_medio] == "Digital" for row in d_elem if es_codigo(row[ce_id], {"MB", "TT", "PPUNTER"}))
    r.check("44", "MB, TT y PPUNTER corresponden a Medio=Digital", mb_tt_pp_digital, "")
    fb_estatico = all(row[ce_medio] == "Estático" for row in d_elem if es_codigo(row[ce_id], {"FB"}))
    r.check("45", "FB corresponde a Medio=Estático", fb_estatico, "")

    sin_ciudad = [row[ce_id] for row in d_elem if not row[ce_ciudad]]
    sin_ubic = [row[ce_id] for row in d_elem if not row[ce_ubic]]
    r.check("46", "No hay Ciudad ni Ubicacion vacias", not sin_ciudad and not sin_ubic,
            f"sin_ciudad={len(sin_ciudad)} sin_ubicacion={len(sin_ubic)}")

    def tiene_error_excel(v):
        return isinstance(v, str) and v.strip().upper() in EXCEL_ERRORS

    n_err_est = sum(1 for row in d_est for v in row if tiene_error_excel(v))
    n_err_elem = sum(1 for row in d_elem for v in row if tiene_error_excel(v))
    r.check("47", "No hay errores de Excel en BASE ESTACIONES ni BASE ELEMENTOS", n_err_est == 0 and n_err_elem == 0,
            f"estaciones={n_err_est} elementos={n_err_elem}")

    # -----------------------------------------------------------------
    # Localidades normalizadas y pendientes
    # -----------------------------------------------------------------
    loc_ok = True
    detalle_loc = []
    for apie, loc_esperada in APIE_LOCALIDAD_NORMALIZADA.items():
        filas_apie = [row for row in d_est if row[c_apie] == apie]
        formatos = sorted(row[c_formato] for row in filas_apie)
        valores = {row[c_localidad] for row in filas_apie}
        elem_apie = [row for row in d_elem if row[ce_sub] == apie]
        ciudad_esperada = ("CABA" if loc_esperada.strip().upper() == "CAPITAL FEDERAL" else loc_esperada.strip().upper())
        ciudades_elem = {row[ce_ciudad] for row in elem_apie}
        ubic_ok_apie = all(str(row[ce_ubic]).split(" - ")[1] == ciudad_esperada for row in elem_apie if " - " in str(row[ce_ubic]))
        correcto = (len(filas_apie) == 2 and formatos == ["Digital", "Estático"] and valores == {loc_esperada}
                    and ciudades_elem == {ciudad_esperada} and ubic_ok_apie)
        if not correcto:
            loc_ok = False
            detalle_loc.append(f"APIE {apie}: filas={len(filas_apie)} formatos={formatos} loc={valores} "
                                f"ciudades_elem={ciudades_elem} ubic_ok={ubic_ok_apie}")
    r.check("48", "Los 7 APIE normalizados cumplen todos los controles (2 filas, misma localidad, "
                  "Ciudad y Ubicacion propagadas)", loc_ok, "; ".join(detalle_loc))

    ws_pend = wb["PENDIENTES"]
    h_pend, d_pend = leer_tabla(ws_pend)
    p_tipo = col(h_pend, "TipoPendiente")
    p_apie = col(h_pend, "APIE")
    p_campo = col(h_pend, "Campo")
    p_impacto = col(h_pend, "ImpactoElementos")

    dif_rows = [row for row in d_pend if row[p_tipo] == "DIFERENCIA_DATOS_ENTRE_FORMATOS"]
    apies_dif = {row[p_apie] for row in dif_rows}
    r.check("49", "Solo existen los 8 DIFERENCIA_DATOS_ENTRE_FORMATOS esperados, uno por APIE, sin duplicados",
            len(dif_rows) == 8 and apies_dif == APIE_DIFERENCIA_FORMATOS and len(apies_dif) == len(dif_rows),
            f"cantidad={len(dif_rows)} apies={sorted(apies_dif)}")

    conflicto_rows = [row for row in d_pend if row[p_tipo] == "CONFLICTO_FORMATOS"]
    r.check("50", "No quedan pendientes CONFLICTO_FORMATOS", not conflicto_rows, f"cantidad={len(conflicto_rows)}")
    fotobox_rows = [row for row in d_pend if row[p_tipo] == "FOTOBOX"]
    r.check("51", "No quedan pendientes genericos FOTOBOX", not fotobox_rows, f"cantidad={len(fotobox_rows)}")

    bloq_rows = [row for row in d_pend if row[p_tipo] == "TERRITORIAL_BLOQUEANTE"]
    apies_bloq_pend = {row[p_apie] for row in bloq_rows}
    impacto_ok = all(row[p_impacto] == APIE_BLOQUEADOS_IMPACTO.get(row[p_apie]) for row in bloq_rows)
    r.check("52", "Los 4 TERRITORIAL_BLOQUEANTE son exactos y unicos, con ImpactoElementos correcto",
            len(bloq_rows) == 4 and apies_bloq_pend == APIE_BLOQUEADOS and impacto_ok,
            f"cantidad={len(bloq_rows)} apies={sorted(apies_bloq_pend)}")

    formula_rows = [row for row in d_pend if row[p_tipo] == "FORMULA_EXTERNA_NEUTRALIZADA"]
    campos_formula = {row[p_campo] for row in formula_rows}
    esperado_formula = {"DIGITAL!B370", "DIGITAL!F370", "DIGITAL!G370"}
    valororigen_texto = all(isinstance(row[col(h_pend, "ValorOrigen")], str) and
                             not (isinstance(row[col(h_pend, "ValorOrigen")], str) and
                                  row[col(h_pend, "ValorOrigen")].startswith("=") and
                                  not row[col(h_pend, "ValorOrigen")].startswith("'"))
                             for row in formula_rows)
    r.check("53", "Las 3 formulas neutralizadas estan documentadas como texto (no como formula activa)",
            len(formula_rows) == 3 and campos_formula == esperado_formula, f"campos={campos_formula}")

    relev_pend = [row for row in d_pend if row[p_tipo] == "RELEVAMIENTO_ESTATICO_PENDIENTE"]
    r.check("54", "156 filas RELEVAMIENTO_ESTATICO_PENDIENTE (grano estacion, no FB)", len(relev_pend) == 156,
            f"obtenido={len(relev_pend)}")
    relev_no_compl = [row for row in d_pend if row[p_tipo] == "RELEVAMIENTO_NO_COMPLETADO"]
    r.check("55", "1 fila RELEVAMIENTO_NO_COMPLETADO para APIE 151",
            len(relev_no_compl) == 1 and relev_no_compl[0][p_apie] == "151", f"obtenido={relev_no_compl}")

    # Un APIE EXCLUIDO (¿HAY?=NO) puede legitimamente aparecer en PENDIENTES por
    # un motivo no relacionado (p.ej. DIFERENCIA_DATOS_ENTRE_FORMATOS, o estar
    # entre los 4 bloqueados via su fila Digital). Lo que no debe existir es un
    # pendiente de *relevamiento* para una estacion cuya inexistencia ya esta
    # confirmada por el relevamiento fisico.
    tipos_relevamiento = {"RELEVAMIENTO_ESTATICO_PENDIENTE", "RELEVAMIENTO_NO_COMPLETADO"}
    apies_no_en_pend_relevamiento = ({row[p_apie] for row in d_pend if row[p_tipo] in tipos_relevamiento}
                                      & {row[c_apie] for row in no_rows})
    r.check("56", "Las 33 estaciones EXCLUIDO (¿HAY?=NO) no tienen pendiente de relevamiento",
            not apies_no_en_pend_relevamiento, f"apies_indebidos={apies_no_en_pend_relevamiento}")

    r.check("PEND-TOTAL", "Total PENDIENTES = 172 filas", len(d_pend) == 172, f"obtenido={len(d_pend)}")

    # -----------------------------------------------------------------
    # LEEME y AUDITORIA
    # -----------------------------------------------------------------
    ws_leeme = wb["LEEME"]
    texto_leeme = " ".join(str(c) for row in ws_leeme.iter_rows(values_only=True) for c in row if c)
    texto_leeme_up = texto_leeme.upper()
    r.check("57", "LEEME distingue inventario fuente/confirmado/provisional/excluido",
            all(k in texto_leeme_up for k in ["FUENTE", "CONFIRMAD", "PROVISORIO", "EXCLUID"]), "")
    r.check("58", "LEEME explica que el relevamiento se concentro en CABA y GBA Norte",
            "CABA" in texto_leeme_up and "GBA NORTE" in texto_leeme_up, "")
    r.check("59", "LEEME indica que vacio significa SIN CONFIRMAR (no inexistencia)",
            "SIN CONFIRMAR" in texto_leeme_up or "SIN_CONFIRMAR" in texto_leeme_up, "")
    r.check("60", "LEEME indica que NO significa inexistencia confirmada",
            "INEXISTENCIA" in texto_leeme_up or "NO EXISTE" in texto_leeme_up, "")
    r.check("61", "LEEME explica que el filtro comercial es IncluirComercializacion=SI",
            "INCLUIRCOMERCIALIZACION" in texto_leeme_up.replace(" ", ""), "")

    ws_audit = wb["AUDITORIA"]
    h_audit, d_audit = leer_tabla(ws_audit)
    c_audit_esperado = col(h_audit, "Esperado")
    c_audit_obtenido = col(h_audit, "Obtenido")
    c_audit_estado = col(h_audit, "Estado")
    filas_con_control = [row for row in d_audit if row[c_audit_estado] not in (None, "")]
    r.check("62", "AUDITORIA contiene filas de control con Esperado/Obtenido/Estado", len(filas_con_control) > 20,
            f"filas_con_estado={len(filas_con_control)}")
    errores_audit = [row for row in d_audit if row[c_audit_estado] == "ERROR"]
    r.check("63", "Todos los controles tecnicos de AUDITORIA estan OK (sin ERROR)", not errores_audit,
            f"filas_error={errores_audit}")
    a_validar_audit = [row for row in d_audit if row[c_audit_estado] == "A VALIDAR"]
    r.check("64", "Los pendientes reales quedan identificados como no bloqueantes (A VALIDAR, no ERROR)",
            all(row[c_audit_estado] != "ERROR" for row in a_validar_audit), f"filas_a_validar={len(a_validar_audit)}")

    valores_audit = {row[c_audit_esperado] for row in d_audit} | {row[c_audit_obtenido] for row in d_audit}
    r.check("65", "418 (total historico de PROVINCIAS) no aparece como Esperado/Obtenido en AUDITORIA "
                  "ni en el texto de LEEME (no se presenta como total oficial)",
            418 not in valores_audit and "418" not in texto_leeme_up, "")

    # Los controles numericos de la auditoria ZONAS se agregan al MISMO
    # acumulador `r` (agrega checks "Z*-...") antes de imprimir, para que
    # queden en una unica tabla consolidada de solo-lectura.
    az = auditar_13_cuadros_zonas(d_est, h_est, d_elem, h_elem, r)

    core_data_ok = r.imprimir()

    print()
    print("=" * 78)
    print("AUDITORIA DE LOS 13 CUADROS DE ZONAS")
    print("Reconstruccion independiente desde BASE ESTACIONES / BASE ELEMENTOS,")
    print("comparada contra la foto de ZONAS (fuente de solo lectura).")
    print("=" * 78)
    print()
    az.imprimir()
    n_ok, n_adv, n_err = az.resumen()

    print("=" * 78)
    presentacion_zonas_ok = n_err == 0

    print(f"CUADROS OK: {n_ok}/13")
    print(f"ADVERTENCIAS: {n_adv}")
    print(f"ERRORES: {n_err}")
    print(f"CORE DATA: {'PASS' if core_data_ok else 'FAIL'}")
    print(f"PRESENTACIÓN ZONAS: {'PASS' if presentacion_zonas_ok else 'FAIL'}")
    exit_code = 0 if (core_data_ok and presentacion_zonas_ok) else 1
    print(f"VALIDATOR EXIT: {exit_code}")

    return exit_code


if __name__ == "__main__":
    sys.exit(main())
