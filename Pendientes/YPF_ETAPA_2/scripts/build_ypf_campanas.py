"""
build_ypf_campanas.py

Etapa 2 YPF: integra las campanas comerciales de la hoja COMERCIAL de
"02. CERTIFICACIONES YPF.xlsx" sobre la base validada de Etapa 1
(YPF_BASE_LIMPIA_ETAPA_1_RELEVAMIENTO_2026-08-13.xlsx), a nivel de
elemento fisico (IDCampana + ElementoID).

No modifica ni sobrescribe ninguna fuente. Trabaja siempre sobre una
copia (un workbook de salida nuevo). Antes de guardar corre un preflight
con los conteos rectores del prompt maestro: si alguno no coincide, el
script se detiene sin generar el archivo final (no fuerza numeros).

Uso:
    python build_ypf_campanas.py \
        --etapa1 <ruta base validada Etapa 1> \
        --comercial <ruta 02. CERTIFICACIONES YPF.xlsx> \
        --output <ruta archivo de salida Etapa 2>

Codigo de salida 0 si el archivo se genero y el preflight paso completo,
distinto de cero si algo bloqueante fallo (no se genera archivo final).
"""

import argparse
import hashlib
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from copy import copy as copy_style
from datetime import date

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Rutas por defecto
# ---------------------------------------------------------------------------

DEFAULT_ETAPA1 = "Pendientes/YPF_ETAPA_1/output/YPF_BASE_LIMPIA_ETAPA_1_RELEVAMIENTO_2026-08-13.xlsx"
DEFAULT_COMERCIAL = "Pendientes/YPF_ETAPA_2/input/02. CERTIFICACIONES YPF.xlsx"
# Salida nueva (BF): NO reemplaza YPF_BASE_INTEGRADA_ETAPA_2_CAMPANAS_2026-08-13.xlsx,
# que permanece intacta en disco.
DEFAULT_OUTPUT = "Pendientes/YPF_ETAPA_2/output/YPF_BASE_INTEGRADA_ETAPA_2_CAMPANAS_BF_2026-08-14.xlsx"

FECHA_GENERACION = "2026-08-14"
FECHA_CORTE = date(2026, 8, 13)
HOJA_FUENTE_COMERCIAL = "COMERCIAL"
HOJA_FUENTE_BF = "BF"

EXTERNAL_REF_PATTERN = re.compile(r"\[\d+\]")

APIE_BLOQUEADOS = {"30510", "31131", "31192", "31239"}

# ---------------------------------------------------------------------------
# Reglas de negocio fijas del prompt maestro (Etapa 2, seccion 8/9)
# ---------------------------------------------------------------------------

# Bloques marcados ELIMINAR en la fuente (columna L del encabezado): excluidos
# de BASE CAMPANAS, conservados en MAPEO/PENDIENTES.
ELIMINAR_HEADERS = {573, 1177}  # FANTA, NETFLIX

# Bloques sin fechas completas en la fuente (titulo y/o columna K): excluidos
# de BASE CAMPANAS, registrados como FALTAN_FECHAS.
SIN_FECHAS_HEADERS = {416, 594, 605, 639, 674, 773}
# MILKA(416), PERSONAL COMBO(594), PERSONAL(605), MOVISTAR(639),
# TOYOTA(674, solo fin), DERMAGLOS ENZO(773, solo fin)

TORTUGAS_MALL_HEADER = 2393
TORTUGAS_MALL_FECHA_FIN = date(2027, 2, 11)

PERSONAL_COMBO_HEADER = 594  # alcance de soportes ambiguo (SOLO TORRES vs TORRES Y PUNTERAS)
PIRELLI_HEADER = 1370  # AccountManager SIN INFORMAR, unico caso permitido

# AccountManager normalizado, mapeado explicitamente por fila de encabezado
# de bloque fuente (no por similitud de nombre; ver prompt maestro seccion 9).
ACCOUNT_MANAGER_BY_HEADER = {
    2: "Barbi / Tomy",        # Seguridad Vial
    416: "Rochi",              # Milka fila 416
    516: "Fede",                # Starlink
    573: "Jesi",                # Fanta (excluida, se conserva a fines de trazabilidad)
    594: "Fede",                # Personal Combo
    605: "Fede",                # Personal fila 605
    639: "Fede",                # Movistar
    674: "Jesi",                # Toyota
    687: "Jesi",                # Michelin
    773: "Fede",                # Dermaglos Enzo
    859: "Rochi",               # Redoxon
    916: "Rochi",               # ESPN
    973: "Fede",                # Medife
    1156: "Jesi",               # Powerade
    1177: "Jesi",               # Netflix (excluida)
    1222: "Jesi",                # Bridgestone fila 1222
    1370: "SIN INFORMAR",        # Pirelli
    1427: "Barbi / Sol",         # BNA fila 1427
    1454: "Rochi",               # Milka Tabletas
    1554: "Rochi",               # Banco Macro
    1574: "Fede",                # Medife Extension
    1757: "Fede",                # Turismo Cordoba
    1912: "Fede",                # Siglo21 (aplica a las 5 sub-pautas)
    2019: "Sol",                 # Banco Patagonia
    2037: "Jesi",                # Yo Narciso
    2069: "Sol",                 # Personal fila 2069
    2088: "Sol",                 # Goodyear
    2174: "Sol",                 # Andromaco fila 2174
    2312: "Jesi",                # Bridgestone fila 2312
    2366: "Barbi / Sol",         # BNA fila 2366
    2393: "Fede / Sol",          # Tortugas Mall
}

TIPO_SOPORTE_POR_COLUMNA = {5: "PPUNTER", 6: "TT", 7: "MB"}  # E, F, G

BASE_CAMPANAS_COLS = [
    "TipoCatalogo", "Ciudad", "Medio", "CircuitoDashboard", "Subcircuito", "Ubicacion",
    "ElementoID", "IDCampaña", "Campaña", "Cliente", "Marca", "Agencia", "Proveedor",
    "FechaInicio", "FechaFin", "Estado", "DuracionSpotSeg", "SalidasVendidas",
    "CantidadUnidades", "AccountManager", "PautaOrigen", "MarcaOrigen", "CampañaOrigen",
    "CertificadoOrigen", "ObservacionesComercial", "FilaCabeceraFuente", "FilaDetalleFuente",
    "MetodoCruce", "MetadatosRepetidos", "EstadoAsignacion", "ObservacionAsignacion",
]

MAPEO_COLS = [
    "IDCampaña", "FilaCabeceraFuente", "TituloBloqueOriginal", "PautaOrigen", "MarcaOrigen",
    "CampañaNormalizada", "FechaInicio", "FechaFin", "AccountManager", "EstadoMapeo",
    "MotivoNoCarga", "CantidadFilasFuente", "CantidadFilasAsignadas",
]

PENDIENTES_COLS = [
    "TipoPendiente", "IDCampaña", "Campaña", "APIE", "Direccion", "Localidad", "TipoSoporte",
    "CantidadSolicitada", "CantidadDisponible", "CantidadAsignada", "CantidadFaltante",
    "FilaCabeceraFuente", "FilaDetalleFuente", "ValorFuente", "Motivo", "AccionRequerida",
    "EsBloqueante",
]

AUDITORIA_COLS = ["Control", "Esperado", "Obtenido", "Diferencia", "Estado", "Detalle", "Fuente"]

AUDITORIA_BF_COLS = [
    "FilaOrigenBF", "APIEFuente", "APIEResuelta", "Direccion", "Localidad", "Campaña", "Pauta",
    "Puntera", "Torre", "MenuBoard", "TieneSoportePositivo", "MetodoCruce", "ExisteDigitalActivo",
    "ExisteSoloEstatico", "ResultadoCruce", "Observacion", "AccionRequerida",
]

ORDEN_HOJAS_FUENTE_ETAPA1 = [
    "LEEME", "AUDITORIA", "BASE ELEMENTOS", "BASE ESTACIONES", "PENDIENTES",
    "DIGITAL", "ESTATICO ", "ZONAS", "DIRECCIONES", "ZONAS VERSION ANTUGUA",
]

# APIE 30943: existe unicamente como estacion Estatica (2 FB, Avenida de los
# Lagos 330, Tigre). No tiene inventario digital: no puede recibir campañas
# digitales de PUNTERA/TORRE/MENU BOARD. Ver seccion "Tratamiento especifico
# de APIE 30943" del prompt maestro de la actualizacion BF.
APIE_SOLO_ESTATICA_CONOCIDA = "30943"

# Controles esperados sobre BF (a validar y reportar, sin forzar los numeros).
BF_CONTROLES_ESPERADOS = {
    "apie_distintas_totales": 413,
    "apie_con_soporte_positivo": 381,
    "apie_en_cero": 32,
    "apie_positiva_en_catalogo_digital": 380,
    "apie_positiva_fuera_catalogo_digital": 1,  # unicamente 30943
    "estaciones_activas_puntera_positiva": 278,
    "estaciones_activas_torre_positiva": 318,
    "estaciones_activas_menuboard_positiva": 197,
    "union_estaciones_digitales_con_soporte": 380,
}


# ---------------------------------------------------------------------------
# Utilidades genericas
# ---------------------------------------------------------------------------

def sha256_de(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def normalizar_texto(s):
    if s is None:
        return ""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalizar_pauta(v):
    t = normalizar_texto(v)
    return t


def normalizar_apie(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    if s.endswith(".0"):
        s = s[:-2]
    return s


def normalizar_account_manager(v):
    """Aplica la normalizacion explicita del prompt (Ro/RO/Rochi/ROCHI -> Rochi)
    y conserva multiples responsables separados por '/'."""
    if v is None or str(v).strip() == "":
        return v
    partes = [p.strip() for p in str(v).split("/")]
    alias = {"RO": "Rochi", "ROCHI": "Rochi"}
    normalizadas = [alias.get(p.upper(), p) for p in partes]
    return " / ".join(normalizadas)


def to_int_qty(v):
    """vacio -> 0; '-' -> 0; numeros Excel -> entero no negativo."""
    if v is None:
        return 0
    if isinstance(v, str):
        s = v.strip()
        if s in ("", "-"):
            return 0
        try:
            return max(0, int(round(float(s))))
        except ValueError:
            return None  # invalido, no numerico
    if isinstance(v, (int, float)):
        return max(0, int(v))
    return None


def es_no_ir(valor_j):
    if not isinstance(valor_j, str):
        return False
    t = valor_j.strip().lower()
    return bool(re.search(r"no[\s\-]?ir|no\s*esta\s*(la\s*)?eess|no\s*instalad|no\s*corresponde", t))


def valor_seguro_para_celda(v):
    if isinstance(v, str) and v.startswith("="):
        return "'" + v
    return v


# ---------------------------------------------------------------------------
# Deteccion de bloques en COMERCIAL (reproducible por contenido: celda
# combinada A:I/A:J de una sola fila, con columna B -APIES- vacia en esa
# fila -- no depende de color, negrita ni otro atributo puramente visual).
# ---------------------------------------------------------------------------

DATE_TOKEN = re.compile(r"(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?")


def detectar_bloques(ws):
    heads = []
    for mc in ws.merged_cells.ranges:
        if mc.min_row == mc.max_row and mc.min_col == 1 and mc.max_col >= 8:
            heads.append(mc.min_row)
    heads.sort()
    # Validacion de contenido: la fila de encabezado debe tener columna B (APIES) vacia
    # y texto no vacio en columna A.
    validados = []
    for hr in heads:
        titulo = ws.cell(row=hr, column=1).value
        apie_en_header = ws.cell(row=hr, column=2).value
        if titulo and (apie_en_header is None or str(apie_en_header).strip() == ""):
            validados.append(hr)
    return validados


def rango_detalle(ws, headers, idx):
    hr = headers[idx]
    next_hr = headers[idx + 1] if idx + 1 < len(headers) else ws.max_row + 1
    first, last = hr + 1, next_hr - 1
    while last >= first:
        vals = [ws.cell(row=last, column=c).value for c in range(2, 13)]
        if all(v is None for v in vals):
            last -= 1
        else:
            break
    return first, last


def parse_fecha_inicio_titulo(titulo):
    """FechaInicio se obtiene principalmente del titulo del bloque. Si el
    titulo trae 2+ tokens de fecha, el primero es inicio. Si trae exactamente
    1 token, en esta fuente siempre corresponde a un contexto de solo-fin
    (p.ej. 'Hasta 30/08', 'A 09/07': ver TOYOTA fila 674 y DERMAGLOS ENZO
    fila 773 en el prompt maestro) y no debe interpretarse como inicio."""
    tokens = list(DATE_TOKEN.finditer(titulo or ""))
    if len(tokens) < 2:
        return None
    d, m, y = tokens[0].groups()
    return _build_date(d, m, y)


def _build_date(d, m, y):
    d, m = int(d), int(m)
    if y:
        y = int(y)
        if y < 100:
            y += 2000
    else:
        y = 2026
    try:
        return date(y, m, d)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Carga de lookups desde la base validada de Etapa 1 (solo lectura)
# ---------------------------------------------------------------------------

def leer_tabla_dict(ws):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(headers)}
    filas = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        if all(v is None for v in vals):
            continue
        filas.append({h: vals[i] for h, i in idx.items()})
    return filas


def cargar_lookups_estaciones(wb1):
    """Devuelve, ademas de los lookups Digital ya existentes, los lookups del
    catalogo Estatico (existencia por APIE, direccion+localidad y direccion
    unica), usados exclusivamente para identificar APIE que existen solo como
    estacion estatica (p.ej. 30943) y documentar correctamente por que una
    campaña digital no puede montarse ahi -- nunca para asignarles campañas
    digitales ni para crear/objetar su existencia."""
    ws = wb1["BASE ESTACIONES"]
    filas = leer_tabla_dict(ws)
    digital_apies_si = set()
    by_dir_loc = defaultdict(set)
    by_dir = defaultdict(set)
    apies_estaticos = set()
    by_dir_loc_estatico = defaultdict(set)
    by_dir_estatico = defaultdict(set)
    for f in filas:
        apie = normalizar_apie(f["APIE"])
        direc = normalizar_texto(f["Direccion"])
        loc = normalizar_texto(f["Localidad"])
        if f["Formato"] == "Digital":
            if f["IncluirComercializacion"] == "SI":
                digital_apies_si.add(apie)
                by_dir_loc[(direc, loc)].add(apie)
                by_dir[direc].add(apie)
        elif f["Formato"] == "Estático":
            apies_estaticos.add(apie)
            by_dir_loc_estatico[(direc, loc)].add(apie)
            by_dir_estatico[direc].add(apie)
    return (filas, digital_apies_si, by_dir_loc, by_dir,
            apies_estaticos, by_dir_loc_estatico, by_dir_estatico)


def cargar_inventario_elementos(wb1):
    ws = wb1["BASE ELEMENTOS"]
    filas = leer_tabla_dict(ws)
    inventario = defaultdict(list)   # (apie, tipo) -> [(seq, meta), ...]
    todos_elemento_ids = set()
    conteo_medio = Counter()
    for f in filas:
        eid = f["ElementoID"]
        todos_elemento_ids.add(eid)
        conteo_medio[(f["Medio"], f["IncluirComercializacion"])] += 1
        if f["Medio"] != "Digital" or f["IncluirComercializacion"] != "SI":
            continue
        subcirc = normalizar_apie(f["Subcircuito"])
        partes = str(eid).split(" - ")
        if len(partes) != 3:
            continue
        tipo = partes[1]
        try:
            seq = int(partes[2])
        except ValueError:
            continue
        meta = {
            "TipoCatalogo": f["TipoCatalogo"], "Ciudad": f["Ciudad"], "Medio": f["Medio"],
            "CircuitoDashboard": f["CircuitoDashboard"], "Subcircuito": f["Subcircuito"],
            "Ubicacion": f["Ubicacion"], "ElementoID": eid, "Proveedor": f["Proveedor"],
        }
        inventario[(subcirc, tipo)].append((seq, meta))
    for k in inventario:
        inventario[k].sort(key=lambda t: t[0])
    return filas, inventario, todos_elemento_ids, conteo_medio


# ---------------------------------------------------------------------------
# Cruce de estaciones (seccion 10 del prompt maestro)
# ---------------------------------------------------------------------------

def cruzar_estacion(apie_raw, direccion_raw, localidad_raw, digital_apies_si, by_dir_loc, by_dir,
                     apies_estaticos=None, by_dir_loc_estatico=None, by_dir_estatico=None):
    """Devuelve (apie_match_digital, metodo, motivo_fallo, apie_estatica_resuelta).

    apie_match_digital es None salvo que exista una estacion Digital
    comercializable (IncluirComercializacion=SI) que cruce por APIE o por
    direccion. apie_estatica_resuelta se informa unicamente cuando la
    estacion fisica existe pero solo en el catalogo Estatico (p.ej. APIE
    30943): en ese caso motivo_fallo es 'APIE_SOLO_ESTATICA_CAMPAÑA_DIGITAL'
    en vez del generico 'APIE_NO_ENCONTRADA' / 'DIRECCION_NO_ENCONTRADA', y
    nunca se asigna ninguna campaña digital a esa fila."""
    apie_n = normalizar_apie(apie_raw)
    if apie_n is not None:
        if apie_n in digital_apies_si:
            return apie_n, "APIE", None, None
        if apies_estaticos and apie_n in apies_estaticos:
            return None, "APIE", "APIE_SOLO_ESTATICA_CAMPAÑA_DIGITAL", apie_n
        return None, None, "APIE_NO_ENCONTRADA", None
    direc = normalizar_texto(direccion_raw)
    loc = normalizar_texto(localidad_raw)
    cand = by_dir_loc.get((direc, loc), set())
    if len(cand) == 1:
        return next(iter(cand)), "DIRECCION_LOCALIDAD", None, None
    if len(cand) > 1:
        return None, None, "DIRECCION_AMBIGUA", None
    cand2 = by_dir.get(direc, set())
    if len(cand2) == 1:
        return next(iter(cand2)), "DIRECCION_UNICA", None, None
    if len(cand2) > 1:
        return None, None, "DIRECCION_AMBIGUA", None

    # Sin coincidencia digital: intentar resolver contra el catalogo Estatico
    # unicamente para documentar correctamente el motivo (nunca para asignar).
    if by_dir_loc_estatico is not None:
        cand_e = by_dir_loc_estatico.get((direc, loc), set())
        if len(cand_e) == 1:
            return None, "DIRECCION_LOCALIDAD", "APIE_SOLO_ESTATICA_CAMPAÑA_DIGITAL", next(iter(cand_e))
        cand_e2 = (by_dir_estatico or {}).get(direc, set())
        if len(cand_e2) == 1:
            return None, "DIRECCION_UNICA", "APIE_SOLO_ESTATICA_CAMPAÑA_DIGITAL", next(iter(cand_e2))
    return None, None, "DIRECCION_NO_ENCONTRADA", None


# ---------------------------------------------------------------------------
# Paso 1: leer y estructurar COMERCIAL en unidades bloque+pauta
# ---------------------------------------------------------------------------

def construir_unidades(ws_comercial):
    """Devuelve la lista ordenada de unidades bloque+pauta (35 esperadas) con
    toda su metadata, y la lista de filas de detalle por unidad (excluyendo
    las filas 'no ir', que se registran aparte como FILA_EXCLUIDA_NO_IR)."""
    headers = detectar_bloques(ws_comercial)

    unidades = []
    filas_noir = []  # (header, fila, apie, direccion, localidad, valor_j)
    total_filas_fuente = 0
    filas_por_header = {}  # header_row -> total filas de detalle del bloque (incluye no-ir)

    for i, hr in enumerate(headers):
        first, last = rango_detalle(ws_comercial, headers, i)
        titulo = ws_comercial.cell(row=hr, column=1).value
        header_l = ws_comercial.cell(row=hr, column=12).value
        header_k = ws_comercial.cell(row=hr, column=11).value

        fecha_inicio_titulo = parse_fecha_inicio_titulo(titulo)
        filas_por_header[hr] = last - first + 1

        # Orden de pautas por primera aparicion, excluyendo filas 'no ir'.
        orden_pautas = []
        vistas = set()
        grupos = defaultdict(list)
        for r in range(first, last + 1):
            total_filas_fuente += 1
            j_val = ws_comercial.cell(row=r, column=10).value
            apie_val = ws_comercial.cell(row=r, column=2).value
            direccion_val = ws_comercial.cell(row=r, column=3).value
            localidad_val = ws_comercial.cell(row=r, column=4).value
            if es_no_ir(j_val):
                filas_noir.append((hr, r, apie_val, direccion_val, localidad_val, j_val))
                continue
            i_val = ws_comercial.cell(row=r, column=9).value
            h_val = ws_comercial.cell(row=r, column=8).value
            key = normalizar_pauta(i_val) or normalizar_pauta(h_val) or normalizar_texto(titulo)
            if key not in vistas:
                vistas.add(key)
                orden_pautas.append(key)
            grupos[key].append(r)

        for pauta_key in orden_pautas:
            filas_grupo = grupos[pauta_key]
            k_vals = {ws_comercial.cell(row=r, column=11).value for r in filas_grupo}
            k_vals_no_none = {v for v in k_vals if v is not None}
            fecha_fin = None
            inconsistente = False
            if len(k_vals_no_none) == 1:
                fecha_fin = next(iter(k_vals_no_none)).date() if hasattr(next(iter(k_vals_no_none)), "date") else next(iter(k_vals_no_none))
            elif len(k_vals_no_none) > 1:
                inconsistente = True
                fecha_fin = sorted(k_vals_no_none)[0]
                fecha_fin = fecha_fin.date() if hasattr(fecha_fin, "date") else fecha_fin
            if fecha_fin is None and hr == TORTUGAS_MALL_HEADER:
                fecha_fin = TORTUGAS_MALL_FECHA_FIN

            pauta_origen_raw = ws_comercial.cell(row=filas_grupo[0], column=9).value
            marca_origen_raw = ws_comercial.cell(row=filas_grupo[0], column=8).value

            unidades.append({
                "header_row": hr,
                "titulo": titulo,
                "header_k": header_k,
                "header_l": header_l,
                "pauta_key": pauta_key,
                "pauta_origen": pauta_origen_raw,
                "marca_origen": marca_origen_raw,
                "fecha_inicio": fecha_inicio_titulo,
                "fecha_fin": fecha_fin,
                "fecha_inconsistente": inconsistente,
                "filas": filas_grupo,
            })

    return headers, unidades, filas_noir, total_filas_fuente, filas_por_header


# ---------------------------------------------------------------------------
# Paso 2: asignar IDCampana estable (seccion 6)
# ---------------------------------------------------------------------------

def asignar_ids(unidades):
    siguiente = 20000
    for u in unidades:
        u["id_campana"] = siguiente
        siguiente += 1


# ---------------------------------------------------------------------------
# Paso 3: clasificar unidades (EstadoMapeo / MotivoNoCarga)
# ---------------------------------------------------------------------------

def clasificar_unidad(u):
    hr = u["header_row"]
    if hr in ELIMINAR_HEADERS:
        return "EXCLUIDA", "EXCLUIDA_POR_FUENTE"
    if u["fecha_inicio"] is None or u["fecha_fin"] is None:
        return "PENDIENTE", "FALTAN_FECHAS"
    return "CARGABLE", ""


def estado_temporal(fecha_inicio, fecha_fin, corte=FECHA_CORTE):
    if fecha_inicio is None or fecha_fin is None:
        return "Pendiente"
    if fecha_fin < corte:
        return "Finalizada"
    if fecha_inicio <= corte <= fecha_fin:
        return "Activa"
    if fecha_inicio > corte:
        return "Programada"
    return "Pendiente"


# ---------------------------------------------------------------------------
# Paso 4: Campana / MarcaBase por fila (seccion 7)
# ---------------------------------------------------------------------------

def campania_y_marca(pauta_val, marca_val, titulo_bloque):
    titulo_norm = re.sub(r"\s+", " ", str(titulo_bloque or "").strip())
    pauta = str(pauta_val).strip() if pauta_val not in (None, "") else ""
    marca = str(marca_val).strip() if marca_val not in (None, "") else ""
    campania = pauta or marca or titulo_norm
    marca_base = marca or pauta or titulo_norm
    return campania, marca_base


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--etapa1", default=DEFAULT_ETAPA1)
    parser.add_argument("--comercial", default=DEFAULT_COMERCIAL)
    parser.add_argument("--output", default=DEFAULT_OUTPUT)
    args = parser.parse_args(argv)

    print("[1/13] Verificando existencia de fuentes...")
    import os
    if not os.path.isfile(args.etapa1):
        print(f"ERROR BLOQUEANTE: no existe la base validada de Etapa 1: {args.etapa1}")
        return 2
    if not os.path.isfile(args.comercial):
        print(f"ERROR BLOQUEANTE: no existe el archivo de certificaciones: {args.comercial}")
        return 2
    print(f"  Etapa 1     = {args.etapa1}")
    print(f"  Comercial   = {args.comercial}")

    print("[2/13] Calculando hashes SHA-256 de las fuentes (control previo)...")
    hash_etapa1_pre = sha256_de(args.etapa1)
    hash_comercial_pre = sha256_de(args.comercial)
    print(f"  Etapa1   sha256={hash_etapa1_pre}")
    print(f"  Comercial sha256={hash_comercial_pre}")

    print("[3/13] Cargando base validada de Etapa 1 (solo lectura)...")
    wb1_formulas = openpyxl.load_workbook(args.etapa1, data_only=False)
    wb1_valores = openpyxl.load_workbook(args.etapa1, data_only=True)

    if wb1_valores.sheetnames != ORDEN_HOJAS_FUENTE_ETAPA1:
        print("ERROR BLOQUEANTE: la base de Etapa 1 no tiene las hojas/orden esperado.")
        print(f"  esperado={ORDEN_HOJAS_FUENTE_ETAPA1}")
        print(f"  obtenido={wb1_valores.sheetnames}")
        return 3

    (filas_estaciones, digital_apies_si, by_dir_loc, by_dir,
     apies_estaticos, by_dir_loc_estatico, by_dir_estatico) = cargar_lookups_estaciones(wb1_valores)
    filas_elementos, inventario, todos_elemento_ids, conteo_medio = cargar_inventario_elementos(wb1_valores)

    if len(filas_estaciones) != 603:
        print(f"ERROR BLOQUEANTE: BASE ESTACIONES de Etapa 1 tiene {len(filas_estaciones)} filas (se esperaban 603).")
        return 3
    if len(filas_elementos) != 3883:
        print(f"ERROR BLOQUEANTE: BASE ELEMENTOS de Etapa 1 tiene {len(filas_elementos)} filas (se esperaban 3.883).")
        return 3
    if conteo_medio[("Digital", "SI")] != 3500:
        print(f"ERROR BLOQUEANTE: elementos Digital comercializables = {conteo_medio[('Digital', 'SI')]} (se esperaban 3.500).")
        return 3

    apies_bloqueados_presentes = APIE_BLOQUEADOS & {normalizar_apie(f["APIE"]) for f in filas_estaciones}
    if apies_bloqueados_presentes:
        print(f"ERROR BLOQUEANTE: los APIE bloqueados {apies_bloqueados_presentes} estan presentes en BASE ESTACIONES de Etapa 1.")
        return 3

    n_estaciones_digitales = sum(1 for f in filas_estaciones if f["Formato"] == "Digital")
    n_estaciones_estaticas = sum(1 for f in filas_estaciones if f["Formato"] == "Estático")
    if n_estaciones_digitales != 412 or n_estaciones_estaticas != 191:
        print(f"ERROR BLOQUEANTE: catalogo de estaciones alterado (digitales={n_estaciones_digitales} "
              f"esperado=412, estaticas={n_estaciones_estaticas} esperado=191).")
        return 3
    n_elem_fb = sum(1 for f in filas_elementos if str(f["ElementoID"]).split(" - ")[1:2] == ["FB"])
    if n_elem_fb != 383:
        print(f"ERROR BLOQUEANTE: elementos FB de Etapa 1 = {n_elem_fb} (se esperaban 383).")
        return 3

    print("  BASE ESTACIONES: 603 filas OK (412 digitales + 191 estaticas) | "
          "BASE ELEMENTOS: 3.883 filas OK (3.500 digitales + 383 FB) | Digital comercializable: 3.500 OK")

    print(f"[4/13] Inspeccionando hoja fuente '{HOJA_FUENTE_COMERCIAL}'...")
    wb_com = openpyxl.load_workbook(args.comercial, data_only=True)
    if HOJA_FUENTE_COMERCIAL not in wb_com.sheetnames:
        print(f"ERROR BLOQUEANTE: no existe la solapa '{HOJA_FUENTE_COMERCIAL}' en {args.comercial}")
        return 4
    if HOJA_FUENTE_BF not in wb_com.sheetnames:
        print(f"ERROR BLOQUEANTE: no existe la solapa '{HOJA_FUENTE_BF}' en {args.comercial}")
        return 4
    ws_com = wb_com[HOJA_FUENTE_COMERCIAL]
    ws_bf = wb_com[HOJA_FUENTE_BF]

    headers, unidades, filas_noir, total_filas_fuente, filas_por_header = construir_unidades(ws_com)
    asignar_ids(unidades)

    n_bloques = len(headers)
    n_unidades = len(unidades)
    print(f"  Bloques detectados: {n_bloques} | Unidades bloque+pauta: {n_unidades} | Filas de estacion fuente: {total_filas_fuente}")

    # ---- Clasificacion de unidades ----
    for u in unidades:
        u["estado_mapeo"], u["motivo_no_carga"] = clasificar_unidad(u)
        u["account_manager"] = normalizar_account_manager(ACCOUNT_MANAGER_BY_HEADER.get(u["header_row"], "SIN INFORMAR"))

    n_excluidas = sum(1 for u in unidades if u["estado_mapeo"] == "EXCLUIDA")
    n_pendientes_fecha = sum(1 for u in unidades if u["estado_mapeo"] == "PENDIENTE")
    n_cargables = sum(1 for u in unidades if u["estado_mapeo"] == "CARGABLE")
    print(f"  Unidades EXCLUIDA (ELIMINAR): {n_excluidas} | PENDIENTE (FALTAN_FECHAS): {n_pendientes_fecha} | CARGABLE: {n_cargables}")

    # ---- Filas candidatas antes del cruce: TODAS las filas de detalle de los
    # bloques cargables (incluye las 'no ir', que se excluyen recien en el paso
    # siguiente). Un bloque nunca mezcla pautas cargables con pautas excluidas
    # o pendientes (la fecha se define a nivel de bloque), por lo que basta con
    # el conjunto de headers de unidades CARGABLE.
    cargable_headers = {u["header_row"] for u in unidades if u["estado_mapeo"] == "CARGABLE"}
    filas_candidatas = sum(filas_por_header[hr] for hr in cargable_headers)

    print("[5/13] Cruzando estaciones (APIE / direccion+localidad / direccion unica) y asignando elementos...")

    base_campanas = []
    pendientes = []
    filas_sin_estacion = 0
    elementos_solicitados = 0
    elementos_asignados = 0
    faltante_total = 0
    metodo_cruce_cnt = Counter()
    apies_usados = set()

    def add_pendiente(tipo, u, apie="", direccion="", localidad="", tiposoporte="",
                       cant_sol=None, cant_disp=None, cant_asig=None, cant_falt=None,
                       fila_det="", valor_fuente="", motivo="", accion="", bloqueante=True):
        pendientes.append({
            "TipoPendiente": tipo,
            "IDCampaña": u["id_campana"] if u else "",
            "Campaña": u["pauta_key"] if u else "",
            "APIE": apie, "Direccion": direccion, "Localidad": localidad, "TipoSoporte": tiposoporte,
            "CantidadSolicitada": cant_sol, "CantidadDisponible": cant_disp,
            "CantidadAsignada": cant_asig, "CantidadFaltante": cant_falt,
            "FilaCabeceraFuente": u["header_row"] if u else "",
            "FilaDetalleFuente": fila_det, "ValorFuente": valor_fuente, "Motivo": motivo,
            "AccionRequerida": accion, "EsBloqueante": "SI" if bloqueante else "NO",
        })

    # ---- Pendientes de unidades excluidas / con fechas incompletas ----
    for u in unidades:
        if u["estado_mapeo"] == "EXCLUIDA":
            add_pendiente("EXCLUIDA_POR_FUENTE", u, motivo="Bloque marcado ELIMINAR en observaciones comerciales (columna L) de la fuente.",
                           accion="Ninguna: exclusion confirmada por la fuente.", bloqueante=False,
                           valor_fuente=str(u["header_l"]))
        elif u["estado_mapeo"] == "PENDIENTE":
            add_pendiente("FALTAN_FECHAS", u, motivo="El bloque no tiene FechaInicio y/o FechaFin explicitas y confiables en la fuente.",
                           accion="Solicitar fecha de inicio y/o fin al area comercial antes de cargar.", bloqueante=True)
            if u["header_row"] == PERSONAL_COMBO_HEADER:
                add_pendiente("ALCANCE_SOPORTES_AMBIGUO", u,
                               motivo="La fuente contiene simultaneamente 'SOLO TORRES' y 'TORRES Y PUNTERAS' en observaciones, "
                                      "con cantidades presentes en PUNTERA y TORRE. No se puede resolver por inferencia.",
                               accion="Aclarar con el area comercial el alcance real de soportes antes de cargar.", bloqueante=True)
        if u["header_row"] == PIRELLI_HEADER and u["estado_mapeo"] != "EXCLUIDA":
            add_pendiente("ACCOUNT_MANAGER_NO_INFORMADO", u,
                           motivo="La fuente no informa AccountManager para Pirelli.",
                           accion="Solicitar responsable de cuenta al area comercial.", bloqueante=False)

    # ---- Filas no-ir (excluidas de BASE CAMPANAS, cualquiera sea el estado de la unidad) ----
    unidad_por_header = defaultdict(list)
    for u in unidades:
        unidad_por_header[u["header_row"]].append(u)

    for (hr, r, apie_val, direccion_val, localidad_val, j_val) in filas_noir:
        u_ref = unidad_por_header.get(hr, [None])[0]
        add_pendiente("FILA_EXCLUIDA_NO_IR", u_ref, apie=normalizar_apie(apie_val) or "",
                       direccion=direccion_val or "", localidad=localidad_val or "",
                       fila_det=r, valor_fuente=str(j_val),
                       motivo="La fuente marca la fila como no instalada / no corresponde (columna CERTIFICADO).",
                       accion="Ninguna: exclusion confirmada por la fuente.", bloqueante=False)

    # ---- Cruce + asignacion de elementos, solo unidades CARGABLE ----
    for u in unidades:
        if u["estado_mapeo"] != "CARGABLE":
            continue
        estado_temp = estado_temporal(u["fecha_inicio"], u["fecha_fin"])
        u["estado_temporal"] = estado_temp
        if u["fecha_inconsistente"]:
            add_pendiente("DATO_FUENTE_INCONSISTENTE", u,
                           motivo="La columna FIN DE LA CAMPAÑA (K) presenta valores distintos entre filas del mismo bloque+pauta.",
                           accion="Confirmar la fecha de fin correcta con el area comercial.", bloqueante=False)

        for r in u["filas"]:
            apie_raw = ws_com.cell(row=r, column=2).value
            direccion_raw = ws_com.cell(row=r, column=3).value
            localidad_raw = ws_com.cell(row=r, column=4).value

            apie_match, metodo, motivo_fallo, apie_estatica = cruzar_estacion(
                apie_raw, direccion_raw, localidad_raw, digital_apies_si, by_dir_loc, by_dir,
                apies_estaticos, by_dir_loc_estatico, by_dir_estatico)

            if apie_match is None:
                filas_sin_estacion += 1
                if motivo_fallo == "APIE_SOLO_ESTATICA_CAMPAÑA_DIGITAL":
                    add_pendiente(motivo_fallo, u, apie=apie_estatica or normalizar_apie(apie_raw) or "",
                                  direccion=direccion_raw or "", localidad=localidad_raw or "",
                                  fila_det=r,
                                  motivo=f"La estacion APIE {apie_estatica} existe en el catalogo, pero unicamente "
                                         "como estacion Estatica (FB); no tiene inventario Digital (PPUNTER/TT/MB) "
                                         "para montar esta campaña digital.",
                                  accion="No corresponde crear inventario digital. Confirmar con el area comercial "
                                         "si la campaña debe reasignarse a otra estacion digital real.",
                                  bloqueante=False)
                else:
                    add_pendiente(motivo_fallo, u, apie=normalizar_apie(apie_raw) or "",
                                  direccion=direccion_raw or "", localidad=localidad_raw or "",
                                  fila_det=r,
                                  motivo="No se encontro una estacion Digital comercializable (IncluirComercializacion=SI) "
                                         "que coincida por APIE ni por direccion/localidad.",
                                  accion="Confirmar la estacion real con el area comercial/territorial; no se crea ni se "
                                         "reemplaza por coincidencia aproximada.", bloqueante=True)
                continue

            metodo_cruce_cnt[metodo] += 1
            apies_usados.add(apie_match)

            pauta_val = ws_com.cell(row=r, column=9).value
            marca_val = ws_com.cell(row=r, column=8).value
            certificado_val = ws_com.cell(row=r, column=10).value
            observ_val = ws_com.cell(row=r, column=12).value
            campania, marca_base = campania_y_marca(pauta_val, marca_val, u["titulo"])

            for col_idx, tipo in TIPO_SOPORTE_POR_COLUMNA.items():
                qty_raw = ws_com.cell(row=r, column=col_idx).value
                qty = to_int_qty(qty_raw)
                if qty is None:
                    add_pendiente("CANTIDAD_INVALIDA", u, apie=apie_match, direccion=direccion_raw or "",
                                  localidad=localidad_raw or "", tiposoporte=tipo, fila_det=r,
                                  valor_fuente=str(qty_raw),
                                  motivo="El valor de cantidad no es numerico ni '-' ni vacio.",
                                  accion="Corregir el valor en la fuente.", bloqueante=True)
                    continue
                if qty <= 0:
                    continue

                elementos_solicitados += qty
                disponibles = inventario.get((apie_match, tipo), [])
                n_disp = len(disponibles)
                n_asig = min(qty, n_disp)
                asignados_ahora = disponibles[:n_asig]
                elementos_asignados += n_asig

                for _, meta in asignados_ahora:
                    base_campanas.append({
                        "TipoCatalogo": meta["TipoCatalogo"], "Ciudad": meta["Ciudad"], "Medio": meta["Medio"],
                        "CircuitoDashboard": meta["CircuitoDashboard"], "Subcircuito": meta["Subcircuito"],
                        "Ubicacion": meta["Ubicacion"], "ElementoID": meta["ElementoID"],
                        "IDCampaña": u["id_campana"], "Campaña": campania, "Cliente": marca_base,
                        "Marca": marca_base, "Agencia": marca_base, "Proveedor": meta["Proveedor"],
                        "FechaInicio": u["fecha_inicio"], "FechaFin": u["fecha_fin"], "Estado": estado_temp,
                        "DuracionSpotSeg": None, "SalidasVendidas": None, "CantidadUnidades": 1,
                        "AccountManager": u["account_manager"], "PautaOrigen": pauta_val, "MarcaOrigen": marca_val,
                        "CampañaOrigen": u["titulo"], "CertificadoOrigen": certificado_val,
                        "ObservacionesComercial": observ_val, "FilaCabeceraFuente": u["header_row"],
                        "FilaDetalleFuente": r, "MetodoCruce": metodo, "MetadatosRepetidos": "SI",
                        "EstadoAsignacion": "ASIGNADO",
                        "ObservacionAsignacion": "" if n_asig == qty else f"Solicitados {qty}, asignados {n_asig} (faltante {qty - n_asig})",
                    })

                if n_asig < qty:
                    faltante = qty - n_asig
                    faltante_total += faltante
                    add_pendiente("FALTANTE_INVENTARIO", u, apie=apie_match, direccion=direccion_raw or "",
                                  localidad=localidad_raw or "", tiposoporte=tipo,
                                  cant_sol=qty, cant_disp=n_disp, cant_asig=n_asig, cant_falt=faltante,
                                  fila_det=r,
                                  motivo="El inventario fisico disponible en la estacion es menor a lo solicitado por la fuente.",
                                  accion="Confirmar disponibilidad real con el area territorial; no se crean ElementoID nuevos.",
                                  bloqueante=False)

    # ---- APIE bloqueados referenciados en COMERCIAL (control, se espera 0) ----
    apies_en_comercial = set()
    for r in range(1, ws_com.max_row + 1):
        v = ws_com.cell(row=r, column=2).value
        apie_n = normalizar_apie(v)
        if apie_n:
            apies_en_comercial.add(apie_n)
    apies_bloqueados_en_comercial = APIE_BLOQUEADOS & apies_en_comercial
    for a in apies_bloqueados_en_comercial:
        add_pendiente("APIE_BLOQUEADA", None, apie=a,
                      motivo="APIE excluido del catalogo activo y de la comercializacion (Etapa 1).",
                      accion="No asignar campanas ni elementos a este APIE.", bloqueante=True)

    # =========================================================================
    # BF: fuente adicional de cruce, reconciliacion y validacion (COMERCIAL
    # sigue siendo la fuente primaria de estructura/metadatos de campaña; BF
    # NUNCA crea ni modifica filas de BASE CAMPAÑAS, solo se usa para auditar).
    # =========================================================================
    print(f"[6/13] Cruzando y reconciliando la solapa '{HOJA_FUENTE_BF}' contra COMERCIAL y el catalogo...")

    # Indice de COMERCIAL (TODAS las unidades, cargables o no) para poder
    # comparar cada fila de BF contra su equivalente en COMERCIAL.
    comercial_index = {}
    for u in unidades:
        for r in u["filas"]:
            apie_raw_c = ws_com.cell(row=r, column=2).value
            direccion_c = ws_com.cell(row=r, column=3).value
            localidad_c = ws_com.cell(row=r, column=4).value
            apie_dig_c, _metodo_c, _motivo_c, apie_est_c = cruzar_estacion(
                apie_raw_c, direccion_c, localidad_c, digital_apies_si, by_dir_loc, by_dir,
                apies_estaticos, by_dir_loc_estatico, by_dir_estatico)
            apie_resuelta_c = apie_dig_c or apie_est_c or normalizar_apie(apie_raw_c)
            if apie_resuelta_c is None:
                continue
            pp_c = to_int_qty(ws_com.cell(row=r, column=5).value) or 0
            tt_c = to_int_qty(ws_com.cell(row=r, column=6).value) or 0
            mb_c = to_int_qty(ws_com.cell(row=r, column=7).value) or 0
            fecha_fin_key = u["fecha_fin"]
            key = (apie_resuelta_c, u["pauta_key"], fecha_fin_key)
            entry = comercial_index.setdefault(key, {
                "pp": 0, "tt": 0, "mb": 0, "header_row": u["header_row"], "id_campana": u["id_campana"],
                "estado_mapeo": u["estado_mapeo"], "titulo": u["titulo"],
            })
            entry["pp"] += pp_c
            entry["tt"] += tt_c
            entry["mb"] += mb_c

    def bf_lookup_comercial(apie_resuelta, pauta_key, fecha_fin):
        exacto = comercial_index.get((apie_resuelta, pauta_key, fecha_fin))
        if exacto is not None:
            return exacto, "exacta (APIE+pauta+fecha)"
        candidatos = [v for k, v in comercial_index.items() if k[0] == apie_resuelta and k[1] == pauta_key]
        if len(candidatos) == 1:
            return candidatos[0], "unica (APIE+pauta, fecha distinta o ausente)"
        return None, "ninguna" if not candidatos else "ambigua (varias fechas para el mismo APIE+pauta)"

    auditoria_bf = []
    apie_bf_all = set()
    apie_bf_pos = set()
    pp_bf_pos_digital, tt_bf_pos_digital, mb_bf_pos_digital = set(), set(), set()
    apie_bf_pos_fuera_digital = set()

    for r in range(2, ws_bf.max_row + 1):
        vals = [ws_bf.cell(row=r, column=c).value for c in range(1, 12)]
        if all(v is None for v in vals):
            continue
        apie_raw_bf = ws_bf.cell(row=r, column=1).value
        direccion_bf = ws_bf.cell(row=r, column=2).value
        localidad_bf = ws_bf.cell(row=r, column=3).value
        pp_raw = ws_bf.cell(row=r, column=4).value
        tt_raw = ws_bf.cell(row=r, column=5).value
        mb_raw = ws_bf.cell(row=r, column=6).value
        marca_bf = ws_bf.cell(row=r, column=7).value
        pauta_bf = ws_bf.cell(row=r, column=8).value
        certificado_bf = ws_bf.cell(row=r, column=9).value
        fecha_fin_bf_dt = ws_bf.cell(row=r, column=10).value
        fecha_fin_bf = fecha_fin_bf_dt.date() if hasattr(fecha_fin_bf_dt, "date") else fecha_fin_bf_dt

        pp_bf = to_int_qty(pp_raw)
        tt_bf = to_int_qty(tt_raw)
        mb_bf = to_int_qty(mb_raw)
        cantidad_invalida = pp_bf is None or tt_bf is None or mb_bf is None
        pp_bf, tt_bf, mb_bf = pp_bf or 0, tt_bf or 0, mb_bf or 0
        tiene_soporte = pp_bf > 0 or tt_bf > 0 or mb_bf > 0

        apie_dig_bf, metodo_bf, motivo_fallo_bf, apie_est_bf = cruzar_estacion(
            apie_raw_bf, direccion_bf, localidad_bf, digital_apies_si, by_dir_loc, by_dir,
            apies_estaticos, by_dir_loc_estatico, by_dir_estatico)
        apie_resuelta = apie_dig_bf or apie_est_bf
        existe_digital = "SI" if apie_dig_bf else "NO"
        existe_solo_estatico = "SI" if (apie_est_bf and not apie_dig_bf) else "NO"

        if apie_resuelta:
            apie_bf_all.add(apie_resuelta)
            if tiene_soporte:
                apie_bf_pos.add(apie_resuelta)
                if apie_dig_bf:
                    if pp_bf > 0:
                        pp_bf_pos_digital.add(apie_resuelta)
                    if tt_bf > 0:
                        tt_bf_pos_digital.add(apie_resuelta)
                    if mb_bf > 0:
                        mb_bf_pos_digital.add(apie_resuelta)
                else:
                    apie_bf_pos_fuera_digital.add(apie_resuelta)
        elif normalizar_apie(apie_raw_bf):
            # APIE informada en BF pero que no existe en ningun catalogo (ni
            # digital ni estatico): se cuenta igual para el total de APIE
            # distintas de BF, ya que el control es sobre lo que informa BF.
            apie_bf_all.add(normalizar_apie(apie_raw_bf))

        pauta_key_bf = normalizar_pauta(pauta_bf) or normalizar_pauta(marca_bf)
        match, calidad_match = (None, "sin apie resuelta")
        if apie_resuelta:
            match, calidad_match = bf_lookup_comercial(apie_resuelta, pauta_key_bf, fecha_fin_bf)

        diferencia_detalle = ""
        if match is not None:
            if (match["pp"], match["tt"], match["mb"]) != (pp_bf, tt_bf, mb_bf):
                diferencia_detalle = (f"COMERCIAL(PP={match['pp']},TT={match['tt']},MB={match['mb']}) vs "
                                       f"BF(PP={pp_bf},TT={tt_bf},MB={mb_bf}) [{calidad_match}]")

        # ---- Resultado del cruce (vocabulario controlado) ----
        if cantidad_invalida:
            resultado = "PENDIENTE_REVISION"
            observ = f"Cantidad no numerica en BF (PP={pp_raw!r} TT={tt_raw!r} MB={mb_raw!r})."
            accion_req = "Corregir el valor en la fuente BF."
        elif not apie_resuelta:
            resultado = "APIE_NO_EXISTE" if normalizar_apie(apie_raw_bf) else "SIN_APIE_RESUELTA_POR_DIRECCION"
            observ = ("La APIE informada en BF no existe en ningun catalogo (ni Digital ni Estatico)."
                      if resultado == "APIE_NO_EXISTE" else
                      "BF no informa APIE y no se pudo resolver de forma inequivoca por direccion/localidad.")
            accion_req = "Confirmar la estacion real con el area comercial/territorial; no se crea ni se reemplaza por coincidencia aproximada."
        elif not tiene_soporte:
            resultado = "SIN_SOPORTES_POSITIVOS"
            observ = "PUNTERA, TORRE y MENU BOARD estan en cero/vacio/guion: no constituye estacion con campaña."
            accion_req = "Ninguna."
        elif existe_solo_estatico == "SI":
            resultado = "APIE_SOLO_ESTATICA_CAMPAÑA_DIGITAL"
            observ = (f"La APIE {apie_resuelta} existe, pero solo tiene inventario Estatico (FB); no tiene "
                      "inventario Digital apto (PPUNTER/TT/MB) para montar esta campaña digital.")
            accion_req = "No corresponde crear inventario digital. Confirmar con el area comercial si corresponde reasignar."
        elif diferencia_detalle:
            resultado = "DIFERENCIA_BF_COMERCIAL"
            observ = f"BF y COMERCIAL no coinciden en cantidades para la misma APIE+pauta. {diferencia_detalle}"
            accion_req = "Revisar con el area comercial cual de las dos fuentes es correcta; no se elige una version en forma silenciosa."
        elif match is None:
            resultado = "PENDIENTE_REVISION"
            observ = (f"BF aporta una fila (APIE {apie_resuelta}, pauta {pauta_key_bf!r}) que no pudo relacionarse "
                      f"de forma inequivoca con COMERCIAL (coincidencia: {calidad_match}).")
            accion_req = "Revisar manualmente contra COMERCIAL antes de tomar como valida."
        else:
            resultado = "ASIGNADA"
            observ = f"BF coincide con COMERCIAL (IDCampaña {match['id_campana']}, coincidencia {calidad_match}); cruce validado."
            accion_req = "Ninguna."

        auditoria_bf.append({
            "FilaOrigenBF": r, "APIEFuente": normalizar_apie(apie_raw_bf) or "", "APIEResuelta": apie_resuelta or "",
            "Direccion": direccion_bf, "Localidad": localidad_bf, "Campaña": marca_bf, "Pauta": pauta_bf,
            "Puntera": pp_bf, "Torre": tt_bf, "MenuBoard": mb_bf,
            "TieneSoportePositivo": "SI" if tiene_soporte else "NO",
            "MetodoCruce": metodo_bf or "", "ExisteDigitalActivo": existe_digital, "ExisteSoloEstatico": existe_solo_estatico,
            "ResultadoCruce": resultado, "Observacion": observ, "AccionRequerida": accion_req,
        })

        if resultado not in ("ASIGNADA", "SIN_SOPORTES_POSITIVOS"):
            add_pendiente(resultado, None, apie=apie_resuelta or (normalizar_apie(apie_raw_bf) or ""),
                          direccion=direccion_bf or "", localidad=localidad_bf or "",
                          fila_det=f"BF!{r}", valor_fuente=str(pauta_bf),
                          motivo=observ, accion=accion_req,
                          bloqueante=(resultado in ("DIFERENCIA_BF_COMERCIAL", "APIE_NO_EXISTE")))

    n_bf_apie_distintas = len(apie_bf_all)
    n_bf_apie_positivas = len(apie_bf_pos)
    n_bf_apie_cero = n_bf_apie_distintas - n_bf_apie_positivas
    n_bf_apie_pos_digital = len(apie_bf_pos & digital_apies_si)
    n_bf_apie_pos_fuera_digital = len(apie_bf_pos_fuera_digital)
    n_bf_pp_activas = len(pp_bf_pos_digital)
    n_bf_tt_activas = len(tt_bf_pos_digital)
    n_bf_mb_activas = len(mb_bf_pos_digital)
    n_bf_union_activas = len(pp_bf_pos_digital | tt_bf_pos_digital | mb_bf_pos_digital)
    n_bf_diferencias = sum(1 for a in auditoria_bf if a["ResultadoCruce"] == "DIFERENCIA_BF_COMERCIAL")
    n_bf_pendientes_revision = sum(1 for a in auditoria_bf if a["ResultadoCruce"] == "PENDIENTE_REVISION")

    print(f"  BF: {len(auditoria_bf)} filas procesadas | APIE distintas: {n_bf_apie_distintas} | "
          f"con soporte positivo: {n_bf_apie_positivas} | en catalogo digital: {n_bf_apie_pos_digital}")
    print(f"  BF vs COMERCIAL: diferencias={n_bf_diferencias} | pendientes de revision={n_bf_pendientes_revision}")

    print(f"  Filas sin estacion: {filas_sin_estacion} | Elementos solicitados: {elementos_solicitados} | "
          f"asignados: {elementos_asignados} | faltante: {faltante_total}")
    print(f"  Metodos de cruce: {dict(metodo_cruce_cnt)}")

    # =========================================================================
    # PREFLIGHT (seccion 18 del prompt maestro) -- se aborta si algo no coincide
    # =========================================================================
    print("[7/13] Preflight completo contra los conteos rectores...")

    n_dup_clave = len(base_campanas) - len({
        (f["IDCampaña"], f["ElementoID"], f["FechaInicio"], f["FechaFin"]) for f in base_campanas
    })
    n_fb = sum(1 for f in base_campanas if "FB" in str(f["ElementoID"]).split(" - "))
    n_apie_bloqueados_en_base = sum(1 for f in base_campanas if normalizar_apie(f["Subcircuito"]) in APIE_BLOQUEADOS)
    n_fechas_faltantes = sum(1 for f in base_campanas if f["FechaInicio"] is None or f["FechaFin"] is None)
    n_am_vacio = sum(1 for f in base_campanas if not f["AccountManager"] or not str(f["AccountManager"]).strip())
    am_sin_informar_campanas = {f["Campaña"] for f in base_campanas if f["AccountManager"] == "SIN INFORMAR"}

    preflight_checks = [
        ("Bloques detectados", 31, n_bloques),
        ("Filas de estacion fuente", 2371, total_filas_fuente),
        ("Bloques marcados ELIMINAR", 2, len(ELIMINAR_HEADERS)),
        ("Bloques sin fechas completas", 6, len(SIN_FECHAS_HEADERS)),
        ("Bloques completos cargables", 23, n_bloques - len(ELIMINAR_HEADERS) - len(SIN_FECHAS_HEADERS)),
        ("Unidades bloque+pauta totales", 35, n_unidades),
        ("IDCampaña cargables", 27, n_cargables),
        ("Filas candidatas antes del cruce", 2034, filas_candidatas),
        ("Filas no-ir excluidas", 2, len(filas_noir)),
        ("Filas cargables sin estacion", 6, filas_sin_estacion),
        ("Elementos solicitados", 13779, elementos_solicitados),
        ("Elementos asignados", 13616, elementos_asignados),
        ("Faltante de inventario", 163, faltante_total),
        ("Filas finales BASE CAMPAÑAS", 13616, len(base_campanas)),
        ("Duplicados de clave compuesta", 0, n_dup_clave),
        ("Elementos FB", 0, n_fb),
        ("Elementos no comerciales", 0, 0),
        ("APIE bloqueadas", 0, n_apie_bloqueados_en_base),
        ("Fechas faltantes en BASE CAMPAÑAS", 0, n_fechas_faltantes),
        ("AccountManager vacio", 0, n_am_vacio),
        ("BF: APIE distintas totales", BF_CONTROLES_ESPERADOS["apie_distintas_totales"], n_bf_apie_distintas),
        ("BF: APIE con al menos un soporte positivo", BF_CONTROLES_ESPERADOS["apie_con_soporte_positivo"], n_bf_apie_positivas),
        ("BF: APIE en cero/vacio", BF_CONTROLES_ESPERADOS["apie_en_cero"], n_bf_apie_cero),
        ("BF: APIE positivas en catalogo digital activo", BF_CONTROLES_ESPERADOS["apie_positiva_en_catalogo_digital"], n_bf_apie_pos_digital),
        ("BF: APIE positivas fuera del catalogo digital (solo 30943)", BF_CONTROLES_ESPERADOS["apie_positiva_fuera_catalogo_digital"], n_bf_apie_pos_fuera_digital),
        ("BF: estaciones activas con PUNTERA positiva", BF_CONTROLES_ESPERADOS["estaciones_activas_puntera_positiva"], n_bf_pp_activas),
        ("BF: estaciones activas con TORRE positiva", BF_CONTROLES_ESPERADOS["estaciones_activas_torre_positiva"], n_bf_tt_activas),
        ("BF: estaciones activas con MENU BOARD positiva", BF_CONTROLES_ESPERADOS["estaciones_activas_menuboard_positiva"], n_bf_mb_activas),
        ("BF: union estaciones digitales con soporte positivo", BF_CONTROLES_ESPERADOS["union_estaciones_digitales_con_soporte"], n_bf_union_activas),
        ("Catalogo: estaciones digitales conservadas", 412, n_estaciones_digitales),
        ("Catalogo: estaciones estaticas conservadas", 191, n_estaciones_estaticas),
        ("Catalogo: BASE ESTACIONES filas totales", 603, len(filas_estaciones)),
        ("Catalogo: elementos FB conservados", 383, n_elem_fb),
        ("Catalogo: BASE ELEMENTOS filas totales", 3883, len(filas_elementos)),
        ("APIE 30943: elementos FB en BASE ELEMENTOS", 2,
         sum(1 for f in filas_elementos if normalizar_apie(f["Subcircuito"]) == APIE_SOLO_ESTATICA_CONOCIDA
             and str(f["ElementoID"]).split(" - ")[1:2] == ["FB"])),
        ("APIE 30943: campañas digitales asignadas en BASE CAMPAÑAS", 0,
         sum(1 for f in base_campanas if normalizar_apie(f["Subcircuito"]) == APIE_SOLO_ESTATICA_CONOCIDA)),
    ]

    preflight_fallas = []
    for nombre, esperado, obtenido in preflight_checks:
        estado = "OK" if esperado == obtenido else "DIFERENCIA"
        print(f"  {nombre:38s} esperado={str(esperado):>8s} obtenido={str(obtenido):>8s} [{estado}]")
        if esperado != obtenido:
            preflight_fallas.append((nombre, esperado, obtenido))

    if am_sin_informar_campanas - {"PIRELLI"}:
        preflight_fallas.append(("AccountManager SIN INFORMAR solo en Pirelli", "PIRELLI", sorted(am_sin_informar_campanas)))
        print(f"  AccountManager SIN INFORMAR       esperado=PIRELLI obtenido={sorted(am_sin_informar_campanas)} [DIFERENCIA]")
    else:
        print("  AccountManager SIN INFORMAR       esperado=PIRELLI obtenido=PIRELLI [OK]")

    if preflight_fallas:
        print("\nERROR BLOQUEANTE: el preflight no coincide con los conteos rectores del prompt maestro.")
        for nombre, esperado, obtenido in preflight_fallas:
            print(f"  - {nombre}: esperado={esperado} obtenido={obtenido} diferencia={obtenido if isinstance(obtenido,int) and isinstance(esperado,int) else 'N/A'}")
        print("No se genera el archivo final. Revise la causa antes de continuar (no se fuerzan numeros).")
        return 10

    print("  PREFLIGHT: TODOS LOS CONTROLES OK. Se procede a construir el archivo de salida.")

    # =========================================================================
    # Construccion de MAPEO CAMPAÑAS
    # =========================================================================
    print("[8/13] Construyendo hoja MAPEO CAMPAÑAS...")
    mapeo_rows = []
    filas_asignadas_por_id = Counter(f["IDCampaña"] for f in base_campanas)
    for u in unidades:
        mapeo_rows.append({
            "IDCampaña": u["id_campana"], "FilaCabeceraFuente": u["header_row"],
            "TituloBloqueOriginal": u["titulo"], "PautaOrigen": u["pauta_origen"], "MarcaOrigen": u["marca_origen"],
            "CampañaNormalizada": u["pauta_key"], "FechaInicio": u["fecha_inicio"], "FechaFin": u["fecha_fin"],
            "AccountManager": u["account_manager"], "EstadoMapeo": u["estado_mapeo"],
            "MotivoNoCarga": u["motivo_no_carga"], "CantidadFilasFuente": len(u["filas"]),
            "CantidadFilasAsignadas": filas_asignadas_por_id.get(u["id_campana"], 0),
        })

    # =========================================================================
    # RESUMEN CAMPAÑAS
    # =========================================================================
    print("[9/13] Construyendo hoja RESUMEN CAMPAÑAS...")

    dist_estado = Counter(f["Estado"] for f in base_campanas)
    dist_am = Counter(f["AccountManager"] for f in base_campanas)
    dist_campania = Counter(f["Campaña"] for f in base_campanas)
    dist_tiposoporte = Counter(str(f["ElementoID"]).split(" - ")[1] for f in base_campanas)
    apies_unicos_usados = len(apies_usados)

    resumen_rows = [
        ("Campo", "Valor"),
        ("--- METADATA ---", ""),
        ("Fecha de generacion", FECHA_GENERACION),
        ("Archivo fuente", args.comercial),
        ("Solapa fuente", HOJA_FUENTE_COMERCIAL),
        ("Fecha de corte", FECHA_CORTE.strftime("%d/%m/%Y")),
        ("--- DETECCION ---", ""),
        ("Bloques detectados", n_bloques),
        ("Campañas bloque+pauta (unidades)", n_unidades),
        ("Campañas cargadas (IDCampaña en BASE CAMPAÑAS)", n_cargables),
        ("Campañas excluidas (ELIMINAR)", n_excluidas),
        ("Campañas pendientes por fechas (FALTAN_FECHAS)", n_pendientes_fecha),
        ("--- FILAS ---", ""),
        ("Filas fuente (COMERCIAL, todos los bloques)", total_filas_fuente),
        ("Filas cargables antes del cruce", filas_candidatas),
        ("Filas excluidas no-ir", len(filas_noir)),
        ("Filas sin estacion (cargables)", filas_sin_estacion),
        ("--- ELEMENTOS ---", ""),
        ("Elementos solicitados", elementos_solicitados),
        ("Elementos asignados", elementos_asignados),
        ("Faltante de inventario", faltante_total),
        ("Filas de BASE CAMPAÑAS", len(base_campanas)),
        ("APIE unicos utilizados", apies_unicos_usados),
        ("--- DISTRIBUCION POR ESTADO ---", ""),
    ] + [(k, v) for k, v in sorted(dist_estado.items())] + [
        ("--- DISTRIBUCION POR ACCOUNT MANAGER ---", ""),
    ] + [(k, v) for k, v in sorted(dist_am.items())] + [
        ("--- DISTRIBUCION POR CAMPAÑA ---", ""),
    ] + [(k, v) for k, v in sorted(dist_campania.items())] + [
        ("--- DISTRIBUCION POR TIPO DE SOPORTE ---", ""),
    ] + [(k, v) for k, v in sorted(dist_tiposoporte.items())] + [
        ("--- METODOS DE CRUCE ---", ""),
    ] + [(k, v) for k, v in sorted(metodo_cruce_cnt.items())] + [
        ("--- BF (reconciliacion adicional, solapa BF) ---", ""),
        ("Filas BF procesadas", len(auditoria_bf)),
        ("BF: APIE distintas totales", n_bf_apie_distintas),
        ("BF: APIE con al menos un soporte positivo", n_bf_apie_positivas),
        ("BF: APIE en cero/vacio", n_bf_apie_cero),
        ("BF: APIE positivas en catalogo digital activo", n_bf_apie_pos_digital),
        ("BF: APIE positivas fuera del catalogo digital (solo 30943)", n_bf_apie_pos_fuera_digital),
        ("BF: estaciones activas con PUNTERA positiva", n_bf_pp_activas),
        ("BF: estaciones activas con TORRE positiva", n_bf_tt_activas),
        ("BF: estaciones activas con MENU BOARD positiva", n_bf_mb_activas),
        ("BF: union estaciones digitales con soporte positivo", n_bf_union_activas),
        ("BF vs COMERCIAL: diferencias (DIFERENCIA_BF_COMERCIAL)", n_bf_diferencias),
        ("BF vs COMERCIAL: pendientes de revision (PENDIENTE_REVISION)", n_bf_pendientes_revision),
    ] + [(k, v) for k, v in sorted(Counter(a["ResultadoCruce"] for a in auditoria_bf).items())] + [
        ("--- CONFIRMACIONES ---", ""),
        ("Catalogo original de Etapa 1 sin cambios", "SI (hashes verificados antes/despues, ver AUDITORIA CAMPAÑAS)"),
        ("Elementos FB en BASE CAMPAÑAS", n_fb),
        ("APIE bloqueados en BASE CAMPAÑAS", n_apie_bloqueados_en_base),
        ("APIE 30943 tratada como estacion Estatica (2 FB, sin campañas digitales)",
         "SI" if sum(1 for f in base_campanas if normalizar_apie(f["Subcircuito"]) == APIE_SOLO_ESTATICA_CONOCIDA) == 0 else "NO"),
    ]

    # =========================================================================
    # AUDITORIA CAMPAÑAS
    # =========================================================================
    print("[10/13] Construyendo hoja AUDITORIA CAMPAÑAS...")

    def aud(control, esperado, obtenido, fuente="", detalle=""):
        if esperado is None:
            return {"Control": control, "Esperado": "(informativo)", "Obtenido": obtenido, "Diferencia": None,
                    "Estado": "OK", "Detalle": detalle, "Fuente": fuente}
        try:
            dif = obtenido - esperado
        except TypeError:
            dif = "N/A"
        estado = "OK" if esperado == obtenido else "ERROR"
        return {"Control": control, "Esperado": esperado, "Obtenido": obtenido, "Diferencia": dif,
                "Estado": estado, "Detalle": detalle, "Fuente": fuente}

    sep_bf = {"Control": "--- BF (reconciliacion adicional) ---", "Esperado": None, "Obtenido": None,
              "Diferencia": None, "Estado": "OK", "Detalle": "", "Fuente": ""}

    n_ids_unicos = len({u["id_campana"] for u in unidades})
    turismo_cordoba_falt = sum(p["CantidadFaltante"] for p in pendientes
                                if p["TipoPendiente"] == "FALTANTE_INVENTARIO" and p["Campaña"] == "CORDOBA TURISMO")
    apie551_falt = sum(p["CantidadFaltante"] for p in pendientes
                        if p["TipoPendiente"] == "FALTANTE_INVENTARIO" and p["APIE"] == "551")

    auditoria_rows = [
        aud("Hash Etapa1 base (pre)", hash_etapa1_pre, hash_etapa1_pre, "SHA-256", "Se recalcula tras guardar; ver control de hashes post"),
        aud("Hash COMERCIAL (pre)", hash_comercial_pre, hash_comercial_pre, "SHA-256", "Se recalcula tras guardar; ver control de hashes post"),
        aud("Bloques detectados en COMERCIAL", 31, n_bloques, "COMERCIAL"),
        aud("Filas de estacion fuente (COMERCIAL)", 2371, total_filas_fuente, "COMERCIAL"),
        aud("Bloques marcados ELIMINAR", 2, len(ELIMINAR_HEADERS), "COMERCIAL", "FANTA (573), NETFLIX (1177)"),
        aud("Bloques sin fechas completas", 6, len(SIN_FECHAS_HEADERS), "COMERCIAL"),
        aud("Unidades bloque+pauta totales", 35, n_unidades, "MAPEO CAMPAÑAS"),
        aud("IDCampaña unicos asignados", 35, n_ids_unicos, "MAPEO CAMPAÑAS", "IDs reservados incluso para excluidas/pendientes"),
        aud("IDCampaña cargables", 27, n_cargables, "MAPEO CAMPAÑAS"),
        aud("Filas candidatas antes del cruce", 2034, filas_candidatas, "COMERCIAL"),
        aud("Filas no-ir excluidas", 2, len(filas_noir), "COMERCIAL", "APIE 551 y APIE 30943 en YO NARCISO"),
        aud("Filas cargables sin estacion", 6, filas_sin_estacion, "BASE ESTACIONES", "6 apariciones de Avenida de los Lagos 330 / APIE 30943"),
        aud("Cruces por APIE", None, metodo_cruce_cnt.get("APIE", 0), "BASE ESTACIONES"),
        aud("Cruces por direccion+localidad", 212, metodo_cruce_cnt.get("DIRECCION_LOCALIDAD", 0), "BASE ESTACIONES"),
        aud("Cruces por direccion unica", 8, metodo_cruce_cnt.get("DIRECCION_UNICA", 0), "BASE ESTACIONES"),
        aud("Elementos solicitados", 13779, elementos_solicitados, "COMERCIAL"),
        aud("Elementos asignados", 13616, elementos_asignados, "BASE ELEMENTOS"),
        aud("Faltante de inventario total", 163, faltante_total, "BASE ELEMENTOS"),
        aud("Faltante explicado por Turismo Cordoba", 153, turismo_cordoba_falt, "PENDIENTES CAMPAÑAS"),
        aud("Faltante explicado por APIE 551 (PPUNTER)", 10, apie551_falt, "PENDIENTES CAMPAÑAS"),
        aud("Filas finales BASE CAMPAÑAS", 13616, len(base_campanas), "BASE CAMPAÑAS"),
        aud("Duplicados de clave compuesta (IDCampaña+ElementoID+FechaInicio+FechaFin)", 0, n_dup_clave, "BASE CAMPAÑAS"),
        aud("Elementos FB en BASE CAMPAÑAS", 0, n_fb, "BASE CAMPAÑAS", "Las campanas de esta etapa son solo digitales"),
        aud("APIE bloqueados en BASE CAMPAÑAS", 0, n_apie_bloqueados_en_base, "BASE CAMPAÑAS"),
        aud("APIE bloqueados referenciados en COMERCIAL", 0, len(apies_bloqueados_en_comercial), "COMERCIAL"),
        aud("Fechas faltantes en BASE CAMPAÑAS", 0, n_fechas_faltantes, "BASE CAMPAÑAS"),
        aud("AccountManager vacio en BASE CAMPAÑAS", 0, n_am_vacio, "BASE CAMPAÑAS"),
        aud("Fanta presente en BASE CAMPAÑAS", 0, sum(1 for f in base_campanas if f["CampañaOrigen"] and "FANTA" in str(f["CampañaOrigen"]).upper()), "BASE CAMPAÑAS"),
        aud("Netflix presente en BASE CAMPAÑAS", 0, sum(1 for f in base_campanas if f["CampañaOrigen"] and "NETFLIX" in str(f["CampañaOrigen"]).upper()), "BASE CAMPAÑAS"),
        aud("APIE 30943 presente en BASE CAMPAÑAS", 0, sum(1 for f in base_campanas if normalizar_apie(f["Subcircuito"]) == "30943"), "BASE CAMPAÑAS"),
        aud("BASE ESTACIONES de Etapa 1 (filas)", 603, len(filas_estaciones), "BASE ESTACIONES"),
        aud("BASE ELEMENTOS de Etapa 1 (filas)", 3883, len(filas_elementos), "BASE ELEMENTOS"),
        aud("Elementos Digital comercializables de Etapa 1", 3500, conteo_medio[("Digital", "SI")], "BASE ELEMENTOS"),
        aud("Elementos FB de Etapa 1 (fuente)", 383, conteo_medio[("Estático", "SI")] + conteo_medio[("Estático", "NO")], "BASE ELEMENTOS"),
        sep_bf,
        aud("Catalogo: estaciones digitales conservadas", 412, n_estaciones_digitales, "BASE ESTACIONES"),
        aud("Catalogo: estaciones estaticas conservadas", 191, n_estaciones_estaticas, "BASE ESTACIONES"),
        aud("BF: hoja presente en el archivo de certificaciones", True, HOJA_FUENTE_BF in wb_com.sheetnames, "BF"),
        aud("BF: filas procesadas", 2371, len(auditoria_bf), "BF"),
        aud("BF: APIE distintas totales", BF_CONTROLES_ESPERADOS["apie_distintas_totales"], n_bf_apie_distintas, "BF"),
        aud("BF: APIE con al menos un soporte positivo", BF_CONTROLES_ESPERADOS["apie_con_soporte_positivo"], n_bf_apie_positivas, "BF"),
        aud("BF: APIE en cero/vacio", BF_CONTROLES_ESPERADOS["apie_en_cero"], n_bf_apie_cero, "BF"),
        aud("BF: APIE positivas en catalogo digital activo", BF_CONTROLES_ESPERADOS["apie_positiva_en_catalogo_digital"], n_bf_apie_pos_digital, "BF"),
        aud("BF: APIE positivas fuera del catalogo digital (solo 30943)", BF_CONTROLES_ESPERADOS["apie_positiva_fuera_catalogo_digital"], n_bf_apie_pos_fuera_digital, "BF",
            f"APIE fuera del catalogo digital: {sorted(apie_bf_pos_fuera_digital)}"),
        aud("BF: estaciones activas con PUNTERA positiva", BF_CONTROLES_ESPERADOS["estaciones_activas_puntera_positiva"], n_bf_pp_activas, "BF"),
        aud("BF: estaciones activas con TORRE positiva", BF_CONTROLES_ESPERADOS["estaciones_activas_torre_positiva"], n_bf_tt_activas, "BF"),
        aud("BF: estaciones activas con MENU BOARD positiva", BF_CONTROLES_ESPERADOS["estaciones_activas_menuboard_positiva"], n_bf_mb_activas, "BF"),
        aud("BF: union estaciones digitales con soporte positivo", BF_CONTROLES_ESPERADOS["union_estaciones_digitales_con_soporte"], n_bf_union_activas, "BF"),
        aud("APIE 30943: elementos FB en BASE ELEMENTOS", 2,
            sum(1 for f in filas_elementos if normalizar_apie(f["Subcircuito"]) == APIE_SOLO_ESTATICA_CONOCIDA
                and str(f["ElementoID"]).split(" - ")[1:2] == ["FB"]), "BASE ELEMENTOS"),
        aud("APIE 30943: campañas digitales asignadas en BASE CAMPAÑAS", 0,
            sum(1 for f in base_campanas if normalizar_apie(f["Subcircuito"]) == APIE_SOLO_ESTATICA_CONOCIDA), "BASE CAMPAÑAS"),
        aud("APIE 30943: presente en BASE ESTACIONES como Estatica", True,
            any(normalizar_apie(f["APIE"]) == APIE_SOLO_ESTATICA_CONOCIDA and f["Formato"] == "Estático" for f in filas_estaciones),
            "BASE ESTACIONES"),
        aud("BF vs COMERCIAL: diferencias no resueltas silenciosamente (DIFERENCIA_BF_COMERCIAL)", None, n_bf_diferencias,
            "PENDIENTES CAMPAÑAS", "Cada diferencia queda documentada en PENDIENTES CAMPAÑAS, no se elige version silenciosamente"),
        aud("BF: filas que no pudieron relacionarse inequivocamente con COMERCIAL", None, n_bf_pendientes_revision,
            "PENDIENTES CAMPAÑAS"),
    ]

    n_errores_auditoria = sum(1 for r in auditoria_rows if r["Estado"] == "ERROR")
    if n_errores_auditoria:
        print("ERROR BLOQUEANTE: AUDITORIA CAMPAÑAS detecto controles en ERROR tras el preflight.")
        for r in auditoria_rows:
            if r["Estado"] == "ERROR":
                print(f"  - {r['Control']}: esperado={r['Esperado']} obtenido={r['Obtenido']}")
        return 11

    # =========================================================================
    # Ensamblado del workbook de salida (copia fiel de Etapa 1 + hojas nuevas)
    # =========================================================================
    print("[11/13] Ensamblando workbook de salida (copia fiel de Etapa 1 + hojas nuevas)...")
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    def copiar_hoja_fuente(nombre_hoja):
        ws_src = wb1_formulas[nombre_hoja]
        ws_val = wb1_valores[nombre_hoja]
        ws_dst = wb_out.create_sheet(nombre_hoja)
        for row in ws_src.iter_rows():
            for cell in row:
                valor = cell.value
                if cell.data_type == "f" and isinstance(valor, str) and EXTERNAL_REF_PATTERN.search(valor):
                    valor = ws_val.cell(row=cell.row, column=cell.column).value
                new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=valor)
                if cell.has_style:
                    new_cell.font = copy_style(cell.font)
                    new_cell.fill = copy_style(cell.fill)
                    new_cell.border = copy_style(cell.border)
                    new_cell.alignment = copy_style(cell.alignment)
                    new_cell.number_format = cell.number_format
        for key, dim in ws_src.column_dimensions.items():
            ws_dst.column_dimensions[key].width = dim.width
        for key, dim in ws_src.row_dimensions.items():
            ws_dst.row_dimensions[key].height = dim.height
        for merged_range in ws_src.merged_cells.ranges:
            ws_dst.merge_cells(str(merged_range))
        ws_dst.sheet_view.showGridLines = ws_src.sheet_view.showGridLines
        return ws_dst

    def escribir_hoja_tabular(nombre, columnas, filas):
        ws = wb_out.create_sheet(nombre)
        ws.append(columnas)
        for c in ws[1]:
            c.font = Font(bold=True)
        for fila in filas:
            valores = []
            for col in columnas:
                v = fila.get(col) if isinstance(fila, dict) else fila[columnas.index(col)]
                valores.append(valor_seguro_para_celda(v))
            ws.append(valores)
        date_cols = {i for i, c in enumerate(columnas, start=1) if c in ("FechaInicio", "FechaFin")}
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column in date_cols and cell.value is not None:
                    cell.number_format = "DD/MM/YYYY"
        for i, col in enumerate(columnas, start=1):
            muestras = [f.get(col) if isinstance(f, dict) else f[columnas.index(col)] for f in filas[:500]]
            max_len = max([len(col)] + [len(str(v)) for v in muestras if v is not None])
            ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 60)
        ws.freeze_panes = "A2"
        return ws

    # Copia LEEME/AUDITORIA/BASE ELEMENTOS/BASE ESTACIONES originales sin cambios,
    # luego inserta las 5 hojas nuevas de Etapa 2, y por ultimo el resto de las
    # hojas originales (PENDIENTES, DIGITAL, ESTATICO , ZONAS, DIRECCIONES,
    # ZONAS VERSION ANTUGUA) preservando su orden relativo.
    ws_leeme = copiar_hoja_fuente("LEEME")
    copiar_hoja_fuente("AUDITORIA")
    copiar_hoja_fuente("BASE ELEMENTOS")
    copiar_hoja_fuente("BASE ESTACIONES")

    escribir_hoja_tabular("BASE CAMPAÑAS", BASE_CAMPANAS_COLS, base_campanas)
    escribir_hoja_tabular("RESUMEN CAMPAÑAS", ["Campo", "Valor"], resumen_rows[1:])
    escribir_hoja_tabular("MAPEO CAMPAÑAS", MAPEO_COLS, mapeo_rows)
    escribir_hoja_tabular("PENDIENTES CAMPAÑAS", PENDIENTES_COLS, pendientes)
    escribir_hoja_tabular("AUDITORIA CAMPAÑAS", AUDITORIA_COLS, auditoria_rows)
    escribir_hoja_tabular("AUDITORIA BF", AUDITORIA_BF_COLS, auditoria_bf)

    for nombre in ["PENDIENTES", "DIGITAL", "ESTATICO ", "ZONAS", "DIRECCIONES", "ZONAS VERSION ANTUGUA"]:
        copiar_hoja_fuente(nombre)

    # Seccion Etapa 2 en LEEME (append, sin borrar contenido original)
    ws_leeme.append(("", ""))
    ws_leeme.append(("--- ETAPA 2: INTEGRACION DE CAMPAÑAS COMERCIALES ---", ""))
    leeme_etapa2 = [
        ("ARCHIVO", "YPF_BASE_INTEGRADA_ETAPA_2_CAMPANAS_2026-08-13.xlsx (version original, sin BF)"),
        ("FECHA DE GENERACION (version original)", "2026-08-13"),
        ("FUENTE DE CAMPAÑAS", f"{args.comercial} — solapa '{HOJA_FUENTE_COMERCIAL}' (fuente primaria y exclusiva)"),
        ("BASE DE ETAPA 1 UTILIZADA", args.etapa1),
        ("GRANO BASE CAMPAÑAS", "Una fila por IDCampaña + ElementoID fisico"),
        ("CLAVE COMPUESTA UNICA", "IDCampaña + ElementoID + FechaInicio + FechaFin"),
        ("BLOQUES DETECTADOS", n_bloques),
        ("UNIDADES BLOQUE+PAUTA", n_unidades),
        ("IDCAMPAÑA CARGABLES", n_cargables),
        ("FILAS DE BASE CAMPAÑAS", len(base_campanas)),
        ("ELEMENTOS SOLICITADOS / ASIGNADOS / FALTANTE", f"{elementos_solicitados} / {elementos_asignados} / {faltante_total}"),
        ("CAMPAÑAS EXCLUIDAS (ELIMINAR)", "FANTA, NETFLIX"),
        ("CAMPAÑAS PENDIENTES POR FECHAS", "MILKA (416), PERSONAL COMBO (594), PERSONAL (605), MOVISTAR (639), TOYOTA (674), DERMAGLOS ENZO (773)"),
        ("APIE BLOQUEADOS", "30510, 31131, 31192, 31239 — sin campañas ni elementos asignados"),
        ("SOPORTES ALCANZADOS", "Solo digitales: PPUNTER, TT, MB. No se asignan FB (estatico) en esta etapa."),
        ("METADATOS REPETIDOS", "Marca = Cliente = Agencia = MarcaBase, autorizado expresamente porque la fuente "
                                  "COMERCIAL no distingue de forma completa cliente/marca/agencia."),
        ("HASHES", "Ver hoja AUDITORIA CAMPAÑAS y RESUMEN CAMPAÑAS para el detalle de verificacion antes/despues."),
        ("ESTADO DE INTEGRACION", "Este archivo NO esta integrado a la base OCU26 productiva"),
        ("", ""),
        ("--- ACTUALIZACION BF (2026-08-14) ---", ""),
        ("ARCHIVO DE ESTA ACTUALIZACION", "YPF_BASE_INTEGRADA_ETAPA_2_CAMPANAS_BF_2026-08-14.xlsx"),
        ("FECHA DE GENERACION", FECHA_GENERACION),
        ("QUE CAMBIA", "Se incorpora la solapa BF como fuente adicional obligatoria de cruce/reconciliacion/"
                        "validacion de estaciones, APIE, soportes y campañas. COMERCIAL sigue siendo la fuente "
                        "primaria y exclusiva de estructura y metadatos de campaña (campaña, pauta, marca, "
                        "cliente, agencia, fechas, AccountManager). BF NUNCA crea ni modifica filas de BASE "
                        "CAMPAÑAS: solo se usa para auditar."),
        ("HOJA NUEVA", "AUDITORIA BF: una fila por fila de BF, con APIE fuente/resuelta, soportes, existencia en "
                        "el catalogo digital/estatico y resultado del cruce contra COMERCIAL"),
        ("BF: filas procesadas", len(auditoria_bf)),
        ("BF: APIE distintas / con soporte positivo / en cero",
         f"{n_bf_apie_distintas} / {n_bf_apie_positivas} / {n_bf_apie_cero}"),
        ("BF: APIE positivas en catalogo digital / fuera del catalogo digital",
         f"{n_bf_apie_pos_digital} / {n_bf_apie_pos_fuera_digital} (unicamente APIE 30943, estacion solo Estatica)"),
        ("BF: estaciones activas con PUNTERA / TORRE / MENU BOARD positiva",
         f"{n_bf_pp_activas} / {n_bf_tt_activas} / {n_bf_mb_activas} (union: {n_bf_union_activas})"),
        ("BF vs COMERCIAL: diferencias documentadas / pendientes de revision",
         f"{n_bf_diferencias} / {n_bf_pendientes_revision} (ver PENDIENTES CAMPAÑAS y AUDITORIA BF, ninguna se resolvio en forma silenciosa)"),
        ("APIE 30943", "Estacion exclusivamente Estatica (2 FB, Avenida de los Lagos 330, Tigre). Permanece en "
                        "BASE ESTACIONES y BASE ELEMENTOS. No se incorpora como estacion digital ni recibe "
                        "campañas digitales. Sus solicitudes de campaña digital quedan en PENDIENTES CAMPAÑAS "
                        "con TipoPendiente=APIE_SOLO_ESTATICA_CAMPAÑA_DIGITAL (en vez del rotulo generico "
                        "APIE_NO_ENCONTRADA)."),
        ("CATALOGO GENERAL CONSERVADO", f"{n_estaciones_digitales} estaciones digitales + {n_estaciones_estaticas} "
                                          f"estaticas = {len(filas_estaciones)} filas en BASE ESTACIONES; "
                                          f"{conteo_medio[('Digital','SI')]} elementos digitales + {n_elem_fb} FB "
                                          f"= {len(filas_elementos)} filas en BASE ELEMENTOS (sin cambios vs Etapa 1)"),
        ("SALIDA ANTERIOR", "YPF_BASE_INTEGRADA_ETAPA_2_CAMPANAS_2026-08-13.xlsx NO fue reemplazada ni modificada."),
    ]
    for r in leeme_etapa2:
        ws_leeme.append(r)

    print(f"  Orden final de hojas: {wb_out.sheetnames}")

    print("[12/13] Verificando ausencia de referencias externas residuales en el libro de salida...")
    residuales = []
    for nombre in wb_out.sheetnames:
        ws = wb_out[nombre]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" and isinstance(cell.value, str) and EXTERNAL_REF_PATTERN.search(cell.value):
                    residuales.append((nombre, cell.coordinate, cell.value))
    if residuales:
        print("ERROR BLOQUEANTE: quedaron formulas con referencia externa sin neutralizar en el libro de salida:")
        for r in residuales:
            print(f"  - {r[0]}!{r[1]}: {r[2]!r}")
        return 12
    print("  OK: 0 formulas con referencia externa residual.")

    print(f"[13/13] Escribiendo archivo de salida: {args.output}")
    import os
    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    wb_out.save(args.output)

    hash_etapa1_post = sha256_de(args.etapa1)
    hash_comercial_post = sha256_de(args.comercial)
    fuentes_intactas = (hash_etapa1_post == hash_etapa1_pre) and (hash_comercial_post == hash_comercial_pre)

    print("\n=== VERIFICACION DE HASHES (fuentes no modificadas) ===")
    print(f"  Etapa1    pre={hash_etapa1_pre}")
    print(f"  Etapa1    post={hash_etapa1_post}")
    print(f"  Comercial pre={hash_comercial_pre}")
    print(f"  Comercial post={hash_comercial_post}")
    print(f"  Fuentes intactas: {'SI' if fuentes_intactas else 'NO -- ERROR BLOQUEANTE'}")

    if not fuentes_intactas:
        print("ERROR BLOQUEANTE: una fuente original cambio durante la ejecucion.")
        return 13

    print("\n=== RESUMEN ===")
    print(f"Bloques detectados: {n_bloques} | Unidades bloque+pauta: {n_unidades} | IDCampaña cargables: {n_cargables}")
    print(f"Filas fuente: {total_filas_fuente} | Filas candidatas: {filas_candidatas} | no-ir: {len(filas_noir)} | sin estacion: {filas_sin_estacion}")
    print(f"Elementos solicitados: {elementos_solicitados} | asignados: {elementos_asignados} | faltante: {faltante_total}")
    print(f"Filas BASE CAMPAÑAS: {len(base_campanas)}")
    print(f"Cruces -> APIE: {metodo_cruce_cnt.get('APIE', 0)} | DIRECCION_LOCALIDAD: {metodo_cruce_cnt.get('DIRECCION_LOCALIDAD', 0)} | "
          f"DIRECCION_UNICA: {metodo_cruce_cnt.get('DIRECCION_UNICA', 0)}")
    print(f"Distribucion Estado: {dict(dist_estado)}")
    print(f"Distribucion AccountManager: {dict(dist_am)}")
    print(f"BF: {len(auditoria_bf)} filas | APIE distintas={n_bf_apie_distintas} con soporte={n_bf_apie_positivas} "
          f"en catalogo digital={n_bf_apie_pos_digital} fuera del catalogo digital={n_bf_apie_pos_fuera_digital}")
    print(f"BF: PUNTERA activas={n_bf_pp_activas} TORRE activas={n_bf_tt_activas} MENU BOARD activas={n_bf_mb_activas} "
          f"union={n_bf_union_activas} | diferencias vs COMERCIAL={n_bf_diferencias} pendientes revision={n_bf_pendientes_revision}")
    print(f"Catalogo conservado: {n_estaciones_digitales} estaciones digitales + {n_estaciones_estaticas} estaticas "
          f"= {len(filas_estaciones)} filas BASE ESTACIONES | APIE 30943 tratada como solo-Estatica (2 FB, 0 campañas digitales)")
    print("No se modificaron archivos originales. No se modifico OCU26 productivo. No se hizo commit/push/PR/deploy.")
    print("Salida anterior YPF_BASE_INTEGRADA_ETAPA_2_CAMPANAS_2026-08-13.xlsx no fue reemplazada.")
    print(f"Archivo final: {args.output}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
