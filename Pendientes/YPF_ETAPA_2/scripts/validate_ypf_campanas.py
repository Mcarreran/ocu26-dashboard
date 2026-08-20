"""
validate_ypf_campanas.py

Validador de solo lectura para
YPF_BASE_INTEGRADA_ETAPA_2_CAMPANAS_2026-08-13.xlsx.

No escribe, no borra, no renombra, no mueve ni sobrescribe ningun archivo.
No ejecuta operaciones de Git ni accede a la red. Abre el Excel de salida
con openpyxl y el paquete XLSX con zipfile / xml.etree.ElementTree para
verificar integridad. Ademas recalcula de forma independiente, a partir de
las fuentes originales (COMERCIAL y la base validada de Etapa 1), los
conteos de solicitud/asignacion de elementos, reutilizando unicamente las
funciones puras de deteccion/cruce de build_ypf_campanas.py (deteccion de
bloques, normalizacion, cruce de estaciones): no vuelve a escribir nada, ni
depende de los resultados ya guardados en el Excel de salida para ese
recalculo.

Uso:
    python validate_ypf_campanas.py [--output <ruta.xlsx>]

Codigo de salida 0 si todos los controles bloqueantes pasan, distinto de
cero si alguno falla.
"""

import argparse
import hashlib
import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import date

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_ypf_campanas as B  # noqa: E402  (funciones puras, solo lectura)

DEFAULT_OUTPUT = B.DEFAULT_OUTPUT

# Hashes SHA-256 registrados de las fuentes originales antes de generar
# Etapa 2 (ver Pendientes/YPF_ETAPA_2/output/HASHES_PRE_ETAPA2.txt). Deben
# permanecer identicos: estas fuentes no deben modificarse nunca.
ARCHIVOS_ORIGINALES = {
    B.DEFAULT_ETAPA1: "4f05f6c5615f3dcd5eb15ef931c410282c2b2f689f963acf01493265c8f1a60c",
    B.DEFAULT_COMERCIAL: "4073e6d1cbf7675099df0a54154090ca0cccc1be289185d43a4091470798c1c6",
}

EXTERNAL_REF_PATTERN = re.compile(r"\[\d+\]")

HOJAS_ESPERADAS = [
    "LEEME", "AUDITORIA", "BASE ELEMENTOS", "BASE ESTACIONES",
    "BASE CAMPAÑAS", "RESUMEN CAMPAÑAS", "MAPEO CAMPAÑAS", "PENDIENTES CAMPAÑAS", "AUDITORIA CAMPAÑAS",
    "PENDIENTES", "DIGITAL", "ESTATICO ", "ZONAS", "DIRECCIONES", "ZONAS VERSION ANTUGUA",
]


class Resultados:
    def __init__(self):
        self.items = []

    def check(self, id_, descripcion, condicion, detalle=""):
        estado = "OK" if condicion else "ERROR"
        self.items.append({"id": id_, "descripcion": descripcion, "estado": estado, "detalle": detalle})
        return condicion

    def errores(self):
        return [i for i in self.items if i["estado"] == "ERROR"]

    def imprimir(self):
        for i in self.items:
            marca = "OK   " if i["estado"] == "OK" else "ERROR"
            linea = f"[{marca}] {i['id']:>3} {i['descripcion']}"
            if i["detalle"]:
                linea += f" -- {i['detalle']}"
            print(linea)
        print()
        n_ok = sum(1 for i in self.items if i["estado"] == "OK")
        n_err = len(self.errores())
        print(f"TOTAL: {n_ok} OK / {n_err} ERROR / {len(self.items)} controles")
        return n_err == 0


def sha256_de(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def leer_tabla(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    data = [r for r in rows[1:] if any(c is not None for c in r)]
    return header, data


def col(header, nombre):
    return header.index(nombre)


def es_fecha(v):
    return hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day")


def as_date(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    return v


def recomputar_solicitud_asignacion_independiente():
    """Reconstruye, desde las fuentes originales (solo lectura), la cantidad
    de elementos solicitados/asignados/faltantes y las filas candidatas,
    usando las mismas funciones puras de deteccion/cruce del constructor
    (build_ypf_campanas.py), sin leer nada del Excel de salida ya generado."""
    wb1 = openpyxl.load_workbook(B.DEFAULT_ETAPA1, data_only=True)
    _, digital_apies_si, by_dir_loc, by_dir = B.cargar_lookups_estaciones(wb1)
    _, inventario, _, _ = B.cargar_inventario_elementos(wb1)

    wb_com = openpyxl.load_workbook(B.DEFAULT_COMERCIAL, data_only=True)
    ws_com = wb_com[B.HOJA_FUENTE_COMERCIAL]
    headers, unidades, filas_noir, total_filas_fuente, filas_por_header = B.construir_unidades(ws_com)
    B.asignar_ids(unidades)
    for u in unidades:
        u["estado_mapeo"], u["motivo_no_carga"] = B.clasificar_unidad(u)

    cargable_headers = {u["header_row"] for u in unidades if u["estado_mapeo"] == "CARGABLE"}
    filas_candidatas = sum(filas_por_header[hr] for hr in cargable_headers)

    solicitados = 0
    asignados = 0
    faltante = 0
    filas_sin_estacion = 0
    faltante_por_campania = Counter()
    faltante_por_apie_tipo = Counter()

    for u in unidades:
        if u["estado_mapeo"] != "CARGABLE":
            continue
        for r in u["filas"]:
            apie_raw = ws_com.cell(row=r, column=2).value
            direccion_raw = ws_com.cell(row=r, column=3).value
            localidad_raw = ws_com.cell(row=r, column=4).value
            apie_match, _metodo, _motivo = B.cruzar_estacion(
                apie_raw, direccion_raw, localidad_raw, digital_apies_si, by_dir_loc, by_dir)
            if apie_match is None:
                filas_sin_estacion += 1
                continue
            for col_idx, tipo in B.TIPO_SOPORTE_POR_COLUMNA.items():
                qty = B.to_int_qty(ws_com.cell(row=r, column=col_idx).value) or 0
                if qty <= 0:
                    continue
                solicitados += qty
                disponibles = inventario.get((apie_match, tipo), [])
                n_asig = min(qty, len(disponibles))
                asignados += n_asig
                if n_asig < qty:
                    faltante += (qty - n_asig)
                    faltante_por_campania[u["pauta_key"]] += (qty - n_asig)
                    faltante_por_apie_tipo[(apie_match, tipo)] += (qty - n_asig)

    n_cargables = sum(1 for u in unidades if u["estado_mapeo"] == "CARGABLE")
    n_unidades = len(unidades)

    return {
        "bloques": len(headers), "unidades": n_unidades, "cargables": n_cargables,
        "filas_fuente": total_filas_fuente, "filas_candidatas": filas_candidatas,
        "filas_noir": len(filas_noir), "filas_sin_estacion": filas_sin_estacion,
        "solicitados": solicitados, "asignados": asignados, "faltante": faltante,
        "faltante_turismo_cordoba": faltante_por_campania.get("CORDOBA TURISMO", 0),
        "faltante_apie551_ppunter": faltante_por_apie_tipo.get(("551", "PPUNTER"), 0),
    }


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", default=DEFAULT_OUTPUT)
    args = parser.parse_args(argv)
    path = args.output

    r = Resultados()

    # -----------------------------------------------------------------
    # 1. Integridad del archivo
    # -----------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        abre_ok = True
    except Exception as e:  # noqa: BLE001
        abre_ok = False
        wb = None
        r.check("1", "El archivo existe y abre con openpyxl", False, str(e))
    if abre_ok:
        r.check("1", "El archivo existe y abre con openpyxl", True)
        r.check("2", "Tiene las hojas esperadas (incluye las 5 nuevas de Etapa 2, en orden)",
                wb.sheetnames == HOJAS_ESPERADAS, f"obtenido={wb.sheetnames}")

    if not abre_ok:
        r.imprimir()
        print("VALIDATOR EXIT: 1")
        return 1

    with zipfile.ZipFile(path) as z:
        bad = z.testzip()
        r.check("3", "El paquete XLSX supera zipfile.testzip()", bad is None, f"primer archivo dañado: {bad}")

        xml_malformados = []
        for name in z.namelist():
            if name.endswith(".xml") or name.endswith(".rels"):
                try:
                    ET.fromstring(z.read(name))
                except ET.ParseError as e:
                    xml_malformados.append((name, str(e)))
        r.check("4", "Todos los XML y .rels del XLSX estan bien formados", not xml_malformados, f"{xml_malformados[:5]}")

        ext_links = [n for n in z.namelist() if "externallink" in n.lower()]
        r.check("5", "No existen partes externalLinks", not ext_links, f"encontrados: {ext_links}")

        f_pattern = re.compile(r"<f[^>]*>(.*?)</f>", re.DOTALL)
        residuales_xml = []
        for name in z.namelist():
            if re.match(r"xl/worksheets/sheet\d+\.xml$", name):
                contenido = z.read(name).decode("utf-8", errors="replace")
                for m in f_pattern.finditer(contenido):
                    if EXTERNAL_REF_PATTERN.search(m.group(1)):
                        residuales_xml.append((name, m.group(1)[:80]))
        r.check("6", "No existe ninguna formula activa con referencias externas [n] (XML crudo)",
                not residuales_xml, f"{residuales_xml[:5]}")

    # -----------------------------------------------------------------
    # 2. Hashes de las fuentes originales (no deben cambiar)
    # -----------------------------------------------------------------
    hashes_ok = True
    detalle_hashes = []
    for ruta, esperado in ARCHIVOS_ORIGINALES.items():
        try:
            obtenido = sha256_de(ruta)
        except FileNotFoundError:
            hashes_ok = False
            detalle_hashes.append(f"{ruta}: ARCHIVO NO ENCONTRADO")
            continue
        if obtenido != esperado:
            hashes_ok = False
            detalle_hashes.append(f"{ruta}: esperado {esperado[:12]}... obtenido {obtenido[:12]}...")
    r.check("7", "Los hashes SHA-256 de las fuentes originales (Etapa 1 y COMERCIAL) no cambiaron",
            hashes_ok, "; ".join(detalle_hashes))

    # -----------------------------------------------------------------
    # 3. Invariantes de Etapa 1 preservados dentro del archivo de salida
    # -----------------------------------------------------------------
    h_est, d_est = leer_tabla(wb["BASE ESTACIONES"])
    h_elem, d_elem = leer_tabla(wb["BASE ELEMENTOS"])

    r.check("8", "BASE ESTACIONES: 603 filas", len(d_est) == 603, f"obtenido={len(d_est)}")
    r.check("9", "BASE ELEMENTOS: 3.883 filas", len(d_elem) == 3883, f"obtenido={len(d_elem)}")

    c_medio = col(h_elem, "Medio")
    c_incluir_elem = col(h_elem, "IncluirComercializacion")
    c_eid = col(h_elem, "ElementoID")
    n_digital = sum(1 for row in d_elem if row[c_medio] == "Digital" and row[c_incluir_elem] == "SI")
    n_fb_total = sum(1 for row in d_elem if str(row[c_eid]).split(" - ")[1:2] == ["FB"])
    r.check("10", "BASE ELEMENTOS: 3.500 elementos Digital comercializables", n_digital == 3500, f"obtenido={n_digital}")
    r.check("11", "BASE ELEMENTOS: 383 elementos FB", n_fb_total == 383, f"obtenido={n_fb_total}")

    c_apie_est = col(h_est, "APIE")
    apies_bloqueados_presentes = B.APIE_BLOQUEADOS & {B.normalizar_apie(row[c_apie_est]) for row in d_est}
    r.check("12", "Los 4 APIE bloqueados no estan presentes en BASE ESTACIONES", not apies_bloqueados_presentes,
            f"encontrados={apies_bloqueados_presentes}")

    # -----------------------------------------------------------------
    # 4. BASE CAMPAÑAS
    # -----------------------------------------------------------------
    h_bc, d_bc = leer_tabla(wb["BASE CAMPAÑAS"])
    c = {name: col(h_bc, name) for name in B.BASE_CAMPANAS_COLS}

    r.check("13", "BASE CAMPAÑAS: 13.616 filas", len(d_bc) == 13616, f"obtenido={len(d_bc)}")

    ids_cargables = {row[c["IDCampaña"]] for row in d_bc}
    r.check("14", "BASE CAMPAÑAS: 27 IDCampaña unicos cargados", len(ids_cargables) == 27, f"obtenido={len(ids_cargables)}")

    claves = [(row[c["IDCampaña"]], row[c["ElementoID"]], as_date(row[c["FechaInicio"]]), as_date(row[c["FechaFin"]]))
              for row in d_bc]
    n_dup = len(claves) - len(set(claves))
    r.check("15", "0 claves compuestas duplicadas (IDCampaña+ElementoID+FechaInicio+FechaFin)", n_dup == 0, f"duplicados={n_dup}")

    eids_validos = {row[c_eid] for row in d_elem}
    eids_bc = {row[c["ElementoID"]] for row in d_bc}
    eids_invalidos = eids_bc - eids_validos
    r.check("16", "Todos los ElementoID de BASE CAMPAÑAS existen en BASE ELEMENTOS", not eids_invalidos,
            f"invalidos={list(eids_invalidos)[:10]}")

    medios_no_digital = {row[c["Medio"]] for row in d_bc} - {"Digital"}
    r.check("17", "Todas las filas de BASE CAMPAÑAS son Medio=Digital", not medios_no_digital, f"obtenido={medios_no_digital}")

    elem_incl = {row[c_eid]: row[c_incluir_elem] for row in d_elem}
    elem_medio = {row[c_eid]: row[c_medio] for row in d_elem}
    no_comercial = [eid for eid in eids_bc if elem_incl.get(eid) != "SI"]
    r.check("18", "Todos los ElementoID de BASE CAMPAÑAS tienen IncluirComercializacion=SI", not no_comercial,
            f"invalidos={no_comercial[:10]}")

    n_fb_bc = sum(1 for row in d_bc if str(row[c["ElementoID"]]).split(" - ")[1:2] == ["FB"])
    r.check("19", "0 elementos FB en BASE CAMPAÑAS (las campañas de Etapa 2 son solo digitales)", n_fb_bc == 0, f"obtenido={n_fb_bc}")

    apie_bloq_en_bc = sum(1 for row in d_bc if B.normalizar_apie(row[c["Subcircuito"]]) in B.APIE_BLOQUEADOS)
    r.check("20", "0 filas de BASE CAMPAÑAS pertenecen a los 4 APIE bloqueados", apie_bloq_en_bc == 0, f"obtenido={apie_bloq_en_bc}")

    n_fecha_vacia = sum(1 for row in d_bc if row[c["FechaInicio"]] is None or row[c["FechaFin"]] is None)
    r.check("21", "Ninguna FechaInicio/FechaFin vacia en BASE CAMPAÑAS", n_fecha_vacia == 0, f"obtenido={n_fecha_vacia}")

    n_fecha_invertida = sum(1 for row in d_bc
                             if row[c["FechaInicio"]] is not None and row[c["FechaFin"]] is not None
                             and as_date(row[c["FechaFin"]]) < as_date(row[c["FechaInicio"]]))
    r.check("22", "Ninguna FechaFin anterior a FechaInicio", n_fecha_invertida == 0, f"obtenido={n_fecha_invertida}")

    n_am_vacio = sum(1 for row in d_bc if not row[c["AccountManager"]] or not str(row[c["AccountManager"]]).strip())
    r.check("23", "Ningun AccountManager vacio en BASE CAMPAÑAS", n_am_vacio == 0, f"obtenido={n_am_vacio}")

    campanias_sin_informar = {row[c["Campaña"]] for row in d_bc if row[c["AccountManager"]] == "SIN INFORMAR"}
    r.check("24", "Solo Pirelli puede tener AccountManager=SIN INFORMAR", campanias_sin_informar <= {"PIRELLI"},
            f"obtenido={campanias_sin_informar}")

    errores_estado_temporal = 0
    for row in d_bc:
        fi, ff = as_date(row[c["FechaInicio"]]), as_date(row[c["FechaFin"]])
        esperado = B.estado_temporal(fi, ff)
        if row[c["Estado"]] != esperado:
            errores_estado_temporal += 1
    r.check("25", "El Estado temporal de cada fila coincide con FechaInicio/FechaFin y la fecha de corte 13/08/2026",
            errores_estado_temporal == 0, f"filas_incorrectas={errores_estado_temporal}")

    inventario_por_apie_tipo = Counter()
    for row in d_elem:
        if row[c_medio] != "Digital" or row[c_incluir_elem] != "SI":
            continue
        partes = str(row[c_eid]).split(" - ")
        if len(partes) == 3:
            inventario_por_apie_tipo[(B.normalizar_apie(row[col(h_elem, "Subcircuito")]), partes[1])] += 1

    uso_por_campania_apie_tipo = defaultdict(set)
    for row in d_bc:
        partes = str(row[c["ElementoID"]]).split(" - ")
        if len(partes) == 3:
            key = (row[c["IDCampaña"]], B.normalizar_apie(row[c["Subcircuito"]]), partes[1])
            uso_por_campania_apie_tipo[key].add(row[c["ElementoID"]])

    excede_inventario = [k for k, v in uso_por_campania_apie_tipo.items()
                          if len(v) > inventario_por_apie_tipo.get((k[1], k[2]), 0)]
    r.check("26", "Ningun agregado IDCampaña+APIE+TipoSoporte supera el inventario fisico disponible",
            not excede_inventario, f"excedidos={excede_inventario[:10]}")

    # -----------------------------------------------------------------
    # 5. MAPEO CAMPAÑAS / PENDIENTES CAMPAÑAS
    # -----------------------------------------------------------------
    h_mp, d_mp = leer_tabla(wb["MAPEO CAMPAÑAS"])
    r.check("27", "MAPEO CAMPAÑAS: 35 unidades bloque+pauta con IDCampaña reservado", len(d_mp) == 35, f"obtenido={len(d_mp)}")

    c_mp_estado = col(h_mp, "EstadoMapeo")
    c_mp_titulo = col(h_mp, "TituloBloqueOriginal")
    c_mp_id = col(h_mp, "IDCampaña")
    ids_excluidas = {row[c_mp_id] for row in d_mp if row[c_mp_estado] == "EXCLUIDA"}
    ids_pendientes = {row[c_mp_id] for row in d_mp if row[c_mp_estado] == "PENDIENTE"}
    r.check("28", "Las campañas EXCLUIDA/PENDIENTE de MAPEO no aparecen en BASE CAMPAÑAS",
            not ((ids_excluidas | ids_pendientes) & ids_cargables),
            f"interseccion={((ids_excluidas | ids_pendientes) & ids_cargables)}")

    fanta_netflix = [row[c_mp_titulo] for row in d_mp
                     if row[c_mp_estado] == "EXCLUIDA"
                     and ("FANTA" in str(row[c_mp_titulo]).upper() or "NETFLIX" in str(row[c_mp_titulo]).upper())]
    r.check("29", "Fanta y Netflix figuran como EXCLUIDA en MAPEO CAMPAÑAS", len(fanta_netflix) == 2, f"obtenido={fanta_netflix}")

    h_pc, d_pc = leer_tabla(wb["PENDIENTES CAMPAÑAS"])
    c_pc_tipo = col(h_pc, "TipoPendiente")
    c_pc_apie = col(h_pc, "APIE")
    c_pc_falt = col(h_pc, "CantidadFaltante")
    c_pc_camp = col(h_pc, "Campaña")

    faltante_total_pend = sum(row[c_pc_falt] for row in d_pc
                               if row[c_pc_tipo] == "FALTANTE_INVENTARIO" and row[c_pc_falt] is not None)
    r.check("30", "PENDIENTES CAMPAÑAS: faltante total de inventario = 163", faltante_total_pend == 163,
            f"obtenido={faltante_total_pend}")

    faltante_turismo = sum(row[c_pc_falt] for row in d_pc
                            if row[c_pc_tipo] == "FALTANTE_INVENTARIO" and row[c_pc_camp] == "CORDOBA TURISMO"
                            and row[c_pc_falt] is not None)
    r.check("31", "Turismo Cordoba explica 153 del faltante", faltante_turismo == 153, f"obtenido={faltante_turismo}")

    faltante_apie551 = sum(row[c_pc_falt] for row in d_pc
                            if row[c_pc_tipo] == "FALTANTE_INVENTARIO" and row[c_pc_apie] == "551"
                            and row[c_pc_falt] is not None)
    r.check("32", "APIE 551 explica 10 PPUNTER de faltante", faltante_apie551 == 10, f"obtenido={faltante_apie551}")

    n_noir_pend = sum(1 for row in d_pc if row[c_pc_tipo] == "FILA_EXCLUIDA_NO_IR")
    r.check("33", "PENDIENTES CAMPAÑAS: 2 filas FILA_EXCLUIDA_NO_IR", n_noir_pend == 2, f"obtenido={n_noir_pend}")

    r.check("34", "APIE 30943 no aparece en BASE CAMPAÑAS",
            "30943" not in {B.normalizar_apie(row[c["Subcircuito"]]) for row in d_bc}, "")

    n_apie_bloqueados_pend = sum(1 for row in d_pc if row[c_pc_tipo] == "APIE_BLOQUEADA")
    r.check("35", "0 pendientes APIE_BLOQUEADA (ningun APIE bloqueado referenciado en COMERCIAL)",
            n_apie_bloqueados_pend == 0, f"obtenido={n_apie_bloqueados_pend}")

    # -----------------------------------------------------------------
    # 6. Reconciliacion independiente contra las fuentes originales
    # -----------------------------------------------------------------
    print("Recalculando de forma independiente desde las fuentes originales (solo lectura)...")
    indep = recomputar_solicitud_asignacion_independiente()

    r.check("36", "Bloques detectados (recalculo independiente) = 31", indep["bloques"] == 31, f"obtenido={indep['bloques']}")
    r.check("37", "Unidades bloque+pauta (recalculo independiente) = 35", indep["unidades"] == 35, f"obtenido={indep['unidades']}")
    r.check("38", "IDCampaña cargables (recalculo independiente) = 27", indep["cargables"] == 27, f"obtenido={indep['cargables']}")
    r.check("39", "Filas de estacion fuente (recalculo independiente) = 2.371", indep["filas_fuente"] == 2371,
            f"obtenido={indep['filas_fuente']}")
    r.check("40", "Filas candidatas antes del cruce (recalculo independiente) = 2.034", indep["filas_candidatas"] == 2034,
            f"obtenido={indep['filas_candidatas']}")
    r.check("41", "Filas no-ir (recalculo independiente) = 2", indep["filas_noir"] == 2, f"obtenido={indep['filas_noir']}")
    r.check("42", "Filas cargables sin estacion (recalculo independiente) = 6", indep["filas_sin_estacion"] == 6,
            f"obtenido={indep['filas_sin_estacion']}")
    r.check("43", "Elementos solicitados (recalculo independiente) = 13.779", indep["solicitados"] == 13779,
            f"obtenido={indep['solicitados']}")
    r.check("44", "Elementos asignados (recalculo independiente) = 13.616", indep["asignados"] == 13616,
            f"obtenido={indep['asignados']}")
    r.check("45", "Faltante de inventario (recalculo independiente) = 163", indep["faltante"] == 163,
            f"obtenido={indep['faltante']}")

    r.check("46", "El recalculo independiente de elementos asignados coincide con BASE CAMPAÑAS",
            indep["asignados"] == len(d_bc), f"independiente={indep['asignados']} BASE CAMPAÑAS={len(d_bc)}")
    r.check("47", "El recalculo independiente de faltante coincide con PENDIENTES CAMPAÑAS",
            indep["faltante"] == faltante_total_pend, f"independiente={indep['faltante']} PENDIENTES={faltante_total_pend}")

    # -----------------------------------------------------------------
    # 7. RESUMEN CAMPAÑAS / AUDITORIA CAMPAÑAS presentes y sin ERROR
    # -----------------------------------------------------------------
    h_ac, d_ac = leer_tabla(wb["AUDITORIA CAMPAÑAS"])
    c_ac_estado = col(h_ac, "Estado")
    errores_ac = [row for row in d_ac if row[c_ac_estado] == "ERROR"]
    r.check("48", "AUDITORIA CAMPAÑAS no contiene ninguna fila en estado ERROR", not errores_ac,
            f"filas_error={len(errores_ac)}")

    h_rc, d_rc = leer_tabla(wb["RESUMEN CAMPAÑAS"])
    r.check("49", "RESUMEN CAMPAÑAS contiene filas de resumen", len(d_rc) > 20, f"obtenido={len(d_rc)}")

    core_data_ok = r.imprimir()
    print()
    print(f"CORE DATA: {'PASS' if core_data_ok else 'FAIL'}")
    exit_code = 0 if core_data_ok else 1
    print(f"VALIDATOR EXIT: {exit_code}")
    return exit_code


if __name__ == "__main__":
    sys.exit(main())
