"""Capa de datos para el dashboard TV5 - OCU26 (YPF).

Se ejecuta DESPUES de scripts/semantic_model.py y scripts/metrics_engine.py
(Gate 3B). Reutiliza export_data.load_pipeline (Gate 4A) y
MetricsEngine._campanas_overlap / MetricsEngine.query (mismo patron que
build_tv1_dashboard.py y build_tv6_dashboard.py): no reabre el Excel, no
reimplementa Gate 1/2/3.

Universo TV5 = exclusivamente CircuitoNegocio YPF (CM1 A13). No incorpora
Core no YPF, APSA ni London Supply.

Estaciones YPF: no existe columna APIE en la fuente. Se reutiliza el
surrogate ya auditado y vigente en TV1 (CM1 A13 / build_tv1_dashboard.py
_ypf_station_key): StationKey = prefijo numerico de ElementoID + localidad
normalizada de Ubicacion. Se llama explicitamente "StationKey (surrogate
temporal)", nunca "APIE".

Definiciones de negocio obligatorias (prompt Sec.4):
- Campanas unicas = distinct IDCampana en el periodo.
- Elementos activos = distinct ElementoID con >=1 campana en el periodo
  (metrica canonica del motor: elementos_con_actividad).
- Activaciones = distinct (IDCampana, ElementoID) en el periodo.
- Estaciones activas = distinct StationKey con >=1 activacion en el periodo.

Uso:
    python scripts/build_tv5_dashboard.py
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent))
import validate_input as vi  # noqa: E402
from semantic_model import filter_universe  # noqa: E402
from metrics_engine import MetricsEngine  # noqa: E402
from export_data import load_pipeline  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "tv5_template.html"
REFERENCE_PATH = REPO_ROOT / "audit_sources" / "TV5_REFERENCE.html"
DEFAULT_OUTPUT_HTML = REPO_ROOT / "tv5.html"
DEFAULT_OUTPUT_JSON = REPO_ROOT / "output" / "tv5_data.json"

# Cartografia real (Natural Earth, admin 1:10m), descargada UNA VEZ con
# autorizacion explicita del usuario y recortada a Argentina (ver docstring
# de compute_mapa mas abajo). Se inlinea en tv5.html en build time para que
# el kiosk quede 100% offline (sin fetch/CORS en runtime).
ARG_PAIS_GEOJSON_PATH = Path(__file__).resolve().parent / "templates" / "assets" / "argentina_pais.geojson"
ARG_PROVINCIAS_GEOJSON_PATH = Path(__file__).resolve().parent / "templates" / "assets" / "argentina_provincias.geojson"

# Fuente geografica auxiliar (ajuste post-entrega): NO forma parte de
# input/OCU26_BASE_DATOS.xlsx ni de Gate 1/2/3 (validate_input/transform_
# data/semantic_model/metrics_engine). Se lee de forma aislada, solo dentro
# de este builder, exclusivamente para el join geografico del mapa de TV5.
GEO_AUX_PATH = REPO_ROOT / "input_aux" / "YPF_GEO_COORDENADAS.xlsx.xlsx"

MESES_ES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]
MESES_ES_ABR = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

# Periodo de referencia TV5 (mismo mes de reporte vigente en TV1-6).
REPORT_YEAR = 2026
REPORT_MONTH = 7

# Formatos principales de negocio YPF (config/business_semantics.json
# formato_negocio.ypf_elemento_id_token_map): unicos 3 pedidos por spec.
# YPF_MUPI_FOTOBOX existe en catalogo pero no es uno de los 3 pedidos; se
# reporta aparte (Sec.6 "otros formatos ... reportar como A VALIDAR").
FORMATOS_PRINCIPALES = ["YPF_MENU_BOARD", "YPF_TORRE", "YPF_PUNTERA"]
FORMATO_LABELS = {
    "YPF_MENU_BOARD": "Menu Board",
    "YPF_TORRE": "Torres",
    "YPF_PUNTERA": "Punteras",
    "YPF_MUPI_FOTOBOX": "Mupi Fotobox",
}

class BuildError(Exception):
    """Error bloqueante al construir el dashboard TV5."""


def _period_bounds(year: int, month: int) -> tuple[str, str]:
    start = pd.Timestamp(year=year, month=month, day=1)
    end = start + pd.offsets.MonthEnd(0)
    return str(start.date()), str(end.date())


def _previous_month(year: int, month: int) -> tuple[int, int]:
    ts = pd.Timestamp(year=year, month=month, day=1) - pd.DateOffset(months=1)
    return ts.year, ts.month


def _fmt_es_int(n: int) -> str:
    return f"{n:,}".replace(",", ".")


def _fmt_es_pct(v: float | None) -> str:
    if v is None:
        return "S/D"
    return f"{v:.1f}".replace(".", ",")


def _round1(value: Any) -> float | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    return round(float(value), 1)


def _normalize_token(s: Any) -> str | None:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return None
    s = str(s).strip().upper()
    return re.sub(r"\s+", " ", s)


def _ypf_station_key(elemento_id: Any, ubicacion: Any) -> str | None:
    """Grano comercial de venta YPF (CM1 A13): 1 estacion = 1 unidad
    comercial. No existe columna APIE en la fuente -- surrogate ya auditado
    y vigente en TV1: (prefijo numerico de ElementoID, localidad normalizada
    de Ubicacion). Nunca cae a ElementoID en silencio: si no se puede
    derivar, devuelve None y el llamador debe bloquear el build (mismo
    contrato que build_tv1_dashboard.py)."""
    if elemento_id is None or (isinstance(elemento_id, float) and pd.isna(elemento_id)):
        return None
    prefix = str(elemento_id).split(" - ")[0].strip()
    if not prefix:
        return None
    parts = str(ubicacion).split(" - ") if ubicacion is not None and not (isinstance(ubicacion, float) and pd.isna(ubicacion)) else []
    town = _normalize_token(parts[1]) if len(parts) >= 2 else None
    if town is None:
        return None
    return f"{prefix}|{town}"


# ---------------------------------------------------------------------------
# Universo TV5 (exclusivamente YPF)
# ---------------------------------------------------------------------------


def build_tv5_universe(semantic_result: dict[str, Any]) -> dict[str, Any]:
    maestro = semantic_result["maestro"]
    config = semantic_result["config"]

    op = filter_universe(maestro, "OPERATIVO_GENERAL", config)
    ypf_maestro = op[op["CircuitoNegocio"] == "YPF"].copy()
    if ypf_maestro.empty:
        raise BuildError("Universo TV5 vacio: no hay elementos YPF en OPERATIVO_GENERAL")

    ypf_maestro["StationKey"] = [
        _ypf_station_key(eid, ubic) for eid, ubic in zip(ypf_maestro["ElementoID"], ypf_maestro["Ubicacion"])
    ]
    null_station = ypf_maestro[ypf_maestro["StationKey"].isna()]
    if len(null_station):
        raise BuildError(
            f"{len(null_station)} elemento(s) YPF sin estacion (StationKey surrogate) derivable a partir de "
            f"ElementoID/Ubicacion: {sorted(null_station['ElementoID'].unique().tolist())[:10]}. "
            f"No se aplica fallback silencioso a ElementoID."
        )

    station_map = dict(zip(ypf_maestro["ElementoID"], ypf_maestro["StationKey"]))
    city_map = dict(zip(ypf_maestro["ElementoID"], ypf_maestro["Ciudad"]))
    station_city_map = dict(zip(ypf_maestro["StationKey"], ypf_maestro["Ciudad"]))
    fmt_map = dict(zip(ypf_maestro["ElementoID"], ypf_maestro["FormatoNegocio"]))
    element_ids = ypf_maestro["ElementoID"].tolist()

    # Prefijo numerico de estacion (= Subcircuito = primer token de
    # ElementoID; identico por construccion para todos los ElementoID de
    # una misma StationKey): clave de join contra la fuente geografica
    # auxiliar (ver compute_mapa / load_geo_aux).
    station_prefix_map = dict(zip(ypf_maestro["StationKey"], ypf_maestro["Subcircuito"].astype(str)))

    fmt_counts = ypf_maestro["FormatoNegocio"].value_counts().to_dict()
    formatos_no_principales = {k: int(v) for k, v in fmt_counts.items() if k not in FORMATOS_PRINCIPALES}

    return {
        "maestro": ypf_maestro,
        "element_ids": element_ids,
        "station_map": station_map,
        "city_map": city_map,
        "station_city_map": station_city_map,
        "station_prefix_map": station_prefix_map,
        "fmt_map": fmt_map,
        "estaciones_catalogo": int(ypf_maestro["StationKey"].nunique()),
        "elementos_catalogo": int(ypf_maestro["ElementoID"].nunique()),
        "formatos_no_principales": formatos_no_principales,
    }


def _campanas_overlap_validas(engine: MetricsEngine, element_ids: list[Any], start: str, end: str) -> pd.DataFrame:
    """CAMPANAS con solapamiento valido contra [start,end] y IDCampana
    informado (no vacio): base comun para campanas unicas y activaciones
    (prompt Sec.4)."""
    overlap = engine._campanas_overlap(element_ids, start, end)
    if overlap.empty:
        return overlap
    return overlap[overlap["IDCampaña"].notna() & (overlap["IDCampaña"].astype(str).str.strip() != "")]


def _active_stations(overlap: pd.DataFrame, station_map: dict[Any, str]) -> set[str]:
    active_ids = overlap["ElementoID"].dropna().unique().tolist()
    return {station_map[eid] for eid in active_ids if eid in station_map}


# ---------------------------------------------------------------------------
# Card 1 - Catalogo YPF (estructural, no depende de periodo)
# ---------------------------------------------------------------------------


def compute_catalogo(universe: dict[str, Any]) -> dict[str, Any]:
    return {
        "estaciones_registradas": universe["estaciones_catalogo"],
        "elementos_totales": universe["elementos_catalogo"],
        "status": "catalogo_actual",
        "metric_status": "NO_APLICA",
        "nota": (
            "Catálogo actual. Sin histórico de catálogo disponible: la fuente es un snapshot vigente, "
            "no permite reconstruir cuántas estaciones/elementos existían en junio."
        ),
    }


# ---------------------------------------------------------------------------
# Card 2 - Estaciones activas
# ---------------------------------------------------------------------------


def compute_estaciones_activas(
    engine: MetricsEngine, universe: dict[str, Any], period: tuple[str, str], previous: tuple[str, str],
) -> dict[str, Any]:
    element_ids = universe["element_ids"]
    station_map = universe["station_map"]
    catalogo = universe["estaciones_catalogo"]

    overlap_actual = _campanas_overlap_validas(engine, element_ids, period[0], period[1])
    overlap_anterior = _campanas_overlap_validas(engine, element_ids, previous[0], previous[1])

    actual = len(_active_stations(overlap_actual, station_map))
    anterior = len(_active_stations(overlap_anterior, station_map))

    pct_actual = _round1(actual / catalogo * 100.0) if catalogo else None
    pct_anterior = _round1(anterior / catalogo * 100.0) if catalogo else None
    delta_pp = _round1(pct_actual - pct_anterior) if pct_actual is not None and pct_anterior is not None else None

    return {
        "actual": actual,
        "anterior": anterior,
        "catalogo": catalogo,
        "pct_actual": pct_actual,
        "pct_anterior": pct_anterior,
        "delta_abs": actual - anterior,
        "delta_pp": delta_pp,
    }


# ---------------------------------------------------------------------------
# Card 3 - Campanas (campanas unicas + activaciones)
# ---------------------------------------------------------------------------


def compute_campanas(
    engine: MetricsEngine, universe: dict[str, Any], period: tuple[str, str], previous: tuple[str, str],
) -> dict[str, Any]:
    element_ids = universe["element_ids"]

    overlap_actual = _campanas_overlap_validas(engine, element_ids, period[0], period[1])
    overlap_anterior = _campanas_overlap_validas(engine, element_ids, previous[0], previous[1])

    campanas_actual = int(overlap_actual["IDCampaña"].nunique())
    campanas_anterior = int(overlap_anterior["IDCampaña"].nunique())
    activaciones_actual = int(overlap_actual.drop_duplicates(subset=["IDCampaña", "ElementoID"]).shape[0])
    activaciones_anterior = int(overlap_anterior.drop_duplicates(subset=["IDCampaña", "ElementoID"]).shape[0])

    promedio_actual = _round1(activaciones_actual / campanas_actual) if campanas_actual else None
    promedio_anterior = _round1(activaciones_anterior / campanas_anterior) if campanas_anterior else None

    return {
        "campanas_unicas_actual": campanas_actual,
        "campanas_unicas_anterior": campanas_anterior,
        "delta_campanas": campanas_actual - campanas_anterior,
        "activaciones_actual": activaciones_actual,
        "activaciones_anterior": activaciones_anterior,
        "delta_activaciones": activaciones_actual - activaciones_anterior,
        "promedio_actual": promedio_actual,
        "promedio_anterior": promedio_anterior,
    }


# ---------------------------------------------------------------------------
# Card 4 - Elementos activos (metrica canonica del motor)
# ---------------------------------------------------------------------------


def compute_elementos_activos(
    engine: MetricsEngine, period: tuple[str, str], previous: tuple[str, str],
) -> dict[str, Any]:
    actual = engine.query("elementos_con_actividad", filters={"CircuitoNegocio": "YPF"}, start_date=period[0], end_date=period[1])
    anterior = engine.query("elementos_con_actividad", filters={"CircuitoNegocio": "YPF"}, start_date=previous[0], end_date=previous[1])
    val_actual = int(actual["Value"].iloc[0])
    val_anterior = int(anterior["Value"].iloc[0])
    return {
        "actual": val_actual,
        "anterior": val_anterior,
        "delta": val_actual - val_anterior,
        "metric_status": actual["MetricStatus"].iloc[0],
    }


# ---------------------------------------------------------------------------
# Card 5 - Formato lider (por campanas unicas, NUNCA por activaciones)
# ---------------------------------------------------------------------------


def _formato_campanas_unicas(overlap: pd.DataFrame, fmt_map: dict[Any, str]) -> dict[str, int]:
    if overlap.empty:
        return {f: 0 for f in FORMATOS_PRINCIPALES}
    tagged = overlap.copy()
    tagged["_fmt"] = tagged["ElementoID"].map(fmt_map)
    tagged = tagged[tagged["_fmt"].isin(FORMATOS_PRINCIPALES)]
    counts = tagged.groupby("_fmt")["IDCampaña"].nunique().to_dict()
    return {f: int(counts.get(f, 0)) for f in FORMATOS_PRINCIPALES}


def _formato_elementos_activos(overlap: pd.DataFrame, fmt_map: dict[Any, str]) -> dict[str, int]:
    if overlap.empty:
        return {f: 0 for f in FORMATOS_PRINCIPALES}
    tagged = overlap.copy()
    tagged["_fmt"] = tagged["ElementoID"].map(fmt_map)
    tagged = tagged[tagged["_fmt"].isin(FORMATOS_PRINCIPALES)]
    counts = tagged.groupby("_fmt")["ElementoID"].nunique().to_dict()
    return {f: int(counts.get(f, 0)) for f in FORMATOS_PRINCIPALES}


def _pick_formato_lider(campanas_unicas: dict[str, int], elementos_activos: dict[str, int]) -> dict[str, Any]:
    max_val = max(campanas_unicas.values()) if campanas_unicas else 0
    empatados = [f for f, v in campanas_unicas.items() if v == max_val and max_val > 0]
    empate = len(empatados) > 1
    if not empatados:
        ganador = None
    elif empate:
        ganador = max(empatados, key=lambda f: elementos_activos.get(f, 0))
    else:
        ganador = empatados[0]
    return {
        "formato": ganador,
        "label": FORMATO_LABELS.get(ganador, "Sin dato") if ganador else "Sin dato",
        "campanas_unicas": campanas_unicas.get(ganador, 0) if ganador else 0,
        "elementos_activos": elementos_activos.get(ganador, 0) if ganador else 0,
        "empate_resuelto_por_elementos_activos": empate,
        "detalle_campanas_unicas": campanas_unicas,
    }


def compute_formato_lider(
    engine: MetricsEngine, universe: dict[str, Any], period: tuple[str, str], previous: tuple[str, str],
) -> dict[str, Any]:
    element_ids = universe["element_ids"]
    fmt_map = universe["fmt_map"]

    overlap_actual = _campanas_overlap_validas(engine, element_ids, period[0], period[1])
    overlap_anterior = _campanas_overlap_validas(engine, element_ids, previous[0], previous[1])

    camp_actual = _formato_campanas_unicas(overlap_actual, fmt_map)
    camp_anterior = _formato_campanas_unicas(overlap_anterior, fmt_map)
    elem_actual = _formato_elementos_activos(overlap_actual, fmt_map)
    elem_anterior = _formato_elementos_activos(overlap_anterior, fmt_map)

    return {
        "actual": _pick_formato_lider(camp_actual, elem_actual),
        "anterior": _pick_formato_lider(camp_anterior, elem_anterior),
    }


# ---------------------------------------------------------------------------
# Historico Ene-Jul: activaciones por formato principal
# ---------------------------------------------------------------------------


def compute_historico(engine: MetricsEngine, universe: dict[str, Any], report_year: int, report_month: int) -> dict[str, Any]:
    element_ids = universe["element_ids"]
    fmt_map = universe["fmt_map"]
    meses = MESES_ES_ABR[:report_month]
    series: dict[str, list[int]] = {f: [] for f in FORMATOS_PRINCIPALES}
    otros_mensual: list[int] = []

    for m in range(1, report_month + 1):
        start, end = _period_bounds(report_year, m)
        overlap = _campanas_overlap_validas(engine, element_ids, start, end)
        activ = overlap.drop_duplicates(subset=["IDCampaña", "ElementoID"]).copy()
        if activ.empty:
            for f in FORMATOS_PRINCIPALES:
                series[f].append(0)
            otros_mensual.append(0)
            continue
        activ["_fmt"] = activ["ElementoID"].map(fmt_map)
        counts = activ.groupby("_fmt").size().to_dict()
        for f in FORMATOS_PRINCIPALES:
            series[f].append(int(counts.get(f, 0)))
        otros = sum(v for k, v in counts.items() if k not in FORMATOS_PRINCIPALES)
        otros_mensual.append(int(otros))

    return {
        "meses": meses,
        "series": {FORMATO_LABELS[f]: series[f] for f in FORMATOS_PRINCIPALES},
        "otros_formatos_activaciones_mensual": otros_mensual,
    }


# ---------------------------------------------------------------------------
# Mapa de intensidad comercial (mapa geografico real, ajuste post-entrega).
#
# input/OCU26_BASE_DATOS.xlsx NO tiene latitud/longitud (auditado: headers
# crudos de las 3 hojas + columnas resueltas del maestro semantico). El
# usuario incorporo una fuente geografica auxiliar REAL fuera del pipeline
# central: input_aux/YPF_GEO_COORDENADAS.xlsx.xlsx (hoja "Hoja1", columnas
# CODIGO/LATITUD/LONGITUD/CIUDAD*/DIRECCION*). Se lee de forma aislada, solo
# en este builder (load_geo_aux): NO se modifica input/OCU26_BASE_DATOS.xlsx
# ni ningun Gate, y NO se incorpora todavia al pipeline central.
#
# Join: prefijo numerico de CODIGO (ej. "107-FB 1" -> "107") contra
# Subcircuito/prefijo de ElementoID de OCU26 (identicos por construccion,
# ver build_tv5_universe.station_prefix_map). Auditoria de cobertura sobre
# el archivo productivo (10/8/2026): de 1.508 filas de CODIGO con prefijo
# numerico, 429 tienen lat/lon numericamente validos (rango -90..90/
# -180..180, no 0,0); las 1.079 restantes tienen valores claramente mal
# formateados (ej. "-346158286" en vez de "-34.6158286") -- se EXCLUYEN,
# nunca se corrigen por suposicion. Con eso: 203 de las 451 estaciones del
# catalogo YPF (45.0%) y 79 de las 305 estaciones activas de julio (25.9%)
# obtienen coordenada valida. Cobertura parcial real, documentada como tal
# (MetricStatus PARTIAL): nunca se inventa una posicion para el resto.
#
# Base cartografica (ajuste post-entrega, autorizado explicitamente por el
# usuario): la silueta dibujada a mano fue rechazada por no ser reconocible
# como Argentina/AMBA/CABA. Se descargo UNA VEZ (con permiso) la fuente
# publica Natural Earth "Admin 1: States, Provinces" 1:10m (dominio publico)
# desde https://github.com/nvkelso/natural-earth-vector (mismo dataset
# oficial de Natural Earth, geojson pre-convertido), junto con "Admin 0:
# Countries" 1:10m para el contorno pais. Se filtro exclusivamente
# Argentina (24 features admin1 = 23 provincias + CABA; 1 feature admin0) y
# se guardo localmente en scripts/templates/assets/argentina_pais.geojson y
# argentina_provincias.geojson (coordenadas redondeadas a 4 decimales, sin
# alterar la forma real). render_html() los inlinea en tv5.html: el kiosk
# no depende de internet, tiles ni librerias de mapas en runtime.
# ---------------------------------------------------------------------------

_GEO_PREFIX_RE = re.compile(r"^\s*(\d+)\s*-")


def load_geo_aux(path: Path = GEO_AUX_PATH) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    """Lee la fuente geografica auxiliar YPF (aislada de Gate 1/2/3) y arma
    prefijo_numerico -> {lat, lon, ciudad, direccion}, quedandose con la
    primera fila valida por prefijo (auditado: 0 prefijos con coordenadas
    validas en conflicto entre si). Devuelve tambien el detalle de calidad
    del join para reportarlo en el payload (Sec.3F)."""
    if not path.exists():
        raise BuildError(f"Fuente geografica auxiliar no encontrada: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Hoja1"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i for i, name in enumerate(headers)}
    for required in ("CODIGO", "LATITUD", "LONGITUD"):
        if required not in col_idx:
            raise BuildError(f"Columna requerida {required!r} no encontrada en {path} (Hoja1)")

    total_filas = 0
    con_prefijo = 0
    filas_validas = 0
    filas_invalidas = 0
    geo_map: dict[str, dict[str, Any]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = row[col_idx["CODIGO"]]
        if codigo is None:
            continue
        total_filas += 1
        m = _GEO_PREFIX_RE.match(str(codigo))
        if not m:
            continue
        con_prefijo += 1
        prefix = m.group(1)

        lat_raw = row[col_idx["LATITUD"]]
        lon_raw = row[col_idx["LONGITUD"]]
        try:
            lat, lon = float(lat_raw), float(lon_raw)
        except (TypeError, ValueError):
            filas_invalidas += 1
            continue
        if not (-90.0 <= lat <= 90.0 and -180.0 <= lon <= 180.0) or (lat == 0 and lon == 0):
            filas_invalidas += 1
            continue

        filas_validas += 1
        if prefix in geo_map:
            continue  # primera fila valida del prefijo gana (auditado: sin conflicto real)
        geo_map[prefix] = {
            "lat": lat,
            "lon": lon,
            "ciudad": row[col_idx.get("CIUDAD*", -1)] if "CIUDAD*" in col_idx else None,
            "direccion": row[col_idx.get("DIRECCION*", -1)] if "DIRECCION*" in col_idx else None,
        }

    calidad = {
        "fuente": str(path.relative_to(REPO_ROOT)),
        "filas_totales": total_filas,
        "filas_con_prefijo_numerico": con_prefijo,
        "filas_lat_lon_validas": filas_validas,
        "filas_lat_lon_invalidas_excluidas": filas_invalidas,
        "prefijos_distintos_con_coordenada_valida": len(geo_map),
    }
    return geo_map, calidad


def compute_mapa(
    engine: MetricsEngine, universe: dict[str, Any], report_year: int, report_month: int, geo_map: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    element_ids = universe["element_ids"]
    station_map = universe["station_map"]
    station_city_map = universe["station_city_map"]
    station_prefix_map = universe["station_prefix_map"]

    activ_frames = []
    for m in range(1, report_month + 1):
        start, end = _period_bounds(report_year, m)
        overlap = _campanas_overlap_validas(engine, element_ids, start, end)
        activ = overlap.drop_duplicates(subset=["IDCampaña", "ElementoID"]).copy()
        if activ.empty:
            continue
        activ["_station"] = activ["ElementoID"].map(station_map)
        activ_frames.append(activ[["IDCampaña", "ElementoID", "_station"]])

    if not activ_frames:
        acumulado = pd.DataFrame(columns=["IDCampaña", "ElementoID", "_station"])
    else:
        acumulado = pd.concat(activ_frames, ignore_index=True)
    acumulado = acumulado[acumulado["_station"].notna()].copy()

    by_station = acumulado.groupby("_station").size()
    n_estaciones_con_actividad = int(by_station.shape[0])

    puntos = []
    for station, activaciones in by_station.items():
        prefix = station_prefix_map.get(station)
        geo = geo_map.get(prefix) if prefix is not None else None
        if geo is None:
            continue
        puntos.append({
            "lat": geo["lat"],
            "lon": geo["lon"],
            "activaciones": int(activaciones),
            "ciudad": geo.get("ciudad") or station_city_map.get(station),
        })

    n_con_geo = len(puntos)
    pct_cobertura = _round1(n_con_geo / n_estaciones_con_actividad * 100.0) if n_estaciones_con_actividad else None

    bounds = None
    if puntos:
        lats = [p["lat"] for p in puntos]
        lons = [p["lon"] for p in puntos]
        bounds = {"lat_min": min(lats), "lat_max": max(lats), "lon_min": min(lons), "lon_max": max(lons)}

    return {
        "titulo": "Mapa de intensidad comercial",
        "subtitulo": "Activaciones acumuladas por estación / territorio",
        "periodo_label": f"Ene–{MESES_ES_ABR[report_month-1]} {report_year}",
        "estado": "PARTIAL",
        "metric_status": "PARTIAL",
        "nota": (
            f"{_fmt_es_int(n_con_geo)} de {_fmt_es_int(n_estaciones_con_actividad)} estaciones con actividad "
            f"Ene–{MESES_ES_ABR[report_month-1]} cuentan con ubicación geográfica disponible (fuente auxiliar "
            f"YPF_GEO_COORDENADAS, aún fuera del pipeline central; join por prefijo numérico de estación). "
            f"El resto no tiene coordenada válida en esa fuente: no se inventan posiciones ni se geocodifica "
            f"externamente."
        ),
        "leyenda_min": "Menor actividad",
        "leyenda_max": "Mayor actividad",
        "puntos": puntos,
        "bounds": bounds,
        "n_estaciones_con_actividad": n_estaciones_con_actividad,
        "n_estaciones_con_geo": n_con_geo,
        "pct_cobertura_geo": pct_cobertura,
    }


# ---------------------------------------------------------------------------
# Calidad de datos
# ---------------------------------------------------------------------------


def compute_calidad(
    universe: dict[str, Any], semantic_result: dict[str, Any], geo_calidad: dict[str, Any], mapa: dict[str, Any],
) -> dict[str, Any]:
    warnings = semantic_result.get("warnings", [])
    ypf_warnings = [str(w) for w in warnings if "YPF" in str(w)]
    return {
        "estaciones_sin_clave_derivable": 0,
        "formatos_no_principales_en_catalogo": {
            FORMATO_LABELS.get(k, k): v for k, v in universe["formatos_no_principales"].items()
        },
        "warnings_semanticos_ypf": ypf_warnings,
        "nota": (
            "Sin nulos en ElementoID/Ubicacion/Ciudad para el universo YPF ni advertencias de formato "
            "no reconocido en el snapshot auditado."
            if not ypf_warnings
            else f"{len(ypf_warnings)} advertencia(s) semántica(s) YPF detectada(s) en el snapshot."
        ),
        "geo_aux": {
            **geo_calidad,
            "estaciones_con_actividad_ene_jul": mapa["n_estaciones_con_actividad"],
            "estaciones_con_actividad_con_geo_valida": mapa["n_estaciones_con_geo"],
            "pct_cobertura_geo_sobre_activas": mapa["pct_cobertura_geo"],
            "nota": (
                f"{_fmt_es_int(geo_calidad['filas_lat_lon_invalidas_excluidas'])} filas de la fuente auxiliar "
                f"tienen lat/lon claramente mal formateados (ej. sin punto decimal) y se excluyeron del join; "
                f"nunca se corrigieron por suposición."
            ),
        },
    }


# ---------------------------------------------------------------------------
# Insights (Lectura / Punto positivo / A atender)
# ---------------------------------------------------------------------------


def compute_insights(
    catalogo: dict[str, Any], estaciones: dict[str, Any], campanas: dict[str, Any],
    elementos: dict[str, Any], formato_lider: dict[str, Any], mapa: dict[str, Any],
    report_month_label: str, previous_month_label: str,
) -> dict[str, str]:
    fl_actual = formato_lider["actual"]
    fl_anterior = formato_lider["anterior"]

    lectura = (
        f"{report_month_label} registra <b>{_fmt_es_int(estaciones['actual'])} estaciones activas</b> "
        f"({_fmt_es_pct(estaciones['pct_actual'])}% del catálogo de {_fmt_es_int(catalogo['estaciones_registradas'])}), "
        f"con <b>{_fmt_es_int(campanas['campanas_unicas_actual'])} campañas únicas</b> que generaron "
        f"<b>{_fmt_es_int(campanas['activaciones_actual'])} activaciones</b> sobre "
        f"{_fmt_es_int(elementos['actual'])} elementos activos. El formato líder es "
        f"<b>{fl_actual['label']}</b> ({_fmt_es_int(fl_actual['campanas_unicas'])} campañas únicas), "
        f"frente a {previous_month_label.lower()} ({_fmt_es_int(estaciones['anterior'])} estaciones, "
        f"{_fmt_es_int(campanas['campanas_unicas_anterior'])} campañas, "
        f"{_fmt_es_int(campanas['activaciones_anterior'])} activaciones, formato líder {fl_anterior['label']})."
    )

    # PUNTO POSITIVO: prioridad 1) crecimiento estaciones activas.
    if estaciones["delta_abs"] > 0:
        punto_positivo = (
            f"Las estaciones activas crecieron de <b>{_fmt_es_int(estaciones['anterior'])} a "
            f"{_fmt_es_int(estaciones['actual'])}</b> ({_signed_int(estaciones['delta_abs'])} estaciones), "
            f"con una penetración sobre catálogo que subió de {_fmt_es_pct(estaciones['pct_anterior'])}% a "
            f"{_fmt_es_pct(estaciones['pct_actual'])}% ({_signed_pp(estaciones['delta_pp'])})."
        )
    elif campanas["delta_campanas"] > 0:
        punto_positivo = (
            f"Las campañas únicas crecieron de {_fmt_es_int(campanas['campanas_unicas_anterior'])} a "
            f"<b>{_fmt_es_int(campanas['campanas_unicas_actual'])}</b> respecto a {previous_month_label.lower()}."
        )
    elif elementos["delta"] > 0:
        punto_positivo = (
            f"Los elementos activos crecieron de {_fmt_es_int(elementos['anterior'])} a "
            f"<b>{_fmt_es_int(elementos['actual'])}</b> respecto a {previous_month_label.lower()}."
        )
    else:
        punto_positivo = (
            f"La red YPF mantiene <b>{_fmt_es_int(estaciones['actual'])} estaciones activas</b> "
            f"sobre un catálogo de {_fmt_es_int(catalogo['estaciones_registradas'])}."
        )

    # A ATENDER: prioridad 1) concentracion excesiva en un formato.
    total_camp = fl_actual["detalle_campanas_unicas"]
    total_camp_sum = sum(total_camp.values()) or 1
    lider_share = round(fl_actual["campanas_unicas"] / total_camp_sum * 100.0, 1)
    if lider_share >= 55:
        a_atender = (
            f"El formato <b>{fl_actual['label']}</b> concentra {_fmt_es_pct(lider_share)}% de las campañas "
            f"únicas de {report_month_label.lower()} ({_fmt_es_int(fl_actual['campanas_unicas'])} de "
            f"{_fmt_es_int(total_camp_sum)}): la actividad queda poco diversificada entre formatos."
        )
    elif estaciones["pct_actual"] is not None and estaciones["pct_actual"] < 70:
        a_atender = (
            f"Solo el {_fmt_es_pct(estaciones['pct_actual'])}% del catálogo de estaciones tuvo campaña en "
            f"{report_month_label.lower()}: queda margen de penetración comercial sobre la red YPF."
        )
    else:
        a_atender = (
            f"El mapa de intensidad comercial solo geolocaliza <b>{_fmt_es_pct(mapa['pct_cobertura_geo'])}%</b> "
            f"de las estaciones con actividad ({_fmt_es_int(mapa['n_estaciones_con_geo'])} de "
            f"{_fmt_es_int(mapa['n_estaciones_con_actividad'])}): falta ampliar la cobertura de la fuente "
            f"geográfica auxiliar sobre el resto de la red."
        )

    return {"lectura": lectura, "punto_positivo": punto_positivo, "a_atender": a_atender}


def _signed_int(v: int) -> str:
    return (f"+{v}" if v > 0 else str(v))


def _signed_pp(v: float | None) -> str:
    if v is None:
        return "S/D"
    s = "+" if v > 0 else ""
    return f"{s}{_fmt_es_pct(v)} pp"


# ---------------------------------------------------------------------------
# Logo (reutilizado byte-a-byte, igual patron que TV1-4/6)
# ---------------------------------------------------------------------------


def extract_logo_img_tag(reference_path: Path) -> str:
    html = reference_path.read_text(encoding="utf-8")
    marker = '<img src="data:image/png;base64,'
    start = html.find(marker)
    if start == -1:
        raise BuildError(f"No se encontro el logo embebido en {reference_path}")
    end = html.find(">", start)
    if end == -1:
        raise BuildError(f"Tag <img> de logo mal formado en {reference_path}")
    return html[start : end + 1]


# ---------------------------------------------------------------------------
# Orquestacion
# ---------------------------------------------------------------------------


def build_tv5_data(path: str | Path = vi.DEFAULT_INPUT_PATH) -> dict[str, Any]:
    path = Path(path)
    sha_before = vi.calculate_sha256(path)

    _transform_result, semantic_result, engine = load_pipeline(path)
    universe = build_tv5_universe(semantic_result)

    period = _period_bounds(REPORT_YEAR, REPORT_MONTH)
    prev_year, prev_month = _previous_month(REPORT_YEAR, REPORT_MONTH)
    previous = _period_bounds(prev_year, prev_month)

    catalogo = compute_catalogo(universe)
    estaciones = compute_estaciones_activas(engine, universe, period, previous)
    campanas = compute_campanas(engine, universe, period, previous)
    elementos = compute_elementos_activos(engine, period, previous)
    formato_lider = compute_formato_lider(engine, universe, period, previous)
    historico = compute_historico(engine, universe, REPORT_YEAR, REPORT_MONTH)
    geo_map, geo_calidad = load_geo_aux()
    mapa = compute_mapa(engine, universe, REPORT_YEAR, REPORT_MONTH, geo_map)
    calidad = compute_calidad(universe, semantic_result, geo_calidad, mapa)

    report_month_label = MESES_ES[REPORT_MONTH - 1]
    previous_month_label = MESES_ES[prev_month - 1]
    insights = compute_insights(catalogo, estaciones, campanas, elementos, formato_lider, mapa, report_month_label, previous_month_label)

    sha_after = vi.calculate_sha256(path)
    if sha_after != sha_before:
        raise BuildError(
            f"ERROR CRITICO: el SHA-256 del input cambio durante la construccion del dashboard "
            f"(antes={sha_before}, despues={sha_after})."
        )

    data = {
        "meta": {
            "generado": dt.datetime.now().strftime("%d/%m/%Y %H:%M"),
            "report_year": REPORT_YEAR,
            "report_month": REPORT_MONTH,
            "report_month_label": report_month_label,
            "previous_month": prev_month,
            "previous_month_label": previous_month_label,
            "periodo_inicio_iso": period[0],
            "periodo_fin_iso": period[1],
            "previous_inicio_iso": previous[0],
            "previous_fin_iso": previous[1],
            "hist_label": f"Ene–{MESES_ES_ABR[REPORT_MONTH-1]} {REPORT_YEAR}",
            "fuente": "OCU26 · Base maestra + base campañas",
            "station_key_method": (
                "StationKey (surrogate temporal): prefijo numérico de ElementoID + localidad normalizada "
                "de Ubicación. No existe columna APIE en la fuente."
            ),
        },
        "universo": {
            "circuito": "YPF",
            "estaciones_catalogo": universe["estaciones_catalogo"],
            "elementos_catalogo": universe["elementos_catalogo"],
        },
        "kpis": {
            "catalogo": catalogo,
            "estaciones_activas": estaciones,
            "campanas": campanas,
            "elementos_activos": elementos,
            "formato_lider": formato_lider,
        },
        "historico": historico,
        "mapa": mapa,
        "calidad": calidad,
        "insights": insights,
    }

    return {"data": data, "sha256": sha_after, "universe": {"circuito": "YPF", "estaciones_catalogo": universe["estaciones_catalogo"], "elementos_catalogo": universe["elementos_catalogo"]}}


def render_html(data: dict[str, Any]) -> str:
    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    logo_tag = extract_logo_img_tag(REFERENCE_PATH)
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    if not ARG_PAIS_GEOJSON_PATH.exists() or not ARG_PROVINCIAS_GEOJSON_PATH.exists():
        raise BuildError(
            f"Cartografia real de Argentina no encontrada en {ARG_PAIS_GEOJSON_PATH} / "
            f"{ARG_PROVINCIAS_GEOJSON_PATH}."
        )
    arg_pais_geojson = ARG_PAIS_GEOJSON_PATH.read_text(encoding="utf-8")
    arg_provincias_geojson = ARG_PROVINCIAS_GEOJSON_PATH.read_text(encoding="utf-8")
    html = template.replace("{{LOGO_IMG_TAG}}", logo_tag)
    html = html.replace("{{TV5_DATA_JSON}}", payload)
    html = html.replace("{{ARG_PAIS_GEOJSON}}", arg_pais_geojson)
    html = html.replace("{{ARG_PROVINCIAS_GEOJSON}}", arg_provincias_geojson)
    return html


def build_and_write(
    path: str | Path = vi.DEFAULT_INPUT_PATH,
    output_html: str | Path = DEFAULT_OUTPUT_HTML,
    output_json: str | Path = DEFAULT_OUTPUT_JSON,
) -> dict[str, Any]:
    result = build_tv5_data(path)
    html = render_html(result["data"])

    output_html = Path(output_html)
    output_html.write_text(html, encoding="utf-8")

    output_json = Path(output_json)
    output_json.parent.mkdir(parents=True, exist_ok=True)
    with open(output_json, "w", encoding="utf-8") as fh:
        json.dump(result["data"], fh, ensure_ascii=False, indent=2)

    return result


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Construye tv5.html (dashboard TV5 OCU26, YPF) con datos reales.")
    parser.add_argument("--file", default=str(vi.DEFAULT_INPUT_PATH), help="Ruta al archivo .xlsx a leer")
    parser.add_argument("--output-html", default=str(DEFAULT_OUTPUT_HTML), help="Ruta del HTML productivo generado")
    parser.add_argument("--output-json", default=str(DEFAULT_OUTPUT_JSON), help="Ruta del snapshot JSON generado")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = build_and_write(args.file, args.output_html, args.output_json)
    except BuildError as exc:
        print("TV5_BUILD_ERROR:", exc)
        return 1

    print("TV5_BUILD_OK")
    print(json.dumps(result["data"]["meta"], ensure_ascii=False, indent=2))
    print(json.dumps(result["data"]["kpis"], ensure_ascii=False, indent=2))
    print("UNIVERSE:", json.dumps(result["universe"], ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
