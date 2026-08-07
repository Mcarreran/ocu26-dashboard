"""Capa de transformación en memoria para el pipeline OCU26.

Se ejecuta DESPUÉS de scripts/validate_input.py (Gate 1) y ANTES de cualquier
exportador (Power BI / JSON / HTML). Reutiliza validate_input() como gate
obligatorio: si el input es INVALID, la transformación aborta.

Estrictamente READ-ONLY sobre el archivo Excel: nunca usa Workbook.save(),
ExcelWriter, ni escribe celdas. Todo el trabajo ocurre en memoria y se
verifica al final que el SHA-256 del archivo fuente no haya cambiado.

No depende en runtime de audit_sources/: la regla de TipoInventario se
reprodujo a partir de la fórmula documentada en
audit_sources/INFORME_RECONCILIACION_MIGRACION.md (Sección 7), pero ese
archivo no se lee ni se referencia en este script.

Uso:
    python scripts/transform_data.py
    python scripts/transform_data.py --file ruta/al/archivo.xlsx
    python scripts/transform_data.py --json
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

sys.path.insert(0, str(Path(__file__).resolve().parent))
import validate_input as vi  # noqa: E402

DEFAULT_INPUT_PATH = vi.DEFAULT_INPUT_PATH

EXPECTED_TABLES = vi.EXPECTED_TABLES
TIPO_INVENTARIO_VALUES = vi.TIPO_INVENTARIO_VALUES
APLICA_CANTIDAD_VALUES = vi.APLICA_CANTIDAD_VALUES

# Reglas recuperadas de la V3 (ver audit_sources/INFORME_RECONCILIACION_MIGRACION.md,
# Sección 7). No se leen en runtime; se reproducen aquí de forma literal.
TIPO_INVENTARIO_KEYWORDS = ("CARRO", "STOPPER", "FLOORGRAPHIC", "CUBRE", "ALARMA")
TIPO_INVENTARIO_SEARCH_FIELDS = ("Ubicacion", "Descripcion", "Subcircuito", "CircuitoDashboard")


class TransformError(Exception):
    """Error bloqueante durante la transformación (gate fallido, contradicción, invariante rota)."""


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if value is pd.NA:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def read_excel_table(path: str | Path, sheet_name: str, table_name: str) -> pd.DataFrame:
    """Lee una tabla estructurada de Excel usando su rango real (ws.tables),
    sin depender de cantidades de filas hardcodeadas. Estrictamente lectura."""
    path = Path(path)
    wb = load_workbook(path, data_only=True, read_only=False, keep_vba=False)
    try:
        if sheet_name not in wb.sheetnames:
            raise TransformError(f"La hoja '{sheet_name}' no existe en {path}")
        ws = wb[sheet_name]
        if table_name not in ws.tables:
            raise TransformError(f"La tabla '{table_name}' no existe en la hoja '{sheet_name}' de {path}")

        table = ws.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)

        headers = list(
            next(
                ws.iter_rows(
                    min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_col, values_only=True
                )
            )
        )
        data_rows = list(
            ws.iter_rows(
                min_row=min_row + 1, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
            )
        )
    finally:
        wb.close()

    return pd.DataFrame(data_rows, columns=headers)


def derive_tipo_inventario(row: Any) -> str:
    """Regla determinística recuperada de la V3 (celda W2 de BASE_MAESTRA_ELEMENTOS).

    Búsqueda case-insensitive de palabras clave en Ubicacion/Descripcion/
    Subcircuito/CircuitoDashboard. None/NaN se tratan como texto vacío.
    """
    if _is_blank(row["ElementoID"]):
        return ""

    haystack = " ".join(
        "" if _is_blank(row[field]) else str(row[field]) for field in TIPO_INVENTARIO_SEARCH_FIELDS
    ).upper()

    if any(keyword in haystack for keyword in TIPO_INVENTARIO_KEYWORDS):
        return "Flexible gráfico"
    if row["Medio"] == "Digital":
        return "Digital"
    return "Físico estático"


def derive_aplica_cantidad(tipo_inventario: str, elemento_id: Any) -> str:
    """AplicaCantidad depende únicamente del TipoInventario ya derivado."""
    if _is_blank(elemento_id):
        return ""
    return "SI" if tipo_inventario == "Flexible gráfico" else "NO"


def _reconcile_derived_column(
    source: pd.Series, computed: pd.Series, elemento_ids: pd.Series, column_name: str
) -> pd.Series:
    """Aplica la política: fuente vacía -> usar calculado; fuente igual -> aceptar;
    fuente distinta -> TransformError (nunca sobrescribir silenciosamente)."""
    result = computed.copy()
    for idx in source.index:
        source_value = source.loc[idx]
        computed_value = computed.loc[idx]
        if _is_blank(source_value):
            result.loc[idx] = computed_value
        elif source_value == computed_value:
            result.loc[idx] = source_value
        else:
            raise TransformError(
                f"{column_name} contradictorio para ElementoID={elemento_ids.loc[idx]!r}: "
                f"fuente={source_value!r} calculado={computed_value!r}"
            )
    return result


def _classify_and_convert_numeric(value: Any, column_name: str) -> tuple[str, float | None]:
    if _is_blank(value):
        return "empty", None
    if isinstance(value, bool):
        raise TransformError(f"{column_name}: valor booleano no válido para columna numérica: {value!r}")
    if isinstance(value, (int, float)):
        return "native", float(value)
    if isinstance(value, str):
        try:
            return "text", float(value.strip())
        except ValueError as exc:
            raise TransformError(
                f"{column_name}: valor no convertible a número: {value!r}"
            ) from exc
    raise TransformError(f"{column_name}: tipo de valor inesperado ({type(value)!r}) = {value!r}")


def normalize_numeric_column(
    series: pd.Series, column_name: str, warnings: list[str]
) -> tuple[pd.Series, dict[str, Any]]:
    """Normaliza una columna numérica potencialmente mixta (int/texto numérico)
    a un dtype nullable de pandas (Int64 si todos los valores son enteros,
    Float64 si aparece algún decimal). Nunca trunca decimales."""
    kinds: list[str] = []
    numeric_values: list[float | None] = []

    for value in series:
        kind, number = _classify_and_convert_numeric(value, column_name)
        kinds.append(kind)
        numeric_values.append(number)

    non_empty = [n for n in numeric_values if n is not None]
    all_integer = all(float(n).is_integer() for n in non_empty) if non_empty else True

    if all_integer:
        converted = pd.array(
            [None if n is None else int(n) for n in numeric_values],
            dtype="Int64",
        )
        dtype_final = "Int64"
    else:
        converted = pd.array(numeric_values, dtype="Float64")
        dtype_final = "Float64"
        warnings.append(
            f"{column_name}: se detectaron valores no enteros; normalizado a Float64 en vez de Int64"
        )

    result_series = pd.Series(converted, index=series.index, name=column_name)

    stats = {
        "total": len(series),
        "empty": kinds.count("empty"),
        "originally_native": kinds.count("native"),
        "originally_text_numeric": kinds.count("text"),
        "converted": len(non_empty),
        "min": (int(min(non_empty)) if all_integer else min(non_empty)) if non_empty else None,
        "max": (int(max(non_empty)) if all_integer else max(non_empty)) if non_empty else None,
        "dtype_final": dtype_final,
    }
    return result_series, stats


def _assert_maestro_non_target_columns_unchanged(raw: pd.DataFrame, transformed: pd.DataFrame) -> None:
    target_columns = {"TipoInventario", "AplicaCantidad", "CapacidadSlotsReel", "SegundosDia"}
    other_columns = [c for c in raw.columns if c not in target_columns]
    if not raw[other_columns].equals(transformed[other_columns]):
        raise TransformError(
            "MAESTRO_ELEMENTOS: se detectó una modificación en columnas que no debían tocarse"
        )


def _assert_dataframe_equivalent(raw: pd.DataFrame, transformed: pd.DataFrame, label: str) -> None:
    if not raw.equals(transformed):
        raise TransformError(f"{label}: se detectó una modificación de datos (debe ser passthrough)")


def transform_data(path: str | Path = DEFAULT_INPUT_PATH) -> dict[str, Any]:
    """Punto de entrada reutilizable del pipeline OCU26 - Gate 2.

    1. Ejecuta validate_input(path) (Gate 1). Aborta con TransformError si INVALID.
    2. Lee MAESTRO_ELEMENTOS / CAMPANAS / PARAMETROS por rango real de tabla.
    3. Reconstruye TipoInventario / AplicaCantidad de forma determinística.
    4. Normaliza CapacidadSlotsReel / SegundosDia a tipo numérico nullable.
    5. Verifica invariantes de filas/orden/dominio y que ninguna otra columna
       haya cambiado.
    6. Verifica que el SHA-256 del archivo fuente sea idéntico antes y después.

    Devuelve dict con: validation, maestro, campanas, parametros, stats, warnings.
    """
    path = Path(path)

    validation = vi.validate_input(path)
    if validation["result"] == "INVALID":
        raise TransformError(
            f"El input no pasó el gate de validación (validate_input): {path}. "
            f"Errores: {validation['errors']}"
        )

    sha_before = validation["sha256"] or vi.calculate_sha256(path)

    maestro_raw = read_excel_table(path, "MAESTRO_ELEMENTOS", EXPECTED_TABLES["MAESTRO_ELEMENTOS"])
    campanas_raw = read_excel_table(path, "CAMPANAS", EXPECTED_TABLES["CAMPANAS"])
    parametros_raw = read_excel_table(path, "PARAMETROS", EXPECTED_TABLES["PARAMETROS"])

    warnings: list[str] = list(validation["warnings"])

    maestro = maestro_raw.copy(deep=True)
    campanas = campanas_raw.copy(deep=True)
    parametros = parametros_raw.copy(deep=True)

    if len(maestro) != len(maestro_raw):
        raise TransformError("MAESTRO_ELEMENTOS: la cantidad de filas cambió durante la transformación")
    if len(campanas) != len(campanas_raw):
        raise TransformError("CAMPANAS: la cantidad de filas cambió durante la transformación")
    if len(parametros) != len(parametros_raw):
        raise TransformError("PARAMETROS: la cantidad de filas cambió durante la transformación")

    # --- TipoInventario / AplicaCantidad ---
    computed_tipo = maestro.apply(derive_tipo_inventario, axis=1)
    final_tipo = _reconcile_derived_column(
        maestro["TipoInventario"], computed_tipo, maestro["ElementoID"], "TipoInventario"
    )
    maestro["TipoInventario"] = final_tipo

    computed_aplica = pd.Series(
        [derive_aplica_cantidad(t, eid) for t, eid in zip(final_tipo, maestro["ElementoID"])],
        index=maestro.index,
    )
    final_aplica = _reconcile_derived_column(
        maestro["AplicaCantidad"], computed_aplica, maestro["ElementoID"], "AplicaCantidad"
    )
    maestro["AplicaCantidad"] = final_aplica

    # --- Normalización numérica ---
    numeric_stats: dict[str, Any] = {}
    for column in ("CapacidadSlotsReel", "SegundosDia"):
        normalized, stats = normalize_numeric_column(maestro[column], column, warnings)
        maestro[column] = normalized
        numeric_stats[column] = stats

    # --- Invariantes de orden / identidad ---
    if list(maestro["ElementoID"]) != list(maestro_raw["ElementoID"]):
        raise TransformError("MAESTRO_ELEMENTOS: el orden/contenido de ElementoID cambió")
    if list(campanas["CargaID"]) != list(campanas_raw["CargaID"]):
        raise TransformError("CAMPANAS: el orden/contenido de CargaID cambió")
    if list(campanas["FilaOrigen"]) != list(campanas_raw["FilaOrigen"]):
        raise TransformError("CAMPANAS: el orden/contenido de FilaOrigen cambió")
    if list(campanas["ElementoID"]) != list(campanas_raw["ElementoID"]):
        raise TransformError("CAMPANAS: el orden/contenido de ElementoID cambió")

    # --- Invariantes de dominio ---
    non_blank_mask = maestro["ElementoID"].apply(lambda v: not _is_blank(v))

    if maestro.loc[non_blank_mask, "TipoInventario"].apply(_is_blank).any():
        raise TransformError("MAESTRO_ELEMENTOS: TipoInventario vacío en fila(s) con ElementoID informado")
    if maestro.loc[non_blank_mask, "AplicaCantidad"].apply(_is_blank).any():
        raise TransformError("MAESTRO_ELEMENTOS: AplicaCantidad vacío en fila(s) con ElementoID informado")

    invalid_tipo = set(maestro.loc[non_blank_mask, "TipoInventario"]) - TIPO_INVENTARIO_VALUES
    if invalid_tipo:
        raise TransformError(f"MAESTRO_ELEMENTOS: TipoInventario con valor(es) fuera de dominio: {invalid_tipo}")

    invalid_aplica = set(maestro.loc[non_blank_mask, "AplicaCantidad"]) - APLICA_CANTIDAD_VALUES
    if invalid_aplica:
        raise TransformError(f"MAESTRO_ELEMENTOS: AplicaCantidad con valor(es) fuera de dominio: {invalid_aplica}")

    # --- Ninguna otra columna debe cambiar ---
    _assert_maestro_non_target_columns_unchanged(maestro_raw, maestro)
    _assert_dataframe_equivalent(campanas_raw, campanas, "CAMPANAS")
    _assert_dataframe_equivalent(parametros_raw, parametros, "PARAMETROS")

    # --- Controles de calidad (informativos; no tocan RevisionMaestro) ---
    ids_non_blank = maestro.loc[non_blank_mask, "ElementoID"]
    duplicate_element_ids = int(ids_non_blank.duplicated().sum())

    digital_mask = maestro["TipoInventario"] == "Digital"
    digital_capacity_zero = int(((maestro["CapacidadSlotsReel"] == 0) & digital_mask).sum())

    flexible_mask = maestro["TipoInventario"] == "Flexible gráfico"
    flexible_with_nonzero_slots = int(((maestro["CapacidadSlotsReel"] != 0) & flexible_mask).sum())

    # --- SHA-256 antes/después: el Excel fuente no debe haber sido tocado ---
    sha_after = vi.calculate_sha256(path)
    if sha_after != sha_before:
        raise TransformError(
            "ERROR CRÍTICO: el SHA-256 del input cambió durante la transformación "
            f"(antes={sha_before}, después={sha_after}). El archivo pudo haber sido modificado."
        )

    stats = {
        "source_file": str(path),
        "source_sha256": sha_after,
        "validation_result": validation["result"],
        "rows": {
            "maestro": len(maestro),
            "campanas": len(campanas),
            "parametros": len(parametros),
        },
        "tipo_inventario_counts": {k: int(v) for k, v in maestro["TipoInventario"].value_counts().items()},
        "aplica_cantidad_counts": {k: int(v) for k, v in maestro["AplicaCantidad"].value_counts().items()},
        "numeric_normalization": numeric_stats,
        "quality_controls": {
            "duplicate_element_ids": duplicate_element_ids,
            "digital_capacity_zero": digital_capacity_zero,
            "flexible_with_nonzero_slots": flexible_with_nonzero_slots,
        },
    }

    return {
        "validation": validation,
        "maestro": maestro,
        "campanas": campanas,
        "parametros": parametros,
        "stats": stats,
        "warnings": warnings,
    }


def _format_number(n: int) -> str:
    return f"{n:,}".replace(",", ".")


def print_human_report(result: dict[str, Any]) -> None:
    line = "=" * 50
    print(line)
    print("OCU26 DATA TRANSFORMATION")
    print(line)
    print()

    print("Input validation:")
    print(f"  {result['validation']['result']}")
    print()

    print("Source:")
    print(f"  {result['stats']['source_file']}")
    print()

    rows = result["stats"]["rows"]
    print("Rows preserved:")
    print(f"  MAESTRO_ELEMENTOS: {_format_number(rows['maestro'])} -> {_format_number(rows['maestro'])}")
    print(f"  CAMPANAS: {_format_number(rows['campanas'])} -> {_format_number(rows['campanas'])}")
    print(f"  PARAMETROS: {_format_number(rows['parametros'])} -> {_format_number(rows['parametros'])}")
    print()

    print("Derived TipoInventario:")
    tipo_counts = result["stats"]["tipo_inventario_counts"]
    for key in ("Digital", "Físico estático", "Flexible gráfico"):
        print(f"  {key}: {_format_number(tipo_counts.get(key, 0))}")
    print()

    print("Derived AplicaCantidad:")
    aplica_counts = result["stats"]["aplica_cantidad_counts"]
    for key in ("SI", "NO"):
        print(f"  {key}: {_format_number(aplica_counts.get(key, 0))}")
    print()

    print("Numeric normalization:")
    for column in ("CapacidadSlotsReel", "SegundosDia"):
        s = result["stats"]["numeric_normalization"][column]
        print(f"  {column}:")
        print(f"    original numeric: {_format_number(s['originally_native'])}")
        print(f"    original numeric text: {_format_number(s['originally_text_numeric'])}")
        print(f"    empty: {_format_number(s['empty'])}")
        print(f"    min: {s['min']}")
        print(f"    max: {s['max']}")
        print(f"    final dtype: {s['dtype_final']}")
    print()

    qc = result["stats"]["quality_controls"]
    print("Quality controls:")
    print(f"  Duplicate ElementoID: {_format_number(qc['duplicate_element_ids'])}")
    print(f"  Digital capacity = 0: {_format_number(qc['digital_capacity_zero'])}")
    print(f"  Flexible with slots != 0: {_format_number(qc['flexible_with_nonzero_slots'])}")
    print()

    print("Warnings:")
    if result["warnings"]:
        for warning in result["warnings"]:
            print(f"  - {warning}")
    else:
        print("  none")
    print()

    print("RESULT:")
    print("  TRANSFORM_OK")
    print(line)


def _print_json_summary(result: dict[str, Any]) -> None:
    summary = {
        "result": "TRANSFORM_OK",
        "validation_result": result["validation"]["result"],
        "source": result["stats"]["source_file"],
        "source_sha256": result["stats"]["source_sha256"],
        "warnings": result["warnings"],
        "stats": result["stats"],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Transforma input/OCU26_BASE_DATOS.xlsx en memoria tras pasar el gate de validación."
    )
    parser.add_argument("--file", default=str(DEFAULT_INPUT_PATH), help="Ruta al archivo .xlsx a transformar")
    parser.add_argument("--json", action="store_true", help="Imprime un resumen JSON en stdout")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    try:
        result = transform_data(args.file)
    except TransformError as exc:
        if args.json:
            print(json.dumps({"result": "TRANSFORM_ERROR", "error": str(exc)}, ensure_ascii=False, indent=2))
        else:
            line = "=" * 50
            print(line)
            print("OCU26 DATA TRANSFORMATION")
            print(line)
            print()
            print("RESULT:")
            print("  TRANSFORM_ERROR")
            print(f"  {exc}")
            print(line)
        return 1

    if args.json:
        _print_json_summary(result)
    else:
        print_human_report(result)

    return 0


if __name__ == "__main__":
    sys.exit(main())
