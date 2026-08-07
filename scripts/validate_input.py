"""Validador de entrada para el pipeline OCU26.

Puerta de entrada obligatoria antes de transform_data.py / Power BI / JSON / HTML.

Este script es estrictamente READ-ONLY sobre el archivo Excel validado: nunca
escribe, corrige ni recalcula nada dentro del workbook. Solo lee, valida y
reporta.

Uso:
    python scripts/validate_input.py
    python scripts/validate_input.py --file ruta/al/archivo.xlsx
    python scripts/validate_input.py --json
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
import sys
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl
from dateutil import parser as dateutil_parser
from openpyxl.utils import range_boundaries
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_INPUT_PATH = Path("input/OCU26_BASE_DATOS.xlsx")

EXPECTED_SHEETS = ["MAESTRO_ELEMENTOS", "CAMPANAS", "PARAMETROS"]

EXPECTED_TABLES = {
    "MAESTRO_ELEMENTOS": "tblElementos",
    "CAMPANAS": "tblCampanas",
    "PARAMETROS": "tblParametros",
}

MAESTRO_HEADERS = [
    "TipoCatalogo",
    "Ciudad",
    "Medio",
    "CircuitoDashboard",
    "Subcircuito",
    "Ubicacion",
    "ElementoID",
    "Nivel",
    "Descripcion",
    "Resolucion",
    "DimensionOptico",
    "DimensionTotal",
    "Observaciones",
    "Material",
    "TipoInstalacion",
    "Original",
    "b",
    "h",
    "q",
    "m2",
    "CapacidadSlotsReel",
    "SegundosDia",
    "TipoInventario",
    "AplicaCantidad",
    "RevisionMaestro",
    "Proveedor",
]

CAMPANAS_HEADERS = [
    "CargaID",
    "ClaveNegocio",
    "FechaHoraCarga",
    "UsuarioCarga",
    "FuenteCarga",
    "EstadoValidacion",
    "ObservacionValidacion",
    "IDCampaña",
    "Campaña",
    "Cliente",
    "Marca",
    "Agencia",
    "Proveedor",
    "ElementoID",
    "TipoCargaDeclarado",
    "FechaInicio",
    "FechaFin",
    "FechaIndefinida",
    "Estado",
    "CantidadUnidades",
    "DuracionSpotSeg",
    "SalidasVendidas",
    "ModalidadPauta",
    "PROGRAMATICA",
    "TipoExclusividad",
    "HoraInicio",
    "HoraFin",
    "CANJE",
    "Observaciones",
    "FilaOrigen",
]

PARAMETROS_HEADERS = ["Categoria", "Valor"]

TIPO_INVENTARIO_VALUES = {"Digital", "Físico estático", "Flexible gráfico"}
APLICA_CANTIDAD_VALUES = {"SI", "NO"}
MEDIO_VALUES = {"Digital", "Estático"}
SI_NO_VALUES = {"Si", "No"}
ESTADO_VALIDACION_VALUES = {
    "OK",
    "PENDIENTE_HISTORICO",
    "PENDIENTE_DUPLICADO",
    "PENDIENTE_COLISION",
}

EXCEL_ERROR_STRINGS = {
    "#REF!",
    "#VALUE!",
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#GETTING_DATA",
    "#SPILL!",
    "#CALC!",
}

# CAMPANAS: campos que pueden estar incompletos y solo generan advertencia,
# sin ningún otro control de contenido específico.
CAMPANAS_GENERIC_BLANK_FIELDS = [
    "IDCampaña",
    "Campaña",
    "Cliente",
    "Marca",
    "Agencia",
    "Proveedor",
    "Estado",
    "CantidadUnidades",
    "TipoExclusividad",
]

# MAESTRO_ELEMENTOS: campos sin vocabulario, cuyos vacíos son solo advertencia.
MAESTRO_GENERIC_BLANK_FIELDS = ["RevisionMaestro", "Proveedor"]


class ValidationContext:
    """Acumula errores, advertencias y estadísticas de una corrida de validación."""

    def __init__(self) -> None:
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.stats: dict[str, Any] = {}

    def error(self, message: str) -> None:
        self.errors.append(message)

    def warning(self, message: str) -> None:
        self.warnings.append(message)


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def calculate_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def validate_zip(path: Path, ctx: ValidationContext) -> bool:
    """Verifica la integridad física del .xlsx como contenedor ZIP.

    Devuelve False si el archivo no es utilizable y la validación debe
    detenerse antes de intentar abrirlo con openpyxl.
    """
    ok = True

    if path.suffix.lower() != ".xlsx":
        ctx.error(f"Extensión de archivo inválida (se esperaba .xlsx): '{path.suffix}'")
        ok = False

    try:
        with zipfile.ZipFile(path) as zf:
            bad_entry = zf.testzip()
            if bad_entry is not None:
                ctx.error(f"El contenedor ZIP interno está corrupto en: {bad_entry}")
                return False
            names = zf.namelist()
            content_types = ""
            if "[Content_Types].xml" in names:
                content_types = zf.read("[Content_Types].xml").decode("utf-8", errors="replace")
    except zipfile.BadZipFile as exc:
        ctx.error(f"El archivo no es un ZIP/XLSX válido: {exc}")
        return False
    except OSError as exc:
        ctx.error(f"No se pudo abrir el archivo: {exc}")
        return False

    if any(re.search(r"vbaProject\.bin$", name, re.IGNORECASE) for name in names):
        ctx.error("El workbook contiene macros (se encontró vbaProject.bin)")
        ok = False

    if any(name.startswith("xl/externalLinks/") for name in names):
        ctx.error("El workbook contiene vínculos externos (xl/externalLinks)")
        ok = False

    if "macroEnabled" in content_types:
        ctx.error("El workbook está marcado como macro-habilitado en [Content_Types].xml")
        ok = False

    return ok


def load_workbook_safe(path: Path, ctx: ValidationContext) -> Workbook | None:
    try:
        return openpyxl.load_workbook(path, data_only=False, read_only=False, keep_vba=False)
    except Exception as exc:  # noqa: BLE001 - cualquier fallo de apertura es bloqueante
        ctx.error(f"openpyxl no pudo abrir el workbook: {exc}")
        return None


def validate_workbook_structure(wb: Workbook, ctx: ValidationContext) -> bool:
    """Valida hojas esperadas, orden exacto y visibilidad."""
    ok = True
    actual = wb.sheetnames

    if actual != EXPECTED_SHEETS:
        ctx.error(
            "Las hojas del workbook no coinciden con lo esperado (nombre/orden/cantidad). "
            f"Esperado: {EXPECTED_SHEETS}. Encontrado: {actual}"
        )
        ok = False

    for name in EXPECTED_SHEETS:
        if name in wb.sheetnames:
            ws = wb[name]
            if ws.sheet_state != "visible":
                ctx.error(f"La hoja '{name}' no está visible (estado='{ws.sheet_state}')")
                ok = False

    return ok


def validate_tables(wb: Workbook, ctx: ValidationContext) -> dict[str, Table | None]:
    """Valida que cada hoja tenga exactamente la tabla esperada, que empiece en
    A1, que no existan datos fuera de la tabla y que no haya filas
    completamente vacías dentro de ella."""
    result: dict[str, Table | None] = {}

    for sheet_name, table_name in EXPECTED_TABLES.items():
        if sheet_name not in wb.sheetnames:
            result[sheet_name] = None
            continue

        ws: Worksheet = wb[sheet_name]
        table_names = list(ws.tables.keys())

        if table_names != [table_name]:
            ctx.error(
                f"La hoja '{sheet_name}' no contiene exactamente la tabla '{table_name}' "
                f"(encontrado: {table_names})"
            )
            result[sheet_name] = None
            continue

        table = ws.tables[table_name]
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)

        if (min_row, min_col) != (1, 1):
            ctx.error(f"La tabla '{table_name}' no comienza en A1 (ref={table.ref})")

        extra_cells: list[str] = []
        if ws.max_row > max_row:
            for row in ws.iter_rows(min_row=max_row + 1, max_row=ws.max_row):
                for cell in row:
                    if cell.value is not None:
                        extra_cells.append(cell.coordinate)
        if ws.max_column > max_col:
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=max_col + 1, max_col=ws.max_column):
                for cell in row:
                    if cell.value is not None:
                        extra_cells.append(cell.coordinate)
        if extra_cells:
            ctx.error(
                f"Hay datos fuera de la tabla '{table_name}' en la hoja '{sheet_name}': "
                f"{len(extra_cells)} celda(s), ej. {extra_cells[:5]}"
            )

        empty_rows: list[int] = []
        if max_row > min_row:
            for r_idx, row_values in enumerate(
                ws.iter_rows(
                    min_row=min_row + 1,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True,
                ),
                start=min_row + 1,
            ):
                if all(v is None for v in row_values):
                    empty_rows.append(r_idx)
        if empty_rows:
            ctx.error(
                f"La tabla '{table_name}' contiene {len(empty_rows)} fila(s) completamente "
                f"vacía(s) dentro de su rango: {empty_rows[:10]}"
            )

        ctx.stats.setdefault("tables", {})[sheet_name] = {
            "table_name": table_name,
            "ref": table.ref,
            "data_rows": max_row - min_row,
        }
        result[sheet_name] = table

    return result


def validate_headers(
    ws: Worksheet,
    table: Table,
    expected_headers: list[str],
    ctx: ValidationContext,
    sheet_name: str,
) -> bool:
    min_col, min_row, max_col, _ = range_boundaries(table.ref)
    actual = list(
        next(
            ws.iter_rows(
                min_row=min_row,
                max_row=min_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        )
    )

    if len(expected_headers) != (max_col - min_col + 1):
        ctx.error(
            f"La tabla de '{sheet_name}' tiene {max_col - min_col + 1} columnas, "
            f"se esperaban {len(expected_headers)}"
        )
        return False

    if actual != expected_headers:
        diffs = [
            f"col {i + 1}: esperado='{e}' encontrado='{a}'"
            for i, (e, a) in enumerate(zip(expected_headers, actual))
            if e != a
        ]
        ctx.error(f"Encabezados de '{sheet_name}' no coinciden con lo esperado: {diffs}")
        return False

    return True


def extract_table_rows(ws: Worksheet, table: Table, headers: list[str]) -> list[dict[str, Any]]:
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    rows: list[dict[str, Any]] = []
    for r_idx, values in enumerate(
        ws.iter_rows(
            min_row=min_row + 1,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ),
        start=min_row + 1,
    ):
        record = dict(zip(headers, values))
        record["_row"] = r_idx
        rows.append(record)
    return rows


def validate_formulas(wb: Workbook, ctx: ValidationContext) -> None:
    """Confirma que la base sea 100% de valores: 0 fórmulas y 0 errores de Excel."""
    formula_cells: list[str] = []
    error_cells: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.data_type == "f":
                    formula_cells.append(f"{sheet_name}!{cell.coordinate}")
                elif cell.data_type == "e":
                    error_cells.append(f"{sheet_name}!{cell.coordinate}={cell.value}")
                elif isinstance(cell.value, str) and cell.value.strip() in EXCEL_ERROR_STRINGS:
                    error_cells.append(f"{sheet_name}!{cell.coordinate}={cell.value}")

    ctx.stats["formula_cells_found"] = len(formula_cells)
    ctx.stats["error_cells_found"] = len(error_cells)

    if formula_cells:
        ctx.error(f"Se encontraron {len(formula_cells)} celda(s) con fórmulas: {formula_cells[:10]}")
    if error_cells:
        ctx.error(f"Se encontraron {len(error_cells)} celda(s) con errores de Excel: {error_cells[:10]}")


def validate_element_ids(maestro_rows: list[dict[str, Any]], ctx: ValidationContext) -> set[Any]:
    """Valida MAESTRO_ELEMENTOS.ElementoID: vacíos, duplicados exactos y
    colisiones al normalizar con strip()+casefold(). No modifica los IDs."""
    total = len(maestro_rows)
    raw_ids = [row["ElementoID"] for row in maestro_rows]

    empty_count = sum(1 for v in raw_ids if is_blank(v))

    non_blank = [v for v in raw_ids if not is_blank(v)]
    counts = Counter(non_blank)
    duplicates_exact = {v: c for v, c in counts.items() if c > 1}

    normalized_groups: dict[str, set[Any]] = defaultdict(set)
    for v in non_blank:
        key = str(v).strip().casefold()
        normalized_groups[key].add(v)
    normalized_conflicts = {k: vs for k, vs in normalized_groups.items() if len(vs) > 1}

    unique_ids = set(non_blank)

    ctx.stats["maestro_elementos"] = {
        "total_rows": total,
        "elemento_id_unique": len(unique_ids),
        "elemento_id_empty": empty_count,
        "elemento_id_duplicates_exact": len(duplicates_exact),
        "elemento_id_duplicates_normalized": len(normalized_conflicts),
    }

    if empty_count:
        ctx.error(f"MAESTRO_ELEMENTOS: {empty_count} ElementoID vacío(s)")
    if duplicates_exact:
        examples = list(duplicates_exact.items())[:10]
        ctx.error(
            f"MAESTRO_ELEMENTOS: {len(duplicates_exact)} ElementoID duplicado(s) exacto(s): {examples}"
        )
    if normalized_conflicts:
        examples = list(normalized_conflicts.items())[:10]
        ctx.error(
            f"MAESTRO_ELEMENTOS: {len(normalized_conflicts)} grupo(s) de ElementoID que colisionan "
            f"solo al normalizar (espacios/mayúsculas): {examples}"
        )

    return unique_ids


def validate_referential_integrity(
    campanas_rows: list[dict[str, Any]], maestro_ids: set[Any], ctx: ValidationContext
) -> None:
    total = len(campanas_rows)
    empty = 0
    orphans: list[tuple[int, Any]] = []
    distinct_used: set[Any] = set()

    for row in campanas_rows:
        v = row["ElementoID"]
        if is_blank(v):
            empty += 1
            continue
        distinct_used.add(v)
        if v not in maestro_ids:
            orphans.append((row["_row"], v))

    stats = ctx.stats.setdefault("campanas", {})
    stats.update(
        {
            "total_rows": total,
            "elemento_id_distinct": len(distinct_used),
            "elemento_id_empty": empty,
            "elemento_id_orphans": len(orphans),
        }
    )

    if empty:
        ctx.error(f"CAMPANAS: {empty} ElementoID vacío(s)")
    if orphans:
        ctx.error(
            f"CAMPANAS: {len(orphans)} ElementoID huérfano(s) (no existen en MAESTRO_ELEMENTOS): "
            f"{orphans[:10]}"
        )


def validate_carga_id(campanas_rows: list[dict[str, Any]], ctx: ValidationContext) -> None:
    empty = 0
    counts: Counter = Counter()
    for row in campanas_rows:
        v = row["CargaID"]
        if is_blank(v):
            empty += 1
        else:
            counts[v] += 1
    duplicates = {v: c for v, c in counts.items() if c > 1}

    stats = ctx.stats.setdefault("campanas", {})
    stats["carga_id_empty"] = empty
    stats["carga_id_duplicates"] = len(duplicates)

    if empty:
        ctx.error(f"CAMPANAS: {empty} CargaID vacío(s)")
    if duplicates:
        ctx.error(f"CAMPANAS: {len(duplicates)} CargaID duplicado(s): {list(duplicates.items())[:10]}")


def validate_fila_origen(campanas_rows: list[dict[str, Any]], ctx: ValidationContext) -> None:
    empty = 0
    counts: Counter = Counter()
    invalid: list[tuple[int, Any]] = []

    for row in campanas_rows:
        v = row["FilaOrigen"]
        if is_blank(v):
            empty += 1
            continue
        if isinstance(v, bool):
            invalid.append((row["_row"], v))
            continue
        if isinstance(v, int) and v > 0:
            counts[v] += 1
            continue
        if isinstance(v, float) and v.is_integer() and v > 0:
            counts[int(v)] += 1
            continue
        if isinstance(v, str) and v.strip().isdigit() and int(v.strip()) > 0:
            counts[int(v.strip())] += 1
            continue
        invalid.append((row["_row"], v))

    duplicates = {v: c for v, c in counts.items() if c > 1}

    stats = ctx.stats.setdefault("campanas", {})
    stats["fila_origen_empty"] = empty
    stats["fila_origen_duplicates"] = len(duplicates)
    stats["fila_origen_invalid"] = len(invalid)

    if empty:
        ctx.warning(f"CAMPANAS: {empty} FilaOrigen vacío(s)")
    if invalid:
        ctx.error(
            f"CAMPANAS: {len(invalid)} FilaOrigen no interpretable(s) como entero positivo: {invalid[:10]}"
        )
    if duplicates:
        ctx.error(
            f"CAMPANAS: {len(duplicates)} FilaOrigen duplicado(s) entre valores no vacíos: "
            f"{list(duplicates.items())[:10]}"
        )


def validate_tipo_inventario(maestro_rows: list[dict[str, Any]], ctx: ValidationContext) -> None:
    for field, allowed in (
        ("TipoInventario", TIPO_INVENTARIO_VALUES),
        ("AplicaCantidad", APLICA_CANTIDAD_VALUES),
    ):
        empty = 0
        invalid: list[tuple[int, Any]] = []
        for row in maestro_rows:
            v = row[field]
            if is_blank(v):
                empty += 1
                continue
            if v not in allowed:
                invalid.append((row["_row"], v))
        if empty:
            ctx.warning(f"MAESTRO_ELEMENTOS: {empty} {field} vacío(s)")
        if invalid:
            ctx.error(f"MAESTRO_ELEMENTOS: {len(invalid)} {field} con valor no permitido: {invalid[:10]}")

    for field in MAESTRO_GENERIC_BLANK_FIELDS:
        empty = sum(1 for row in maestro_rows if is_blank(row[field]))
        if empty:
            ctx.warning(f"MAESTRO_ELEMENTOS: {empty} {field} vacío(s)")


def _classify_numeric(value: Any) -> tuple[str, float | None]:
    if is_blank(value):
        return "empty", None
    if isinstance(value, bool):
        return "invalid", None
    if isinstance(value, (int, float)):
        return "native", float(value)
    if isinstance(value, str):
        try:
            return "text", float(value.strip())
        except ValueError:
            return "invalid", None
    return "invalid", None


def validate_numeric_fields(maestro_rows: list[dict[str, Any]], ctx: ValidationContext) -> None:
    """CapacidadSlotsReel y SegundosDia pueden venir como número o como texto
    numérico (ya auditado, no implica pérdida de información)."""
    for field in ("CapacidadSlotsReel", "SegundosDia"):
        native = text = empty = invalid = 0
        values: list[float] = []
        invalid_examples: list[tuple[int, Any]] = []

        for row in maestro_rows:
            kind, num = _classify_numeric(row[field])
            if kind == "empty":
                empty += 1
            elif kind == "native":
                native += 1
                values.append(num)  # type: ignore[arg-type]
            elif kind == "text":
                text += 1
                values.append(num)  # type: ignore[arg-type]
            else:
                invalid += 1
                invalid_examples.append((row["_row"], row[field]))

        ctx.stats.setdefault("numeric_fields", {})[field] = {
            "native": native,
            "text_numeric": text,
            "empty": empty,
            "invalid": invalid,
            "min": min(values) if values else None,
            "max": max(values) if values else None,
        }

        if text:
            ctx.warning(f"MAESTRO_ELEMENTOS: {text} {field} almacenado(s) como texto numérico")
        if invalid:
            ctx.error(
                f"MAESTRO_ELEMENTOS: {invalid} {field} con valor no convertible a número: "
                f"{invalid_examples[:10]}"
            )


def _classify_date(value: Any) -> tuple[str, dt.date | None]:
    if is_blank(value):
        return "empty", None
    if isinstance(value, dt.datetime):
        return "ok", value
    if isinstance(value, dt.date):
        return "ok", value
    if isinstance(value, str):
        try:
            return "ok_text", dateutil_parser.parse(value.strip(), dayfirst=True)
        except (ValueError, OverflowError, TypeError):
            return "invalid", None
    return "invalid", None


def validate_dates(campanas_rows: list[dict[str, Any]], ctx: ValidationContext) -> None:
    field_stats: dict[str, Any] = {}

    for field in ("FechaHoraCarga", "FechaInicio", "FechaFin"):
        empty = 0
        invalid: list[tuple[int, Any]] = []
        for row in campanas_rows:
            kind, _ = _classify_date(row[field])
            if kind == "empty":
                empty += 1
            elif kind == "invalid":
                invalid.append((row["_row"], row[field]))

        field_stats[field] = {"empty": empty, "invalid": len(invalid)}

        if empty:
            ctx.warning(f"CAMPANAS: {empty} {field} vacío(s)")
        if invalid:
            ctx.error(
                f"CAMPANAS: {len(invalid)} {field} con valor no interpretable como fecha: {invalid[:10]}"
            )

    ctx.stats["fechas"] = field_stats

    invalid_indefinida: list[tuple[int, Any]] = []
    conflicts: list[int] = []
    for row in campanas_rows:
        v = row["FechaIndefinida"]
        if is_blank(v):
            continue
        if v not in SI_NO_VALUES:
            invalid_indefinida.append((row["_row"], v))
            continue
        if v == "Si" and not is_blank(row["FechaFin"]):
            conflicts.append(row["_row"])

    if invalid_indefinida:
        ctx.error(
            f"CAMPANAS: {len(invalid_indefinida)} FechaIndefinida con valor no permitido: "
            f"{invalid_indefinida[:10]}"
        )
    if conflicts:
        ctx.error(
            f"CAMPANAS: {len(conflicts)} fila(s) con FechaIndefinida='Si' pero FechaFin no vacía: "
            f"{conflicts[:10]}"
        )


def _classify_time(value: Any) -> tuple[str, Any]:
    if is_blank(value):
        return "empty", None
    if isinstance(value, dt.time):
        return "ok", value
    if isinstance(value, dt.datetime):
        return "ok", value.time()
    if isinstance(value, str):
        try:
            return "ok_text", dateutil_parser.parse(value.strip()).time()
        except (ValueError, OverflowError, TypeError):
            return "invalid", None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if 0 <= value < 1:
            return "ok_text", None
        return "invalid", None
    return "invalid", None


def validate_horas(campanas_rows: list[dict[str, Any]], ctx: ValidationContext) -> None:
    stats: dict[str, Any] = {}
    for field in ("HoraInicio", "HoraFin"):
        empty = 0
        invalid: list[tuple[int, Any]] = []
        for row in campanas_rows:
            kind, _ = _classify_time(row[field])
            if kind == "empty":
                empty += 1
            elif kind == "invalid":
                invalid.append((row["_row"], row[field]))

        stats[field] = {"empty": empty, "with_data": len(campanas_rows) - empty}

        if empty:
            ctx.warning(f"CAMPANAS: {empty} {field} vacío(s)")
        if invalid:
            ctx.error(
                f"CAMPANAS: {len(invalid)} {field} con valor no interpretable como hora: {invalid[:10]}"
            )
    ctx.stats["horas"] = stats


def validate_canje_programatica(campanas_rows: list[dict[str, Any]], ctx: ValidationContext) -> None:
    for field in ("CANJE", "PROGRAMATICA"):
        empty = 0
        invalid: list[tuple[int, Any]] = []
        for row in campanas_rows:
            v = row[field]
            if is_blank(v):
                empty += 1
                continue
            if v not in SI_NO_VALUES:
                invalid.append((row["_row"], v))
        if empty:
            ctx.warning(f"CAMPANAS: {empty} {field} vacío(s)")
        if invalid:
            ctx.error(f"CAMPANAS: {len(invalid)} {field} con valor no permitido: {invalid[:10]}")


def validate_estado_validacion(campanas_rows: list[dict[str, Any]], ctx: ValidationContext) -> None:
    counts: Counter = Counter()
    invalid: list[tuple[int, Any]] = []

    for row in campanas_rows:
        v = row["EstadoValidacion"]
        if v not in ESTADO_VALIDACION_VALUES:
            invalid.append((row["_row"], v))
        else:
            counts[v] += 1

    ctx.stats["estado_validacion"] = dict(counts)

    if invalid:
        ctx.error(f"CAMPANAS: {len(invalid)} EstadoValidacion con valor no permitido: {invalid[:10]}")

    pendientes = sum(c for k, c in counts.items() if k != "OK")
    if pendientes:
        ctx.warning(f"CAMPANAS: {pendientes} fila(s) con EstadoValidacion pendiente (distinto de OK)")


def validate_vocabularies(
    maestro_rows: list[dict[str, Any]], campanas_rows: list[dict[str, Any]], ctx: ValidationContext
) -> None:
    """Cruza MAESTRO_ELEMENTOS.Medio con CAMPANAS.TipoCargaDeclarado."""
    invalid_medio = [(row["_row"], row["Medio"]) for row in maestro_rows if row["Medio"] not in MEDIO_VALUES]
    if invalid_medio:
        ctx.error(f"MAESTRO_ELEMENTOS: {len(invalid_medio)} Medio con valor no permitido: {invalid_medio[:10]}")

    invalid_tipocarga = [
        (row["_row"], row["TipoCargaDeclarado"])
        for row in campanas_rows
        if row["TipoCargaDeclarado"] not in MEDIO_VALUES
    ]
    if invalid_tipocarga:
        ctx.error(
            f"CAMPANAS: {len(invalid_tipocarga)} TipoCargaDeclarado con valor no permitido: "
            f"{invalid_tipocarga[:10]}"
        )

    medio_by_elemento: dict[Any, Any] = {}
    for row in maestro_rows:
        eid = row["ElementoID"]
        if not is_blank(eid) and eid not in medio_by_elemento:
            medio_by_elemento[eid] = row["Medio"]

    mismatches: list[tuple[int, Any, Any, Any]] = []
    for row in campanas_rows:
        eid = row["ElementoID"]
        medio = medio_by_elemento.get(eid)
        if medio is None:
            continue  # ElementoID huérfano, ya reportado en validate_referential_integrity
        if row["TipoCargaDeclarado"] != medio:
            mismatches.append((row["_row"], eid, row["TipoCargaDeclarado"], medio))

    if mismatches:
        ctx.error(
            f"CAMPANAS: {len(mismatches)} fila(s) donde TipoCargaDeclarado no coincide con el Medio "
            f"del ElementoID en MAESTRO_ELEMENTOS: {mismatches[:10]}"
        )


def validate_completeness_warnings(campanas_rows: list[dict[str, Any]], ctx: ValidationContext) -> None:
    """Reporta como advertencia los campos incompletos no bloqueantes de CAMPANAS
    que no tienen ningún otro control de contenido (sección 19)."""
    for field in CAMPANAS_GENERIC_BLANK_FIELDS:
        count = sum(1 for row in campanas_rows if is_blank(row[field]))
        if count:
            ctx.warning(f"CAMPANAS: {count} {field} vacío(s)")


def _finalize(path: Path, file_info: dict[str, Any], ctx: ValidationContext) -> dict[str, Any]:
    if ctx.errors:
        result = "INVALID"
    elif ctx.warnings:
        result = "VALID_WITH_WARNINGS"
    else:
        result = "VALID"

    return {
        "result": result,
        "file": file_info["path"],
        "file_name": file_info["name"],
        "size_bytes": file_info["size_bytes"],
        "sha256": file_info["sha256"],
        "errors": ctx.errors,
        "warnings": ctx.warnings,
        "stats": ctx.stats,
    }


def validate_input(path: str | Path) -> dict[str, Any]:
    """Punto de entrada reutilizable: valida `path` y devuelve un dict con
    result / file / sha256 / errors / warnings / stats.

    100% read-only: nunca escribe, corrige ni recalcula el Excel.
    """
    path = Path(path)
    ctx = ValidationContext()
    file_info: dict[str, Any] = {"path": str(path), "name": path.name, "size_bytes": None, "sha256": None}

    if not path.exists():
        ctx.error(f"El archivo no existe: {path}")
        return _finalize(path, file_info, ctx)

    file_info["size_bytes"] = path.stat().st_size
    file_info["sha256"] = calculate_sha256(path)

    if not validate_zip(path, ctx):
        return _finalize(path, file_info, ctx)

    wb = load_workbook_safe(path, ctx)
    if wb is None:
        return _finalize(path, file_info, ctx)

    if not validate_workbook_structure(wb, ctx):
        return _finalize(path, file_info, ctx)

    tables = validate_tables(wb, ctx)
    validate_formulas(wb, ctx)

    maestro_rows: list[dict[str, Any]] = []
    campanas_rows: list[dict[str, Any]] = []

    if tables.get("MAESTRO_ELEMENTOS") is not None:
        if validate_headers(wb["MAESTRO_ELEMENTOS"], tables["MAESTRO_ELEMENTOS"], MAESTRO_HEADERS, ctx, "MAESTRO_ELEMENTOS"):
            maestro_rows = extract_table_rows(wb["MAESTRO_ELEMENTOS"], tables["MAESTRO_ELEMENTOS"], MAESTRO_HEADERS)

    if tables.get("CAMPANAS") is not None:
        if validate_headers(wb["CAMPANAS"], tables["CAMPANAS"], CAMPANAS_HEADERS, ctx, "CAMPANAS"):
            campanas_rows = extract_table_rows(wb["CAMPANAS"], tables["CAMPANAS"], CAMPANAS_HEADERS)

    if tables.get("PARAMETROS") is not None:
        validate_headers(wb["PARAMETROS"], tables["PARAMETROS"], PARAMETROS_HEADERS, ctx, "PARAMETROS")

    maestro_ids: set[Any] = set()
    if maestro_rows:
        maestro_ids = validate_element_ids(maestro_rows, ctx)
        validate_tipo_inventario(maestro_rows, ctx)
        validate_numeric_fields(maestro_rows, ctx)

    if campanas_rows:
        validate_referential_integrity(campanas_rows, maestro_ids, ctx)
        validate_carga_id(campanas_rows, ctx)
        validate_fila_origen(campanas_rows, ctx)
        validate_dates(campanas_rows, ctx)
        validate_horas(campanas_rows, ctx)
        validate_canje_programatica(campanas_rows, ctx)
        validate_estado_validacion(campanas_rows, ctx)
        validate_completeness_warnings(campanas_rows, ctx)

    if maestro_rows and campanas_rows:
        validate_vocabularies(maestro_rows, campanas_rows, ctx)

    return _finalize(path, file_info, ctx)


def _format_number(n: int) -> str:
    return f"{n:,}".replace(",", ".")


def print_human_report(result: dict[str, Any]) -> None:
    line = "=" * 50
    print(line)
    print("OCU26 INPUT VALIDATION")
    print(line)
    print()
    print("File:")
    print(f"  {result['file']}")
    if result.get("size_bytes") is not None:
        print(f"  {_format_number(result['size_bytes'])} bytes")
    print()
    print("SHA-256:")
    print(f"  {result.get('sha256')}")
    print()

    stats = result.get("stats", {})

    maestro = stats.get("maestro_elementos")
    if maestro:
        print("MAESTRO_ELEMENTOS:")
        print(f"  {_format_number(maestro['total_rows'])} rows")
        print(f"  {_format_number(maestro['elemento_id_unique'])} ElementoID unique")
        print()

    campanas = stats.get("campanas")
    if campanas:
        print("CAMPANAS:")
        print(f"  {_format_number(campanas['total_rows'])} rows")
        print(f"  {_format_number(campanas.get('elemento_id_orphans', 0))} orphan ElementoID")
        print()

    print("WARNINGS:")
    if result["warnings"]:
        for warning in result["warnings"]:
            print(f"  - {warning}")
    else:
        print("  none")
    print()

    print("ERRORS:")
    if result["errors"]:
        for error in result["errors"]:
            print(f"  - {error}")
    else:
        print("  none")
    print()

    print("RESULT:")
    print(f"  {result['result']}")
    print(line)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Valida input/OCU26_BASE_DATOS.xlsx antes de transformarlo.")
    parser.add_argument("--file", default=str(DEFAULT_INPUT_PATH), help="Ruta al archivo .xlsx a validar")
    parser.add_argument("--json", action="store_true", help="Imprime el resultado como JSON en stdout")
    args = parser.parse_args(argv)

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    result = validate_input(args.file)

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    else:
        print_human_report(result)

    return 1 if result["result"] == "INVALID" else 0


if __name__ == "__main__":
    sys.exit(main())
