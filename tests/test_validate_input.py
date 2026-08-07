"""Pruebas para scripts/validate_input.py.

Ninguna prueba modifica input/OCU26_BASE_DATOS.xlsx. Los workbooks corruptos
o inválidos se generan en tmp_path.
"""

from __future__ import annotations

import copy
import datetime as dt
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "scripts"))

import validate_input as vi  # noqa: E402

PRODUCTION_FILE = REPO_ROOT / "input" / "OCU26_BASE_DATOS.xlsx"


def _add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, name=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False
    )
    ws.add_table(table)


def _write_workbook(
    path: Path,
    maestro_rows: list[list],
    campanas_rows: list[list],
    parametros_rows: list[list] | None = None,
) -> None:
    """Construye un workbook mínimo con el esquema productivo esperado."""
    wb = openpyxl.Workbook()
    ws_maestro = wb.active
    ws_maestro.title = "MAESTRO_ELEMENTOS"
    ws_maestro.append(vi.MAESTRO_HEADERS)
    for row in maestro_rows:
        ws_maestro.append(row)
    last_row = len(maestro_rows) + 1
    _add_table(ws_maestro, "tblElementos", f"A1:{openpyxl.utils.get_column_letter(len(vi.MAESTRO_HEADERS))}{last_row}")

    ws_campanas = wb.create_sheet("CAMPANAS")
    ws_campanas.append(vi.CAMPANAS_HEADERS)
    for row in campanas_rows:
        ws_campanas.append(row)
    last_row_c = len(campanas_rows) + 1
    _add_table(
        ws_campanas, "tblCampanas", f"A1:{openpyxl.utils.get_column_letter(len(vi.CAMPANAS_HEADERS))}{last_row_c}"
    )

    ws_parametros = wb.create_sheet("PARAMETROS")
    ws_parametros.append(vi.PARAMETROS_HEADERS)
    for row in parametros_rows or [["Estado", "Activa"]]:
        ws_parametros.append(row)
    last_row_p = len(parametros_rows or [["Estado", "Activa"]]) + 1
    _add_table(ws_parametros, "tblParametros", f"A1:B{last_row_p}")

    wb.save(path)


def _base_maestro_row(elemento_id: str = "ELEM-0001") -> list:
    row = [None] * len(vi.MAESTRO_HEADERS)
    idx = {h: i for i, h in enumerate(vi.MAESTRO_HEADERS)}
    row[idx["TipoCatalogo"]] = "Via Publica"
    row[idx["Ciudad"]] = "Buenos Aires"
    row[idx["Medio"]] = "Digital"
    row[idx["CircuitoDashboard"]] = "Circuito 1"
    row[idx["Subcircuito"]] = "Sub 1"
    row[idx["Ubicacion"]] = "Av. Siempre Viva 123"
    row[idx["ElementoID"]] = elemento_id
    row[idx["Nivel"]] = 1
    row[idx["Descripcion"]] = "Pantalla LED"
    row[idx["CapacidadSlotsReel"]] = 20
    row[idx["SegundosDia"]] = 100800
    return row


def _base_campana_row(carga_id: str, elemento_id: str = "ELEM-0001") -> list:
    row = [None] * len(vi.CAMPANAS_HEADERS)
    idx = {h: i for i, h in enumerate(vi.CAMPANAS_HEADERS)}
    row[idx["CargaID"]] = carga_id
    row[idx["EstadoValidacion"]] = "OK"
    row[idx["ElementoID"]] = elemento_id
    row[idx["TipoCargaDeclarado"]] = "Digital"
    row[idx["FechaIndefinida"]] = "No"
    row[idx["FilaOrigen"]] = 1
    return row


def _minimal_valid_workbook(path: Path) -> None:
    _write_workbook(
        path,
        maestro_rows=[_base_maestro_row("ELEM-0001")],
        campanas_rows=[_base_campana_row("HIST-0001")],
    )


def test_real_production_file_is_valid_with_warnings():
    assert PRODUCTION_FILE.exists(), "input/OCU26_BASE_DATOS.xlsx no existe"
    result = vi.validate_input(PRODUCTION_FILE)
    assert result["errors"] == []
    assert result["result"] == "VALID_WITH_WARNINGS"

    proc = subprocess.run(
        [sys.executable, str(REPO_ROOT / "scripts" / "validate_input.py")],
        cwd=REPO_ROOT,
        capture_output=True,
    )
    assert proc.returncode == 0


def test_minimal_valid_workbook_has_no_errors(tmp_path):
    path = tmp_path / "valid.xlsx"
    _minimal_valid_workbook(path)
    result = vi.validate_input(path)
    assert result["errors"] == []


def test_duplicate_elemento_id_is_invalid(tmp_path):
    path = tmp_path / "dup_elemento_id.xlsx"
    _write_workbook(
        path,
        maestro_rows=[_base_maestro_row("ELEM-0001"), _base_maestro_row("ELEM-0001")],
        campanas_rows=[_base_campana_row("HIST-0001")],
    )
    result = vi.validate_input(path)
    assert result["result"] == "INVALID"
    assert any("duplicado" in e.lower() for e in result["errors"])


def test_workbook_with_formula_is_invalid(tmp_path):
    path = tmp_path / "with_formula.xlsx"
    _minimal_valid_workbook(path)

    wb = openpyxl.load_workbook(path)
    ws = wb["MAESTRO_ELEMENTOS"]
    ws["I2"] = "=1+1"  # Descripcion, dentro del rango de la tabla
    wb.save(path)

    result = vi.validate_input(path)
    assert result["result"] == "INVALID"
    assert any("fórmula" in e.lower() for e in result["errors"])


def test_orphan_elemento_id_in_campanas_is_invalid(tmp_path):
    path = tmp_path / "orphan.xlsx"
    _write_workbook(
        path,
        maestro_rows=[_base_maestro_row("ELEM-0001")],
        campanas_rows=[_base_campana_row("HIST-0001", elemento_id="NO-EXISTE")],
    )
    result = vi.validate_input(path)
    assert result["result"] == "INVALID"
    assert any("huérfano" in e for e in result["errors"])


def test_non_numeric_capacidad_slots_reel_is_invalid(tmp_path):
    path = tmp_path / "bad_numeric.xlsx"
    row = _base_maestro_row("ELEM-0001")
    idx = {h: i for i, h in enumerate(vi.MAESTRO_HEADERS)}
    row[idx["CapacidadSlotsReel"]] = "veinte"
    _write_workbook(
        path,
        maestro_rows=[row],
        campanas_rows=[_base_campana_row("HIST-0001")],
    )
    result = vi.validate_input(path)
    assert result["result"] == "INVALID"
    assert any("CapacidadSlotsReel" in e and "no convertible" in e for e in result["errors"])


def test_numeric_text_capacidad_slots_reel_is_warning_not_error(tmp_path):
    path = tmp_path / "text_numeric.xlsx"
    row = _base_maestro_row("ELEM-0001")
    idx = {h: i for i, h in enumerate(vi.MAESTRO_HEADERS)}
    row[idx["CapacidadSlotsReel"]] = "20"
    _write_workbook(
        path,
        maestro_rows=[row],
        campanas_rows=[_base_campana_row("HIST-0001")],
    )
    result = vi.validate_input(path)
    assert result["errors"] == []
    assert any("texto numérico" in w for w in result["warnings"])
    assert result["result"] == "VALID_WITH_WARNINGS"


def test_cli_exit_code_zero_for_valid(tmp_path):
    path = tmp_path / "valid.xlsx"
    _minimal_valid_workbook(path)
    proc = subprocess.run(
        [sys.executable, str(REPO_ROOT / "scripts" / "validate_input.py"), "--file", str(path)],
        capture_output=True,
    )
    assert proc.returncode == 0


def test_cli_exit_code_one_for_invalid(tmp_path):
    path = tmp_path / "invalid.xlsx"
    _write_workbook(
        path,
        maestro_rows=[_base_maestro_row("ELEM-0001"), _base_maestro_row("ELEM-0001")],
        campanas_rows=[_base_campana_row("HIST-0001")],
    )
    proc = subprocess.run(
        [sys.executable, str(REPO_ROOT / "scripts" / "validate_input.py"), "--file", str(path)],
        capture_output=True,
    )
    assert proc.returncode == 1


def test_cli_json_output_is_valid_json(tmp_path):
    import json

    path = tmp_path / "valid.xlsx"
    _minimal_valid_workbook(path)
    proc = subprocess.run(
        [sys.executable, str(REPO_ROOT / "scripts" / "validate_input.py"), "--file", str(path), "--json"],
        capture_output=True,
    )
    data = json.loads(proc.stdout.decode("utf-8"))
    assert data["result"] == "VALID_WITH_WARNINGS" or data["result"] == "VALID"
    assert data["errors"] == []
